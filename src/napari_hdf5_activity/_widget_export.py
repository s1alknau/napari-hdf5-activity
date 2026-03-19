"""_widget_export.py — ExportMixin for HDF5AnalysisWidget.

Handles all results export: Excel, CSV, HDF5, and MATLAB-compatible
formats.  Mixed into HDF5AnalysisWidget so all methods share ``self``.
"""

from __future__ import annotations

import csv
import json
import os
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd


class ExportMixin:
    """Mixin providing all results save/export functionality.

    Requires that the host class provides:
    - self.merged_results, self.movement_data, self.fraction_data, etc.
    - self.file_path (str)
    - self._log_message(msg: str)
    - UI widgets: self.frame_interval, self.bin_size_seconds, etc.
    """

    # --- paste methods here ---

    def export_results_for_matlab_compatibility(self):
        """Export analysis results for MATLAB compatibility."""
        # 1) Check preconditions
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.results_label.setText("No results to export.")
            return

        directory = QFileDialog.getExistingDirectory(
            self, "Select Directory for MATLAB Export"
        )
        if not directory:
            return

        try:
            # 2) Prepare analysis object
            analysis_results = {
                "merged_results": getattr(self, "merged_results", {}),
                "baseline_means": getattr(self, "roi_baseline_means", {}),
                "upper_thresholds": getattr(self, "roi_upper_thresholds", {}),
                "lower_thresholds": getattr(self, "roi_lower_thresholds", {}),
                "movement_data": getattr(self, "movement_data", {}),
                "fraction_data": getattr(self, "fraction_data", {}),
                "sleep_data": getattr(self, "sleep_data", {}),
                "parameters": {
                    "threshold_method": self._get_current_threshold_method_display(),
                    "frame_interval": self.frame_interval.value(),
                    "enable_matlab_norm": True,
                    "enable_detrending": self.enable_detrending.isChecked(),
                },
            }

            # Add method-specific parameters
            if hasattr(self, "baseline_duration_minutes"):
                analysis_results["parameters"][
                    "baseline_duration_minutes"
                ] = self.baseline_duration_minutes.value()
            if hasattr(self, "threshold_multiplier"):
                analysis_results["parameters"][
                    "threshold_multiplier"
                ] = self.threshold_multiplier.value()
            if hasattr(self, "calibration_multiplier"):
                analysis_results["parameters"][
                    "calibration_multiplier"
                ] = self.calibration_multiplier.value()

            # 3) Try modern export
            created_files = []
            try:
                from ._calc_integration import export_results_for_matlab as _export_fn

                created_files = _export_fn(analysis_results, directory)
            except Exception:
                _export_fn = None

            # 4) Fallback export if modern system not available
            if not created_files:
                created_files = self._create_basic_matlab_export(
                    analysis_results, directory
                )

            # 5) UI feedback
            if created_files:
                self.results_label.setText(
                    f"Exported {len(created_files)} files for MATLAB"
                )
                self._log_message(
                    f"MATLAB export completed: {len(created_files)} files"
                )
                for p in created_files:
                    self._log_message(f"  Created: {os.path.basename(p)}")
            else:
                self.results_label.setText("Export failed")
                self._log_message("MATLAB export failed")

        except Exception as e:
            err = f"Error exporting for MATLAB: {e}"
            self.results_label.setText(err)
            self._log_message(f"ERROR: {err}")

    def save_results_consolidated_complete(self):
        """
        UPDATED VERSION: Complete consolidated save function with all sheets.
        This replaces the previous save_results_consolidated method.
        """
        import os
        import time

        # Check if analysis results are available
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.results_label.setText(
                "❌ No analysis results to save. Run analysis first."
            )
            self._log_message("Save failed: No analysis results available")
            return

        # Check if we have behavioral analysis data
        has_behavioral_data = (
            hasattr(self, "movement_data")
            and self.movement_data
            and hasattr(self, "fraction_data")
            and self.fraction_data
        )

        if not has_behavioral_data:
            self.results_label.setText("⚠️ Incomplete analysis data detected.")
            self._log_message("Warning: Saving with incomplete behavioral analysis")

        # Get base filename from user
        from qtpy.QtWidgets import QFileDialog

        base_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Analysis Results",
            f"analysis_results_{int(time.time())}",  # No extension - we'll add them
            "All Files (*)",
        )

        if not base_path:
            self._log_message("Save cancelled by user")
            return

        # Remove any extension from base_path to ensure clean naming
        base_path = os.path.splitext(base_path)[0]

        saved_files = []
        sheets_created = []

        try:
            # === SAVE CSV VERSION ===
            csv_path = f"{base_path}.csv"
            self._log_message(f"Saving CSV version: {csv_path}")

            try:
                self._save_results_csv(csv_path)
                saved_files.append(("CSV", csv_path))
                self._log_message("✅ CSV saved successfully")
            except Exception as e:
                self._log_message(f"❌ CSV save failed: {e}")

            # === SAVE COMPLETE EXCEL VERSION (if possible) ===
            try:
                import pandas as pd
                import openpyxl

                excel_path = f"{base_path}.xlsx"
                self._log_message(
                    f"Saving complete Excel version with all sheets: {excel_path}"
                )

                # Use the complete Excel save method
                self._save_results_excel_to_path(excel_path)
                saved_files.append(("Excel", excel_path))

                # Count sheets created
                wb = openpyxl.load_workbook(excel_path)
                sheets_created = wb.sheetnames
                wb.close()

                self._log_message(
                    f"✅ Excel saved with {len(sheets_created)} sheets: {', '.join(sheets_created)}"
                )

            except ImportError:
                self._log_message(
                    "⚠️ Excel export not available (missing pandas/openpyxl)"
                )
                self._log_message("   Install with: pip install pandas openpyxl")
            except Exception as e:
                self._log_message(f"❌ Excel save failed: {e}")
                import traceback

                self._log_message(f"Traceback: {traceback.format_exc()}")

            # === SHOW THRESHOLD STATS IN LOG ===
            if hasattr(self, "roi_baseline_means") and self.roi_baseline_means:
                self._log_message("\n" + "=" * 50)
                self._log_message("THRESHOLD STATISTICS (included with save)")
                self._log_message("=" * 50)
                self._show_threshold_stats_in_log()

            # === UPDATE UI WITH RESULTS ===
            if saved_files:
                if sheets_created:
                    file_list = f"CSV + Excel ({len(sheets_created)} sheets: {', '.join(sheets_created)})"
                else:
                    file_list = ", ".join(
                        [
                            f"{fmt} ({os.path.basename(path)})"
                            for fmt, path in saved_files
                        ]
                    )

                self.results_label.setText(f"✅ Saved: {file_list}")
                self._log_message(
                    f"\n🎉 SAVE COMPLETE: {len(saved_files)} files created"
                )

                # Show success dialog with file details
                self._show_save_success_dialog_complete(saved_files, sheets_created)
            else:
                self.results_label.setText(
                    "❌ All save attempts failed - check log for details"
                )
                self._log_message("❌ No files were saved successfully")

        except Exception as e:
            error_msg = f"Save operation failed: {e}"
            self.results_label.setText(error_msg)
            self._log_message(f"❌ {error_msg}")
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def _show_save_success_dialog_complete(self, saved_files, sheets_created):
        """Show success dialog with complete file and sheet details."""
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)
        msg.setWindowTitle("Save Complete")
        msg.setText(
            f"Successfully saved analysis results in {len(saved_files)} format(s):"
        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                file_size = "Unknown size"

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        if sheets_created:
            file_details.append("")
            file_details.append("Excel Sheets Created:")
            for sheet in sheets_created:
                file_details.append(f"  - {sheet}")

        msg.setDetailedText("\n".join(file_details))
        msg.setInformativeText("Analysis results saved successfully")
        msg.exec_()

    def add_nematostella_analysis_to_widget(widget_instance):
        """
        Add Nematostella-specific analysis capabilities to the existing widget.
        This function can be called from the widget to enable enhanced analysis.

        Args:
            widget_instance: Instance of HDF5AnalysisWidget
        """
        # Add new button to widget if it doesn't exist
        if not hasattr(widget_instance, "btn_nematostella_analysis"):
            from qtpy.QtWidgets import QPushButton

            widget_instance.btn_nematostella_analysis = QPushButton(
                "Nematostella Timeseries Analysis"
            )
            widget_instance.btn_nematostella_analysis.setToolTip(
                "Run specialized Nematostella timeseries analysis"
            )
            widget_instance.btn_nematostella_analysis.setStyleSheet(
                "QPushButton { background-color: #9C27B0; color: white; font-weight: bold; }"
            )

            # Add to the existing results tab layout
            if hasattr(widget_instance, "tab_results"):
                layout = widget_instance.tab_results.layout()
                if layout:
                    # Find the plot buttons group and add after it
                    for i in range(layout.count()):
                        item = layout.itemAt(i)
                        if (
                            item
                            and hasattr(item.widget(), "title")
                            and "Controls" in item.widget().title()
                        ):
                            # Insert after plot controls group
                            layout.insertWidget(
                                i + 1, widget_instance.btn_nematostella_analysis
                            )
                            break
                    else:
                        # Fallback: add at the end
                        layout.addWidget(widget_instance.btn_nematostella_analysis)

            # Connect the button
            widget_instance.btn_nematostella_analysis.clicked.connect(
                lambda: run_nematostella_analysis_from_widget(widget_instance)
            )

    def run_nematostella_analysis_from_widget(widget_instance):
        """
        Run Nematostella analysis from within the napari widget.

        Args:
            widget_instance: Instance of HDF5AnalysisWidget
        """
        if not hasattr(widget_instance, "file_path") or not widget_instance.file_path:
            widget_instance._log_message(
                "No HDF5 file loaded for Nematostella analysis"
            )
            widget_instance.results_label.setText("Error: No HDF5 file loaded")
            return

        try:
            widget_instance._log_message("Starting Nematostella timeseries analysis...")
            widget_instance.results_label.setText("Running Nematostella analysis...")

            # Get quick summary first
            summary = get_nematostella_timeseries_summary(widget_instance.file_path)
            widget_instance._log_message("Timeseries summary:")
            for line in summary.split("\n"):
                if line.strip():
                    widget_instance._log_message(f"  {line}")

            # Run full analysis
            results = analyze_nematostella_hdf5_file(widget_instance.file_path)

            if results["success"]:
                widget_instance._log_message(
                    "Nematostella analysis completed successfully!"
                )
                widget_instance._log_message(
                    f"Excel file created: {results['excel_file']}"
                )
                widget_instance._log_message(
                    f"Report file created: {results['report_file']}"
                )
                widget_instance._log_message(
                    f"Sheets created: {', '.join(results['sheets_created'])}"
                )

                # Update results display
                widget_instance.results_label.setText(
                    f"Nematostella analysis complete: {len(results['sheets_created'])} Excel sheets created"
                )

                # Log key findings from report
                widget_instance._log_message("Key Analysis Results:")
                report_lines = results["report"].split("\n")
                in_important_section = False
                for line in report_lines:
                    if any(
                        section in line
                        for section in [
                            "## Timing Analysis",
                            "## LED System Analysis",
                            "## Environmental Conditions",
                        ]
                    ):
                        in_important_section = True
                        widget_instance._log_message(line)
                    elif line.startswith("##") and in_important_section:
                        in_important_section = False
                    elif in_important_section and line.strip().startswith("-"):
                        widget_instance._log_message(f"  {line.strip()}")

            else:
                widget_instance._log_message(
                    f"Nematostella analysis failed: {results['error']}"
                )
                widget_instance.results_label.setText(
                    f"Analysis failed: {results['error']}"
                )

        except Exception as e:
            error_msg = f"Nematostella analysis error: {e}"
            widget_instance._log_message(error_msg)
            widget_instance.results_label.setText(error_msg)

    def _show_save_success_dialog_with_metadata(
        self, saved_files, metadata_dict, nematostella_results=None
    ):
        """Show success dialog with metadata details and optional Nematostella analysis."""
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)

        # Adjust title and message based on whether Nematostella analysis was performed
        if nematostella_results and nematostella_results["success"]:
            msg.setWindowTitle("Save Complete: Metadata + Nematostella Analysis")
            msg.setText(
                f"Successfully saved metadata with specialized Nematostella timeseries analysis in {len(saved_files)} format(s):"
            )
        else:
            msg.setWindowTitle("Save with Metadata Complete")
            msg.setText(
                f"Successfully saved analysis results with metadata in {len(saved_files)} format(s):"
            )

        # Count metadata statistics
        total_static_params = 0
        total_timeseries_params = 0
        total_timeseries_points = 0

        for source_name, metadata in metadata_dict.items():
            # Skip Nematostella analysis entry for counting (it's not traditional metadata)
            if source_name == "nematostella_analysis":
                continue

            # Count static parameters
            static_data = {k: v for k, v in metadata.items() if k != "timeseries_data"}
            total_static_params += len(static_data)

            # Count time-series parameters
            if "timeseries_data" in metadata and metadata["timeseries_data"]:
                ts_data = metadata["timeseries_data"]
                total_timeseries_params += len(ts_data)
                for param_data in ts_data.values():
                    if hasattr(param_data, "__len__"):
                        total_timeseries_points = max(
                            total_timeseries_points, len(param_data)
                        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            file_size = "Unknown size"
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:  # > 1MB
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:  # > 1KB
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                pass

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        # Add Nematostella analysis files if available
        if nematostella_results and nematostella_results["success"]:
            file_details.append("")
            file_details.append("Nematostella Analysis Files:")

            # Add Excel analysis file
            excel_file = nematostella_results["excel_file"]
            try:
                excel_size_bytes = os.path.getsize(excel_file)
                if excel_size_bytes > 1024 * 1024:
                    excel_size = f"{excel_size_bytes/(1024*1024):.1f} MB"
                elif excel_size_bytes > 1024:
                    excel_size = f"{excel_size_bytes/1024:.1f} KB"
                else:
                    excel_size = f"{excel_size_bytes} bytes"
            except:
                excel_size = "Unknown size"

            file_details.append(
                f"• Excel Analysis: {os.path.basename(excel_file)} ({excel_size})"
            )

            # Add text report file
            report_file = nematostella_results["report_file"]
            try:
                report_size_bytes = os.path.getsize(report_file)
                if report_size_bytes > 1024:
                    report_size = f"{report_size_bytes/1024:.1f} KB"
                else:
                    report_size = f"{report_size_bytes} bytes"
            except:
                report_size = "Unknown size"

            file_details.append(
                f"• Text Report: {os.path.basename(report_file)} ({report_size})"
            )
            file_details.append(
                f"• Analysis Sheets: {len(nematostella_results['sheets_created'])}"
            )

        # Add metadata summary
        file_details.append("")
        file_details.append("Metadata Summary:")
        file_details.append(f"• Static parameters: {total_static_params}")
        file_details.append(f"• Time-series parameters: {total_timeseries_params}")
        if total_timeseries_points > 0:
            file_details.append(
                f"• Time-series length: {total_timeseries_points} time points"
            )
            duration_min = (
                total_timeseries_points * self.frame_interval.value()
            ) / 60.0
            file_details.append(f"• Total duration: {duration_min:.1f} minutes")

        # Add Nematostella analysis summary if available
        if nematostella_results and nematostella_results["success"]:
            file_details.append("")
            file_details.append("Nematostella Analysis Summary:")

            # Extract key metrics from the analysis results
            if (
                "analysis_results" in nematostella_results
                and nematostella_results["analysis_results"]
            ):
                analysis_results = nematostella_results["analysis_results"]

                # Timing analysis summary
                if "timing_analysis" in analysis_results:
                    timing = analysis_results["timing_analysis"]
                    if "timing" in timing:
                        accuracy = timing["timing"]["timing_accuracy"]
                        file_details.append(f"• Timing Accuracy: {accuracy:.1%}")

                # Environmental stability
                if (
                    "env_analysis" in analysis_results
                    and analysis_results["env_analysis"]
                ):
                    env = analysis_results["env_analysis"]["environment"]
                    if "temperature" in env:
                        temp_range = env["temperature"]["range"]
                        file_details.append(
                            f"• Temperature Stability: ±{temp_range/2:.1f}°C"
                        )

                # LED system performance
                if (
                    "led_analysis" in analysis_results
                    and analysis_results["led_analysis"]
                ):
                    led = analysis_results["led_analysis"]
                    if "led_sync" in led:
                        sync_rate = led["led_sync"]["success_rate"]
                        file_details.append(f"• LED Sync Success: {sync_rate:.1%}")

        msg.setDetailedText("\n".join(file_details))

        # Adjust informative text based on analysis type
        if nematostella_results and nematostella_results["success"]:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata AND specialized Nematostella timeseries analysis with timing, environmental, and LED system evaluation."
            )
        else:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata in time-series format matching analysis data structure."
            )

        msg.exec_()
        # Adjust informative text based on analysis type
        if nematostella_results and nematostella_results["success"]:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata AND specialized Nematostella timeseries analysis with timing, environmental, and LED system evaluation."
            )
        else:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata in time-series format matching analysis data structure."
            )

        msg.exec_()

    def _show_threshold_stats_in_log(self):
        """
        HELPER METHOD: Show threshold statistics in the log.
        This replaces the separate "Show Threshold Stats" button functionality.
        """
        if not hasattr(self, "roi_baseline_means") or not self.roi_baseline_means:
            self._log_message("No threshold statistics available")
            return

        # Generate threshold statistics for log
        method = self._get_current_threshold_method_display()
        self._log_message(f"Method: {method}")

        if hasattr(self, "baseline_duration_minutes"):
            self._log_message(
                f"Baseline Duration: {self.baseline_duration_minutes.value():.1f} minutes"
            )
        if hasattr(self, "threshold_multiplier"):
            self._log_message(
                f"Hysteresis Multiplier: {self.threshold_multiplier.value():.2f}"
            )
        elif hasattr(self, "calibration_multiplier"):
            self._log_message(
                f"Calibration Multiplier: {self.calibration_multiplier.value():.2f}"
            )

        self._log_message(
            f"Detrending: {'Enabled' if getattr(self, 'enable_detrending', False) and self.enable_detrending.isChecked() else 'Disabled'}"
        )

        roi_band_widths = getattr(self, "roi_band_widths", {})
        roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
        roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})

        # Show statistics for first 5 ROIs to avoid log spam
        rois_to_show = sorted(self.roi_baseline_means.keys())[:5]

        for roi in rois_to_show:
            baseline_mean = self.roi_baseline_means[roi]
            band_width = roi_band_widths.get(roi, 0)
            upper_threshold = roi_upper_thresholds.get(roi, baseline_mean + band_width)
            lower_threshold = roi_lower_thresholds.get(roi, baseline_mean - band_width)

            self._log_message(f"\nROI {roi} HYSTERESIS SYSTEM:")
            self._log_message(f"  Baseline Mean: {baseline_mean:.3f}")
            self._log_message(f"  Band Width: ±{band_width:.3f}")
            self._log_message(
                f"  Upper Threshold: {upper_threshold:.3f} (Movement = TRUE when above)"
            )
            self._log_message(
                f"  Lower Threshold: {lower_threshold:.3f} (Movement = FALSE when below)"
            )
            self._log_message(
                f"  Hysteresis Zone: {lower_threshold:.3f} to {upper_threshold:.3f} (State unchanged)"
            )

        if len(self.roi_baseline_means) > 5:
            self._log_message(f"\n... and {len(self.roi_baseline_means) - 5} more ROIs")

    def _show_save_success_dialog(self, saved_files):
        """
        HELPER METHOD: Show success dialog with list of saved files.
        """
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)
        msg.setWindowTitle("Save Results Complete")
        msg.setText(
            f"Successfully saved analysis results in {len(saved_files)} format(s):"
        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            file_size = "Unknown size"
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:  # > 1MB
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:  # > 1KB
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                pass

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        msg.setDetailedText("\n".join(file_details))
        msg.setInformativeText(
            "Files can be opened in Excel, analyzed further, or imported into other analysis software."
        )
        msg.exec_()

    def _save_results_excel_to_path(self, excel_path: str):
        """
        COMPLETE METHOD: Save Excel results with ALL sheets to a specific file path.
        Creates the same multi-sheet structure as shown in the screenshot.
        """
        try:
            import pandas as pd

            # Create Excel writer
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

                # === SHEET 1: SUMMARY ===
                summary_data = []
                sorted_rois = sorted(self.merged_results.keys())

                for roi in sorted_rois:
                    row_data = {
                        "ROI": roi,
                        "Baseline Mean": getattr(self, "roi_baseline_means", {}).get(
                            roi, 0
                        ),
                        "Upper Threshold": getattr(
                            self, "roi_upper_thresholds", {}
                        ).get(roi, 0),
                        "Lower Threshold": getattr(
                            self, "roi_lower_thresholds", {}
                        ).get(roi, 0),
                        "Threshold Band Width": getattr(
                            self, "roi_band_widths", {}
                        ).get(roi, 0),
                    }

                    # Calculate movement statistics
                    movement_data = getattr(self, "movement_data", {})
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        row_data["Total Movement Events"] = sum(movement_values)
                        row_data["Movement Percentage"] = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                    # Calculate sleep statistics
                    sleep_data = getattr(self, "sleep_data", {})
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        row_data["Total Sleep Bins"] = total_sleep_bins
                        row_data["Sleep Time (min)"] = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                    summary_data.append(row_data)

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
                writer.sheets["Summary"].cell(row=1, column=1).value = (
                    "Summary of ROI statistics: movement events, sleep bins, and detection "
                    "thresholds per ROI. One row per ROI."
                )

                # === SHEET 2: RAW INTENSITY ===
                if hasattr(self, "merged_results") and self.merged_results:
                    intensity_df = self._create_time_series_dataframe(
                        self.merged_results,
                        sorted_rois,
                        "Intensity",
                        convert_to_minutes=True,
                    )
                    intensity_df.to_excel(
                        writer, sheet_name="Raw_Intensity", index=False, startrow=1
                    )
                    writer.sheets["Raw_Intensity"].cell(row=1, column=1).value = (
                        "MinMax-normalized [0–1] raw intensity per ROI (per-pixel mean, "
                        "after MinMax scaling). See 'Real_Amplitude_MATLAB' sheet for "
                        "MATLAB-equivalent pixel sum values. Time column in minutes."
                    )

                # === SHEET 3: MOVEMENT ===
                if hasattr(self, "movement_data") and self.movement_data:
                    movement_df = self._create_time_series_dataframe(
                        self.movement_data,
                        sorted_rois,
                        "Movement",
                        convert_to_minutes=True,
                    )
                    movement_df.to_excel(writer, sheet_name="Movement", index=False, startrow=1)
                    writer.sheets["Movement"].cell(row=1, column=1).value = (
                        "Binary movement detection per ROI (1=moving, 0=stationary). "
                        "Based on hysteresis threshold crossing of raw intensity signal. "
                        "Time column in minutes."
                    )

                # === SHEET 4: FRACTION MOVEMENT ===
                if hasattr(self, "fraction_data") and self.fraction_data:
                    fraction_df = self._create_time_series_dataframe(
                        self.fraction_data,
                        sorted_rois,
                        "Fraction",
                        convert_to_minutes=True,
                    )
                    fraction_df.to_excel(
                        writer, sheet_name="Fraction_Movement", index=False, startrow=1
                    )
                    writer.sheets["Fraction_Movement"].cell(row=1, column=1).value = (
                        "Fraction of time bins with movement per ROI (0–1 scale). "
                        "Derived from binary movement data, binned over the analysis bin size. "
                        "Time column in minutes."
                    )

                # === SHEET 5: SLEEP ===
                if hasattr(self, "sleep_data") and self.sleep_data:
                    sleep_df = self._create_time_series_dataframe(
                        self.sleep_data, sorted_rois, "Sleep", convert_to_minutes=True
                    )
                    sleep_df.to_excel(writer, sheet_name="Sleep", index=False, startrow=1)
                    writer.sheets["Sleep"].cell(row=1, column=1).value = (
                        "Binary sleep state per ROI (1=sleeping, 0=not sleeping). "
                        "Sleep = continuous quiescence exceeding the minimum sleep duration threshold. "
                        "Time column in minutes."
                    )

                # === SHEET 6: QUIESCENCE ===
                if hasattr(self, "quiescence_data") and self.quiescence_data:
                    quiescence_df = self._create_time_series_dataframe(
                        self.quiescence_data,
                        sorted_rois,
                        "Quiescence",
                        convert_to_minutes=True,
                    )
                    quiescence_df.to_excel(
                        writer, sheet_name="Quiescence", index=False, startrow=1
                    )
                    writer.sheets["Quiescence"].cell(row=1, column=1).value = (
                        "Binary quiescence state per ROI (1=quiescent, 0=active). "
                        "Quiescence = fraction movement below the quiescence threshold. "
                        "Time column in minutes."
                    )

                # === SHEET 7: LIGHTING CONDITIONS ===
                if hasattr(self, "fraction_data") and self.fraction_data:
                    # Create lighting conditions data (binned activity for circadian analysis)
                    try:
                        # Use 30-minute bins for lighting analysis
                        from ._calc import bin_activity_data_for_lighting

                        lighting_data = bin_activity_data_for_lighting(
                            self.fraction_data, bin_minutes=30
                        )

                        if lighting_data:
                            lighting_df = self._create_time_series_dataframe(
                                lighting_data,
                                sorted_rois,
                                "Activity_30min_bins",
                                convert_to_minutes=True,
                            )
                            lighting_df.to_excel(
                                writer, sheet_name="Lighting_Conditions", index=False, startrow=1
                            )
                            writer.sheets["Lighting_Conditions"].cell(row=1, column=1).value = (
                                "Fraction movement binned in 30-minute intervals per ROI. "
                                "Suitable for circadian rhythm and lighting condition analysis. "
                                "Time column in minutes."
                            )
                    except Exception as e:
                        self._log_message(
                            f"Warning: Could not create lighting conditions sheet: {e}"
                        )

                # === SHEET 8: REAL AMPLITUDE (MATLAB-EQUIVALENT) ===
                if hasattr(self, "merged_results_raw") and self.merged_results_raw:
                    _norm_factor = getattr(self, "frame_norm_factor", 1.0)
                    _pixel_counts = getattr(self, "roi_pixel_counts", {})
                    _dtype_label = "uint8" if _norm_factor == 255.0 else "uint16"
                    matlab_data = {}
                    for roi, data in self.merged_results_raw.items():
                        _scale = _pixel_counts.get(roi, 1) * _norm_factor
                        matlab_data[roi] = [(t, v * _scale) for t, v in data]
                    matlab_df = self._create_time_series_dataframe(
                        matlab_data, sorted_rois, "Amplitude_raw", convert_to_minutes=True
                    )
                    matlab_df.to_excel(
                        writer, sheet_name="Real_Amplitude_MATLAB", index=False, startrow=1
                    )
                    writer.sheets["Real_Amplitude_MATLAB"].cell(row=1, column=1).value = (
                        f"MATLAB-equivalent pixel sum amplitude: per-pixel-mean x n_pixels x "
                        f"norm_factor (norm_factor={_norm_factor:.0f}, {_dtype_label}). "
                        "Units: sum of absolute pixel differences per frame per ROI "
                        "(equivalent to MATLAB diffs.sum()). Time column in minutes."
                    )

                # === SHEET 9: PARAMETERS ===
                # Determine source type (HDF5 or AVI)
                is_avi = hasattr(self, "avi_batch_paths") and self.avi_batch_paths
                source_label = "AVI Batch" if is_avi else "HDF5"

                params_data = {
                    "Parameter": [
                        "Data Source Type",
                        "Analysis Method",
                        "Frame Interval (s)",
                        "Baseline Duration (min)",
                        "Threshold Multiplier",
                        "Detrending Enabled",
                        "Jump Correction Enabled",
                        "Adaptive Illumination Baseline",
                        "Bin Size (s)",
                        "Quiescence Threshold",
                        "Sleep Threshold (min)",
                        "Number of ROIs",
                        "Total Analysis Time (min)",
                        "Generated Date",
                        "Software Version",
                    ],
                    "Value": [
                        source_label,
                        self._get_current_threshold_method_display(),
                        self.frame_interval.value(),
                        (
                            self.baseline_duration_minutes.value()
                            if hasattr(self, "baseline_duration_minutes")
                            else "N/A"
                        ),
                        (
                            self.threshold_multiplier.value()
                            if hasattr(self, "threshold_multiplier")
                            else "N/A"
                        ),
                        (
                            self.enable_detrending.isChecked()
                            if hasattr(self, "enable_detrending")
                            else "N/A"
                        ),
                        (
                            self.enable_jump_correction.isChecked()
                            if hasattr(self, "enable_jump_correction")
                            else "N/A"
                        ),
                        (
                            self.adaptive_illumination_baseline.isChecked()
                            if hasattr(self, "adaptive_illumination_baseline")
                            else "N/A"
                        ),
                        self.bin_size_seconds.value(),
                        self.quiescence_threshold.value(),
                        self.sleep_threshold_minutes.value(),
                        len(sorted_rois),
                        (
                            self.plot_end_time.value()
                            if hasattr(self, "plot_end_time")
                            else "N/A"
                        ),
                        pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "HDF5 Analysis Widget v1.0",
                    ],
                    "Description": [
                        "Type of data source (HDF5 file or AVI batch)",
                        "Threshold calculation method used",
                        "Time interval between frames",
                        "Duration of baseline period for threshold calculation",
                        "Multiplier for hysteresis band width",
                        "Whether detrending was applied to remove drift",
                        "Whether jump correction was applied",
                        "Whether adaptive per-period illumination baseline was used",
                        "Time bin size for fraction movement calculation",
                        "Threshold below which animal is considered quiescent",
                        "Minimum continuous quiescence duration for sleep",
                        "Total number of ROIs analyzed",
                        "Total duration of analysis",
                        "When this file was generated",
                        "Software version and name",
                    ],
                }

                # Add AVI-specific parameters if applicable
                if is_avi and hasattr(self, "avi_batch_paths"):
                    params_data["Parameter"].extend(
                        [
                            "Number of AVI Files",
                            "AVI Start Time (s)",
                            "AVI End Time (s)",
                            "Total Duration (min)",
                        ]
                    )

                    # Calculate total duration from merged_results
                    total_duration_s = 0
                    start_time_s = 0
                    end_time_s = 0
                    if self.merged_results:
                        first_roi = next(iter(self.merged_results.values()))
                        if first_roi:
                            start_time_s = first_roi[0][0]
                            end_time_s = first_roi[-1][0]
                            total_duration_s = end_time_s - start_time_s

                    params_data["Value"].extend(
                        [
                            len(self.avi_batch_paths),
                            f"{start_time_s:.1f}",
                            f"{end_time_s:.1f}",
                            f"{total_duration_s / 60:.1f}",
                        ]
                    )

                    params_data["Description"].extend(
                        [
                            "Number of AVI video files processed",
                            "Start time of the analysis (in seconds)",
                            "End time of the analysis (in seconds)",
                            "Total duration of the video sequence (in minutes)",
                        ]
                    )

                # Add normalization / pixel-count parameters
                _p_norm = getattr(self, "frame_norm_factor", 1.0)
                _p_pixels = getattr(self, "roi_pixel_counts", {})
                _p_pixel_str = (
                    "; ".join(f"{roi}: {px}" for roi, px in sorted(_p_pixels.items()))
                    if _p_pixels else "N/A"
                )
                params_data["Parameter"].extend(
                    ["Frame Norm Factor", "ROI Pixel Counts"]
                )
                params_data["Value"].extend(
                    [f"{_p_norm:.0f}", _p_pixel_str]
                )
                params_data["Description"].extend(
                    [
                        "Pixel intensity normalization factor (255 for uint8 / 65535 for uint16). "
                        "Used to compute MATLAB-equivalent pixel sum in 'Real_Amplitude_MATLAB' sheet.",
                        "Number of pixels per ROI used for MATLAB-equivalent pixel sum scaling.",
                    ]
                )

                params_df = pd.DataFrame(params_data)
                params_df.to_excel(writer, sheet_name="Parameters", index=False, startrow=1)
                writer.sheets["Parameters"].cell(row=1, column=1).value = (
                    "Analysis parameters and settings used to generate this dataset. "
                    "Refer to these values to reproduce the analysis."
                )

        except Exception as e:
            raise Exception(f"Complete Excel save error: {e}")

    def _save_results_csv(self, file_path: str):
        """Save results in clear CSV format."""
        try:
            sorted_rois = sorted(self.merged_results.keys())

            # Create main data CSV
            with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)

                # === HEADER SECTION ===
                # Determine source type (HDF5 or AVI)
                is_avi = hasattr(self, "avi_batch_paths") and self.avi_batch_paths
                source_label = "AVI" if is_avi else "HDF5"
                writer.writerow([f"{source_label} Analysis Results"])
                writer.writerow(
                    [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                )
                writer.writerow(
                    [f"Analysis Method: {self._get_current_threshold_method_display()}"]
                )
                writer.writerow(
                    [f"Frame Interval: {self.frame_interval.value()} seconds"]
                )
                writer.writerow([f"Number of ROIs: {len(sorted_rois)}"])
                writer.writerow([])  # Empty row

                # === ROI SUMMARY TABLE ===
                writer.writerow(["ROI SUMMARY"])
                summary_headers = [
                    "ROI",
                    "Baseline Mean",
                    "Upper Threshold",
                    "Lower Threshold",
                    "Movement %",
                    "Sleep Time (min)",
                ]
                writer.writerow(summary_headers)

                roi_baseline_means = getattr(self, "roi_baseline_means", {})
                roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
                roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})
                movement_data = getattr(self, "movement_data", {})
                sleep_data = getattr(self, "sleep_data", {})

                for roi in sorted_rois:
                    # Calculate statistics
                    movement_pct = 0
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        movement_pct = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                    sleep_minutes = 0
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        sleep_minutes = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                    writer.writerow(
                        [
                            roi,
                            f"{roi_baseline_means.get(roi, 0):.3f}",
                            f"{roi_upper_thresholds.get(roi, 0):.3f}",
                            f"{roi_lower_thresholds.get(roi, 0):.3f}",
                            f"{movement_pct:.1f}",
                            f"{sleep_minutes:.1f}",
                        ]
                    )

                writer.writerow([])  # Empty row
                writer.writerow([])  # Empty row

                # === TIME SERIES DATA ===
                writer.writerow(["RAW INTENSITY DATA (Time in minutes)"])

                # Create time-aligned data
                all_times = set()
                for roi_data in self.merged_results.values():
                    for time_point, _ in roi_data:
                        all_times.add(round(time_point / 60.0, 2))  # Convert to minutes

                sorted_times = sorted(all_times)

                # Header row: Time, ROI1, ROI2, ROI3, ...
                header = ["Time (min)"] + [f"ROI_{roi}" for roi in sorted_rois]
                writer.writerow(header)

                # Create data rows
                for time_min in sorted_times:
                    row = [f"{time_min:.2f}"]

                    for roi in sorted_rois:
                        # Find value at this time point
                        value = None
                        if roi in self.merged_results:
                            for t, v in self.merged_results[roi]:
                                if abs(t / 60.0 - time_min) < 0.01:  # Within tolerance
                                    value = v
                                    break

                        row.append(f"{value:.6f}" if value is not None else "")

                    writer.writerow(row)

            self._log_message(f"CSV saved: {file_path}")

        except Exception as e:
            self._log_message(f"Error in CSV export: {e}")
            raise

    def _create_time_series_dataframe(
        self,
        data_dict: Dict,
        sorted_rois: List[int],
        data_type: str,
        convert_to_minutes: bool = True,
    ):
        """
        Create a pandas DataFrame with clear time-series structure for Excel export.
        This method already exists in your code but here's the complete version to ensure compatibility.
        """

        # Collect all unique time points
        all_times = set()
        for roi_data in data_dict.values():
            for time_point, _ in roi_data:
                if convert_to_minutes:
                    all_times.add(round(time_point / 60.0, 2))  # Round to 2 decimals
                else:
                    all_times.add(round(time_point, 2))

        sorted_times = sorted(all_times)

        # Create DataFrame structure
        df_data = {"Time (min)" if convert_to_minutes else "Time (s)": sorted_times}

        # Add data for each ROI
        for roi in sorted_rois:
            roi_values = []

            # Create time->value mapping for this ROI
            time_value_map = {}
            if roi in data_dict:
                for time_point, value in data_dict[roi]:
                    if convert_to_minutes:
                        time_key = round(time_point / 60.0, 2)
                    else:
                        time_key = round(time_point, 2)
                    time_value_map[time_key] = value

            # Fill values for all time points
            for time_point in sorted_times:
                if time_point in time_value_map:
                    roi_values.append(time_value_map[time_point])
                else:
                    roi_values.append(None)  # Missing data as None

            # Column name format: ROI_1, ROI_2, etc.
            df_data[f"ROI_{roi}"] = roi_values

        return pd.DataFrame(df_data)

    def save_results(self):
        """Save analysis results - let user choose format."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.results_label.setText("No results to save.")
            return

        # Ask user for format
        file_path, file_type = QFileDialog.getSaveFileName(
            self,
            "Save Results",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
        )

        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx") or "Excel" in file_type:
                # Ensure .xlsx extension
                if not file_path.endswith(".xlsx"):
                    file_path += ".xlsx"
                self._save_results_excel_to_path(file_path)
                self.results_label.setText(
                    f"Results saved to {os.path.basename(file_path)}"
                )
                self._log_message(f"Excel results saved: {file_path}")
            else:
                # Default to CSV
                if not file_path.endswith(".csv"):
                    file_path += ".csv"
                self._save_results_csv(file_path)
                self.results_label.setText(
                    f"Results saved to {os.path.basename(file_path)}"
                )
                self._log_message(f"CSV results saved: {file_path}")

        except Exception as e:
            self.results_label.setText(f"Error saving results: {str(e)}")
            self._log_message(f"ERROR saving results: {str(e)}")
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def show_threshold_statistics(self):
        """Show detailed hysteresis statistics."""
        if not hasattr(self, "roi_baseline_means") or not self.roi_baseline_means:
            self._log_message("No hysteresis statistics available")
            return

        # Create statistics summary
        stats_text = "HYSTERESIS DETECTION SYSTEM STATISTICS\n"
        stats_text += "=" * 50 + "\n\n"

        # Add method parameters
        stats_text += f"Method: {self._get_current_threshold_method_display()}\n"
        if hasattr(self, "baseline_duration_minutes"):
            stats_text += f"Baseline Duration: {self.baseline_duration_minutes.value():.1f} minutes\n"
        if hasattr(self, "threshold_multiplier"):
            stats_text += (
                f"Hysteresis Multiplier: {self.threshold_multiplier.value():.2f}\n"
            )
        elif hasattr(self, "calibration_multiplier"):
            stats_text += (
                f"Calibration Multiplier: {self.calibration_multiplier.value():.2f}\n"
            )
        stats_text += f"Detrending: {'Enabled' if self.enable_detrending.isChecked() else 'Disabled'}\n\n"

        roi_band_widths = getattr(self, "roi_band_widths", {})
        roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
        roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})
        roi_statistics = getattr(self, "roi_statistics", {})

        for roi in sorted(self.roi_baseline_means.keys()):
            baseline_mean = self.roi_baseline_means[roi]
            band_width = roi_band_widths.get(roi, 0)
            stats = roi_statistics.get(roi, {})

            upper_threshold = roi_upper_thresholds.get(roi, baseline_mean + band_width)
            lower_threshold = roi_lower_thresholds.get(roi, baseline_mean - band_width)

            stats_text += f"ROI {roi} - HYSTERESIS SYSTEM:\n"
            stats_text += f"  Baseline Mean: {baseline_mean:.3f}\n"
            stats_text += f"  Band Width: ±{band_width:.3f}\n"
            stats_text += f"  Upper Threshold: {upper_threshold:.3f} (Movement = TRUE when above)\n"
            stats_text += f"  Lower Threshold: {lower_threshold:.3f} (Movement = FALSE when below)\n"
            stats_text += f"  Hysteresis Zone: {lower_threshold:.3f} to {upper_threshold:.3f} (State unchanged)\n"

            if stats.get("was_detrended", False):
                stats_text += "  Detrending: Applied\n"

            # Add method-specific information
            method = stats.get("method", "unknown")
            if "calibration" in method:
                snr = stats.get("signal_to_noise_ratio", 0)
                quality = stats.get("calibration_quality", "unknown")
                stats_text += f"  Calibration Quality: {quality}\n"
                stats_text += f"  Signal-to-Noise Ratio: {snr:.2f}\n"

            stats_text += "\n"

        # Display in log
        self._log_message("DETAILED HYSTERESIS STATISTICS:")
        for line in stats_text.split("\n"):
            if line.strip():
                self._log_message(line)

    def save_results_with_metadata(self):
        """
        Save HDF5 metadata with automatic legacy enhancement and optional Nematostella analysis.
        Enhanced to automatically detect legacy files and add unit documentation.
        """

        # Check if we have a file loaded
        if not hasattr(self, "file_path") or not self.file_path:
            self.results_label.setText("No HDF5 file loaded. Load a file first.")
            self._log_message("Save failed: No HDF5 file loaded")
            return

        # Analysis results are optional for metadata extraction
        has_analysis_results = hasattr(self, "merged_results") and self.merged_results

        if has_analysis_results:
            self._log_message("Saving analysis results with HDF5 metadata...")
        else:
            self._log_message(
                "Saving HDF5 metadata only (no analysis results available)..."
            )

        # NEW: Check for Nematostella timeseries data
        nematostella_results = None
        # Direkte Prüfung statt globaler Variable
        try:
            from ._metadata import analyze_nematostella_hdf5_file

            nematostella_available = True
        except ImportError:
            nematostella_available = False

        if nematostella_available:
            try:
                self._log_message("Checking for Nematostella timeseries data...")

                # Quick check if this is a Nematostella experiment
                with h5py.File(self.file_path, "r") as h5_file:
                    if "timeseries" in h5_file:
                        ts_group = h5_file["timeseries"]
                        # Check for typical Nematostella parameters
                        nematostella_indicators = [
                            "actual_intervals",
                            "expected_intervals",
                            "frame_drift",
                            "temperature",
                            "humidity",
                            "led_power_percent",
                        ]

                        found_indicators = [
                            key
                            for key in ts_group.keys()
                            if key in nematostella_indicators
                        ]

                        if len(found_indicators) >= 2:  # At least 2 indicators found
                            self._log_message(
                                f"Nematostella experiment detected! Found: {', '.join(found_indicators)}"
                            )
                            self._log_message(
                                "Running specialized Nematostella timeseries analysis..."
                            )

                            # Run Nematostella analysis
                            nematostella_results = analyze_nematostella_hdf5_file(
                                self.file_path
                            )

                            if nematostella_results["success"]:
                                self._log_message(
                                    f"Nematostella analysis completed: {len(nematostella_results['sheets_created'])} sheets"
                                )
                            else:
                                self._log_message(
                                    f"Nematostella analysis failed: {nematostella_results['error']}"
                                )
                        else:
                            self._log_message(
                                "No Nematostella-specific timeseries detected"
                            )
            except Exception as e:
                self._log_message(f"Nematostella detection failed: {e}")

        # Get base filename from user
        from qtpy.QtWidgets import QFileDialog

        if nematostella_results and nematostella_results["success"]:
            dialog_title = "Save HDF5 Metadata with Nematostella Analysis"
            default_name = f"nematostella_metadata_{int(time.time())}"
        else:
            dialog_title = "Save HDF5 Metadata" + (
                " with Analysis Results" if has_analysis_results else ""
            )
            default_name = f"hdf5_metadata_{int(time.time())}"

        base_path, _ = QFileDialog.getSaveFileName(
            self, dialog_title, default_name, "All Files (*)"
        )

        if not base_path:
            self._log_message("Save cancelled by user")
            return

        base_path = os.path.splitext(base_path)[0]
        saved_files = []

        try:
            # === AUTOMATIC LEGACY ENHANCEMENT INTEGRATION ===
            self._log_message(
                "Extracting HDF5 metadata with automatic legacy enhancement..."
            )
            metadata_dict = {}

            # Extract from main file with automatic legacy enhancement
            if hasattr(self, "file_path") and self.file_path:
                self._log_message(
                    f"   Extracting from main file: {os.path.basename(self.file_path)}"
                )
                try:
                    # This function now automatically enhances legacy files
                    main_metadata = extract_hdf5_metadata_timeseries(self.file_path)
                    metadata_dict["main_file"] = main_metadata

                    # Log automatic legacy enhancement results
                    if main_metadata.get("legacy_enhanced", False):
                        enhancement_info = main_metadata.get("_enhancement_summary", {})
                        enhanced_params = enhancement_info.get("parameters_enhanced", 0)
                        self._log_message("     ✅ Legacy file automatically enhanced!")
                        self._log_message(
                            f"     📏 Unit documentation added for {enhanced_params} parameters"
                        )
                        self._log_message(
                            f"     🕒 Enhancement timestamp: {main_metadata.get('enhancement_timestamp', 'unknown')}"
                        )
                        self._log_message(
                            "     📊 Unit standard: seconds for timing, celsius for temp, percent for humidity"
                        )
                    elif main_metadata.get("modern_file", False):
                        self._log_message(
                            "     ✅ Modern file with existing unit documentation detected"
                        )
                    else:
                        self._log_message("     ⚠️ File type could not be determined")

                    # Log timeseries data found
                    if (
                        "timeseries_data" in main_metadata
                        and main_metadata["timeseries_data"]
                    ):
                        ts_data = main_metadata["timeseries_data"]
                        # Count non-metadata parameters
                        param_count = len(
                            [k for k in ts_data.keys() if not k.startswith("_")]
                        )
                        self._log_message(
                            f"     📈 Found {param_count} time-series parameters"
                        )

                        # Log unit enhancement details if available
                        unit_info = ts_data.get("_unit_info", {})
                        if unit_info:
                            timing_params = [
                                k
                                for k, v in unit_info.items()
                                if v.get("units") == "seconds"
                            ]
                            environmental_params = [
                                k
                                for k, v in unit_info.items()
                                if v.get("units") in ["celsius", "percent"]
                            ]
                            if timing_params:
                                self._log_message(
                                    f"       ⏱️ Timing parameters: {', '.join(timing_params[:3])}{'...' if len(timing_params) > 3 else ''}"
                                )
                            if environmental_params:
                                self._log_message(
                                    f"       🌡️ Environmental parameters: {', '.join(environmental_params)}"
                                )

                except Exception as e:  # <-- JETZT KORREKT EINGERÜCKT
                    self._log_message(f"     Main file metadata extraction failed: {e}")
                    metadata_dict["main_file"] = {
                        "error": str(e),
                        "timeseries_data": {},
                    }

            # Add analysis metadata (only if we have analysis results)
            if has_analysis_results:
                metadata_dict["analysis_info"] = {
                    "analysis_method": self._get_current_threshold_method_display(),
                    "analysis_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "frame_interval": self.frame_interval.value(),
                    "rois_analyzed": len(self.merged_results),
                    "software_version": "HDF5 Activity Analysis Widget v1.0 (Legacy Enhanced)",
                    "parameters": self._get_analysis_parameters_for_metadata(),
                    "timeseries_data": {},
                    "legacy_compatibility": True,  # Mark as legacy-compatible
                }
            else:
                metadata_dict["file_info_only"] = {
                    "extraction_type": "HDF5 metadata only (Legacy Enhanced)",
                    "extraction_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "source_file": os.path.basename(self.file_path),
                    "software_version": "HDF5 Activity Analysis Widget v1.0 (Legacy Enhanced)",
                    "timeseries_data": {},
                    "legacy_compatibility": True,
                }

            # NEW: Add Nematostella analysis results if available
            if nematostella_results and nematostella_results["success"]:
                metadata_dict["nematostella_analysis"] = {
                    "analysis_type": "Nematostella Timeseries Analysis (Legacy Enhanced)",
                    "excel_file": nematostella_results["excel_file"],
                    "report_file": nematostella_results["report_file"],
                    "sheets_created": nematostella_results["sheets_created"],
                    "analysis_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "timeseries_data": {},
                    "legacy_enhanced": metadata_dict["main_file"].get(
                        "legacy_enhanced", False
                    ),
                }

            self._log_message("Metadata extraction with legacy enhancement completed")

            # Save CSV with enhanced metadata
            csv_path = f"{base_path}_metadata.csv"
            self._log_message(
                f"Saving enhanced CSV with metadata: {os.path.basename(csv_path)}"
            )

            try:
                self._save_results_csv_with_metadata(
                    csv_path, metadata_dict, has_analysis_results
                )
                saved_files.append(("Enhanced CSV with Metadata", csv_path))
                self._log_message("Enhanced CSV with metadata saved successfully")
            except Exception as e:
                self._log_message(f"CSV save failed: {e}")

            # Save Excel with enhanced metadata (if pandas available)
            try:
                import pandas as pd

                excel_path = f"{base_path}_metadata.xlsx"
                self._log_message(
                    f"Saving enhanced Excel with metadata: {os.path.basename(excel_path)}"
                )

                if has_analysis_results:
                    # Step 1: Write all analysis sheets (Movement, Sleep, etc.)
                    self._save_results_excel_to_path(excel_path)
                    # Step 2: Append HDF5 sensor timeseries sheets to same file
                    self._append_hdf5_sheets_to_excel(excel_path, metadata_dict)
                else:
                    # No analysis data — write HDF5 metadata only
                    self._save_results_excel_with_metadata(
                        excel_path, metadata_dict, False
                    )
                saved_files.append(("Enhanced Excel with Metadata", excel_path))
                self._log_message("Enhanced Excel with metadata saved successfully")

            except ImportError:
                self._log_message(
                    "Excel export not available (missing pandas/openpyxl)"
                )
            except Exception as e:
                self._log_message(f"Excel save failed: {e}")

            # Update UI with legacy enhancement information
            if saved_files:
                file_list = ", ".join(
                    [f"{fmt} ({os.path.basename(path)})" for fmt, path in saved_files]
                )

                # Enhanced result message
                is_legacy = metadata_dict.get("main_file", {}).get(
                    "legacy_enhanced", False
                )
                legacy_suffix = " (Legacy Enhanced)" if is_legacy else ""

                if nematostella_results and nematostella_results["success"]:
                    result_msg = f"Saved metadata + Nematostella analysis{legacy_suffix}: {file_list}"
                    result_msg += (
                        f" + {os.path.basename(nematostella_results['excel_file'])}"
                    )
                else:
                    result_msg = f"Saved metadata{legacy_suffix}: {file_list}"

                self.results_label.setText(result_msg)
                self._log_message(
                    f"Save with metadata complete: {len(saved_files)} files created"
                )

                # Log enhancement summary
                if is_legacy:
                    enhancement_summary = metadata_dict["main_file"].get(
                        "_enhancement_summary", {}
                    )
                    enhanced_count = enhancement_summary.get("parameters_enhanced", 0)
                    self._log_message("📋 Legacy Enhancement Summary:")
                    self._log_message(f"   Parameters enhanced: {enhanced_count}")
                    self._log_message(
                        f"   Unit standard applied: {enhancement_summary.get('unit_standard', 'Unknown')}"
                    )
                    self._log_message(
                        "   Files include comprehensive unit documentation"
                    )

                # Log Nematostella results if available
                if nematostella_results and nematostella_results["success"]:
                    self._log_message("Nematostella Analysis Summary:")
                    report_lines = nematostella_results["report"].split("\n")
                    for line in report_lines:
                        if any(
                            section in line
                            for section in [
                                "## Timing Analysis",
                                "## Environmental Conditions",
                                "## LED System",
                            ]
                        ):
                            self._log_message(line)
                        elif line.strip().startswith("-") and any(
                            keyword in line
                            for keyword in ["Mean", "Accuracy", "Success Rate"]
                        ):
                            self._log_message(f"  {line.strip()}")

                # Show enhanced success dialog
                try:
                    self._show_save_success_dialog_with_metadata(
                        saved_files, metadata_dict, nematostella_results
                    )
                except TypeError:
                    # Fallback to old method signature
                    self._show_save_success_dialog_with_metadata(
                        saved_files, metadata_dict
                    )
            else:
                self.results_label.setText("All save attempts failed - check log")

        except Exception as e:
            error_msg = f"Save with metadata failed: {e}"
            self.results_label.setText(error_msg)
            self._log_message(error_msg)
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def _create_legacy_enhanced_sheets(
        self, writer, ts_data: dict, unit_info: dict, source_name: str
    ):
        """Create enhanced sheets for legacy data with automatic unit documentation."""

        # Erstelle DataFrame mit Unit-erweiterten Spalten-Namen
        enhanced_columns = {}

        for param_name, param_data in ts_data.items():
            if param_name.startswith("_"):
                continue  # Skip metadata

            unit = unit_info.get(param_name, {}).get("units", "unknown")

            if unit == "seconds" and "drift" in param_name.lower():
                # Für Timing-Daten: beide Einheiten
                enhanced_columns[f"{param_name}_sec"] = param_data
                enhanced_columns[f"{param_name}_ms"] = [
                    d * 1000 if d else 0 for d in param_data
                ]
            else:
                # Standard Parameter mit Unit-Suffix
                enhanced_columns[f"{param_name}_{unit}"] = param_data

        if enhanced_columns:
            # Frame index hinzufügen
            max_length = max(
                len(data)
                for data in enhanced_columns.values()
                if isinstance(data, (list, tuple))
            )
            enhanced_columns["frame_index"] = list(range(max_length))

            df = pd.DataFrame(enhanced_columns)
            sheet_name = f"Enhanced_{source_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _create_automatic_units_reference_sheet(self, writer, metadata_dict: dict):
        """Automatically create units reference sheet for legacy files."""

        units_found = set()

        # Sammle alle gefundenen Parameter und ihre Units
        for metadata in metadata_dict.values():
            if "timeseries_data" in metadata:
                unit_info = metadata["timeseries_data"].get("_unit_info", {})
                for param, info in unit_info.items():
                    units_found.add(
                        (param, info["units"], info.get("display_hint", ""))
                    )

        if units_found:
            units_data = []
            for param, unit, hint in sorted(units_found):
                units_data.append(
                    {
                        "Parameter": param,
                        "Units": unit,
                        "Display_Hint": hint,
                        "Enhancement": "Automatically added for legacy compatibility",
                    }
                )

            units_df = pd.DataFrame(units_data)
            units_df.to_excel(writer, sheet_name="Auto_Units_Reference", index=False)

    def _save_results_csv_with_metadata(
        self, file_path: str, metadata_dict: dict, has_analysis_results: bool = True
    ):
        """
        Save CSV with HDF5 metadata time-series (with optional analysis results).
        """
        import csv
        from datetime import datetime

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            # === HEADER SECTION ===
            if has_analysis_results:
                writer.writerow(["HDF5 Analysis Results with Time-Series Metadata"])
                sorted_rois = sorted(self.merged_results.keys())
                writer.writerow([f"Number of ROIs: {len(sorted_rois)}"])
            else:
                writer.writerow(["HDF5 Time-Series Metadata Only"])
                writer.writerow(
                    ["No analysis results available - metadata extraction only"]
                )

            writer.writerow(
                [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            )
            writer.writerow([f"Source file: {os.path.basename(self.file_path)}"])
            writer.writerow([])

            # === ANALYSIS RESULTS SUMMARY (only if available) ===
            if has_analysis_results:
                writer.writerow(["=== ANALYSIS RESULTS SUMMARY ==="])
                writer.writerow(
                    [
                        "ROI",
                        "Baseline Mean",
                        "Upper Threshold",
                        "Lower Threshold",
                        "Movement %",
                        "Sleep Time (min)",
                    ]
                )

                # Get analysis data
                roi_baseline_means = getattr(self, "roi_baseline_means", {})
                roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
                roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})
                movement_data = getattr(self, "movement_data", {})
                sleep_data = getattr(self, "sleep_data", {})

                for roi in sorted_rois:
                    # Calculate statistics
                    movement_pct = 0
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        movement_pct = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                    sleep_minutes = 0
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        sleep_minutes = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                    writer.writerow(
                        [
                            roi,
                            f"{roi_baseline_means.get(roi, 0):.3f}",
                            f"{roi_upper_thresholds.get(roi, 0):.3f}",
                            f"{roi_lower_thresholds.get(roi, 0):.3f}",
                            f"{movement_pct:.1f}",
                            f"{sleep_minutes:.1f}",
                        ]
                    )

                writer.writerow([])
                writer.writerow([])

            # === STATIC HDF5 METADATA SECTIONS ===
            for source_name, metadata in metadata_dict.items():
                if source_name in ["analysis_info", "file_info_only"]:
                    continue  # Handle separately

                writer.writerow(
                    [f"=== {source_name.upper().replace('_', ' ')} STATIC METADATA ==="]
                )

                # Write static metadata (excluding timeseries_data)
                static_metadata = {
                    k: v for k, v in metadata.items() if k != "timeseries_data"
                }
                if static_metadata:
                    from ._metadata import write_metadata_to_csv

                    write_metadata_to_csv(writer, static_metadata, source_name.upper())
                else:
                    writer.writerow(["No static metadata available"])

                writer.writerow([])

            # === HDF5 TIME-SERIES METADATA SECTIONS ===
            has_hdf5_timeseries = False
            for source_name, metadata in metadata_dict.items():
                if "timeseries_data" in metadata and metadata["timeseries_data"]:
                    has_hdf5_timeseries = True
                    writer.writerow(
                        [
                            f"=== {source_name.upper().replace('_', ' ')} HDF5 TIME-SERIES METADATA ==="
                        ]
                    )

                    ts_data = metadata["timeseries_data"]

                    # Filter out analysis-related data - only keep actual HDF5 metadata
                    from ._metadata import filter_hdf5_metadata_only

                    hdf5_metadata_only = filter_hdf5_metadata_only(ts_data)

                    if hdf5_metadata_only:
                        max_length = max(
                            len(data) for data in hdf5_metadata_only.values()
                        )

                        # Log what we're including
                        param_names = list(hdf5_metadata_only.keys())
                        self._log_message(
                            f"   Including HDF5 time-series: {param_names}"
                        )

                        # Align time with analysis data (or use generic timing)
                        frame_interval = (
                            self.frame_interval.value() if has_analysis_results else 5.0
                        )

                        # Header: Time (min), parameters...
                        header = ["Time (min)"] + param_names
                        writer.writerow(header)

                        # Data rows
                        for i in range(max_length):
                            time_min = (i * frame_interval) / 60.0
                            row = [f"{time_min:.2f}"]

                            for param_name in param_names:
                                param_data = hdf5_metadata_only[param_name]
                                if i < len(param_data):
                                    value = param_data[i]
                                    if isinstance(value, (int, float)):
                                        row.append(f"{value:.6f}")
                                    else:
                                        row.append(str(value))
                                else:
                                    row.append("")  # Missing data

                            writer.writerow(row)

                        writer.writerow([])
                        writer.writerow(
                            [
                                f"HDF5 time-series metadata: {len(hdf5_metadata_only)} parameters, {max_length} time points"
                            ]
                        )

                        # List all parameters
                        writer.writerow(["Parameters included:"])
                        for param in param_names:
                            writer.writerow([f"  - {param}"])

                    else:
                        writer.writerow(["No HDF5 time-series metadata found"])

                    writer.writerow([])
                    writer.writerow([])

            if not has_hdf5_timeseries:
                writer.writerow(["=== NO HDF5 TIME-SERIES METADATA FOUND ==="])
                writer.writerow(
                    ["Your HDF5 file may not contain time-series metadata."]
                )
                writer.writerow([])

            # === ANALYSIS/FILE INFO PARAMETERS ===
            info_section = metadata_dict.get("analysis_info") or metadata_dict.get(
                "file_info_only"
            )
            if info_section:
                section_name = (
                    "ANALYSIS PARAMETERS"
                    if has_analysis_results
                    else "FILE INFORMATION"
                )
                writer.writerow([f"=== {section_name} ==="])

                # Write parameters
                writer.writerow(["Parameter", "Value", "Description"])

                param_descriptions = {
                    "analysis_method": "Threshold calculation method used",
                    "analysis_timestamp": "When analysis was performed",
                    "extraction_timestamp": "When metadata was extracted",
                    "frame_interval": "Time interval between frames (seconds)",
                    "rois_analyzed": "Total number of ROIs analyzed",
                    "extraction_type": "Type of extraction performed",
                    "source_file": "Source HDF5 file name",
                    "software_version": "Analysis software version",
                }

                for param, value in info_section.items():
                    if param != "parameters":
                        description = param_descriptions.get(param, "Parameter")
                        writer.writerow([param, str(value), description])

                # Write nested parameters (if analysis results available)
                if "parameters" in info_section:
                    writer.writerow([])
                    writer.writerow(["Detailed Parameters:"])
                    writer.writerow(["Parameter", "Value"])

                    for param, value in info_section["parameters"].items():
                        writer.writerow([param, str(value)])

            # === FOOTER ===
            writer.writerow([])
            writer.writerow(["=== END OF FILE ==="])
            writer.writerow(
                [f"Generation time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            )

    def _show_save_success_dialog_with_metadata(self, saved_files, metadata_dict):
        """Show success dialog with metadata details."""
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)
        msg.setWindowTitle("Save with Metadata Complete")
        msg.setText(
            f"Successfully saved analysis results with metadata in {len(saved_files)} format(s):"
        )

        # Count metadata statistics
        total_static_params = 0
        total_timeseries_params = 0
        total_timeseries_points = 0

        for source_name, metadata in metadata_dict.items():
            # Count static parameters
            static_data = {k: v for k, v in metadata.items() if k != "timeseries_data"}
            total_static_params += len(static_data)

            # Count time-series parameters
            if "timeseries_data" in metadata and metadata["timeseries_data"]:
                ts_data = metadata["timeseries_data"]
                total_timeseries_params += len(ts_data)
                for param_data in ts_data.values():
                    if hasattr(param_data, "__len__"):
                        total_timeseries_points = max(
                            total_timeseries_points, len(param_data)
                        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            file_size = "Unknown size"
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:  # > 1MB
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:  # > 1KB
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                pass

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        # Add metadata summary
        file_details.append("")
        file_details.append("Metadata Summary:")
        file_details.append(f"• Static parameters: {total_static_params}")
        file_details.append(f"• Time-series parameters: {total_timeseries_params}")
        if total_timeseries_points > 0:
            file_details.append(
                f"• Time-series length: {total_timeseries_points} time points"
            )
            duration_min = (
                total_timeseries_points * self.frame_interval.value()
            ) / 60.0
            file_details.append(f"• Total duration: {duration_min:.1f} minutes")

        msg.setDetailedText("\n".join(file_details))
        msg.setInformativeText(
            "Files include comprehensive HDF5 metadata in time-series format matching analysis data structure."
        )
        msg.exec_()

    def _append_hdf5_sheets_to_excel(self, excel_path: str, metadata_dict: dict):
        """
        Append HDF5 sensor timeseries sheets (temperature, humidity, LED, frame drift, etc.)
        to an existing Excel file without overwriting the analysis sheets already in it.
        """
        import pandas as pd

        is_legacy_enhanced = metadata_dict.get("main_file", {}).get(
            "legacy_enhanced", False
        )
        appended_sheets = []

        try:
            with pd.ExcelWriter(
                excel_path, engine="openpyxl", mode="a", if_sheet_exists="new"
            ) as writer:
                for source_name, metadata in metadata_dict.items():
                    if source_name in [
                        "analysis_info",
                        "file_info_only",
                        "nematostella_analysis",
                    ]:
                        continue

                    # HDF5 timeseries sheets
                    if "timeseries_data" in metadata and metadata["timeseries_data"]:
                        ts_data = metadata["timeseries_data"]
                        unit_info = ts_data.get("_unit_info", {})
                        hdf5_metadata_only = self._filter_hdf5_metadata_only(ts_data)

                        if hdf5_metadata_only:
                            frame_interval = self.frame_interval.value()

                            if is_legacy_enhanced and unit_info:
                                created = self._create_unit_enhanced_timeseries_sheets(
                                    writer,
                                    hdf5_metadata_only,
                                    unit_info,
                                    frame_interval,
                                    source_name,
                                )
                            else:
                                try:
                                    from ._metadata import (
                                        create_individual_timeseries_sheets,
                                    )

                                    created = create_individual_timeseries_sheets(
                                        writer, hdf5_metadata_only, frame_interval
                                    )
                                except Exception:
                                    created = self._create_timeseries_sheets_manually(
                                        writer, hdf5_metadata_only, source_name
                                    )

                            appended_sheets.extend(created)
                            for s in created:
                                self._log_message(f"   ✓ HDF5 sheet appended: '{s}'")

                    # Static HDF5 metadata sheet
                    static_metadata = {
                        k: v for k, v in metadata.items() if k != "timeseries_data"
                    }
                    if static_metadata:
                        try:
                            try:
                                from ._metadata import create_metadata_dataframe

                                meta_df = create_metadata_dataframe(
                                    static_metadata, source_name
                                )
                            except ImportError:
                                meta_df = self._create_metadata_dataframe_manually(
                                    static_metadata, source_name
                                )

                            sheet_name = f"Static_{source_name}"[:31]
                            meta_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            appended_sheets.append(sheet_name)
                            self._log_message(
                                f"   ✓ Static metadata sheet appended: '{sheet_name}'"
                            )
                        except Exception as e:
                            self._log_message(
                                f"   Warning: Could not create static sheet: {e}"
                            )

            self._log_message(
                f"HDF5 sensor sheets appended ({len(appended_sheets)} sheets): "
                + ", ".join(appended_sheets)
            )
        except Exception as e:
            self._log_message(f"Warning: Could not append HDF5 sheets: {e}")

    def _save_results_excel_with_metadata(
        self, excel_path: str, metadata_dict: dict, has_analysis_results: bool = True
    ):
        """
        Save Excel with unit-enhanced headers for legacy files and individual HDF5 sheets.
        """
        import pandas as pd

        # Check if this is a legacy enhanced file
        is_legacy_enhanced = metadata_dict.get("main_file", {}).get(
            "legacy_enhanced", False
        )

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

            # === SUMMARY SHEET (with unit-aware columns if legacy enhanced) ===
            if has_analysis_results:
                sorted_rois = sorted(self.merged_results.keys())

                summary_data = []
                for roi in sorted_rois:
                    if is_legacy_enhanced:
                        # Enhanced column names with units
                        row_data = {
                            "ROI": roi,
                            "Baseline_Mean_intensity": getattr(
                                self, "roi_baseline_means", {}
                            ).get(roi, 0),
                            "Upper_Threshold_intensity": getattr(
                                self, "roi_upper_thresholds", {}
                            ).get(roi, 0),
                            "Lower_Threshold_intensity": getattr(
                                self, "roi_lower_thresholds", {}
                            ).get(roi, 0),
                        }
                    else:
                        # Traditional column names
                        row_data = {
                            "ROI": roi,
                            "Baseline Mean": getattr(
                                self, "roi_baseline_means", {}
                            ).get(roi, 0),
                            "Upper Threshold": getattr(
                                self, "roi_upper_thresholds", {}
                            ).get(roi, 0),
                            "Lower Threshold": getattr(
                                self, "roi_lower_thresholds", {}
                            ).get(roi, 0),
                        }

                    # Add movement and sleep statistics with unit-aware names
                    movement_data = getattr(self, "movement_data", {})
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        movement_pct = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                        if is_legacy_enhanced:
                            row_data["Movement_Percentage_0to100"] = movement_pct
                        else:
                            row_data["Movement Percentage"] = movement_pct

                    sleep_data = getattr(self, "sleep_data", {})
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        sleep_minutes = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                        if is_legacy_enhanced:
                            row_data["Sleep_Time_minutes"] = sleep_minutes
                        else:
                            row_data["Sleep Time (min)"] = sleep_minutes

                    summary_data.append(row_data)

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                all_sheets_created = ["Summary"]
            else:
                # Create info sheet with legacy enhancement info
                info_data = [
                    {
                        "Property": "Extraction Type",
                        "Value": (
                            "HDF5 Metadata Only (Legacy Enhanced)"
                            if is_legacy_enhanced
                            else "HDF5 Metadata Only"
                        ),
                        "Description": "No analysis results available",
                    },
                    {
                        "Property": "Source File",
                        "Value": os.path.basename(self.file_path),
                        "Description": "HDF5 file analyzed",
                    },
                    {
                        "Property": "Extraction Date",
                        "Value": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Description": "When metadata was extracted",
                    },
                ]

                if is_legacy_enhanced:
                    info_data.append(
                        {
                            "Property": "Legacy Enhancement",
                            "Value": "Applied",
                            "Description": "Unit documentation added automatically",
                        }
                    )

                info_df = pd.DataFrame(info_data)
                info_df.to_excel(writer, sheet_name="File_Info", index=False)
                all_sheets_created = ["File_Info"]

            # === PROCESS HDF5 TIME-SERIES METADATA WITH UNIT ENHANCEMENT ===
            for source_name, metadata in metadata_dict.items():
                if source_name in [
                    "analysis_info",
                    "file_info_only",
                    "nematostella_analysis",
                ]:
                    continue

                # Process HDF5 time-series metadata
                if "timeseries_data" in metadata and metadata["timeseries_data"]:
                    ts_data = metadata["timeseries_data"]

                    # Get unit information if available
                    unit_info = ts_data.get("_unit_info", {})

                    # Filter to get only HDF5 metadata
                    hdf5_metadata_only = self._filter_hdf5_metadata_only(ts_data)

                    if hdf5_metadata_only:
                        self._log_message(
                            f"Creating unit-enhanced Excel sheets for {len(hdf5_metadata_only)} HDF5 parameters from {source_name}"
                        )

                        # Use frame interval from analysis if available, otherwise default
                        frame_interval = (
                            self.frame_interval.value() if has_analysis_results else 5.0
                        )

                        # CREATE UNIT-ENHANCED SHEETS
                        if is_legacy_enhanced and unit_info:
                            # Use enhanced unit-aware sheet creation
                            created_sheets = (
                                self._create_unit_enhanced_timeseries_sheets(
                                    writer,
                                    hdf5_metadata_only,
                                    unit_info,
                                    frame_interval,
                                    source_name,
                                )
                            )
                            all_sheets_created.extend(created_sheets)

                            self._log_message(
                                f"   ✅ Created {len(created_sheets)} unit-enhanced sheets"
                            )
                            for sheet_name in created_sheets:
                                self._log_message(f"   - {sheet_name}")

                        else:
                            # Fallback to regular sheet creation
                            try:
                                from ._metadata import (
                                    create_individual_timeseries_sheets,
                                    create_combined_timeseries_sheet,
                                )

                                individual_sheets = create_individual_timeseries_sheets(
                                    writer, hdf5_metadata_only, frame_interval
                                )
                                all_sheets_created.extend(individual_sheets)

                                for sheet_name in individual_sheets:
                                    self._log_message(
                                        f"   ✓ Created sheet '{sheet_name}'"
                                    )

                            except ImportError:
                                created_sheets = (
                                    self._create_timeseries_sheets_manually(
                                        writer, hdf5_metadata_only, source_name
                                    )
                                )
                                all_sheets_created.extend(created_sheets)
                            except Exception as e:
                                self._log_message(
                                    f"   Error with metadata functions: {e}"
                                )
                                created_sheets = (
                                    self._create_timeseries_sheets_manually(
                                        writer, hdf5_metadata_only, source_name
                                    )
                                )
                                all_sheets_created.extend(created_sheets)

                # Static HDF5 metadata sheet
                static_metadata = {
                    k: v for k, v in metadata.items() if k != "timeseries_data"
                }
                if static_metadata:
                    try:
                        try:
                            from ._metadata import create_metadata_dataframe

                            meta_df = create_metadata_dataframe(
                                static_metadata, source_name
                            )
                        except ImportError:
                            meta_df = self._create_metadata_dataframe_manually(
                                static_metadata, source_name
                            )

                        sheet_name = f"Static_{source_name}"[:31]
                        meta_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        all_sheets_created.append(sheet_name)
                        self._log_message(
                            f"   ✓ Created static metadata sheet '{sheet_name}'"
                        )
                    except Exception as e:
                        self._log_message(
                            f"   Warning: Could not create static sheet: {e}"
                        )

            # === PARAMETERS SHEET (only if analysis available) ===
            if has_analysis_results and "analysis_info" in metadata_dict:
                try:
                    params_data = []
                    analysis_info = metadata_dict["analysis_info"]

                    for key, value in analysis_info.items():
                        if key != "parameters":
                            params_data.append(
                                {
                                    "Parameter": key,
                                    "Value": str(value),
                                    "Category": "Analysis Info",
                                }
                            )

                    if "parameters" in analysis_info:
                        for key, value in analysis_info["parameters"].items():
                            params_data.append(
                                {
                                    "Parameter": key,
                                    "Value": str(value),
                                    "Category": "Analysis Parameters",
                                }
                            )

                    if params_data:
                        params_df = pd.DataFrame(params_data)
                        params_df.to_excel(
                            writer, sheet_name="Analysis_Parameters", index=False
                        )
                        all_sheets_created.append("Analysis_Parameters")
                except Exception as e:
                    self._log_message(
                        f"   Warning: Could not create parameters sheet: {e}"
                    )

            # Log final summary with enhancement info
            enhancement_info = " (with unit enhancement)" if is_legacy_enhanced else ""
            self._log_message(
                f"Excel file created{enhancement_info} with {len(all_sheets_created)} sheets:"
            )
            for sheet in all_sheets_created:
                self._log_message(f"   - {sheet}")

    def _create_unit_enhanced_timeseries_sheets(
        self,
        writer,
        hdf5_metadata: dict,
        unit_info: dict,
        frame_interval: float,
        source_name: str,
    ):
        """Create individual timeseries sheets with unit-enhanced column headers."""
        sheets_created = []

        # Create individual sheet for each parameter with unit-enhanced names
        for param_name, param_data in hdf5_metadata.items():
            if (
                not isinstance(param_data, (list, tuple, np.ndarray))
                or len(param_data) == 0
            ):
                continue

            try:
                max_length = len(param_data)
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                # Get unit information
                unit = unit_info.get(param_name, {}).get("units", "unknown")

                # Create unit-enhanced column names
                if unit == "seconds" and "drift" in param_name.lower():
                    # For timing parameters: create both seconds and milliseconds columns
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_seconds": param_data,
                        f"{param_name}_milliseconds": [
                            d * 1000 if d else 0 for d in param_data
                        ],
                    }
                elif unit == "celsius":
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_celsius": param_data,
                    }
                elif unit == "percent":
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_percent": param_data,
                    }
                elif unit == "milliseconds":
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_milliseconds": param_data,
                    }
                else:
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_{unit}": param_data,
                    }

                # Create DataFrame and sheet
                param_df = pd.DataFrame(df_data)
                clean_name = self._clean_sheet_name(f"{param_name}_Enhanced")

                # Ensure unique sheet name
                original_clean_name = clean_name
                counter = 1
                while clean_name in sheets_created:
                    clean_name = f"{original_clean_name[:28]}_{counter}"
                    counter += 1

                param_df.to_excel(writer, sheet_name=clean_name, index=False)
                sheets_created.append(clean_name)

            except Exception as e:
                self._log_message(
                    f"   Warning: Could not create enhanced sheet for {param_name}: {e}"
                )
                continue

        # Create combined sheet with all enhanced parameters
        if len(hdf5_metadata) > 1:
            try:
                max_length = max(len(data) for data in hdf5_metadata.values())
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                combined_data = {"Time_minutes": time_minutes}

                for param_name, param_data in hdf5_metadata.items():
                    unit = unit_info.get(param_name, {}).get("units", "unknown")

                    # Pad data if necessary
                    if len(param_data) < max_length:
                        padded_data = list(param_data) + [np.nan] * (
                            max_length - len(param_data)
                        )
                    else:
                        padded_data = param_data

                    # Add with unit-enhanced name
                    if unit == "seconds" and "drift" in param_name.lower():
                        combined_data[f"{param_name}_seconds"] = padded_data
                        combined_data[f"{param_name}_ms"] = [
                            d * 1000 if d and not np.isnan(d) else np.nan
                            for d in padded_data
                        ]
                    else:
                        combined_data[f"{param_name}_{unit}"] = padded_data

                combined_df = pd.DataFrame(combined_data)
                combined_name = f"Enhanced_All_{source_name}"[:31]
                combined_df.to_excel(writer, sheet_name=combined_name, index=False)
                sheets_created.append(combined_name)

            except Exception as e:
                self._log_message(
                    f"   Warning: Could not create enhanced combined sheet: {e}"
                )

        return sheets_created

    def _filter_hdf5_metadata_only(self, ts_data: dict) -> dict:
        """Filter to keep only actual HDF5 metadata, excluding analysis results."""
        hdf5_metadata_only = {}

        # Only exclude specific analysis result patterns
        analysis_result_patterns = [
            "roi_",
            "baseline_",
            "threshold_",
            "upper_threshold",
            "lower_threshold",
            "movement_data",
            "fraction_data",
            "sleep_data",
            "quiescence_data",
            "intensity_roi_",
            "analysis_",
            "processed_",
            "calculated_",
        ]

        for param_name, param_data in ts_data.items():
            param_lower = param_name.lower()

            # Keep the parameter unless it matches specific analysis result patterns
            is_analysis_result = any(
                pattern in param_lower for pattern in analysis_result_patterns
            )

            if not is_analysis_result:
                hdf5_metadata_only[param_name] = param_data

        return hdf5_metadata_only

    def _create_timeseries_sheets_manually(
        self, writer, hdf5_metadata: dict, source_name: str
    ):
        """Manual fallback for creating time-series sheets."""
        import pandas as pd
        import numpy as np

        sheets_created = []

        for param_name, param_data in hdf5_metadata.items():
            try:
                if not hasattr(param_data, "__len__") or len(param_data) == 0:
                    continue

                # Create DataFrame with this parameter
                max_length = len(param_data)
                frame_interval = self.frame_interval.value()

                # Create time column aligned with analysis
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                # Create DataFrame
                df_data = {"Time (min)": time_minutes, param_name: param_data}

                param_df = pd.DataFrame(df_data)

                # Clean parameter name for Excel sheet (max 31 chars)
                clean_name = self._clean_sheet_name(param_name)

                # Ensure unique sheet name
                original_clean_name = clean_name
                counter = 1
                while clean_name in sheets_created:
                    clean_name = f"{original_clean_name[:28]}_{counter}"
                    counter += 1

                # Create the sheet
                param_df.to_excel(writer, sheet_name=clean_name, index=False)
                sheets_created.append(clean_name)
                self._log_message(
                    f"   ✓ Created manual sheet '{clean_name}' for {param_name}"
                )

            except Exception as e:
                self._log_message(
                    f"   Warning: Could not create sheet for {param_name}: {e}"
                )
                continue

        # Also create a combined sheet if we have multiple parameters
        if len(hdf5_metadata) > 1:
            try:
                # Find the maximum length across all parameters
                max_length = max(len(data) for data in hdf5_metadata.values())
                frame_interval = self.frame_interval.value()
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                # Create combined DataFrame
                df_data = {"Time (min)": time_minutes}

                for param_name, param_data in hdf5_metadata.items():
                    # Pad shorter series with NaN
                    if len(param_data) < max_length:
                        padded_data = list(param_data) + [np.nan] * (
                            max_length - len(param_data)
                        )
                    else:
                        padded_data = param_data
                    df_data[param_name] = padded_data

                combined_df = pd.DataFrame(df_data)
                combined_name = f"All_HDF5_{source_name}"[:31]
                combined_df.to_excel(writer, sheet_name=combined_name, index=False)
                sheets_created.append(combined_name)
                self._log_message(
                    f"   ✓ Created combined manual sheet '{combined_name}'"
                )

            except Exception as e:
                self._log_message(f"   Warning: Could not create combined sheet: {e}")

        return sheets_created

    def _clean_sheet_name(self, param_name: str) -> str:
        """Clean parameter name to be valid Excel sheet name."""
        # Remove or replace invalid characters
        clean = param_name.replace("/", "_").replace("\\", "_").replace(":", "_")
        clean = (
            clean.replace("*", "star")
            .replace("?", "q")
            .replace("[", "")
            .replace("]", "")
        )
        clean = clean.replace("<", "lt").replace(">", "gt").replace("|", "_")

        # Truncate to 31 characters max
        if len(clean) > 31:
            # Try to keep meaningful parts
            if "_" in clean:
                parts = clean.split("_")
                if len(parts[0]) <= 25:
                    clean = parts[0] + "_" + "".join(p[0] for p in parts[1:] if p)[:5]
                else:
                    clean = clean[:31]
            else:
                clean = clean[:31]

        # Remove trailing underscores and ensure not empty
        clean = clean.rstrip("_")
        if not clean:
            clean = "unnamed_param"

        return clean

    def _create_metadata_dataframe_manually(self, metadata: dict, source_name: str):
        """Manual fallback for creating metadata DataFrame."""
        import pandas as pd

        rows = []

        def flatten_metadata(d, prefix=""):
            for key, value in d.items():
                full_key = f"{prefix}.{key}" if prefix else key

                if isinstance(value, dict):
                    flatten_metadata(value, full_key)
                else:
                    try:
                        str_value = str(value)
                        data_type = type(value).__name__
                    except:
                        str_value = f"<{type(value).__name__}>"
                        data_type = "Complex"

                    rows.append(
                        {
                            "Category": prefix if prefix else "Root",
                            "Parameter": key,
                            "Value": str_value,
                            "Data_Type": data_type,
                            "Source": source_name,
                        }
                    )

        flatten_metadata(metadata)
        return pd.DataFrame(rows)

    def _create_individual_sheets_fallback(
        self, writer, hdf5_metadata: dict, sheets_created: list
    ):
        """Fallback method using new module functions if available."""
        try:
            from ._metadata import create_individual_timeseries_sheets

            new_sheets = create_individual_timeseries_sheets(
                writer, hdf5_metadata, self.frame_interval.value()
            )
            sheets_created.extend(new_sheets)
            self._log_message(f"   ✓ Created {len(new_sheets)} fallback sheets")
        except Exception as e:
            self._log_message(f"   ✗ Fallback method failed: {e}")

    def _create_metadata_sheet_fallback(
        self, writer, hdf5_metadata: dict, source_name: str
    ):
        """
        Fallback method to create HDF5 metadata sheet when import fails.
        """
        import pandas as pd
        import numpy as np

        try:
            if not hdf5_metadata:
                return

            # Find the maximum length
            max_length = max(
                len(data) if hasattr(data, "__len__") else 1
                for data in hdf5_metadata.values()
            )

            # Create time column
            frame_interval = self.frame_interval.value()
            time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

            # Build DataFrame
            df_data = {"Time (min)": time_minutes}

            for param_name, param_data in hdf5_metadata.items():
                if hasattr(param_data, "__len__") and len(param_data) > 0:
                    # Pad shorter series with NaN
                    padded_data = list(param_data) + [np.nan] * (
                        max_length - len(param_data)
                    )
                    df_data[param_name] = padded_data
                else:
                    # Single value or empty data
                    df_data[param_name] = (
                        [param_data] * max_length if max_length > 0 else []
                    )

            df = pd.DataFrame(df_data)
            sheet_name = f"HDF5_{source_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            self._log_message(f"   Created fallback metadata sheet '{sheet_name}'")

        except Exception as e:
            self._log_message(f"   Error in fallback metadata sheet creation: {e}")

    def _create_static_metadata_sheet_fallback(
        self, writer, static_metadata: dict, source_name: str
    ):
        """
        Fallback method to create static metadata sheet when import fails.
        """
        import pandas as pd

        try:
            rows = []

            def flatten_metadata(d, prefix=""):
                for key, value in d.items():
                    full_key = f"{prefix}.{key}" if prefix else key

                    if isinstance(value, dict):
                        flatten_metadata(value, full_key)
                    else:
                        rows.append(
                            {
                                "Category": prefix if prefix else "Root",
                                "Parameter": key,
                                "Value": str(value),
                                "Data_Type": type(value).__name__,
                                "Source": source_name,
                            }
                        )

            flatten_metadata(static_metadata)

            df = pd.DataFrame(rows)
            sheet_name = f"Static_{source_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            self._log_message(f"   Created fallback static sheet '{sheet_name}'")

        except Exception as e:
            self._log_message(f"   Error in fallback static sheet creation: {e}")

    def _get_analysis_parameters_for_metadata(self) -> dict:
        """Extract analysis parameters for metadata."""
        params = {
            "method": self._get_current_threshold_method_display(),
            "frame_interval_seconds": self.frame_interval.value(),
            "bin_size_seconds": self.bin_size_seconds.value(),
            "quiescence_threshold": self.quiescence_threshold.value(),
            "sleep_threshold_minutes": self.sleep_threshold_minutes.value(),
        }

        # Method-specific parameters
        if hasattr(self, "baseline_duration_minutes"):
            params["baseline_duration_minutes"] = self.baseline_duration_minutes.value()
        if hasattr(self, "threshold_multiplier"):
            params["threshold_multiplier"] = self.threshold_multiplier.value()
        if hasattr(self, "calibration_multiplier"):
            params["calibration_multiplier"] = self.calibration_multiplier.value()
        if hasattr(self, "enable_detrending"):
            params["detrending_enabled"] = self.enable_detrending.isChecked()
        if hasattr(self, "enable_jump_correction"):
            params["jump_correction_enabled"] = self.enable_jump_correction.isChecked()
        if hasattr(self, "adaptive_illumination_baseline"):
            params["adaptive_illumination_baseline"] = self.adaptive_illumination_baseline.isChecked()

        return params
