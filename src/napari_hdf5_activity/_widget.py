"""
widget.py - All methods properly placed within the HDF5AnalysisWidget class
"""

from datetime import datetime
import os
import cv2
import h5py
import time
import csv
import pandas as pd
import psutil
from typing import Dict, List, Tuple, Optional, Any
import numpy as np
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from napari.qt.threading import thread_worker
from qtpy.QtCore import QTimer, Signal, Qt, QSettings
from qtpy.QtWidgets import (
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QProgressBar,
    QPushButton,
    QSizePolicy,
    QSpinBox,
    QTabWidget,
    QVBoxLayout,
    QWidget,
    QTextEdit,
    QCheckBox,
    QSlider,
    QSplitter,
    QScrollArea,
    QTreeWidget,
    QTreeWidgetItem,
    QHeaderView,
    QListWidget,
    QListWidgetItem,
)

try:
    from ._reader import (
        detect_hdf5_structure_type,
        get_first_frame_enhanced,
        get_frame_norm_factor,
        process_single_file_in_parallel_dual_structure,
        process_hdf5_file_dual_structure,
        reader_function_dual_structure,
        # Keep original imports as fallback
        napari_get_reader,
        get_first_frame,
        get_roi_colors,
        merge_results,
        process_hdf5_files,
        sort_circles_left_to_right,
        sort_circles_meandering_auto,  # New function
    )

    DUAL_STRUCTURE_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Dual structure functions not available: {e}")
    DUAL_STRUCTURE_AVAILABLE = False
    # Use original imports only
    from ._reader import (
        napari_get_reader,
        get_first_frame,  # Should be available if left_to_right is
    )
try:
    from ._metadata import (
        extract_hdf5_metadata_timeseries,
        create_hdf5_metadata_timeseries_dataframe,
        write_metadata_to_csv,
        filter_hdf5_metadata_only,
    )

    METADATA_AVAILABLE = True

    # Try to import Nematostella functions separately
    try:
        from ._metadata import (
            analyze_nematostella_hdf5_file,
            NematostellaTimeseriesAnalyzer,
        )

        nematostella_analysis_available = True
    except ImportError:
        nematostella_analysis_available = False

except ImportError as e:
    print(f"Warning: Metadata functions not available: {e}")
    METADATA_AVAILABLE = False
    nematostella_analysis_available = False


def _fmt_p(p: float) -> str:
    """Format a p-value: scientific notation for p < 0.001, otherwise 4 decimal places."""
    if p < 0.001:
        return f"{p:.2e}"
    return f"{p:.4f}"


def _parse_recording_start_datetime(file_path: str):
    """Extract recording start datetime from filename pattern YYYYMMDD_HHMMSS.

    Returns a datetime object or None if the pattern is not found.
    Example: 'nematostella_timelapse_20260126_181210.h5' → datetime(2026,1,26,18,12,10)
    """
    import re
    from datetime import datetime as _dt
    name = os.path.basename(file_path)
    m = re.search(r'(\d{8})_(\d{6})', name)
    if m:
        try:
            return _dt.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")
        except ValueError:
            return None
    return None


# Import calculation functions with clear fallbacks
def validate_analysis_parameters(
    frame_interval: float, chunk_size: int, baseline_duration_minutes: float
) -> Tuple[bool, str]:
    """Validate analysis parameters before starting analysis."""
    if frame_interval <= 0:
        return False, "Frame interval must be positive"
    if chunk_size <= 0:
        return False, "Chunk size must be positive"
    if baseline_duration_minutes <= 0:
        return False, "Baseline duration must be positive"
    if baseline_duration_minutes > 100000:
        return False, "Baseline duration seems unreasonably long (>100000 minutes)"
    if frame_interval > 300:
        return False, "Frame interval seems unreasonably long (>5 minutes)"
    return True, ""


# Define bin_quiescence fallback function
def bin_quiescence_fallback(fraction_data, threshold):
    """Fallback quiescence calculation when main function not available."""
    quiescence_data = {}
    for roi, data in fraction_data.items():
        quiescence_data[roi] = [
            (t, 1 if fraction < threshold else 0) for t, fraction in data
        ]
    return quiescence_data


# Try integrated system first, then legacy, then fallbacks
CALC_SYSTEM = "none"
bin_quiescence = None  # Initialize as None

try:
    from ._calc_integration import (
        run_analysis_with_method,
        get_analysis_summary,
        quick_method_test,
        validate_hdf5_timing_in_data,
        export_results_for_matlab,
    )

    # Try to get bin_quiescence from _calc since it's not in _calc_integration
    try:
        from ._calc import bin_quiescence
    except ImportError:
        bin_quiescence = bin_quiescence_fallback
        print("Warning: Using fallback bin_quiescence function")

    try:
        if CALC_SYSTEM == "integrated":
            from ._calc_integration import get_performance_metrics
        else:
            from ._calc import get_performance_metrics
    except ImportError:

        def get_performance_metrics(start_time, total_frames):
            import time

            elapsed = time.time() - start_time
            return {
                "elapsed_time": elapsed,
                "fps": total_frames / elapsed if elapsed > 0 else 0,
                "cpu_percent": 0,
                "memory_percent": 0,
                "total_frames": total_frames,
            }

    def run_complete_hdf5_compatible_analysis(merged_results, **kwargs):
        method = kwargs.pop("threshold_method", "baseline")
        return run_analysis_with_method(merged_results, method, **kwargs)

    quick_analysis_test = quick_method_test
    CALC_SYSTEM = "integrated"
    print("Using integrated calculation system")

except ImportError:
    try:
        from ._calc import (
            run_complete_hdf5_compatible_analysis,
            get_analysis_summary,
            quick_analysis_test,
            validate_hdf5_timing_in_data,
            export_results_for_matlab,
            bin_quiescence,
        )

        CALC_SYSTEM = "legacy"
        print("Using legacy calculation system")

    except ImportError as e:
        print(f"Warning: No calculation system available: {e}")
        CALC_SYSTEM = "fallback"

        # Use fallback bin_quiescence
        bin_quiescence = bin_quiescence_fallback

        def run_complete_hdf5_compatible_analysis(merged_results, **kwargs):
            return {
                "method": "fallback",
                "baseline_means": {roi: 0.0 for roi in merged_results.keys()},
                "upper_thresholds": {roi: 1.0 for roi in merged_results.keys()},
                "lower_thresholds": {roi: -1.0 for roi in merged_results.keys()},
                "movement_data": {roi: [] for roi in merged_results.keys()},
                "fraction_data": {roi: [] for roi in merged_results.keys()},
                "sleep_data": {roi: [] for roi in merged_results.keys()},
                "quiescence_data": {roi: [] for roi in merged_results.keys()},
                "roi_statistics": {roi: {} for roi in merged_results.keys()},
                "error": "No calculation system available",
            }

        def get_analysis_summary(results):
            return "No calculation system available for summary"

        def quick_analysis_test(merged_results):
            return "No calculation system available for testing"

        def validate_hdf5_timing_in_data(merged_results, frame_interval=5.0):
            return {"timing_type": "unknown", "needs_correction": False}

        def export_results_for_matlab(results, output_dir):
            print("Export function not available - no calculation system")
            return []


# Import plotting functions
try:
    from ._plot import (
        PlotGenerator,
        create_plot_config,
        create_hysteresis_kwargs,
        save_plot,
        save_all_plot_types,
    )
except ImportError as e:
    print(f"Warning: Could not import plot functions: {e}")


class HDF5AnalysisWidget(QWidget):
    """
    Simplified widget for analyzing activity in HDF5 files.
    Coordinates between _calc.py and _plot.py modules.
    Handles only UI interactions and file operations.
    """

    # Qt Signals
    progress_updated = Signal(int)
    status_updated = Signal(str)
    performance_updated = Signal(str)

    def __init__(self, napari_viewer):
        super().__init__()
        self.viewer = napari_viewer

        # Initialize all attributes first
        self._initialize_attributes()

        # Setup UI after all attributes are initialized
        self.setup_ui()

        # Connect signals
        self._connect_signals()

        # Restore last ROI detection settings from previous session
        self._load_roi_settings()

    def _initialize_attributes(self):
        """Initialize all class attributes."""
        # Performance monitoring
        self.cpu_count = psutil.cpu_count()
        # Add dataset state management
        self._initialize_dataset_state()
        # Memory-aware process count for Windows (workers import napari/numba ~600MB each)
        if os.name == "nt":  # Windows
            try:
                available_mb = psutil.virtual_memory().available / (1024 * 1024)
                usable_mb = max(0, available_mb - 2048)  # Reserve 2GB for main process
                max_from_memory = max(1, int(usable_mb / 600))
                cpu_based = max(1, min(4, int(self.cpu_count * 0.6)))
                self.optimal_processes = min(cpu_based, max_from_memory)
            except Exception:
                self.optimal_processes = max(1, min(2, int(self.cpu_count * 0.5)))
        else:  # Unix/Linux/Mac
            self.optimal_processes = max(1, int(self.cpu_count * 0.9))

        # Analysis state variables
        self.directory: Optional[str] = None
        self.file_path: Optional[str] = None
        self.recording_start_datetime = None  # parsed from filename YYYYMMDD_HHMMSS
        self.masks: List[np.ndarray] = []
        self.labeled_frame: Optional[np.ndarray] = None

        # Analysis results (now populated by _calc.py)
        self.merged_results: Dict[int, List[Tuple[float, float]]] = {}
        self.roi_colors: Dict[int, str] = {}
        self.roi_thresholds: Dict[int, float] = {}
        self.roi_statistics: Dict[int, Dict[str, float]] = {}
        self.movement_data: Dict[int, List[Tuple[float, int]]] = {}
        self.fraction_data: Dict[int, List[Tuple[float, float]]] = {}
        self.quiescence_data: Dict[int, List[Tuple[float, int]]] = {}
        self.sleep_data: Dict[int, List[Tuple[float, int]]] = {}

        # Hysteresis data (populated by _calc.py)
        self.roi_baseline_means: Dict[int, float] = {}
        self.roi_upper_thresholds: Dict[int, float] = {}
        self.roi_lower_thresholds: Dict[int, float] = {}
        self.roi_band_widths: Dict[int, float] = {}

        # Worker handle for background analysis
        self.current_worker = None
        self._cancel_requested = False
        self.analysis_start_time: Optional[float] = None

        # Initialize performance timer
        self.performance_timer = QTimer()
        self.performance_timer.timeout.connect(self._update_performance_metrics)
        self.performance_timer.setInterval(1000)  # Update every second

        # Plot generator (initialized when figure is available)
        self.plot_generator = None

        # Calibration workflow state (add these lines)
        self.current_dataset_type = "main"  # "main" or "calibration"
        self.calibration_file_path_stored = None
        self.calibration_masks = []
        self.calibration_labeled_frame = None
        self.calibration_baseline_processed = False
        self.calibration_baseline_statistics = {}

        # Store main dataset state when switching to calibration
        self.main_dataset_path = None
        self.main_masks = []
        self.main_labeled_frame = None

    def _initialize_dataset_state(self):
        """Initialize dataset state management attributes."""
        self.main_dataset_path = None
        self.main_merged_results = {}
        self.main_masks = []
        self.main_labeled_frame = None
        self.main_dataset_stored = False

        self.calibration_file_path_stored = None
        self.calibration_masks = []
        self.calibration_labeled_frame = None
        self.calibration_baseline_processed = False
        self.calibration_baseline_statistics = {}

        self.current_dataset_type = "main"

    def store_main_dataset_state(self):
        """Store the current main dataset state before calibration operations."""
        try:
            self._log_message("Storing main dataset state...")

            # Check if we have valid main dataset to store
            if not hasattr(self, "file_path") or not self.file_path:
                self._log_message("WARNING: No file_path to store as main dataset")
                return False

            # IMPORTANT: Check if we have processed results OR if file is loaded
            if hasattr(self, "merged_results") and self.merged_results:
                # Case 1: We have processed data (ideal case)
                self.main_dataset_path = self.file_path
                self.main_merged_results = self.merged_results.copy()
                self.main_masks = getattr(self, "masks", []).copy()
                self.main_labeled_frame = getattr(self, "labeled_frame", None)
                self.main_dataset_stored = True

                # Verify storage
                sample_roi = list(self.main_merged_results.keys())[0]
                sample_data = self.main_merged_results[sample_roi]
                if sample_data:
                    main_duration = (sample_data[-1][0] - sample_data[0][0]) / 60
                    self._log_message(
                        "Main dataset stored successfully (PROCESSED DATA):"
                    )
                    self._log_message(
                        f"   Path: {os.path.basename(self.main_dataset_path)}"
                    )
                    self._log_message(f"   ROIs: {len(self.main_merged_results)}")
                    self._log_message(f"   Duration: {main_duration:.1f} minutes")
                    self._log_message(f"   Data points: {len(sample_data)}")
                    return True

            elif (
                hasattr(self, "file_path")
                and self.file_path
                and os.path.exists(self.file_path)
            ):
                # Case 2: We have a file loaded but no processed data yet
                self._log_message("Main dataset file loaded but not yet processed")
                self._log_message(
                    "Storing file path and current state for later processing"
                )

                self.main_dataset_path = self.file_path
                self.main_merged_results = (
                    {}
                )  # Empty for now, will be filled during analysis
                self.main_masks = getattr(self, "masks", []).copy()
                self.main_labeled_frame = getattr(self, "labeled_frame", None)
                self.main_dataset_stored = True

                self._log_message("Main dataset file stored (NOT YET PROCESSED):")
                self._log_message(
                    f"   Path: {os.path.basename(self.main_dataset_path)}"
                )
                self._log_message(f"   Masks: {len(self.main_masks)}")
                self._log_message("   Data will be processed during analysis")

                return True

            else:
                self._log_message("ERROR: No valid main dataset file or data to store")
                return False

        except Exception as e:
            self._log_message(f"ERROR storing main dataset: {e}")
            self.main_dataset_stored = False
            return False

    def restore_main_dataset_for_analysis(self):
        """Restore main dataset before running analysis."""
        self._log_message("=== RESTORING MAIN DATASET FOR ANALYSIS ===")

        # Check if we have stored main dataset
        if not hasattr(self, "main_dataset_stored") or not self.main_dataset_stored:
            self._log_message("ERROR: No main dataset was stored")
            return False

        # Check if stored data is valid
        if not hasattr(self, "main_dataset_path") or not self.main_dataset_path:
            self._log_message("ERROR: No main dataset path stored")
            return False

        try:
            # Restore main dataset state
            self.file_path = self.main_dataset_path
            self.current_dataset_type = "main"

            # Case 1: We have processed data to restore
            if hasattr(self, "main_merged_results") and self.main_merged_results:
                self.merged_results = self.main_merged_results.copy()
                sample_roi = list(self.merged_results.keys())[0]
                sample_data = self.merged_results[sample_roi]
                restored_duration = (sample_data[-1][0] - sample_data[0][0]) / 60

                self._log_message("Main dataset restored (PROCESSED DATA):")
                self._log_message(f"   Path: {os.path.basename(self.file_path)}")
                self._log_message(f"   ROIs: {len(self.merged_results)}")
                self._log_message(f"   Duration: {restored_duration:.1f} minutes")

            # Case 2: We only have file path, need to ensure data gets loaded
            else:
                self._log_message(
                    "Main dataset file restored (WILL PROCESS DURING ANALYSIS):"
                )
                self._log_message(f"   Path: {os.path.basename(self.file_path)}")
                self._log_message("   Data will be loaded/processed during analysis")

                # Ensure merged_results is available for analysis
                if not hasattr(self, "merged_results") or not self.merged_results:
                    self._log_message(
                        "   No processed data available - analysis will process from file"
                    )

            # Restore masks and labeled frame
            if hasattr(self, "main_masks"):
                self.masks = self.main_masks.copy()
            if hasattr(self, "main_labeled_frame"):
                self.labeled_frame = self.main_labeled_frame

            # Update UI to reflect main dataset
            self.lbl_file_info.setText(
                f"MAIN DATASET: {os.path.basename(self.file_path)}"
            )

            return True

        except Exception as e:
            self._log_message(f"ERROR restoring main dataset: {e}")
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")
            return False

    def setup_ui(self):
        """Setup the user interface with all tabs."""
        # Create main layout
        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        # Create scroll area to ensure GUI fits in window
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        # Create container widget for tab widget
        container = QWidget()
        container_layout = QVBoxLayout()
        container.setLayout(container_layout)

        # Create tab widget
        self.tab_widget = QTabWidget()
        self.tab_input = QWidget()
        self.tab_analysis = QWidget()
        self.tab_results = QWidget()
        self.tab_extended = QWidget()
        self.tab_viewer = QWidget()
        self.tab_telemetry = QWidget()

        self.tab_widget.addTab(self.tab_input, "Input")
        self.tab_widget.addTab(self.tab_analysis, "Analysis")
        self.tab_widget.addTab(self.tab_results, "Results")
        self.tab_widget.addTab(self.tab_extended, "Extended Analysis")
        self.tab_widget.addTab(self.tab_viewer, "Frame Viewer")
        self.tab_widget.addTab(self.tab_telemetry, "HDF5 Telemetry")

        container_layout.addWidget(self.tab_widget)

        # Add container to scroll area
        scroll_area.setWidget(container)

        # Add scroll area to main layout
        main_layout.addWidget(scroll_area)

        # Setup individual tabs
        self.setup_input_tab()
        self.setup_analysis_tab()
        self.setup_results_tab()
        self.setup_extended_tab()
        self.setup_viewer_tab()
        self.setup_telemetry_tab()

    def setup_input_tab(self):
        """Setup the input tab with file loading and ROI detection parameters."""
        layout = QVBoxLayout()
        self.tab_input.setLayout(layout)

        # File loading section
        file_group = QGroupBox("Load Data")
        file_layout = QVBoxLayout()
        file_group.setLayout(file_layout)
        # Debug-Button hinzufügen
        self.btn_debug_structure = QPushButton("Debug HDF5 Structure")
        self.btn_debug_structure.setToolTip("Analyze HDF5 file structure")
        self.btn_debug_structure.clicked.connect(self.debug_current_file_structure)
        file_layout.addWidget(self.btn_debug_structure)
        # File loading buttons
        self.btn_load_file = QPushButton("Load File")
        self.btn_load_file.setToolTip("Load HDF5 file or AVI video(s) for analysis")

        self.btn_load_dir = QPushButton("Load Directory")
        self.btn_load_dir.setToolTip("Load all HDF5/AVI files from a directory")

        self.btn_detect_rois = QPushButton("Detect ROIs")
        self.btn_detect_rois.setToolTip(
            "Automatically detect circular ROIs using HoughCircles"
        )

        self.btn_clear_rois = QPushButton("Clear ROI Detection")
        self.btn_clear_rois.setToolTip("Remove ROI detection layers")

        file_layout.addWidget(self.btn_load_file)
        file_layout.addWidget(self.btn_load_dir)
        file_layout.addWidget(self.btn_detect_rois)
        file_layout.addWidget(self.btn_clear_rois)

        self.lbl_file_info = QLabel("No file loaded")
        file_layout.addWidget(self.lbl_file_info)
        layout.addWidget(file_group)

        # ROI Detection Parameters
        roi_group = QGroupBox("ROI Detection Parameters")
        roi_layout = QFormLayout()
        roi_group.setLayout(roi_layout)

        # Radius parameters (defaults for 6-well plate)
        self.min_radius = QSpinBox()
        self.min_radius.setRange(10, 1000)
        self.min_radius.setValue(100)
        self.min_radius.setToolTip("Minimum radius for circle detection (pixels)")
        roi_layout.addRow("Min Radius:", self.min_radius)

        self.max_radius = QSpinBox()
        self.max_radius.setRange(10, 1000)
        self.max_radius.setValue(145)
        self.max_radius.setToolTip("Maximum radius for circle detection (pixels)")
        roi_layout.addRow("Max Radius:", self.max_radius)

        # HoughCircles parameters
        self.dp_param = QDoubleSpinBox()
        self.dp_param.setRange(0.1, 5.0)
        self.dp_param.setValue(1.0)
        self.dp_param.setSingleStep(0.1)
        self.dp_param.setDecimals(1)
        self.dp_param.setToolTip(
            "Inverse ratio of accumulator resolution (1.0 recommended)"
        )
        roi_layout.addRow("DP Parameter:", self.dp_param)

        self.min_dist = QSpinBox()
        self.min_dist.setRange(10, 1000)
        self.min_dist.setValue(300)
        self.min_dist.setToolTip("Minimum distance between circle centers (pixels)")
        roi_layout.addRow("Min Distance:", self.min_dist)

        self.param1 = QSpinBox()
        self.param1.setRange(10, 300)
        self.param1.setValue(30)
        self.param1.setToolTip("Canny edge detection threshold (higher = fewer edges)")
        roi_layout.addRow("Param1 (Edge):", self.param1)

        self.param2 = QSpinBox()
        self.param2.setRange(5, 200)
        self.param2.setValue(60)
        self.param2.setToolTip("Circle detection sensitivity (lower = more circles)")
        roi_layout.addRow("Param2 (Center):", self.param2)

        # Plate presets
        self.chk_6well = QCheckBox("6-Well Plate Preset")
        self.chk_6well.setToolTip(
            "Use preset values for 6-well plates (radius 100-145)"
        )
        self.chk_6well.setChecked(True)
        roi_layout.addRow("", self.chk_6well)

        self.chk_12well = QCheckBox("12-Well Plate Preset")
        self.chk_12well.setToolTip(
            "Use preset values for 12-well plates (radius 70-120)"
        )
        roi_layout.addRow("", self.chk_12well)

        # ROI Scale slider - scale detected ROIs from center
        self.roi_scale = QDoubleSpinBox()
        self.roi_scale.setRange(0.1, 2.0)
        self.roi_scale.setValue(1.0)
        self.roi_scale.setSingleStep(0.05)
        self.roi_scale.setToolTip(
            "Scale ROIs from center (1.0 = original size, 0.8 = 80%)"
        )
        roi_layout.addRow("ROI Scale:", self.roi_scale)

        # Apply scale button
        self.btn_apply_scale = QPushButton("Apply Scale")
        self.btn_apply_scale.setToolTip("Re-scale detected ROIs without re-detecting")
        self.btn_apply_scale.setEnabled(False)
        roi_layout.addRow("", self.btn_apply_scale)

        layout.addWidget(roi_group)

        # ROI Selection (exclude/include)
        roi_select_group = QGroupBox("ROI Selection (exclude empty wells)")
        self.roi_select_layout = QVBoxLayout()
        roi_select_group.setLayout(self.roi_select_layout)

        self.roi_select_info = QLabel("Run ROI detection first")
        self.roi_select_info.setStyleSheet("color: #7f8c8d; font-size: 10px;")
        self.roi_select_layout.addWidget(self.roi_select_info)

        # Buttons for quick select/deselect
        roi_btn_layout = QHBoxLayout()
        self.btn_select_all_rois = QPushButton("Select All")
        self.btn_deselect_all_rois = QPushButton("Deselect All")
        self.btn_select_all_rois.clicked.connect(lambda: self._set_all_roi_checkboxes(True))
        self.btn_deselect_all_rois.clicked.connect(lambda: self._set_all_roi_checkboxes(False))
        roi_btn_layout.addWidget(self.btn_select_all_rois)
        roi_btn_layout.addWidget(self.btn_deselect_all_rois)
        roi_btn_layout.addStretch()
        self.roi_select_layout.addLayout(roi_btn_layout)

        # Container for dynamic checkboxes
        self.roi_checkboxes_layout = QHBoxLayout()
        self.roi_select_layout.addLayout(self.roi_checkboxes_layout)
        self.roi_checkboxes: list = []

        layout.addWidget(roi_select_group)
        layout.addStretch()

    def _populate_roi_checkboxes(self, n_rois: int):
        """Create checkboxes for each detected ROI."""
        # Clear existing
        for cb in self.roi_checkboxes:
            cb.setParent(None)
        self.roi_checkboxes.clear()

        # Create new checkboxes
        for i in range(n_rois):
            cb = QCheckBox(f"ROI {i + 1}")
            cb.setChecked(True)
            cb.setToolTip(f"Include ROI {i + 1} in analysis")
            self.roi_checkboxes_layout.addWidget(cb)
            self.roi_checkboxes.append(cb)

        self.roi_select_info.setText(f"{n_rois} ROIs detected — uncheck to exclude from analysis")

    def _set_all_roi_checkboxes(self, checked: bool):
        """Select or deselect all ROI checkboxes."""
        for cb in self.roi_checkboxes:
            cb.setChecked(checked)

    def _get_excluded_roi_indices(self) -> list:
        """Return list of ROI indices (0-based) that are unchecked."""
        return [i for i, cb in enumerate(self.roi_checkboxes) if not cb.isChecked()]

    def _get_active_masks(self) -> list:
        """Return only the masks for checked (included) ROIs."""
        if not self.roi_checkboxes:
            return self.masks  # No checkboxes = use all
        return [m for i, m in enumerate(self.masks) if i < len(self.roi_checkboxes) and self.roi_checkboxes[i].isChecked()]

    def _get_roi_index_mapping(self) -> dict:
        """Return mapping from sequential reader ROI index (1-based) to original ROI index.

        When ROIs are excluded, the reader assigns sequential indices 1..N to the
        N active masks.  This mapping restores the original ROI numbering.
        Example: if ROI 1 is excluded from 6 ROIs, returns {1:2, 2:3, 3:4, 4:5, 5:6}.
        If no ROIs are excluded, returns identity mapping {1:1, 2:2, ...}.
        """
        if not self.roi_checkboxes:
            return {i + 1: i + 1 for i in range(len(self.masks))}
        active_original_indices = [
            i for i, cb in enumerate(self.roi_checkboxes) if cb.isChecked()
        ]
        # reader_idx is 1-based, original is also 1-based (i+1)
        return {
            reader_idx: orig_0based + 1
            for reader_idx, orig_0based in enumerate(active_original_indices, start=1)
        }

    def setup_analysis_tab(self):
        """Setup the analysis tab with threshold calculation methods."""
        layout = QVBoxLayout()
        self.tab_analysis.setLayout(layout)

        # Basic Analysis Parameters
        analysis_group = QGroupBox("Basic Analysis Parameters")
        analysis_layout = QFormLayout()
        analysis_group.setLayout(analysis_layout)

        self.frame_interval = QDoubleSpinBox()
        self.frame_interval.setRange(0.01, 60.0)
        self.frame_interval.setValue(5.0)
        self.frame_interval.setSingleStep(0.1)
        self.frame_interval.setToolTip("Time interval between frames in seconds")
        analysis_layout.addRow("Frame Interval (s):", self.frame_interval)

        self.time_end = QSpinBox()
        self.time_end.setRange(0, 1000000)
        self.time_end.setValue(0)
        self.time_end.setToolTip("End time for analysis (0 = use full duration)")
        analysis_layout.addRow("End Time (s):", self.time_end)

        self.chunk_size = QSpinBox()
        self.chunk_size.setRange(1, 10000)
        self.chunk_size.setValue(20)
        self.chunk_size.setToolTip("Number of frames to process in each chunk")
        analysis_layout.addRow("Chunk Size:", self.chunk_size)

        self.num_processes = QSpinBox()
        self.num_processes.setRange(1, self.cpu_count)
        self.num_processes.setValue(self.optimal_processes)
        self.num_processes.setToolTip(
            f"Number of parallel processes (recommended: {self.optimal_processes})"
        )
        analysis_layout.addRow("Number of Processes:", self.num_processes)

        layout.addWidget(analysis_group)

        # Threshold Calculation Methods
        threshold_group = QGroupBox("Threshold Calculation Method")
        threshold_layout = QVBoxLayout()
        threshold_group.setLayout(threshold_layout)

        # Method-specific parameters (Tabs control the method selection)
        self.threshold_params_stack = QTabWidget()
        threshold_layout.addWidget(self.threshold_params_stack)

        # === METHOD 1: BASELINE ===
        baseline_tab = QWidget()
        baseline_layout = QFormLayout()
        baseline_tab.setLayout(baseline_layout)

        self.baseline_duration_minutes = QDoubleSpinBox()
        self.baseline_duration_minutes.setRange(1.0, 10000000000.0)
        self.baseline_duration_minutes.setValue(200.0)
        self.baseline_duration_minutes.setSingleStep(1.0)
        self.baseline_duration_minutes.setDecimals(1)
        self.baseline_duration_minutes.setToolTip(
            "Duration of baseline period in minutes"
        )
        baseline_layout.addRow(
            "Baseline Duration (min):", self.baseline_duration_minutes
        )

        self.threshold_multiplier = QDoubleSpinBox()
        self.threshold_multiplier.setRange(0.0, 5.0)
        self.threshold_multiplier.setValue(0.1)
        self.threshold_multiplier.setSingleStep(0.1)
        self.threshold_multiplier.setToolTip(
            "Multiplier for hysteresis band (mean ± multiplier × std)"
        )
        baseline_layout.addRow("Threshold Multiplier:", self.threshold_multiplier)

        # Add detrending option
        self.enable_detrending = QCheckBox("Enable Detrending")
        self.enable_detrending.setChecked(False)
        self.enable_detrending.setToolTip(
            "Remove linear drift from baseline period for more accurate thresholds"
        )
        baseline_layout.addRow("", self.enable_detrending)

        # Add jump correction option
        self.enable_jump_correction = QCheckBox("Enable Jump Correction")
        self.enable_jump_correction.setChecked(False)
        self.enable_jump_correction.setToolTip(
            "Detect and correct sudden jumps/plateaus in baseline data"
        )
        baseline_layout.addRow("", self.enable_jump_correction)

        baseline_info = QLabel(
            "HYSTERESIS METHOD:\n"
            "Uses hysteresis band to prevent flicker.\n"
            "Signal > Upper → Movement = TRUE\n"
            "Signal < Lower → Movement = FALSE\n"
            "Signal between → State unchanged"
        )
        baseline_info.setStyleSheet("color: #666; font-size: 10px;")
        baseline_info.setWordWrap(True)
        baseline_layout.addRow("", baseline_info)

        self.threshold_params_stack.addTab(baseline_tab, "Baseline Method")

        # === METHOD 2: CALIBRATION ===
        calibration_tab = QWidget()
        calibration_layout = QFormLayout()
        calibration_tab.setLayout(calibration_layout)

        # Calibration file selection (existing code)
        cal_file_layout = QHBoxLayout()
        self.calibration_file_path = QLabel("No calibration file selected")
        self.calibration_file_path.setStyleSheet(
            """
            QLabel {
                border: 1px solid #ccc;
                padding: 4px;
                background: #f9f9f9;
                color: #000000;  /* Force black text */
            }
        """
        )
        self.btn_load_calibration = QPushButton("Browse...")
        cal_file_layout.addWidget(self.calibration_file_path, 3)
        cal_file_layout.addWidget(self.btn_load_calibration, 1)
        calibration_layout.addRow("Calibration File:", cal_file_layout)

        # Calibration multiplier (existing code)
        self.calibration_multiplier = QDoubleSpinBox()
        self.calibration_multiplier.setRange(0.01, 5.00)
        self.calibration_multiplier.setValue(1.00)
        self.calibration_multiplier.setSingleStep(0.01)
        self.calibration_multiplier.setDecimals(2)
        self.calibration_multiplier.setToolTip(
            "Multiplier applied to calibration std (same as baseline multiplier)"
        )
        calibration_layout.addRow(
            "Calibration Multiplier:", self.calibration_multiplier
        )

        # NEW: Calibration dataset processing controls
        cal_processing_layout = QVBoxLayout()
        # Load calibration dataset button
        self.btn_load_calibration_dataset = QPushButton("Load Calibration Dataset")
        self.btn_load_calibration_dataset.setToolTip(
            "Load selected calibration file into viewer for ROI detection"
        )
        self.btn_load_calibration_dataset.setStyleSheet(
            "QPushButton { background-color: #2196F3; color: white; font-weight: bold; }"
        )
        self.btn_load_calibration_dataset.setEnabled(
            False
        )  # Enabled when file is selected
        # Process calibration baseline button
        self.btn_process_calibration_baseline = QPushButton(
            "Process Calibration Baseline"
        )
        self.btn_process_calibration_baseline.setToolTip(
            "Process full calibration dataset to create baseline statistics"
        )
        self.btn_process_calibration_baseline.setStyleSheet(
            "QPushButton { background-color: #FF9800; color: white; font-weight: bold; }"
        )
        self.btn_process_calibration_baseline.setEnabled(
            False
        )  # Enabled when calibration ROIs detected

        cal_processing_layout.addWidget(self.btn_load_calibration_dataset)
        cal_processing_layout.addWidget(self.btn_process_calibration_baseline)

        calibration_layout.addRow("Calibration Processing:", cal_processing_layout)
        # NEW: Calibration status display
        self.calibration_status_label = QLabel(
            "1. Select calibration file\n2. Load calibration dataset\n3. Detect ROIs (Input tab)\n4. Process baseline"
        )
        self.calibration_status_label.setStyleSheet(
            """
            QLabel {
                padding: 8px;
                background-color: #f5f5f5;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 10px;
                color: #000000;  /* Force black text */
            }
        """
        )
        self.calibration_status_label.setWordWrap(True)
        calibration_layout.addRow("Status:", self.calibration_status_label)
        # Updated info text
        calibration_info = QLabel(
            "CALIBRATION METHOD:\n"
            "Uses sedated animals to determine noise baseline.\n"
            "Calculates: mean ± multiplier × std from complete calibration dataset.\n"
            "Same hysteresis formula as baseline method."
        )
        calibration_info.setStyleSheet("color: #666; font-size: 10px;")
        calibration_info.setWordWrap(True)
        calibration_layout.addRow("", calibration_info)

        self.threshold_params_stack.addTab(calibration_tab, "Calibration Method")

        # === METHOD 3: ADAPTIVE ===
        adaptive_tab = QWidget()
        adaptive_layout = QFormLayout()
        adaptive_tab.setLayout(adaptive_layout)

        self.adaptive_duration_minutes = QDoubleSpinBox()
        self.adaptive_duration_minutes.setRange(5.0, 120.0)
        self.adaptive_duration_minutes.setValue(15.0)
        self.adaptive_duration_minutes.setSingleStep(1.0)
        self.adaptive_duration_minutes.setDecimals(1)
        self.adaptive_duration_minutes.setToolTip(
            "Duration of initial period for adaptive analysis"
        )
        adaptive_layout.addRow(
            "Analysis Duration (min):", self.adaptive_duration_minutes
        )

        self.adaptive_base_multiplier = QDoubleSpinBox()
        self.adaptive_base_multiplier.setRange(1.0, 5.0)
        self.adaptive_base_multiplier.setValue(2.5)
        self.adaptive_base_multiplier.setSingleStep(0.1)
        self.adaptive_base_multiplier.setToolTip(
            "Base multiplier for adaptive calculation"
        )
        adaptive_layout.addRow("Base Multiplier:", self.adaptive_base_multiplier)

        adaptive_info = QLabel(
            "Automatically adapts threshold based on signal-to-noise ratio."
        )
        adaptive_info.setStyleSheet("color: #666; font-size: 10px;")
        adaptive_info.setWordWrap(True)
        adaptive_layout.addRow("", adaptive_info)

        self.threshold_params_stack.addTab(adaptive_tab, "Adaptive Method")

        layout.addWidget(threshold_group)

        # Behavior Analysis Parameters
        behavior_group = QGroupBox("Behavior Analysis Parameters")
        behavior_layout = QFormLayout()
        behavior_group.setLayout(behavior_layout)

        self.bin_size_seconds = QSpinBox()
        self.bin_size_seconds.setRange(1, 300)
        self.bin_size_seconds.setValue(60)
        self.bin_size_seconds.setToolTip(
            "Bin size for fraction movement (60s recommended)"
        )
        behavior_layout.addRow("Bin Size (seconds):", self.bin_size_seconds)

        self.quiescence_threshold = QDoubleSpinBox()
        self.quiescence_threshold.setRange(0.0, 1.0)
        self.quiescence_threshold.setValue(0.5)
        self.quiescence_threshold.setSingleStep(0.1)
        self.quiescence_threshold.setToolTip(
            "Quiescence threshold for fraction movement:\n"
            "• fraction_movement < threshold → Quiescence = YES (resting)\n"
            "• fraction_movement ≥ threshold → Quiescence = NO (active)\n"
            "Default: 0.5 (less than 50% movement = quiescent)"
        )
        behavior_layout.addRow("Quiescence Threshold:", self.quiescence_threshold)

        self.sleep_threshold_minutes = QSpinBox()
        self.sleep_threshold_minutes.setRange(1, 60)
        self.sleep_threshold_minutes.setValue(8)
        self.sleep_threshold_minutes.setToolTip(
            "Minimum sleep duration in minutes (8 recommended)"
        )
        behavior_layout.addRow("Sleep Threshold (min):", self.sleep_threshold_minutes)

        layout.addWidget(behavior_group)

        # Analysis Control Section
        control_group = QGroupBox("Analysis Control")
        control_layout = QVBoxLayout()
        control_group.setLayout(control_layout)
        reset_layout = QHBoxLayout()
        self.btn_reset_analysis = QPushButton("Reset for New Analysis")
        self.btn_reset_analysis.setToolTip(
            "Clear all data and reset for a new analysis"
        )
        self.btn_reset_analysis.setStyleSheet(
            "QPushButton { background-color: #FF5722; color: white; font-weight: bold; }"
        )
        reset_layout.addWidget(self.btn_reset_analysis)
        control_layout.addLayout(reset_layout)
        # Analysis buttons
        btn_layout = QHBoxLayout()
        self.btn_analyze = QPushButton("Start Analysis")
        self.btn_analyze.setToolTip("Start the analysis with current parameters")
        self.btn_analyze.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }"
        )

        self.btn_stop = QPushButton("Stop Analysis")
        self.btn_stop.setToolTip("Stop the current analysis")
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet(
            "QPushButton { background-color: #f44336; color: white; font-weight: bold; }"
        )

        btn_layout.addWidget(self.btn_analyze)
        btn_layout.addWidget(self.btn_stop)
        control_layout.addLayout(btn_layout)

        # Testing and diagnostics buttons
        test_layout = QHBoxLayout()
        self.btn_quick_test = QPushButton("Quick Test")
        self.btn_quick_test.setToolTip("Run quick analysis test using _calc.py")
        self.btn_validate_timing = QPushButton("Validate HDF5 Timing")
        self.btn_validate_timing.setToolTip("Check HDF5 timing using _calc.py")

        test_layout.addWidget(self.btn_quick_test)
        test_layout.addWidget(self.btn_validate_timing)
        control_layout.addLayout(test_layout)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        control_layout.addWidget(self.progress_bar)
        # Remove extra spacing
        control_layout.setSpacing(0)
        control_layout.setContentsMargins(0, 0, 0, 0)
        # Status label
        self.status_label = QLabel("Ready to start analysis")
        self.status_label.setStyleSheet(
            "QLabel { padding: 5px; background-color: #2b2b2b; border: 1px solid #555; color: #ffffff; }"
        )
        control_layout.addWidget(self.status_label)

        # Performance metrics label
        self.performance_label = QLabel(
            "Performance metrics will appear here during analysis"
        )
        self.performance_label.setStyleSheet(
            "QLabel { padding: 3px; font-size: 10px; color: #FFFFFF; }"
        )
        control_layout.addWidget(self.performance_label)

        layout.addWidget(control_group)

        # Analysis Log Section
        log_group = QGroupBox("Analysis Log")
        log_layout = QVBoxLayout()
        log_group.setLayout(log_layout)

        self.log_text = QTextEdit()
        self.log_text.setMaximumHeight(150)
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet(
            """
            QTextEdit {
                background-color: #000000;
                color: #ffffff;
                font-family: 'Courier New', monospace;
                font-size: 9px;
            }
        """
        )
        log_layout.addWidget(self.log_text)

        layout.addWidget(log_group)

    def setup_results_tab(self):
        """Setup the results tab with plotting and export options."""
        layout = QVBoxLayout()
        self.tab_results.setLayout(layout)

        self.results_label = QLabel("Results will be displayed here.")
        layout.addWidget(self.results_label)

        # Matplotlib figure
        self.figure = Figure(figsize=(10, 6))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)

        try:
            from ._plot import PlotGenerator

            self.plot_generator = PlotGenerator(self.figure)
            self._log_message("✅ Plot generator initialized")
        except Exception as e:
            self.plot_generator = None
            self._log_message(f"⚠️ Plot generator initialization failed: {e}")

        # Threshold Visualization Options
        viz_group = QGroupBox("Threshold Visualization (for Raw Intensity Plots)")
        viz_layout = QFormLayout()
        viz_group.setLayout(viz_layout)

        self.show_baseline_mean = QCheckBox("Show Baseline Mean Line")
        self.show_baseline_mean.setChecked(True)
        self.show_baseline_mean.setToolTip(
            "Show the baseline mean from analysis (red line)"
        )
        viz_layout.addRow("", self.show_baseline_mean)

        self.show_deviation_band = QCheckBox("Show Deviation Band (Hysteresis Zone)")
        self.show_deviation_band.setChecked(True)
        self.show_deviation_band.setToolTip(
            "Show ±σ band around baseline mean (orange area)"
        )
        viz_layout.addRow("", self.show_deviation_band)

        self.show_detection_threshold = QCheckBox("Show Detection Thresholds")
        self.show_detection_threshold.setChecked(True)
        self.show_detection_threshold.setToolTip(
            "Show upper/lower detection boundaries (dashed lines)"
        )
        viz_layout.addRow("", self.show_detection_threshold)

        self.show_threshold_stats = QCheckBox("Show Threshold Statistics")
        self.show_threshold_stats.setChecked(True)
        self.show_threshold_stats.setToolTip(
            "Show threshold calculation details on plot"
        )
        viz_layout.addRow("", self.show_threshold_stats)

        # INFO BOX - SCIENTIFIC EXPLANATION
        baseline_info = QLabel(
            "BASELINE REFERENCE:\n"
            "• Baseline Mean = Fixed reference from analysis baseline period\n"
            "• Detection Thresholds = Used in actual movement detection\n"
            "• These values NEVER change with time range selection\n"
            "• They represent the analysis parameters, not visible data statistics"
        )
        baseline_info.setStyleSheet(
            "color: #0066cc; font-size: 9px; background: #f0f8ff; "
            "padding: 8px; border: 1px solid #ccc; border-radius: 4px;"
        )
        baseline_info.setWordWrap(True)
        viz_layout.addRow("", baseline_info)

        layout.addWidget(viz_group)

        # Plot configuration
        plot_config_group = QGroupBox("Plot Configuration")
        plot_config_layout = QVBoxLayout()
        plot_config_group.setLayout(plot_config_layout)

        # Plot type and basic controls
        basic_row = QHBoxLayout()

        self.plot_type_combo = QComboBox()
        self.plot_type_combo.addItems(
            [
                "Raw Intensity Changes",
                "Movement",
                "Fraction Movement",
                "Quiescence",
                "Sleep",
                "Sleep Quality",
                "Lighting Conditions (dark IR)",
            ]
        )

        self.plot_dpi_spin = QSpinBox()
        self.plot_dpi_spin.setRange(50, 600)
        self.plot_dpi_spin.setValue(100)

        basic_row.addWidget(QLabel("Plot Type:"))
        basic_row.addWidget(self.plot_type_combo)

        # Sleep Quality metric sub-selector (visible only when Sleep Quality selected)
        self.sleep_quality_metric_combo = QComboBox()
        self.sleep_quality_metric_combo.addItems([
            "Sleep min/h",
            "Transitions/h",
            "Bout Length/h",
        ])
        self.sleep_quality_metric_combo.setVisible(False)
        self.sleep_quality_metric_combo.currentIndexChanged.connect(lambda _: self.generate_plot())
        basic_row.addWidget(self.sleep_quality_metric_combo)

        basic_row.addWidget(QLabel("DPI:"))
        basic_row.addWidget(self.plot_dpi_spin)
        basic_row.addStretch()
        plot_config_layout.addLayout(basic_row)

        # Figure size controls
        size_row = QHBoxLayout()

        self.plot_width_spin = QDoubleSpinBox()
        self.plot_width_spin.setRange(1.0, 100.0)
        self.plot_width_spin.setValue(10.0)
        self.plot_width_spin.setSingleStep(0.5)

        self.plot_height_spin = QDoubleSpinBox()
        self.plot_height_spin.setRange(0.1, 10.0)
        self.plot_height_spin.setValue(0.6)
        self.plot_height_spin.setSingleStep(0.1)

        size_row.addWidget(QLabel("Figure Width:"))
        size_row.addWidget(self.plot_width_spin)
        size_row.addWidget(QLabel("Height Per ROI:"))
        size_row.addWidget(self.plot_height_spin)
        size_row.addStretch()
        plot_config_layout.addLayout(size_row)

        # Amplitude mode toggle
        self.show_real_amplitude = QCheckBox("Show Real Amplitude (instead of 0-1 normalized)")
        self.show_real_amplitude.setChecked(False)
        self.show_real_amplitude.setToolTip(
            "Toggle between MinMax-normalized [0,1] view and real amplitude values.\n"
            "Real amplitude shows sum(|Δpixel|) per ROI in raw pixel counts (MATLAB-style)."
        )
        plot_config_layout.addWidget(self.show_real_amplitude)

        self.chk_divide_by_pixels = QCheckBox("÷ ROI pixel count (per-pixel mean)")
        self.chk_divide_by_pixels.setChecked(False)   # default: pixel sum (MATLAB-style)
        self.chk_divide_by_pixels.setEnabled(False)   # enabled only when Real Amplitude is on
        self.chk_divide_by_pixels.setToolTip(
            "Unchecked (default): pixel sum — sum(|Δpixel|) per ROI, MATLAB-equivalent\n"
            "Checked: per-pixel mean — divide by ROI pixel count, ROI-size independent"
        )
        self.chk_divide_by_pixels.toggled.connect(self.generate_plot)
        # Enable/disable together with Real Amplitude toggle
        self.show_real_amplitude.toggled.connect(self.chk_divide_by_pixels.setEnabled)
        plot_config_layout.addWidget(self.chk_divide_by_pixels)

        # Y-Axis scaling controls
        y_axis_group = QGroupBox("Y-Axis Scaling (Per ROI Optimization)")
        y_axis_layout = QVBoxLayout()
        y_axis_group.setLayout(y_axis_layout)

        scaling_mode_layout = QHBoxLayout()
        self.auto_scale_y = QCheckBox("Auto Scale Y-Axis (Recommended)")
        self.auto_scale_y.setChecked(True)
        self.auto_scale_y.setToolTip(
            "Automatically optimize Y-axis for each ROI individually"
        )

        self.robust_scaling = QCheckBox("Robust Scaling (Ignore Outliers)")
        self.robust_scaling.setChecked(True)
        self.robust_scaling.setToolTip(
            "Use percentile-based scaling to ignore outliers and focus on main data"
        )

        scaling_mode_layout.addWidget(self.auto_scale_y)
        scaling_mode_layout.addWidget(self.robust_scaling)
        scaling_mode_layout.addStretch()
        y_axis_layout.addLayout(scaling_mode_layout)

        # Advanced scaling options
        advanced_layout = QHBoxLayout()

        self.adaptive_scaling = QCheckBox("Adaptive Scaling")
        self.adaptive_scaling.setChecked(True)
        self.adaptive_scaling.setToolTip(
            "Automatically adjust scaling strategy based on data characteristics\n"
            "• Low variance data: Tighter scaling to show small changes\n"
            "• Outlier-heavy data: More aggressive filtering\n"
            "• Sparse data: Optimize for non-zero values"
        )

        self.center_around_zero = QCheckBox("Smart Zero Centering")
        self.center_around_zero.setChecked(True)
        self.center_around_zero.setToolTip(
            "Include zero in view when data is centered around zero"
        )

        advanced_layout.addWidget(self.adaptive_scaling)
        advanced_layout.addWidget(self.center_around_zero)
        advanced_layout.addStretch()
        y_axis_layout.addLayout(advanced_layout)

        # Manual Y-axis range controls
        manual_range_layout = QHBoxLayout()

        self.y_min_spin = QDoubleSpinBox()
        self.y_min_spin.setRange(-1e9, 1e9)
        self.y_min_spin.setValue(0.0)
        self.y_min_spin.setEnabled(False)

        self.y_max_spin = QDoubleSpinBox()
        self.y_max_spin.setRange(-1e9, 1e9)
        self.y_max_spin.setValue(1000.0)
        self.y_max_spin.setEnabled(False)

        self.btn_apply_y_range = QPushButton("Apply Manual Range")
        self.btn_apply_y_range.setEnabled(False)
        self.btn_apply_y_range.setToolTip(
            "Use manual Y-axis range instead of automatic optimization"
        )

        manual_range_layout.addWidget(QLabel("Manual Y Min:"))
        manual_range_layout.addWidget(self.y_min_spin)
        manual_range_layout.addWidget(QLabel("Y Max:"))
        manual_range_layout.addWidget(self.y_max_spin)
        manual_range_layout.addWidget(self.btn_apply_y_range)
        manual_range_layout.addStretch()
        y_axis_layout.addLayout(manual_range_layout)

        # Percentile controls for robust scaling
        percentile_layout = QHBoxLayout()

        self.lower_percentile_spin = QDoubleSpinBox()
        self.lower_percentile_spin.setRange(0.0, 50.0)
        self.lower_percentile_spin.setValue(5.0)
        self.lower_percentile_spin.setSingleStep(1.0)

        self.upper_percentile_spin = QDoubleSpinBox()
        self.upper_percentile_spin.setRange(50.0, 100.0)
        self.upper_percentile_spin.setValue(95.0)
        self.upper_percentile_spin.setSingleStep(1.0)

        percentile_layout.addWidget(QLabel("Lower %:"))
        percentile_layout.addWidget(self.lower_percentile_spin)
        percentile_layout.addWidget(QLabel("Upper %:"))
        percentile_layout.addWidget(self.upper_percentile_spin)
        percentile_layout.addStretch()
        y_axis_layout.addLayout(percentile_layout)

        plot_config_layout.addWidget(y_axis_group)

        # Time range selection
        time_range_group = QGroupBox("Time Range Selection")
        time_range_layout = QHBoxLayout()
        time_range_group.setLayout(time_range_layout)

        self.plot_start_time = QDoubleSpinBox()
        self.plot_start_time.setRange(0.0, 1e9)
        self.plot_start_time.setValue(0.0)
        self.plot_start_time.setSuffix(" min")
        self.plot_end_time = QDoubleSpinBox()
        self.plot_end_time.setRange(0.0, 1e9)
        self.plot_end_time.setValue(100000.0)
        self.plot_end_time.setSuffix(" min")
        self.btn_apply_time_range = QPushButton("Apply Time Range")

        time_range_layout.addWidget(QLabel("Start Time (min):"))
        time_range_layout.addWidget(self.plot_start_time)
        time_range_layout.addWidget(QLabel("End Time (min):"))
        time_range_layout.addWidget(self.plot_end_time)
        time_range_layout.addWidget(self.btn_apply_time_range)

        layout.addWidget(time_range_group)

        # Plot binning configuration (separate from analysis binning)
        plot_binning_group = QGroupBox("Plot Binning (Visualization Only)")
        plot_binning_layout = QHBoxLayout()
        plot_binning_group.setLayout(plot_binning_layout)

        self.plot_bin_minutes = QSpinBox()
        self.plot_bin_minutes.setRange(0, 240)  # 0 = no rebinning, 1-240 minutes
        self.plot_bin_minutes.setValue(10)  # Default 10 minutes
        self.plot_bin_minutes.setSuffix(" min")
        self.plot_bin_minutes.setSpecialValueText("Original (1 min)")
        self.plot_bin_minutes.setToolTip(
            "Re-bin data for visualization purposes.\n"
            "0 = Use original binning from analysis (typically 1 min bins)\n"
            "This does NOT affect analysis calculations.\n"
            "Useful for publications (e.g., 60 min for circadian plots).\n"
            "Original analysis uses 'Bin Size' from Behavior Analysis section."
        )

        # Preset buttons for common binning intervals
        self.btn_plot_bin_original = QPushButton("Original")
        self.btn_plot_bin_original.setToolTip("Use original binning (no re-binning)")
        self.btn_plot_bin_10min = QPushButton("10 min")
        self.btn_plot_bin_10min.setToolTip("Set plot binning to 10 minutes")
        self.btn_plot_bin_30min = QPushButton("30 min")
        self.btn_plot_bin_30min.setToolTip("Set plot binning to 30 minutes")
        self.btn_plot_bin_60min = QPushButton("60 min")
        self.btn_plot_bin_60min.setToolTip("Set plot binning to 60 minutes (1 hour)")

        # Connect preset buttons
        self.btn_plot_bin_original.clicked.connect(
            lambda: self.plot_bin_minutes.setValue(0)
        )
        self.btn_plot_bin_10min.clicked.connect(
            lambda: self.plot_bin_minutes.setValue(10)
        )
        self.btn_plot_bin_30min.clicked.connect(
            lambda: self.plot_bin_minutes.setValue(30)
        )
        self.btn_plot_bin_60min.clicked.connect(
            lambda: self.plot_bin_minutes.setValue(60)
        )

        # Auto-refresh plot when bin size changes
        self.plot_bin_minutes.valueChanged.connect(self._on_plot_bin_changed)

        plot_binning_layout.addWidget(QLabel("Plot Bin Size:"))
        plot_binning_layout.addWidget(self.plot_bin_minutes)
        plot_binning_layout.addWidget(QLabel("Presets:"))
        plot_binning_layout.addWidget(self.btn_plot_bin_original)
        plot_binning_layout.addWidget(self.btn_plot_bin_10min)
        plot_binning_layout.addWidget(self.btn_plot_bin_30min)
        plot_binning_layout.addWidget(self.btn_plot_bin_60min)
        plot_binning_layout.addStretch()

        # ZT time axis and lighting overlay options
        self.chk_zt_axis = QCheckBox("ZT time (h)")
        self.chk_zt_axis.setToolTip("X-axis in hours from recording start (ZT 0 = start)")
        self.chk_zt_axis.toggled.connect(lambda _: self.generate_plot())

        self.chk_show_lighting = QCheckBox("Light/Dark")
        self.chk_show_lighting.setToolTip("Overlay day/night shading from HDF5 LED data")
        self.chk_show_lighting.toggled.connect(lambda _: self.generate_plot())

        plot_binning_layout.addWidget(self.chk_zt_axis)
        plot_binning_layout.addWidget(self.chk_show_lighting)

        layout.addWidget(time_range_group)
        layout.addWidget(plot_binning_group)
        layout.addWidget(plot_config_group)

        # ===== SIMPLIFIED PLOT CONTROLS =====
        plot_buttons_group = QGroupBox("Plot Controls")
        plot_buttons_layout = QHBoxLayout()
        plot_buttons_group.setLayout(plot_buttons_layout)

        # Core plotting buttons
        self.btn_plot = QPushButton("Generate Plot")
        self.btn_plot.setToolTip("Generate the selected plot type")
        self.btn_plot.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }"
        )

        self.btn_save_plot = QPushButton("Save Current Plot")
        self.btn_save_plot.setToolTip("Save the currently displayed plot as image file")

        self.btn_save_all_plots = QPushButton("Save All Plots")
        self.btn_save_all_plots.setToolTip(
            "Save all plot types to separate image files"
        )

        # CONSOLIDATED save results button
        self.btn_save_results = QPushButton("Save Results")
        self.btn_save_results.setToolTip(
            "Save analysis results (CSV + Excel + threshold stats)"
        )
        self.btn_save_results.setStyleSheet(
            "QPushButton { background-color: #2196F3; color: white; font-weight: bold; }"
        )
        self.btn_save_with_metadata = QPushButton("Save with HDF5 Metadata")
        self.btn_save_with_metadata.setToolTip(
            "Save analysis results including comprehensive HDF5 metadata"
        )
        self.btn_save_with_metadata.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }"
        )

        plot_buttons_layout.addWidget(self.btn_plot)
        plot_buttons_layout.addWidget(self.btn_save_plot)
        plot_buttons_layout.addWidget(self.btn_save_all_plots)

        plot_buttons_layout.addWidget(self.btn_save_results)
        plot_buttons_layout.addWidget(self.btn_save_with_metadata)
        layout.addWidget(plot_buttons_group)

    def setup_extended_tab(self):
        """Setup the Extended Analysis tab for rhythmic pattern detection."""
        layout = QVBoxLayout()
        self.tab_extended.setLayout(layout)

        # Title and description
        title_label = QLabel("Rhythmic Pattern Analysis")
        title_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        layout.addWidget(title_label)

        desc_label = QLabel(
            "Detect any recurring periodic patterns in activity data: "
            "circadian rhythms (24h), ultradian rhythms (<20h), short activity cycles, "
            "or custom periods using Fischer Z-transformation, FFT, and correlation methods."
        )
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("color: #666; font-size: 10px; margin-bottom: 10px;")
        layout.addWidget(desc_label)

        # Period range parameters
        fisher_params_group = QGroupBox("Period Range Parameters")
        fisher_params_layout = QFormLayout()
        fisher_params_group.setLayout(fisher_params_layout)

        self.fisher_min_period = QDoubleSpinBox()
        self.fisher_min_period.setRange(0.1, 100.0)
        self.fisher_min_period.setValue(0.5)
        self.fisher_min_period.setSingleStep(0.5)
        self.fisher_min_period.setDecimals(1)
        self.fisher_min_period.setSuffix(" hours")
        self.fisher_min_period.setToolTip(
            "Minimum period to test:\n"
            "• 0.5-2h: Very fast ultradian rhythms\n"
            "• 2-8h: Standard ultradian rhythms\n"
            "• 8-12h: Extended ultradian/tidal rhythms\n"
            "• 20-22h: Circadian rhythms (24h cycles)"
        )
        fisher_params_layout.addRow("Minimum Period:", self.fisher_min_period)

        self.fisher_max_period = QDoubleSpinBox()
        self.fisher_max_period.setRange(0.1, 200.0)
        self.fisher_max_period.setValue(8.0)
        self.fisher_max_period.setSingleStep(1.0)
        self.fisher_max_period.setDecimals(1)
        self.fisher_max_period.setSuffix(" hours")
        self.fisher_max_period.setToolTip(
            "Maximum period to test:\n"
            "• 2-8h: Fast ultradian rhythms only\n"
            "• 8-20h: Include infradian/extended rhythms\n"
            "• 20-28h: Include circadian (24h) rhythms\n"
            "• >28h: Longer multi-day cycles (requires 72h+ data)"
        )
        fisher_params_layout.addRow("Maximum Period:", self.fisher_max_period)

        self.fisher_significance = QDoubleSpinBox()
        self.fisher_significance.setRange(0.001, 0.1)
        self.fisher_significance.setValue(0.05)
        self.fisher_significance.setSingleStep(0.01)
        self.fisher_significance.setDecimals(3)
        self.fisher_significance.setToolTip(
            "Statistical significance threshold (p-value)"
        )
        fisher_params_layout.addRow("Significance Level (α):", self.fisher_significance)

        layout.addWidget(fisher_params_group)

        # Quick preset buttons
        preset_group = QGroupBox("Quick Presets")
        preset_layout = QHBoxLayout()
        preset_group.setLayout(preset_layout)

        from qtpy.QtWidgets import QPushButton

        btn_preset_short = QPushButton("Ultradian\n(0.5-8h)")
        btn_preset_short.setToolTip(
            "Ultradian rhythms: Periods shorter than ~12 hours\n"
            "Good for: Fast oscillations, feeding cycles, short activity bouts\n"
            "Requires: Any recording length"
        )
        btn_preset_short.clicked.connect(lambda: self._set_period_preset(0.5, 8.0))
        preset_layout.addWidget(btn_preset_short)

        btn_preset_medium = QPushButton("Infradian\n(8-20h)")
        btn_preset_medium.setToolTip(
            "Infradian rhythms: Periods between ultradian and circadian\n"
            "Good for: Extended activity/rest cycles, tidal rhythms\n"
            "Requires: 24+ hour recordings"
        )
        btn_preset_medium.clicked.connect(lambda: self._set_period_preset(8.0, 20.0))
        preset_layout.addWidget(btn_preset_medium)

        btn_preset_circadian = QPushButton("Circadian\n(20-28h)")
        btn_preset_circadian.setToolTip(
            "Circadian rhythms: ~24-hour day/night cycles\n"
            "Good for: Daily activity patterns, sleep/wake cycles\n"
            "Requires: 48+ hour recordings for reliable detection"
        )
        btn_preset_circadian.clicked.connect(
            lambda: self._set_period_preset(20.0, 28.0)
        )
        preset_layout.addWidget(btn_preset_circadian)

        btn_preset_auto = QPushButton("Auto-Detect\nRange")
        btn_preset_auto.setToolTip(
            "Automatically determine optimal period range based on recording duration"
        )
        btn_preset_auto.clicked.connect(self._auto_detect_period_range)
        preset_layout.addWidget(btn_preset_auto)

        layout.addWidget(preset_group)

        # Cosinor-specific option
        self.chk_cosinor_population = QCheckBox("Show population mean (Cosinor)")
        self.chk_cosinor_population.setToolTip(
            "Add a full-width subplot below the per-ROI panels showing\n"
            "the population-level cosinor fit (mean across all ROIs)."
        )
        layout.addWidget(self.chk_cosinor_population)

        # Analysis method selection
        method_group = QGroupBox("Analysis Method")
        method_layout = QFormLayout()
        method_group.setLayout(method_layout)

        from qtpy.QtWidgets import QComboBox

        self.fisher_method_combo = QComboBox()
        self.fisher_method_combo.addItems(
            [
                "Chi² Periodogram",
                "FFT Power Spectrum",
                "Cosinor Analysis",
                "ROI Similarity Matrix",
                "Coherence Analysis",
                "Phase Clustering",
            ]
        )
        self.fisher_method_combo.setToolTip(
            "Select the rhythmic pattern analysis method:\n\n"
            "• Chi² Periodogram: Most robust. Tests all periods in range using chi-square\n"
            "  statistic. Recommended default. Use Fraction Movement.\n\n"
            "• FFT Power Spectrum: Good for noisy data. Permutation-based significance\n"
            "  (1000 permutations). Use Fraction Movement.\n\n"
            "• Cosinor Analysis: Fits cosine curve — gives MESOR, Amplitude, Acrophase.\n"
            "  Tests specific periods (12, 24, 30 h + min/mid/max of range).\n"
            "  Use Fraction or Normalized Movement. Needs ≥ 2 complete cycles.\n\n"
            "• ROI Similarity Matrix: Pairwise cross-correlation between all ROIs at\n"
            "  optimal time lag (≤ max_period / 2). Hierarchical clustering at r = 0.5.\n"
            "  Period range does not affect result. Use Fraction Movement.\n\n"
            "• Coherence Analysis: Welch magnitude-squared coherence between ROI pairs\n"
            "  at one target period = midpoint of period range.\n"
            "  ⚠ Set period range so that (min + max) / 2 = target period\n"
            "  (e.g. 20–28 h for circadian → target = 24 h).\n"
            "  Needs long recordings (≥ 3× target period) for good frequency resolution.\n"
            "  Recommended bin size: 30–60 min.\n\n"
            "• Phase Clustering: Hilbert-transform phase per ROI → peak activity time.\n"
            "  Bins ROIs into 4 groups: 0–6 h / 6–12 h / 12–18 h / 18–24 h.\n"
            "  ⚠ Set period range so midpoint = target period (same as Coherence).\n"
            "  Reliable only when the target rhythm is strong and dominant in the signal.\n"
            "  Phase is relative to recording start (ZT 0 = start of recording)."
        )
        self.fisher_method_combo.currentIndexChanged.connect(
            self._on_fisher_method_changed
        )
        method_layout.addRow("Method:", self.fisher_method_combo)

        layout.addWidget(method_group)

        # Re-Binning Options for Extended Analysis
        rebinning_group = QGroupBox("Analysis Binning (Optional)")
        rebinning_layout = QVBoxLayout()
        rebinning_group.setLayout(rebinning_layout)

        # Info label
        rebinning_info = QLabel(
            "Re-bin fraction data for extended analysis. "
            "Larger bins = smoother data, better for long periods. "
            "Smaller bins = more detail, better for short periods."
        )
        rebinning_info.setWordWrap(True)
        rebinning_info.setStyleSheet("color: #7f8c8d; font-size: 10px;")
        rebinning_layout.addWidget(rebinning_info)

        # Binning controls layout
        binning_controls = QHBoxLayout()

        binning_controls.addWidget(QLabel("Analysis Bin Size:"))

        # Decrease button
        self.btn_decrease_bin = QPushButton("−")
        self.btn_decrease_bin.setMaximumWidth(30)
        self.btn_decrease_bin.setToolTip("Decrease bin size")
        self.btn_decrease_bin.clicked.connect(
            lambda: self._adjust_analysis_bin_size(-1)
        )
        binning_controls.addWidget(self.btn_decrease_bin)

        # Bin size spinbox
        self.analysis_bin_size = QSpinBox()
        self.analysis_bin_size.setRange(10, 3600)  # 10 sec to 60 min
        self.analysis_bin_size.setValue(
            60
        )  # Default: 60s (matches main analysis default)
        self.analysis_bin_size.setSingleStep(10)
        self.analysis_bin_size.setSuffix(" sec")
        self.analysis_bin_size.setToolTip(
            "Bin size for extended analysis (10 sec – 60 min).\n"
            "Larger bins reduce noise but lose temporal resolution.\n\n"
            "Recommendations by method:\n"
            "• Chi² / FFT: 60–300 s (1–5 min)\n"
            "• Cosinor: 300–600 s (5–10 min) — reduces saturation at 1.0\n"
            "• Similarity: 300–1800 s (5–30 min)\n"
            "• Coherence: 1800–3600 s (30–60 min) — improves frequency resolution\n"
            "• Phase Clustering: 300–1800 s (5–30 min)\n\n"
            "Data will be automatically re-binned if different from original."
        )
        self.analysis_bin_size.setMinimumWidth(100)
        self.analysis_bin_size.valueChanged.connect(self._update_bin_size_info)
        binning_controls.addWidget(self.analysis_bin_size)

        # Increase button
        self.btn_increase_bin = QPushButton("+")
        self.btn_increase_bin.setMaximumWidth(30)
        self.btn_increase_bin.setToolTip("Increase bin size")
        self.btn_increase_bin.clicked.connect(lambda: self._adjust_analysis_bin_size(1))
        binning_controls.addWidget(self.btn_increase_bin)

        binning_controls.addWidget(QLabel("Presets:"))

        # Preset buttons
        self.btn_bin_original = QPushButton("Original")
        self.btn_bin_original.setToolTip(
            "Use original bin size from main analysis (typically 60s)"
        )
        self.btn_bin_original.clicked.connect(
            lambda: self._set_analysis_bin_preset("original")
        )
        binning_controls.addWidget(self.btn_bin_original)

        self.btn_bin_30s = QPushButton("30 sec")
        self.btn_bin_30s.setToolTip(
            "30 second bins - high resolution for short periods"
        )
        self.btn_bin_30s.clicked.connect(lambda: self._set_analysis_bin_preset(30))
        binning_controls.addWidget(self.btn_bin_30s)

        self.btn_bin_1min = QPushButton("1 min")
        self.btn_bin_1min.setToolTip("1 minute bins - standard resolution")
        self.btn_bin_1min.clicked.connect(lambda: self._set_analysis_bin_preset(60))
        binning_controls.addWidget(self.btn_bin_1min)

        self.btn_bin_5min = QPushButton("5 min")
        self.btn_bin_5min.setToolTip("5 minute bins - smooth data for long periods")
        self.btn_bin_5min.clicked.connect(lambda: self._set_analysis_bin_preset(300))
        binning_controls.addWidget(self.btn_bin_5min)

        self.btn_bin_10min = QPushButton("10 min")
        self.btn_bin_10min.setToolTip(
            "10 minute bins - very smooth data for circadian analysis"
        )
        self.btn_bin_10min.clicked.connect(lambda: self._set_analysis_bin_preset(600))
        binning_controls.addWidget(self.btn_bin_10min)

        self.btn_bin_30min = QPushButton("30 min")
        self.btn_bin_30min.setToolTip(
            "30 minute bins - reduces saturation for Cosinor analysis"
        )
        self.btn_bin_30min.clicked.connect(lambda: self._set_analysis_bin_preset(1800))
        binning_controls.addWidget(self.btn_bin_30min)

        binning_controls.addStretch()

        rebinning_layout.addLayout(binning_controls)

        # Display current vs original bin size
        self.bin_size_info_label = QLabel("")
        self.bin_size_info_label.setStyleSheet(
            "color: #27ae60; font-size: 10px; font-style: italic;"
        )
        rebinning_layout.addWidget(self.bin_size_info_label)

        # Data source selection (fraction vs raw movement)
        data_source_layout = QHBoxLayout()
        data_source_layout.addWidget(QLabel("Data Source:"))

        self.data_source_combo = QComboBox()
        self.data_source_combo.addItems([
            "Fraction Movement (0-1)",
            "Raw Intensity (continuous)",
            "Normalized Movement (0-1)",
        ])
        self.data_source_combo.setCurrentIndex(0)  # Default: fraction movement
        self.data_source_combo.setToolTip(
            "Choose data source for extended analysis:\n\n"
            "• Fraction Movement (0-1): Proportion of time in movement state per bin.\n"
            "  Recommended for Chi², FFT, Similarity, Coherence, Phase Clustering.\n\n"
            "• Raw Intensity (continuous): Per-pixel intensity changes, MinMax normalized.\n"
            "  Use for Cosinor when a sinusoidal waveform is expected.\n\n"
            "• Normalized Movement (0-1): Re-binned & min/max normalized per ROI.\n"
            "  Comparable to Aguillon et al. 2023 'Normalized Movement (a.u.)'.\n"
            "  Best for Cosinor and cross-study comparison.\n\n"
            "Recommendation by method:\n"
            "  Chi² / FFT / Similarity / Coherence / Phase Clustering → Fraction Movement\n"
            "  Cosinor → Fraction Movement or Normalized Movement"
        )
        self.data_source_combo.currentIndexChanged.connect(self._on_data_source_changed)
        data_source_layout.addWidget(self.data_source_combo)

        # Checkbox to calculate both activity and sleep phases
        self.chk_calculate_sleep_phase = QCheckBox("Also calculate Sleep Phase")
        self.chk_calculate_sleep_phase.setChecked(True)
        self.chk_calculate_sleep_phase.setToolTip(
            "Calculate both Acrophase (peak activity) and Sleep Phase (peak sleep).\n"
            "Requires main analysis to be run first."
        )
        data_source_layout.addWidget(self.chk_calculate_sleep_phase)

        # Sleep data source selector
        self.sleep_source_combo = QComboBox()
        self.sleep_source_combo.addItems([
            "Quiescence (comparable)",
            "Sleep (≥8min sustained)",
        ])
        self.sleep_source_combo.setCurrentIndex(1)  # Default: Sleep ≥8min (differs from activity spectrum)
        self.sleep_source_combo.setToolTip(
            "Choose data source for sleep rhythm analysis:\n"
            "• Quiescence: Binary rest state (movement < threshold), same temporal\n"
            "  resolution as activity data — best for direct period comparison.\n"
            "• Sleep (≥8min sustained): Only sustained quiescence episodes ≥8 min\n"
            "  are counted as sleep. Acts as a low-pass filter (~16 min cutoff).\n"
            "  More biologically strict, but not directly comparable to activity."
        )
        data_source_layout.addWidget(self.sleep_source_combo)
        data_source_layout.addStretch()

        rebinning_layout.addLayout(data_source_layout)

        layout.addWidget(rebinning_group)

        # Cycle/Period Selection for Post-Hoc Analysis
        cycle_selection_group = QGroupBox("Time Range Selection (Optional)")
        cycle_selection_layout = QFormLayout()
        cycle_selection_group.setLayout(cycle_selection_layout)

        # Enable cycle selection checkbox
        self.enable_cycle_selection = QCheckBox("Analyze specific time range only")
        self.enable_cycle_selection.setToolTip(
            "Enable this to analyze only a specific portion of your recording.\n"
            "Useful for cycle-specific analysis or comparing different time periods."
        )
        self.enable_cycle_selection.stateChanged.connect(
            self._on_cycle_selection_toggled
        )
        cycle_selection_layout.addRow("", self.enable_cycle_selection)

        # Start time selection
        self.cycle_start_time = QDoubleSpinBox()
        self.cycle_start_time.setRange(0.0, 10000.0)
        self.cycle_start_time.setValue(0.0)
        self.cycle_start_time.setSingleStep(1.0)
        self.cycle_start_time.setDecimals(1)
        self.cycle_start_time.setSuffix(" hours")
        self.cycle_start_time.setEnabled(False)
        self.cycle_start_time.setToolTip(
            "Start time of the analysis window (hours from recording start)"
        )
        cycle_selection_layout.addRow("Start Time:", self.cycle_start_time)

        # End time selection
        self.cycle_end_time = QDoubleSpinBox()
        self.cycle_end_time.setRange(0.0, 10000.0)
        self.cycle_end_time.setValue(24.0)
        self.cycle_end_time.setSingleStep(1.0)
        self.cycle_end_time.setDecimals(1)
        self.cycle_end_time.setSuffix(" hours")
        self.cycle_end_time.setEnabled(False)
        self.cycle_end_time.setToolTip(
            "End time of the analysis window (hours from recording start)"
        )
        cycle_selection_layout.addRow("End Time:", self.cycle_end_time)

        # Quick cycle selection buttons
        cycle_buttons_layout = QHBoxLayout()

        self.btn_cycle_first24 = QPushButton("First 24h")
        self.btn_cycle_first24.setToolTip("Analyze first 24 hours of recording")
        self.btn_cycle_first24.clicked.connect(lambda: self._set_cycle_range(0, 24))
        self.btn_cycle_first24.setEnabled(False)
        cycle_buttons_layout.addWidget(self.btn_cycle_first24)

        self.btn_cycle_second24 = QPushButton("Second 24h")
        self.btn_cycle_second24.setToolTip("Analyze hours 24-48 of recording")
        self.btn_cycle_second24.clicked.connect(lambda: self._set_cycle_range(24, 48))
        self.btn_cycle_second24.setEnabled(False)
        cycle_buttons_layout.addWidget(self.btn_cycle_second24)

        self.btn_cycle_last24 = QPushButton("Last 24h")
        self.btn_cycle_last24.setToolTip("Analyze last 24 hours of recording")
        self.btn_cycle_last24.clicked.connect(self._set_cycle_last_24h)
        self.btn_cycle_last24.setEnabled(False)
        cycle_buttons_layout.addWidget(self.btn_cycle_last24)

        self.btn_cycle_reset = QPushButton("Full Recording")
        self.btn_cycle_reset.setToolTip("Reset to analyze entire recording")
        self.btn_cycle_reset.clicked.connect(self._reset_cycle_range)
        self.btn_cycle_reset.setEnabled(False)
        cycle_buttons_layout.addWidget(self.btn_cycle_reset)

        cycle_selection_layout.addRow("", cycle_buttons_layout)

        layout.addWidget(cycle_selection_group)

        # Buttons layout
        buttons_layout = QHBoxLayout()

        # Run analysis button
        self.btn_run_fisher = QPushButton("Run Rhythmic Pattern Analysis")
        self.btn_run_fisher.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; font-weight: bold; "
            "padding: 10px; } QPushButton:hover { background-color: #45a049; }"
        )
        self.btn_run_fisher.clicked.connect(self.run_fisher_analysis)
        buttons_layout.addWidget(self.btn_run_fisher)

        # Load results from HDF5 button (for test data)
        self.btn_load_hdf5_results = QPushButton("Load Results from HDF5")
        self.btn_load_hdf5_results.setStyleSheet(
            "QPushButton { background-color: #2196F3; color: white; font-weight: bold; "
            "padding: 10px; } QPushButton:hover { background-color: #0b7dda; }"
        )
        self.btn_load_hdf5_results.setToolTip(
            "Load pre-computed results directly from HDF5 file.\n"
            "Use this for test data that already contains analysis results."
        )
        self.btn_load_hdf5_results.clicked.connect(self._load_results_from_hdf5)
        buttons_layout.addWidget(self.btn_load_hdf5_results)

        # Save results to HDF5 button
        self.btn_save_hdf5_results = QPushButton("Save Results to HDF5")
        self.btn_save_hdf5_results.setStyleSheet(
            "QPushButton { background-color: #FF9800; color: white; font-weight: bold; "
            "padding: 10px; } QPushButton:hover { background-color: #e68900; }"
        )
        self.btn_save_hdf5_results.setToolTip(
            "Save current analysis results to HDF5 file.\n"
            "This allows you to reload results later without re-running analysis."
        )
        self.btn_save_hdf5_results.clicked.connect(self._save_results_to_hdf5)
        buttons_layout.addWidget(self.btn_save_hdf5_results)

        layout.addLayout(buttons_layout)

        # Create a horizontal splitter for results and plot
        splitter = QSplitter()
        splitter.setOrientation(1)  # Horizontal

        # Results display (left side)
        self.fisher_results_text = QTextEdit()
        self.fisher_results_text.setReadOnly(True)
        self.fisher_results_text.setPlaceholderText(
            "Rhythmic pattern analysis results will appear here...\n\n"
            "Tip: Use 'Auto-Detect Range' to automatically set optimal period range for your recording duration."
        )
        splitter.addWidget(self.fisher_results_text)

        # Plot display (right side)
        self.fisher_plot_widget = QWidget()
        fisher_plot_layout = QVBoxLayout()
        self.fisher_plot_widget.setLayout(fisher_plot_layout)

        # Plot header with title and pop-out button
        plot_header_layout = QHBoxLayout()
        fisher_plot_label = QLabel("Periodogram Plot")
        fisher_plot_label.setStyleSheet("font-weight: bold;")
        plot_header_layout.addWidget(fisher_plot_label)

        plot_header_layout.addStretch()

        # Button to save the current periodogram plot as image
        self.btn_save_fisher_plot = QPushButton("Save Plot...")
        self.btn_save_fisher_plot.setToolTip("Save the current plot as PNG / PDF / SVG")
        self.btn_save_fisher_plot.setMaximumWidth(110)
        self.btn_save_fisher_plot.clicked.connect(self._save_fisher_plot)
        self.btn_save_fisher_plot.setEnabled(False)
        plot_header_layout.addWidget(self.btn_save_fisher_plot)

        # Button to open plot in separate window
        self.btn_popout_plot = QPushButton("🗖 Open in Separate Window")
        self.btn_popout_plot.setToolTip("Open plot in a larger, resizable window")
        self.btn_popout_plot.setMaximumWidth(200)
        self.btn_popout_plot.clicked.connect(self._open_plot_window)
        self.btn_popout_plot.setEnabled(False)  # Enable after first plot
        plot_header_layout.addWidget(self.btn_popout_plot)

        fisher_plot_layout.addLayout(plot_header_layout)

        self.fisher_plot_canvas = QLabel()
        self.fisher_plot_canvas.setMinimumSize(400, 300)
        self.fisher_plot_canvas.setScaledContents(False)  # Maintain aspect ratio
        self.fisher_plot_canvas.setStyleSheet(
            "border: 1px solid #ccc; background-color: white;"
        )
        self.fisher_plot_canvas.setAlignment(Qt.AlignCenter)
        fisher_plot_layout.addWidget(self.fisher_plot_canvas, 1)  # Allow expansion

        splitter.addWidget(self.fisher_plot_widget)

        # Set initial sizes (60% results, 40% plot)
        splitter.setSizes([600, 400])
        layout.addWidget(splitter)

        # Export button
        self.btn_export_fisher = QPushButton("Export Rhythmic Pattern Results")
        self.btn_export_fisher.clicked.connect(self.export_fisher_results)
        self.btn_export_fisher.setEnabled(False)
        layout.addWidget(self.btn_export_fisher)

        layout.addStretch()

    def setup_viewer_tab(self):
        """Setup the Frame Viewer tab for browsing through dataset frames."""
        layout = QVBoxLayout()
        self.tab_viewer.setLayout(layout)

        # Title and description
        title_label = QLabel("Frame Viewer")
        title_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        layout.addWidget(title_label)

        desc_label = QLabel(
            "Browse through the loaded dataset frame-by-frame. "
            "Use the slider or keyboard shortcuts to navigate."
        )
        desc_label.setWordWrap(True)
        desc_label.setStyleSheet("color: #666; font-size: 10px; margin-bottom: 10px;")
        layout.addWidget(desc_label)

        # Status label
        self.viewer_status_label = QLabel(
            "No dataset loaded. Please load a file in the Input tab first."
        )
        self.viewer_status_label.setStyleSheet("color: #999; font-style: italic;")
        layout.addWidget(self.viewer_status_label)

        # Frame navigation controls
        nav_group = QGroupBox("Frame Navigation")
        nav_layout = QVBoxLayout()
        nav_group.setLayout(nav_layout)

        # Frame slider with current frame display
        slider_layout = QHBoxLayout()
        self.viewer_frame_slider = QSlider()
        self.viewer_frame_slider.setOrientation(1)  # Horizontal
        self.viewer_frame_slider.setMinimum(0)
        self.viewer_frame_slider.setMaximum(0)
        self.viewer_frame_slider.setValue(0)
        self.viewer_frame_slider.setEnabled(False)
        self.viewer_frame_slider.valueChanged.connect(self._on_viewer_frame_changed)

        self.viewer_frame_label = QLabel("Frame: 0 / 0")
        self.viewer_frame_label.setMinimumWidth(120)

        slider_layout.addWidget(self.viewer_frame_label)
        slider_layout.addWidget(self.viewer_frame_slider)
        nav_layout.addLayout(slider_layout)

        # Playback controls
        playback_layout = QHBoxLayout()

        self.btn_viewer_first = QPushButton("|◀")
        self.btn_viewer_first.setToolTip("First frame (Home)")
        self.btn_viewer_first.clicked.connect(lambda: self._viewer_goto_frame(0))
        self.btn_viewer_first.setEnabled(False)

        self.btn_viewer_prev = QPushButton("◀")
        self.btn_viewer_prev.setToolTip("Previous frame (←)")
        self.btn_viewer_prev.clicked.connect(lambda: self._viewer_step_frame(-1))
        self.btn_viewer_prev.setEnabled(False)

        self.btn_viewer_play = QPushButton("▶ Play")
        self.btn_viewer_play.setToolTip("Play/Pause (Space)")
        self.btn_viewer_play.setCheckable(True)
        self.btn_viewer_play.clicked.connect(self._viewer_toggle_play)
        self.btn_viewer_play.setEnabled(False)

        self.btn_viewer_next = QPushButton("▶")
        self.btn_viewer_next.setToolTip("Next frame (→)")
        self.btn_viewer_next.clicked.connect(lambda: self._viewer_step_frame(1))
        self.btn_viewer_next.setEnabled(False)

        self.btn_viewer_last = QPushButton("▶|")
        self.btn_viewer_last.setToolTip("Last frame (End)")
        self.btn_viewer_last.clicked.connect(lambda: self._viewer_goto_frame(-1))
        self.btn_viewer_last.setEnabled(False)

        playback_layout.addWidget(self.btn_viewer_first)
        playback_layout.addWidget(self.btn_viewer_prev)
        playback_layout.addWidget(self.btn_viewer_play)
        playback_layout.addWidget(self.btn_viewer_next)
        playback_layout.addWidget(self.btn_viewer_last)
        playback_layout.addStretch()

        nav_layout.addLayout(playback_layout)

        # Playback speed
        speed_layout = QHBoxLayout()
        speed_layout.addWidget(QLabel("Playback FPS:"))

        self.viewer_fps_spin = QSpinBox()
        self.viewer_fps_spin.setRange(1, 60)
        self.viewer_fps_spin.setValue(10)
        self.viewer_fps_spin.setToolTip("Frames per second during playback")
        self.viewer_fps_spin.valueChanged.connect(self._viewer_update_timer_interval)

        speed_layout.addWidget(self.viewer_fps_spin)
        speed_layout.addStretch()
        nav_layout.addLayout(speed_layout)

        layout.addWidget(nav_group)

        # File selector for multiple loaded files
        file_select_layout = QHBoxLayout()
        file_select_layout.addWidget(QLabel("File:"))
        self.viewer_file_combo = QComboBox()
        self.viewer_file_combo.setToolTip("Select which file to view in the Frame Viewer")
        self.viewer_file_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        file_select_layout.addWidget(self.viewer_file_combo)
        layout.addLayout(file_select_layout)

        # Load data button
        self.btn_viewer_load = QPushButton("Load Selected File into Viewer")
        self.btn_viewer_load.clicked.connect(self._viewer_load_data)
        self.btn_viewer_load.setStyleSheet(
            "QPushButton { background-color: #2196F3; color: white; font-weight: bold; "
            "padding: 10px; } QPushButton:hover { background-color: #1976D2; }"
        )
        layout.addWidget(self.btn_viewer_load)

        # ROI overlay option
        overlay_row = QHBoxLayout()
        self.viewer_show_roi_overlay = QCheckBox("Show ROI overlay (circles + numbers)")
        self.viewer_show_roi_overlay.setChecked(True)
        self.viewer_show_roi_overlay.setToolTip(
            "Draw detected ROI circles and numbers on each frame "
            "so you can identify which well corresponds to which ROI"
        )
        self.viewer_show_roi_overlay.toggled.connect(self._on_viewer_roi_overlay_changed)
        overlay_row.addWidget(self.viewer_show_roi_overlay)
        overlay_row.addStretch()
        layout.addLayout(overlay_row)

        # Export video/GIF section
        export_group = QGroupBox("Export Video/GIF")
        export_layout = QVBoxLayout()
        export_group.setLayout(export_layout)

        # Time range selection
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("Time Range:"))

        self.export_start_frame = QSpinBox()
        self.export_start_frame.setPrefix("Start Frame: ")
        self.export_start_frame.setMinimum(0)
        self.export_start_frame.setMaximum(0)
        self.export_start_frame.setValue(0)
        self.export_start_frame.setEnabled(False)
        range_layout.addWidget(self.export_start_frame)

        self.export_end_frame = QSpinBox()
        self.export_end_frame.setPrefix("End Frame: ")
        self.export_end_frame.setMinimum(0)
        self.export_end_frame.setMaximum(0)
        self.export_end_frame.setValue(0)
        self.export_end_frame.setEnabled(False)
        range_layout.addWidget(self.export_end_frame)

        range_layout.addStretch()
        export_layout.addLayout(range_layout)

        # Export buttons
        export_buttons_layout = QHBoxLayout()

        self.btn_export_video = QPushButton("Export as Video (MP4)")
        self.btn_export_video.setToolTip("Export selected frame range as MP4 video")
        self.btn_export_video.clicked.connect(self._export_video)
        self.btn_export_video.setEnabled(False)
        export_buttons_layout.addWidget(self.btn_export_video)

        self.btn_export_gif = QPushButton("Export as GIF")
        self.btn_export_gif.setToolTip("Export selected frame range as animated GIF")
        self.btn_export_gif.clicked.connect(self._export_gif)
        self.btn_export_gif.setEnabled(False)
        export_buttons_layout.addWidget(self.btn_export_gif)

        export_layout.addLayout(export_buttons_layout)

        # Export FPS control
        export_fps_layout = QHBoxLayout()
        export_fps_layout.addWidget(QLabel("Export FPS:"))

        self.export_fps_spin = QSpinBox()
        self.export_fps_spin.setRange(1, 60)
        self.export_fps_spin.setValue(10)
        self.export_fps_spin.setToolTip("Frames per second for exported video/GIF")
        export_fps_layout.addWidget(self.export_fps_spin)
        export_fps_layout.addStretch()

        export_layout.addLayout(export_fps_layout)

        layout.addWidget(export_group)

        # Info display
        self.viewer_info_text = QTextEdit()
        self.viewer_info_text.setReadOnly(True)
        self.viewer_info_text.setMaximumHeight(100)
        self.viewer_info_text.setPlaceholderText(
            "Frame information will appear here..."
        )
        layout.addWidget(self.viewer_info_text)

        # === SYNCHRONIZED ANALYSIS PLOTS ===
        sync_plot_group = QGroupBox("Synchronized Analysis Plots")
        sync_plot_layout = QVBoxLayout()
        sync_plot_group.setLayout(sync_plot_layout)

        # Enable checkbox
        self.sync_plots_enabled = QCheckBox("Show synchronized plots with time marker")
        self.sync_plots_enabled.setChecked(False)
        self.sync_plots_enabled.setToolTip(
            "Display analysis results below with a time marker "
            "synchronized to the current frame position"
        )
        self.sync_plots_enabled.toggled.connect(self._toggle_sync_plots)
        sync_plot_layout.addWidget(self.sync_plots_enabled)

        # Plot type selector
        sync_type_layout = QHBoxLayout()
        sync_type_layout.addWidget(QLabel("Plot Type:"))
        self.sync_plot_type = QComboBox()
        self.sync_plot_type.addItems([
            "Raw Intensity (0-1)",
            "Movement (binary)",
            "Fraction Movement",
            "Quiescence",
            "Sleep",
            "Sleep Quality",
        ])
        self.sync_plot_type.currentIndexChanged.connect(self._update_sync_plot)
        self.sync_plot_type.setCurrentIndex(2)  # Default to Fraction Movement
        sync_type_layout.addWidget(self.sync_plot_type)

        # Binning control for synchronized plots
        sync_type_layout.addWidget(QLabel("Bin:"))
        self.sync_plot_bin = QSpinBox()
        self.sync_plot_bin.setRange(0, 240)
        self.sync_plot_bin.setValue(0)
        self.sync_plot_bin.setSuffix(" min")
        self.sync_plot_bin.setSpecialValueText("Original")
        self.sync_plot_bin.setToolTip("Re-bin data for visualization (0 = original binning)")
        self.sync_plot_bin.valueChanged.connect(self._update_sync_plot)
        sync_type_layout.addWidget(self.sync_plot_bin)

        sync_type_layout.addStretch()
        sync_plot_layout.addLayout(sync_type_layout)

        # Matplotlib canvas for synchronized plots
        from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
        from matplotlib.figure import Figure

        self.sync_figure = Figure(figsize=(12, 6), dpi=100)
        self.sync_canvas = FigureCanvasQTAgg(self.sync_figure)
        self.sync_canvas.setMinimumHeight(350)
        self.sync_canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.sync_canvas.setVisible(False)  # Hidden by default
        sync_plot_layout.addWidget(self.sync_canvas, stretch=1)

        # Store time marker line references
        self.sync_time_markers = []

        # Export buttons for video/GIF with synchronized plots
        export_sync_layout = QHBoxLayout()
        self.btn_export_sync_video = QPushButton("Export Video with Plots (MP4)")
        self.btn_export_sync_video.setToolTip(
            "Export video showing frame + synchronized analysis plots side by side"
        )
        self.btn_export_sync_video.clicked.connect(self._export_video_with_plots)
        self.btn_export_sync_video.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 8px; }"
            "QPushButton:hover { background-color: #45a049; }"
        )
        export_sync_layout.addWidget(self.btn_export_sync_video)

        self.btn_export_sync_gif = QPushButton("Export GIF with Plots")
        self.btn_export_sync_gif.setToolTip(
            "Export animated GIF showing frame + synchronized analysis plots"
        )
        self.btn_export_sync_gif.clicked.connect(self._export_gif_with_plots)
        self.btn_export_sync_gif.setStyleSheet(
            "QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 8px; }"
            "QPushButton:hover { background-color: #1976D2; }"
        )
        export_sync_layout.addWidget(self.btn_export_sync_gif)

        export_sync_layout.addStretch()
        sync_plot_layout.addLayout(export_sync_layout)

        layout.addWidget(sync_plot_group)

        layout.addStretch()

        # Timer for playback
        from qtpy.QtCore import QTimer

        self.viewer_timer = QTimer()
        self.viewer_timer.timeout.connect(self._viewer_play_next_frame)
        self.viewer_is_playing = False

    # =================================================================
    # HDF5 TELEMETRY TAB
    # =================================================================

    # Timeseries categories and units for telemetry display
    TELEMETRY_CATEGORIES = {
        "Timing": [
            "frame_drift", "cumulative_drift", "actual_intervals", "expected_intervals",
        ],
        "Environment": ["temperature", "humidity"],
        "LED": [
            "led_white_power_percent", "led_ir_power_percent", "led_power_percent",
            "led_duration_ms", "led_sync_success",
        ],
        "Frame Stats": ["frame_mean", "frame_max", "frame_min", "frame_std"],
    }

    TIMESERIES_UNITS = {
        "frame_drift": "s", "cumulative_drift": "s", "actual_intervals": "s",
        "expected_intervals": "s", "temperature": "\u00b0C", "humidity": "%",
        "led_white_power_percent": "%", "led_ir_power_percent": "%",
        "led_power_percent": "%", "led_duration_ms": "ms",
        "led_sync_success": "bool", "frame_mean": "px intensity",
        "frame_max": "px intensity", "frame_min": "px intensity",
        "frame_std": "px intensity",
    }

    def setup_telemetry_tab(self):
        """Setup the HDF5 Telemetry tab for viewing file metadata and timeseries."""
        layout = QVBoxLayout()
        self.tab_telemetry.setLayout(layout)

        # --- Section 1: File & Frame Metadata ---
        meta_group = QGroupBox("File & Frame Metadata")
        meta_layout = QVBoxLayout()
        meta_group.setLayout(meta_layout)

        # Load button
        btn_row = QHBoxLayout()
        self.btn_load_telemetry = QPushButton("Load Metadata")
        self.btn_load_telemetry.setToolTip("Read metadata and timeseries info from the loaded HDF5 file")
        self.btn_load_telemetry.setStyleSheet(
            "QPushButton { background-color: #2196F3; color: white; font-weight: bold; "
            "padding: 6px 16px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #1976D2; }"
        )
        self.btn_load_telemetry.clicked.connect(self._load_telemetry_metadata)
        btn_row.addWidget(self.btn_load_telemetry)
        btn_row.addStretch()
        meta_layout.addLayout(btn_row)

        # Tree widget for metadata display
        self.telemetry_tree = QTreeWidget()
        self.telemetry_tree.setHeaderLabels(["Property", "Value"])
        self.telemetry_tree.setAlternatingRowColors(True)
        self.telemetry_tree.setMinimumHeight(200)
        self.telemetry_tree.header().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.telemetry_tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        meta_layout.addWidget(self.telemetry_tree)

        layout.addWidget(meta_group)

        # --- Section 2: Timeseries Plots ---
        ts_group = QGroupBox("Timeseries Data")
        ts_layout = QVBoxLayout()
        ts_group.setLayout(ts_layout)

        # Top row: dataset selector (left) + controls (right)
        selector_and_controls = QHBoxLayout()

        # Left: multi-select list of individual timeseries
        select_panel = QVBoxLayout()
        select_panel.addWidget(QLabel("Select datasets to plot:"))

        self.telemetry_list = QListWidget()
        self.telemetry_list.setSelectionMode(QListWidget.MultiSelection)
        self.telemetry_list.setMaximumHeight(160)
        self.telemetry_list.setToolTip(
            "Click to select/deselect individual timeseries. "
            "Use the category buttons for quick selection."
        )
        select_panel.addWidget(self.telemetry_list)

        # Quick-select buttons for categories
        cat_btn_row = QHBoxLayout()
        btn_sel_all = QPushButton("All")
        btn_sel_all.setToolTip("Select all timeseries")
        btn_sel_all.clicked.connect(lambda: self._select_telemetry_items(all_=True))
        btn_sel_none = QPushButton("None")
        btn_sel_none.setToolTip("Deselect all timeseries")
        btn_sel_none.clicked.connect(lambda: self._select_telemetry_items(all_=False))
        cat_btn_row.addWidget(btn_sel_all)
        cat_btn_row.addWidget(btn_sel_none)

        for cat_name in list(self.TELEMETRY_CATEGORIES.keys()) + ["Other"]:
            btn = QPushButton(cat_name)
            btn.setToolTip(f"Select only {cat_name} timeseries")
            btn.clicked.connect(lambda checked, c=cat_name: self._select_telemetry_category(c))
            cat_btn_row.addWidget(btn)
        cat_btn_row.addStretch()
        select_panel.addLayout(cat_btn_row)

        selector_and_controls.addLayout(select_panel, stretch=1)

        # Right: x-axis control
        right_panel = QVBoxLayout()
        right_panel.addWidget(QLabel("X-Axis:"))
        self.telemetry_xaxis_combo = QComboBox()
        self.telemetry_xaxis_combo.addItems(["Frame Number", "Time (minutes)", "Time (hours)"])
        self.telemetry_xaxis_combo.setCurrentIndex(1)
        right_panel.addWidget(self.telemetry_xaxis_combo)
        right_panel.addStretch()
        selector_and_controls.addLayout(right_panel)

        ts_layout.addLayout(selector_and_controls)

        # Buttons row: Plot + Export
        btn_row2 = QHBoxLayout()
        self.btn_plot_telemetry = QPushButton("Plot Telemetry")
        self.btn_plot_telemetry.setToolTip("Plot selected timeseries")
        self.btn_plot_telemetry.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; font-weight: bold; "
            "padding: 6px 16px; border-radius: 4px; }"
            "QPushButton:hover { background-color: #388E3C; }"
        )
        self.btn_plot_telemetry.clicked.connect(self._plot_telemetry)
        btn_row2.addWidget(self.btn_plot_telemetry)

        self.btn_save_telemetry_plot = QPushButton("Save Current Plot")
        self.btn_save_telemetry_plot.setToolTip("Save the current telemetry plot as PNG/PDF/SVG")
        self.btn_save_telemetry_plot.clicked.connect(self._save_telemetry_plot)
        btn_row2.addWidget(self.btn_save_telemetry_plot)

        self.btn_export_all_telemetry = QPushButton("Export All Plots")
        self.btn_export_all_telemetry.setToolTip(
            "Save each timeseries as individual plot + one combined overview"
        )
        self.btn_export_all_telemetry.clicked.connect(self._export_all_telemetry_plots)
        btn_row2.addWidget(self.btn_export_all_telemetry)

        btn_row2.addStretch()
        ts_layout.addLayout(btn_row2)

        # Matplotlib canvas for timeseries plots
        self.telemetry_figure = Figure(figsize=(10, 8), dpi=100)
        self.telemetry_canvas = FigureCanvas(self.telemetry_figure)
        self.telemetry_canvas.setMinimumHeight(500)
        ts_layout.addWidget(self.telemetry_canvas)

        layout.addWidget(ts_group)

        # Storage for loaded telemetry data
        self.telemetry_timeseries = {}
        self.telemetry_frame_interval = None

    def _load_telemetry_metadata(self):
        """Load and display HDF5 file metadata in the telemetry tree."""
        file_path = getattr(self, "file_path", None)
        if not file_path:
            self._log_message("No HDF5 file loaded. Please load a file first.")
            return

        try:
            import h5py

            self.telemetry_tree.clear()
            self._log_message(f"Loading telemetry from: {os.path.basename(file_path)}")

            with h5py.File(file_path, "r") as f:
                # --- File Info ---
                file_item = QTreeWidgetItem(self.telemetry_tree, ["File Info", ""])
                file_item.setExpanded(True)
                file_size = os.path.getsize(file_path)
                QTreeWidgetItem(file_item, ["Filename", os.path.basename(file_path)])
                QTreeWidgetItem(file_item, ["Size", f"{file_size / (1024*1024):.1f} MB"])
                QTreeWidgetItem(file_item, ["HDF5 Driver", str(f.driver)])

                # --- Root Attributes ---
                if f.attrs:
                    attr_item = QTreeWidgetItem(self.telemetry_tree, ["Root Attributes", f"({len(f.attrs)} entries)"])
                    attr_item.setExpanded(True)
                    for key in sorted(f.attrs.keys()):
                        val = f.attrs[key]
                        if isinstance(val, bytes):
                            val = val.decode("utf-8", errors="replace")
                        QTreeWidgetItem(attr_item, [str(key), str(val)])

                # --- Datasets ---
                datasets_item = QTreeWidgetItem(self.telemetry_tree, ["Datasets", ""])
                datasets_item.setExpanded(True)
                self._add_hdf5_items_to_tree(f, datasets_item)

                # --- Timeseries: read arrays for plotting ---
                self.telemetry_timeseries = {}
                self.telemetry_frame_interval = None

                if "timeseries" in f:
                    ts_group = f["timeseries"]
                    ts_tree_item = QTreeWidgetItem(self.telemetry_tree, [
                        "Timeseries", f"({len(ts_group)} datasets)"
                    ])
                    ts_tree_item.setExpanded(True)

                    # Read frame interval from root attrs
                    for attr_name in ["frame_interval", "interval", "fps"]:
                        if attr_name in f.attrs:
                            val = float(f.attrs[attr_name])
                            if attr_name == "fps" and val > 0:
                                self.telemetry_frame_interval = 1.0 / val
                            else:
                                self.telemetry_frame_interval = val
                            break

                    for ds_name in sorted(ts_group.keys()):
                        ds = ts_group[ds_name]
                        if hasattr(ds, "shape"):
                            # Store array data for plotting
                            try:
                                self.telemetry_timeseries[ds_name] = ds[:]
                            except Exception:
                                pass

                            unit = self.TIMESERIES_UNITS.get(ds_name, "")
                            unit_str = f" [{unit}]" if unit else ""
                            shape_str = f"shape={ds.shape}, dtype={ds.dtype}"
                            QTreeWidgetItem(ts_tree_item, [
                                f"{ds_name}{unit_str}", shape_str
                            ])

                            # Show attributes of timeseries datasets
                            if ds.attrs:
                                ds_tree = ts_tree_item.child(ts_tree_item.childCount() - 1)
                                for ak in ds.attrs:
                                    av = ds.attrs[ak]
                                    if isinstance(av, bytes):
                                        av = av.decode("utf-8", errors="replace")
                                    QTreeWidgetItem(ds_tree, [f"  @{ak}", str(av)])

            n_ts = len(self.telemetry_timeseries)
            self._log_message(f"Telemetry loaded: {n_ts} timeseries datasets found")

            # Populate the dataset selector list
            self._populate_telemetry_list()

            # Auto-plot all loaded timeseries
            if self.telemetry_timeseries:
                self._plot_telemetry()

        except Exception as e:
            self._log_message(f"Error loading telemetry: {e}")
            import traceback
            self._log_message(traceback.format_exc())

    def _add_hdf5_items_to_tree(self, h5_group, parent_item):
        """Recursively add HDF5 groups and datasets to the tree widget."""
        import h5py

        for key in sorted(h5_group.keys()):
            item = h5_group[key]
            if isinstance(item, h5py.Group):
                if key == "timeseries":
                    continue  # Shown separately
                group_node = QTreeWidgetItem(parent_item, [
                    f"{key}/", f"(Group, {len(item)} items)"
                ])
                self._add_hdf5_items_to_tree(item, group_node)
            elif isinstance(item, h5py.Dataset):
                info_parts = [f"shape={item.shape}", f"dtype={item.dtype}"]
                if item.compression:
                    info_parts.append(f"compression={item.compression}")
                QTreeWidgetItem(parent_item, [key, ", ".join(info_parts)])

    def _get_category_for_dataset(self, ds_name: str) -> str:
        """Return the category name for a given timeseries dataset name."""
        for cat_name, ds_list in self.TELEMETRY_CATEGORIES.items():
            if ds_name in ds_list:
                return cat_name
        return "Other"

    def _populate_telemetry_list(self):
        """Populate the dataset selector list with all available timeseries."""
        self.telemetry_list.clear()
        if not self.telemetry_timeseries:
            return

        # Group by category for ordered display
        categorized = {}
        for ds_name in self.telemetry_timeseries:
            cat = self._get_category_for_dataset(ds_name)
            categorized.setdefault(cat, []).append(ds_name)

        # Add items in category order
        cat_order = list(self.TELEMETRY_CATEGORIES.keys()) + ["Other"]
        for cat_name in cat_order:
            if cat_name not in categorized:
                continue
            for ds_name in categorized[cat_name]:
                unit = self.TIMESERIES_UNITS.get(ds_name, "")
                unit_str = f" [{unit}]" if unit else ""
                item = QListWidgetItem(f"[{cat_name}]  {ds_name}{unit_str}")
                item.setData(Qt.UserRole, ds_name)  # store raw name for lookup
                item.setSelected(True)
                self.telemetry_list.addItem(item)

    def _select_telemetry_items(self, all_: bool):
        """Select or deselect all items in the telemetry list."""
        for i in range(self.telemetry_list.count()):
            self.telemetry_list.item(i).setSelected(all_)

    def _select_telemetry_category(self, category: str):
        """Select only items belonging to the given category."""
        for i in range(self.telemetry_list.count()):
            item = self.telemetry_list.item(i)
            ds_name = item.data(Qt.UserRole)
            cat = self._get_category_for_dataset(ds_name)
            item.setSelected(cat == category)

    def _get_active_telemetry_series(self):
        """Return list of (category, dataset_name) for all selected items in the list."""
        active_series = []
        for item in self.telemetry_list.selectedItems():
            ds_name = item.data(Qt.UserRole)
            if ds_name in self.telemetry_timeseries:
                cat = self._get_category_for_dataset(ds_name)
                active_series.append((cat, ds_name))
        return active_series

    def _build_telemetry_xaxis(self, n_points):
        """Build x-axis array and label based on current x-axis mode selection."""
        import numpy as np

        xaxis_mode = self.telemetry_xaxis_combo.currentText()
        if xaxis_mode == "Frame Number":
            return np.arange(n_points), "Frame Number"
        elif xaxis_mode == "Time (hours)":
            interval = self.telemetry_frame_interval or 5.0
            return np.arange(n_points) * interval / 3600.0, "Time (hours)"
        else:  # Time (minutes)
            interval = self.telemetry_frame_interval or 5.0
            return np.arange(n_points) * interval / 60.0, "Time (minutes)"

    # Category colors for telemetry plots
    TELEMETRY_CAT_COLORS = {
        "Timing": "#2196F3",
        "Environment": "#FF9800",
        "LED": "#9C27B0",
        "Frame Stats": "#4CAF50",
        "Other": "#795548",
    }

    def _plot_telemetry(self):
        """Plot selected timeseries categories from the loaded HDF5 telemetry data."""
        if not self.telemetry_timeseries:
            self._log_message("No telemetry data loaded. Click 'Load Metadata' first.")
            return

        active_series = self._get_active_telemetry_series()
        if not active_series:
            self._log_message("No timeseries data available for selected categories.")
            return

        n_plots = len(active_series)

        # Clear and create subplots
        self.telemetry_figure.clear()

        # Dynamically adjust figure height based on number of plots
        height_per_plot = 1.5
        min_height = 4.0
        fig_height = max(min_height, n_plots * height_per_plot)
        self.telemetry_figure.set_size_inches(10, fig_height)
        self.telemetry_canvas.setMinimumHeight(int(fig_height * 100))

        axes = self.telemetry_figure.subplots(n_plots, 1, sharex=True, squeeze=False)

        x_label = "Frame Number"
        for idx, (cat_name, ds_name) in enumerate(active_series):
            ax = axes[idx, 0]
            data = self.telemetry_timeseries[ds_name]
            x, x_label = self._build_telemetry_xaxis(len(data))

            color = self.TELEMETRY_CAT_COLORS.get(cat_name, "#666666")
            unit = self.TIMESERIES_UNITS.get(ds_name, "")
            unit_str = f" [{unit}]" if unit else ""

            mean_val = float(np.mean(data))
            std_val = float(np.std(data))
            stat_label = f"μ = {mean_val:.3g}{' ' + unit if unit else ''},  σ = {std_val:.3g}"

            ax.plot(x, data, color=color, linewidth=0.8, alpha=0.9, label=stat_label)
            ax.legend(fontsize=7, loc="upper right", framealpha=0.7)

            ax.set_ylabel(f"{ds_name}{unit_str}", fontsize=8)
            ax.tick_params(labelsize=7)
            ax.grid(True, alpha=0.3)

            # Category label on right side
            ax.text(
                1.01, 0.5, cat_name, transform=ax.transAxes,
                fontsize=7, color=color, alpha=0.7, va="center", rotation=270,
            )

        # X-axis label on bottom plot only
        if n_plots > 0:
            axes[-1, 0].set_xlabel(x_label, fontsize=9)

        self.telemetry_figure.suptitle(
            f"HDF5 Telemetry — {os.path.basename(getattr(self, 'file_path', '') or '')}",
            fontsize=10, y=1.0,
        )
        self.telemetry_figure.tight_layout()
        self.telemetry_canvas.draw()
        self._log_message(f"Plotted {n_plots} telemetry timeseries ({x_label})")

    def _save_telemetry_plot(self):
        """Save the current telemetry plot to a file."""
        if not self.telemetry_timeseries:
            self._log_message("No telemetry plot to save.")
            return

        try:
            base_name = os.path.splitext(
                os.path.basename(getattr(self, "file_path", "") or "telemetry")
            )[0]
            default_name = f"{base_name}_telemetry.png"
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save Telemetry Plot", default_name,
                "PNG Files (*.png);;PDF Files (*.pdf);;SVG Files (*.svg);;All Files (*)",
            )
            if file_path:
                ext = os.path.splitext(file_path)[1].lower()
                fmt = "pdf" if ext == ".pdf" else ("svg" if ext == ".svg" else "png")
                self.telemetry_figure.savefig(
                    file_path, format=fmt, dpi=300, bbox_inches="tight"
                )
                self._log_message(f"Telemetry plot saved: {file_path}")
        except Exception as e:
            self._log_message(f"Error saving telemetry plot: {e}")

    def _export_all_telemetry_plots(self):
        """Export individual plots for each timeseries + one combined overview."""
        if not self.telemetry_timeseries:
            self._log_message("No telemetry data loaded. Click 'Load Metadata' first.")
            return

        import numpy as np

        # Ask for output directory
        out_dir = QFileDialog.getExistingDirectory(
            self, "Select Output Directory for Telemetry Plots"
        )
        if not out_dir:
            return

        base_name = os.path.splitext(
            os.path.basename(getattr(self, "file_path", "") or "telemetry")
        )[0]

        saved_count = 0

        try:
            # Get all available series (ignore category filters for full export)
            all_known = set()
            for ds_list in self.TELEMETRY_CATEGORIES.values():
                all_known.update(ds_list)

            all_series = []
            for cat_name, ds_names in self.TELEMETRY_CATEGORIES.items():
                for ds_name in ds_names:
                    if ds_name in self.telemetry_timeseries:
                        all_series.append((cat_name, ds_name))
            for ds_name in sorted(self.telemetry_timeseries.keys()):
                if ds_name not in all_known:
                    all_series.append(("Other", ds_name))

            if not all_series:
                self._log_message("No timeseries data to export.")
                return

            # 1. Individual plots for each timeseries
            for cat_name, ds_name in all_series:
                fig, ax = Figure(figsize=(10, 3), dpi=150), None
                ax = fig.add_subplot(111)
                data = self.telemetry_timeseries[ds_name]
                x, x_label = self._build_telemetry_xaxis(len(data))

                color = self.TELEMETRY_CAT_COLORS.get(cat_name, "#666666")
                ax.plot(x, data, color=color, linewidth=0.8, alpha=0.9)

                unit = self.TIMESERIES_UNITS.get(ds_name, "")
                unit_str = f" [{unit}]" if unit else ""
                ax.set_ylabel(f"{ds_name}{unit_str}", fontsize=9)
                ax.set_xlabel(x_label, fontsize=9)
                ax.set_title(f"{ds_name} ({cat_name})", fontsize=10)
                ax.grid(True, alpha=0.3)
                fig.tight_layout()

                out_path = os.path.join(out_dir, f"{base_name}_{ds_name}.png")
                fig.savefig(out_path, format="png", dpi=150, bbox_inches="tight")
                fig.clear()
                import matplotlib.pyplot as plt
                plt.close(fig)
                saved_count += 1

            # 2. Combined overview plot with all timeseries
            n_plots = len(all_series)
            height_per = 1.8
            fig_h = max(6.0, n_plots * height_per)
            fig_combined = Figure(figsize=(12, fig_h), dpi=150)
            axes = fig_combined.subplots(n_plots, 1, sharex=True, squeeze=False)

            x_label = "Frame Number"
            for idx, (cat_name, ds_name) in enumerate(all_series):
                ax = axes[idx, 0]
                data = self.telemetry_timeseries[ds_name]
                x, x_label = self._build_telemetry_xaxis(len(data))

                color = self.TELEMETRY_CAT_COLORS.get(cat_name, "#666666")
                ax.plot(x, data, color=color, linewidth=0.8, alpha=0.9)

                unit = self.TIMESERIES_UNITS.get(ds_name, "")
                unit_str = f" [{unit}]" if unit else ""
                ax.set_ylabel(f"{ds_name}{unit_str}", fontsize=7)
                ax.tick_params(labelsize=6)
                ax.grid(True, alpha=0.3)
                ax.text(
                    1.01, 0.5, cat_name, transform=ax.transAxes,
                    fontsize=6, color=color, alpha=0.7, va="center", rotation=270,
                )

            if n_plots > 0:
                axes[-1, 0].set_xlabel(x_label, fontsize=9)
            fig_combined.suptitle(
                f"HDF5 Telemetry Overview — {base_name}", fontsize=11, y=1.0,
            )
            fig_combined.tight_layout()

            combined_path = os.path.join(out_dir, f"{base_name}_telemetry_all.png")
            fig_combined.savefig(
                combined_path, format="png", dpi=150, bbox_inches="tight"
            )
            fig_combined.clear()
            import matplotlib.pyplot as plt
            plt.close(fig_combined)
            saved_count += 1

            self._log_message(
                f"Exported {saved_count} telemetry plots to: {out_dir}"
            )

        except Exception as e:
            self._log_message(f"Error exporting telemetry plots: {e}")
            import traceback
            self._log_message(traceback.format_exc())

    def debug_current_file_structure(self):
        """Debug the structure of the currently loaded file."""
        if not hasattr(self, "file_path") or not self.file_path:
            self._log_message("No file loaded for structure debugging")
            return

        self._log_message("=== DEBUGGING CURRENT FILE STRUCTURE ===")

        if DUAL_STRUCTURE_AVAILABLE:
            try:
                structure_info = detect_hdf5_structure_type(self.file_path)

                self._log_message(f"Structure type: {structure_info['type']}")

                if structure_info["type"] == "stacked_frames":
                    self._log_message("✅ Stacked frames detected")
                    self._log_message(f"   Dataset: '{structure_info['dataset_name']}'")
                    self._log_message(
                        f"   Frame count: {structure_info['frame_count']}"
                    )
                    self._log_message(
                        f"   Frame shape: {structure_info['frame_shape']}"
                    )
                    self._log_message(f"   Data type: {structure_info['dtype']}")

                elif structure_info["type"] == "individual_frames":
                    self._log_message("✅ Individual frames detected")
                    self._log_message(f"   Group: '{structure_info['group_name']}'")
                    self._log_message(
                        f"   Frame count: {structure_info['frame_count']}"
                    )
                    self._log_message(
                        f"   Frame shape: {structure_info['frame_shape']}"
                    )
                    self._log_message(f"   Data type: {structure_info['dtype']}")

                    # Show sample frame keys
                    if "frame_keys" in structure_info:
                        sample_keys = structure_info["frame_keys"][:10]
                        self._log_message(f"   Sample keys: {sample_keys}")
                        if len(structure_info["frame_keys"]) > 10:
                            self._log_message(
                                f"   ... and {len(structure_info['frame_keys']) - 10} more"
                            )

                elif structure_info["type"] == "error":
                    self._log_message(f"❌ Error: {structure_info['error']}")

                    # Fallback to basic structure info
                    try:
                        import h5py

                        with h5py.File(self.file_path, "r") as f:
                            self._log_message(f"Available keys: {list(f.keys())}")
                    except Exception as e2:
                        self._log_message(f"Cannot read file: {e2}")

            except Exception as e:
                self._log_message(f"Structure debugging failed: {e}")
        else:
            self._log_message(
                "Dual structure support not available - using basic debugging"
            )
            try:
                import h5py

                with h5py.File(self.file_path, "r") as f:
                    self._log_message(f"Root keys: {list(f.keys())}")

                    if "frames" in f:
                        self._log_message(
                            f"Found 'frames' dataset: shape {f['frames'].shape}"
                        )
                    if "images" in f:
                        self._log_message(
                            f"Found 'images' group with {len(f['images'].keys())} items"
                        )
                    if "timeseries" in f:
                        self._log_message(
                            f"Found 'timeseries' group with {len(f['timeseries'].keys())} items"
                        )

            except Exception as e:
                self._log_message(f"Basic debugging failed: {e}")

    def _connect_signals(self):
        """Connect all UI signals to their respective methods."""
        # Progress signals
        self.progress_updated.connect(self._on_progress_update)
        self.status_updated.connect(self._on_status_update)
        self.performance_updated.connect(self._on_performance_update)

        # File operations
        self.btn_load_file.clicked.connect(self.load_file)
        self.btn_load_dir.clicked.connect(self.load_directory)
        self.btn_detect_rois.clicked.connect(self.enhanced_detect_rois)
        self.btn_clear_rois.clicked.connect(self.clear_roi_detection)

        # Plate preset checkboxes - make mutually exclusive
        self.chk_6well.stateChanged.connect(self._on_6well_preset_changed)
        self.chk_12well.stateChanged.connect(self._on_12well_preset_changed)

        # ROI scale button
        self.btn_apply_scale.clicked.connect(self._apply_roi_scale)

        # NEW: Calibration workflow connections
        self.btn_load_calibration.clicked.connect(self.load_calibration_file)
        self.btn_load_calibration_dataset.clicked.connect(
            self.enhanced_load_calibration_dataset
        )
        self.btn_process_calibration_baseline.clicked.connect(
            self.process_calibration_baseline
        )
        # Analysis operations
        self.btn_analyze.clicked.connect(self.run_analysis)
        self.btn_stop.clicked.connect(self.stop_analysis)
        self.btn_reset_analysis.clicked.connect(self.reset_for_new_analysis)
        # Testing and diagnostics
        self.btn_quick_test.clicked.connect(self.run_quick_analysis_test)
        self.btn_validate_timing.clicked.connect(self.validate_hdf5_timing)

        # ===== SIMPLIFIED PLOTTING OPERATIONS =====
        self.plot_type_combo.currentIndexChanged.connect(self._on_plot_type_changed)
        self.btn_plot.clicked.connect(self.generate_plot)
        self.btn_save_plot.clicked.connect(self.save_current_plot)
        self.btn_save_all_plots.clicked.connect(self.save_all_plots)
        self.btn_save_results.clicked.connect(
            self.save_results_consolidated_complete
        )  # NEW CONSOLIDATED METHOD
        self.btn_save_with_metadata.clicked.connect(self.save_results_with_metadata)
        self.btn_apply_time_range.clicked.connect(self.apply_time_range)

        # Amplitude mode toggle
        self.show_real_amplitude.toggled.connect(self.generate_plot)

        # Y-Axis scaling controls
        self.auto_scale_y.toggled.connect(self._on_auto_scale_toggled)
        self.robust_scaling.toggled.connect(self.generate_plot)
        self.adaptive_scaling.toggled.connect(self.generate_plot)
        self.center_around_zero.toggled.connect(self.generate_plot)
        self.lower_percentile_spin.valueChanged.connect(self.generate_plot)
        self.upper_percentile_spin.valueChanged.connect(self.generate_plot)
        self.btn_apply_y_range.clicked.connect(self.generate_plot)

        # Threshold visualization signals
        self.show_baseline_mean.toggled.connect(self.generate_plot)
        self.show_deviation_band.toggled.connect(self.generate_plot)
        self.show_detection_threshold.toggled.connect(self.generate_plot)
        self.show_threshold_stats.toggled.connect(self.generate_plot)

        # UI interactions
        self.tab_widget.currentChanged.connect(self.on_tab_changed)
        self.frame_interval.valueChanged.connect(self.update_end_time)
        self.threshold_params_stack.currentChanged.connect(
            self._on_threshold_tab_changed
        )
        self.chk_6well.toggled.connect(self._on_6well_toggled)
        self.chk_12well.toggled.connect(self._on_12well_toggled)

    # ===================================================================
    # FILE LOADING AND ROI DETECTION METHODS
    # ===================================================================
    def load_file(self):
        """Load HDF5 or AVI file(s) with automatic detection."""
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select File(s)",
            "",
            "Video Files (*.h5 *.hdf5 *.avi);;HDF5 Files (*.h5 *.hdf5);;AVI Files (*.avi);;All Files (*)",
        )
        if not file_paths:
            return

        # Handle single or multiple files
        if len(file_paths) == 1:
            file_path = file_paths[0]

            # Check if single AVI file - load as single video (not batch)
            if file_path.lower().endswith(".avi"):
                self._load_single_avi(file_path)
                return
        else:
            # Multiple files - check if they are AVIs for batch processing
            if all(f.lower().endswith(".avi") for f in file_paths):
                self._load_avi_batch(file_paths)
                return
            else:
                self._log_message(
                    "Multiple file selection only supported for AVI files"
                )
                return

        self.file_path = file_path
        self.recording_start_datetime = _parse_recording_start_datetime(file_path)
        basename = os.path.basename(file_path)

        # === AUTOMATISCHE LEGACY-DETECTION BEIM LADEN ===
        try:
            # Quick legacy check
            with h5py.File(file_path, "r") as f:
                is_legacy = self._quick_legacy_check(f)

            if is_legacy:
                self._log_message(f"Legacy file detected: {basename}")
                self._log_message(
                    "   Will automatically enhance with unit documentation during analysis"
                )
                self.lbl_file_info.setText(
                    f"Loaded LEGACY file: {basename} (auto-enhancement enabled)"
                )
            else:
                self._log_message(f"Modern file detected: {basename}")
                self.lbl_file_info.setText(f"Loaded file: {basename}")

        except Exception as e:
            self._log_message(f"Could not determine file type: {e}")
            self.lbl_file_info.setText(f"Loaded file: {basename}")

        # Clear any existing ROI detection
        self.masks = []

        # Enhanced structure detection and loading
        if DUAL_STRUCTURE_AVAILABLE:
            try:
                # Detect structure first
                structure_info = detect_hdf5_structure_type(file_path)
                self._log_message(f"Detected HDF5 structure: {structure_info['type']}")

                if structure_info["type"] == "error":
                    self._log_message(
                        f"Structure detection failed: {structure_info['error']}"
                    )
                    return

                self._log_message(f"Frame count: {structure_info['frame_count']}")
                self._log_message(f"Frame shape: {structure_info['frame_shape']}")
                self._log_message(f"Data location: {structure_info['data_location']}")

                # Use dual structure reader
                reader = reader_function_dual_structure
                self._log_message("Using enhanced dual structure reader")

            except Exception as e:
                self._log_message(f"Enhanced reader failed, using fallback: {e}")
                reader = napari_get_reader(file_path)
        else:
            # Use original reader
            reader = napari_get_reader(file_path)

        if reader is None:
            self._log_message("No valid HDF5 reader available.")
            return

        try:
            # Clear existing layers
            self.viewer.layers.clear()

            # Load layers from reader
            layers = reader(file_path)
            for data, meta, layer_type in layers:
                name = meta.get("name", basename)
                kwargs = {k: v for k, v in meta.items() if k not in ("name",)}

                if layer_type == "image":
                    self.viewer.add_image(data, name=name, **kwargs)
                elif layer_type == "labels":
                    self.viewer.add_labels(data, name=name, **kwargs)

            # Log structure information if available
            if layers and "metadata" in layers[0][1]:
                metadata = layers[0][1]["metadata"]
                if "structure_type" in metadata:
                    structure_type = metadata["structure_type"]
                    frame_count = metadata.get("frame_count", "unknown")
                    self._log_message(
                        f"Successfully loaded {structure_type} structure with {frame_count} frames"
                    )

        except Exception as e:
            self._log_message(f"Reader error: {e}")
            return

        # Update end time for analysis parameters
        self.update_end_time()
        self.check_hdf5_structure()

        # Update frame viewer file selector
        self._update_viewer_file_combo()

    def _load_single_avi(self, file_path: str):
        """Load a single AVI file (only first frame for ROI detection)."""
        try:
            from ._avi_reader import AVIVideoReader

            self.file_path = file_path
            self.recording_start_datetime = _parse_recording_start_datetime(file_path)
            basename = os.path.basename(file_path)

            self._log_message(f"Loading AVI file: {basename}")

            # Store for later processing
            self.avi_batch_paths = [file_path]  # Single file as batch
            self.avi_batch_interval = 5.0  # Default frame interval

            # Get metadata without loading all frames
            with AVIVideoReader(file_path) as reader:
                # Load ONLY first frame for ROI detection
                first_frame = reader.get_frame(0)
                if first_frame is None:
                    raise ValueError("Could not load first frame")

                # Log frame info for debugging
                self._log_message(f"First frame shape: {first_frame.shape}")
                self._log_message(f"First frame dtype: {first_frame.dtype}")
                self._log_message(
                    f"First frame value range: {first_frame.min()}-{first_frame.max()}"
                )

                # Calculate estimated frames
                video_fps = reader.fps
                target_interval = reader.metadata.get("frame_interval", 5.0)
                self.avi_batch_interval = target_interval
                frames_per_sample = max(1, int(video_fps * target_interval))
                frame_count_estimate = len(
                    range(0, reader.frame_count, frames_per_sample)
                )

                metadata = {
                    "source_type": "avi_single",
                    "fps": reader.fps,
                    "frame_interval": target_interval,
                    "frame_count": reader.frame_count,
                    "frame_count_estimate": frame_count_estimate,
                    "duration": reader.duration,
                    "resolution": {"width": reader.width, "height": reader.height},
                    "source_path": file_path,
                    "frames_per_sample": frames_per_sample,
                }

            # Clear existing layers
            self.viewer.layers.clear()

            # Add only first frame to napari
            self._log_message("Adding first frame to napari viewer...")
            layer = self.viewer.add_image(
                first_frame, name=f"{basename}_first_frame", metadata=metadata
            )
            self._log_message(f"Layer added: {layer.name}, visible: {layer.visible}")

            # Update UI
            duration_min = metadata["duration"] / 60.0
            self.lbl_file_info.setText(
                f"Loaded AVI: {basename} "
                f"({frame_count_estimate} frames estimated, {duration_min:.1f} min, {target_interval}s interval) - First frame only"
            )

            self._log_message("Loaded first frame for ROI detection")
            self._log_message(f"Frames (estimated): {frame_count_estimate}")
            self._log_message(f"Duration: {duration_min:.1f} minutes")
            self._log_message(f"Frame interval: {target_interval}s")
            self._log_message("Note: Full frames will be loaded during processing")

            # Clear any existing ROI detection
            self.masks = []

            # Update end time for analysis
            self.update_end_time()

            # Update frame viewer file selector
            self._update_viewer_file_combo()

        except ImportError:
            self._log_message(
                "Error: AVI support not available. Install opencv-python: pip install opencv-python"
            )
        except Exception as e:
            self._log_message(f"Error loading AVI file: {e}")
            import traceback

            self._log_message(traceback.format_exc())

    def _process_avi_batch_for_analysis(
        self,
        video_paths: List[str],
        masks: List[np.ndarray],
        chunk_size: int,
        progress_callback,
        frame_interval: float,
    ):
        """Process AVI batch with streaming analysis - no need to load all frames."""
        import time
        from ._avi_reader import process_avi_batch_streaming

        start_time = time.time()

        # Stream process all videos - loads and analyzes chunk by chunk
        self._log_message(
            f"Starting streaming analysis of {len(video_paths)} AVI files..."
        )
        self._log_message("Using memory-efficient streaming: load → analyze → discard")

        roi_changes, metadata = process_avi_batch_streaming(
            video_paths,
            masks,
            target_frame_interval=frame_interval,
            chunk_size=chunk_size,
            progress_callback=progress_callback,
        )

        total_duration = metadata["total_duration"]
        proc_time = time.time() - start_time

        # Calculate start and end times from the data
        start_time_data = 0.0  # Start time is always 0
        end_time_data = total_duration

        self._log_message(
            f"✓ AVI batch streaming analysis complete in {proc_time:.2f}s"
        )
        self._log_message(f"  Total frames analyzed: {metadata['total_frames']}")
        self._log_message(
            f"  Total duration: {total_duration:.1f}s ({total_duration/60:.1f}min)"
        )
        self._log_message(f"  Start time: {start_time_data:.1f}s")
        self._log_message(
            f"  End time: {end_time_data:.1f}s ({end_time_data/60:.1f}min)"
        )
        self._log_message(f"  ROIs tracked: {len(roi_changes)}")

        return video_paths[0], roi_changes, total_duration

    def _load_avi_batch(self, file_paths: List[str]):
        """Load multiple AVI files as batch timeseries (only first frame for ROI detection)."""
        self._log_message("=== _load_avi_batch() START ===")
        self._log_message(f"Received {len(file_paths)} file paths")
        for idx, path in enumerate(file_paths):
            self._log_message(f"  [{idx}] {path}")

        try:
            from ._avi_reader import AVIVideoReader

            # Get frame interval from metadata or use default
            target_interval = 5.0  # seconds (same as HDF5)

            self._log_message(f"Loading {len(file_paths)} AVI files as batch...")
            self._log_message(
                f"Target frame interval: {target_interval}s (0.2 FPS effective)"
            )

            # Store batch info for later processing
            self.avi_batch_paths = file_paths
            self.avi_batch_interval = target_interval

            # For memory efficiency: Only open first video to get basic info
            # Full metadata will be calculated during analysis
            self._log_message(
                "Getting metadata from first video only (memory efficient)..."
            )

            batch_metadata = {
                "videos": [],
                "source_type": "avi_batch",
                "target_frame_interval": target_interval,
                "video_count": len(file_paths),
            }

            # Load ONLY first frame from first video for ROI detection
            self._log_message(f"Opening first video: {file_paths[0]}")
            with AVIVideoReader(file_paths[0]) as reader:
                self._log_message("AVIVideoReader opened successfully")
                first_frame = reader.get_frame(0)
                if first_frame is None:
                    raise ValueError("Could not load first frame from first video")

                # Get basic info from first video only
                video_fps = reader.fps
                frames_per_sample = max(1, int(video_fps * target_interval))

                # Store metadata for first video
                batch_metadata["first_video_fps"] = video_fps
                batch_metadata["frames_per_sample"] = frames_per_sample
                batch_metadata["effective_fps"] = 1.0 / target_interval

                # Log frame info for debugging
                self._log_message("First frame loaded successfully")
                self._log_message(f"First frame shape: {first_frame.shape}")
                self._log_message(f"First frame dtype: {first_frame.dtype}")
                self._log_message(
                    f"First frame value range: {first_frame.min()}-{first_frame.max()}"
                )

            # Clear existing layers
            self._log_message(f"Clearing {len(self.viewer.layers)} existing layers...")
            self.viewer.layers.clear()
            self._log_message("Layers cleared")

            # Add only first frame to napari
            self._log_message("Adding first frame to napari viewer...")
            self._log_message(f"Frame data type: {type(first_frame)}")
            self._log_message(
                f"Viewer has {len(self.viewer.layers)} layers before adding"
            )

            layer = self.viewer.add_image(
                first_frame,
                name=f"batch_{len(file_paths)}_videos_first_frame",
                metadata=batch_metadata,
            )

            self._log_message("Layer added successfully!")
            self._log_message(f"Layer name: {layer.name}")
            self._log_message(f"Layer visible: {layer.visible}")
            self._log_message(f"Layer data shape: {layer.data.shape}")
            self._log_message(f"Viewer now has {len(self.viewer.layers)} layers")

            # Store file path (use first file as reference)
            self.file_path = file_paths[0]
            self.recording_start_datetime = _parse_recording_start_datetime(file_paths[0])

            # Update UI - simplified message (full metadata calculated during analysis)
            self.lbl_file_info.setText(
                f"Loaded {len(file_paths)} AVI files as batch "
                f"(~{batch_metadata['effective_fps']:.2f} FPS effective) - First frame only"
            )

            self._log_message("Loaded first frame for ROI detection")
            self._log_message(f"Batch contains {len(file_paths)} video files")
            self._log_message(f"Effective FPS: {batch_metadata['effective_fps']:.2f}")
            self._log_message(f"Frame interval: {target_interval}s")
            self._log_message("Note: Full metadata will be calculated during analysis")
            self._log_message("Note: All frames will be loaded during processing")

            # Clear any existing ROI detection
            self.masks = []

            # Update end time for analysis
            self.update_end_time()

        except ImportError:
            self._log_message(
                "Error: AVI support not available. Install opencv-python: pip install opencv-python"
            )
            import traceback

            self._log_message(traceback.format_exc())
        except Exception as e:
            self._log_message(f"ERROR in _load_avi_batch: {e}")
            import traceback

            self._log_message(traceback.format_exc())
        finally:
            self._log_message("=== _load_avi_batch() END ===")

    def _quick_legacy_check(self, h5_file) -> bool:
        """Quick check if file is legacy (same logic as in _metadata.py)."""

        file_version = h5_file.attrs.get("file_version", "1.0")
        if float(file_version) < 2.2:
            return True

        if "timeseries" in h5_file:
            ts_group = h5_file["timeseries"]
            if not ts_group.attrs.get("expected_intervals_fixed", False):
                return True

        return False

    def load_directory(self):
        """Load a directory containing HDF5 or AVI files."""
        directory = QFileDialog.getExistingDirectory(
            self, "Select Directory Containing Video Files"
        )
        if not directory:
            return

        self.directory = directory
        self.file_path = None
        try:
            # Scan for both HDF5 and AVI files
            h5_files = [
                f for f in os.listdir(directory) if f.lower().endswith((".h5", ".hdf5"))
            ]
            avi_files = [f for f in os.listdir(directory) if f.lower().endswith(".avi")]

            total_files = len(h5_files) + len(avi_files)

            if total_files == 0:
                self.lbl_file_info.setText(
                    f"No HDF5 or AVI files found in: {directory}"
                )
                self._log_message(f"No video files found in directory: {directory}")
                return

            # Build info message
            file_info = []
            if h5_files:
                file_info.append(f"{len(h5_files)} HDF5")
            if avi_files:
                file_info.append(f"{len(avi_files)} AVI")

            files_str = ", ".join(file_info)
            self.lbl_file_info.setText(
                f"Loaded directory: {directory} ({files_str} files)"
            )
            self._log_message(f"Loaded directory with {files_str} files: {directory}")

            if h5_files:
                self._log_message(
                    f"  HDF5 files: {', '.join(h5_files[:5])}{'...' if len(h5_files) > 5 else ''}"
                )
            if avi_files:
                self._log_message(
                    f"  AVI files: {', '.join(avi_files[:5])}{'...' if len(avi_files) > 5 else ''}"
                )

        except Exception as e:
            self.lbl_file_info.setText(f"Error reading directory: {e}")
            self._log_message(f"ERROR reading directory: {e}")
            return

        # If AVI files are found, load them as batch
        if avi_files:
            self._log_message(
                f"Loading {len(avi_files)} AVI files from directory as batch..."
            )
            avi_paths = [os.path.join(directory, f) for f in sorted(avi_files)]
            self._log_message(f"AVI paths to load: {avi_paths}")
            self._log_message("Calling _load_avi_batch()...")
            self._load_avi_batch(avi_paths)
            self._log_message("_load_avi_batch() completed")
            # Update frame viewer file selector
            self._update_viewer_file_combo()
            return

        # Otherwise, use HDF5 reader for directory
        # Clear all existing layers
        self.viewer.layers.clear()

        # Use reader to load directory
        reader = napari_get_reader(directory)
        if reader is None:
            self._log_message("No valid directory for HDF5 reader.")
            return

        try:
            layers = reader(directory)
        except Exception as e:
            self._log_message(f"Reader error: {e}")
            return

        # Add each layer to viewer
        for data, meta, layer_type in layers:
            name = meta.get("name", os.path.basename(directory))
            kwargs = {k: v for k, v in meta.items() if k not in ("name",)}

            if layer_type == "image":
                self.viewer.add_image(data, name=name, **kwargs)
            elif layer_type == "labels":
                self.viewer.add_labels(data, name=name, **kwargs)

        # Update frame viewer file selector for directory
        self._update_viewer_file_combo()

    def update_end_time(self):
        """Enhanced update_end_time method with dual structure support."""
        if self.file_path:
            try:
                # Check if this is an AVI file or AVI batch
                if hasattr(self, "avi_batch_paths") and self.avi_batch_paths:
                    # AVI batch - use metadata from viewer layer
                    if len(self.viewer.layers) > 0:
                        layer = self.viewer.layers[0]
                        if hasattr(layer, "metadata") and layer.metadata:
                            metadata = layer.metadata
                            if "total_duration" in metadata:
                                total_duration_seconds = metadata["total_duration"]
                                frame_count = metadata.get("total_frames_estimate", 0)
                            elif "duration" in metadata:
                                total_duration_seconds = metadata["duration"]
                                frame_count = metadata.get("frame_count_estimate", 0)
                            else:
                                self._log_message("No duration metadata found for AVI")
                                return
                        else:
                            self._log_message("No metadata found in layer")
                            return
                    else:
                        self._log_message("No layers found")
                        return
                elif self.file_path.lower().endswith(".avi"):
                    # Single AVI file
                    from ._avi_reader import AVIVideoReader

                    with AVIVideoReader(self.file_path) as reader:
                        total_duration_seconds = reader.duration
                        video_fps = reader.fps
                        target_interval = reader.metadata.get("frame_interval", 5.0)
                        frames_per_sample = max(1, int(video_fps * target_interval))
                        frame_count = len(
                            range(0, reader.frame_count, frames_per_sample)
                        )
                else:
                    # HDF5 file
                    if DUAL_STRUCTURE_AVAILABLE:
                        # Use structure detection to get frame count
                        structure_info = detect_hdf5_structure_type(self.file_path)
                        if structure_info["type"] != "error":
                            frame_count = structure_info["frame_count"]
                            self._log_message(
                                f"Frame count from structure detection: {frame_count}"
                            )
                        else:
                            raise Exception("Structure detection failed")
                    else:
                        # Fallback to original method
                        with h5py.File(self.file_path, "r") as f:
                            if "frames" in f:
                                frame_count = len(f["frames"])
                            else:
                                raise Exception("No 'frames' dataset found")

                    frame_interval = self.frame_interval.value()
                    total_duration_seconds = frame_count * frame_interval

                total_duration_minutes = total_duration_seconds / 60.0

                self.time_end.setValue(int(total_duration_seconds))
                self.plot_end_time.setRange(0.0, total_duration_minutes)
                self.plot_end_time.setValue(total_duration_minutes)
                self.plot_start_time.setRange(0.0, total_duration_minutes)
                self.plot_start_time.setValue(0.0)

                # Initialize baseline duration spinbox to recording length at load time
                if hasattr(self, "baseline_duration_minutes") and total_duration_minutes > 0:
                    self.baseline_duration_minutes.setMaximum(total_duration_minutes)
                    self.baseline_duration_minutes.setValue(total_duration_minutes)

                self._log_message(
                    f"File contains {frame_count} frames, total duration: {total_duration_minutes:.1f} min"
                )

            except Exception as e:
                self.lbl_file_info.setText(f"Error reading metadata: {str(e)}")
                self._log_message(f"ERROR reading metadata: {str(e)}")

    def enhanced_detect_rois(self):
        """Enhanced ROI detection that properly manages layers for both datasets."""

        # Determine dataset type and log clearly
        current_type = getattr(self, "current_dataset_type", "main")

        # NEW: Ensure main dataset is stored before calibration ROI detection
        if current_type == "calibration":
            if not getattr(self, "main_dataset_stored", False):
                self._log_message(
                    "WARNING: Calibration ROI detection without stored main dataset"
                )
                self._log_message("This may cause issues during analysis")

            # SET CALIBRATION VARIABLES
            current_file = self.calibration_file_path_stored
            dataset_type = "CALIBRATION"
            self._log_message(f"=== ROI DETECTION FOR {dataset_type} DATASET ===")
            self._log_message(f"File: {os.path.basename(current_file)}")
            self._log_message("NOTE: This is for calibration baseline calculation only")
        else:
            # SET MAIN DATASET VARIABLES
            current_file = self.file_path
            dataset_type = "MAIN"
            self._log_message(f"=== ROI DETECTION FOR {dataset_type} DATASET ===")
            self._log_message(f"File: {os.path.basename(current_file)}")
            self._log_message("NOTE: This is for the experimental data analysis")

        if not current_file:
            self.lbl_file_info.setText("Error: No HDF5 file loaded for ROI detection")
            return

        # Get ROI detection parameters based on preset or manual values
        # Always use current spinbox values (presets just set the spinbox values)
        params = {
            "min_radius": self.min_radius.value(),
            "max_radius": self.max_radius.value(),
            "dp": self.dp_param.value(),
            "min_dist": self.min_dist.value(),
            "param1": self.param1.value(),
            "param2": self.param2.value(),
        }

        # Log which preset is active (if any)
        if self.chk_6well.isChecked():
            self._log_message("6-well preset active (values can be adjusted)")
        elif self.chk_12well.isChecked():
            self._log_message("12-well preset active (values can be adjusted)")

        # Log the actual parameters being used
        self._log_message(
            f"ROI Detection: radius={params['min_radius']}-{params['max_radius']}px, "
            f"dp={params['dp']}, minDist={params['min_dist']}, "
            f"param1={params['param1']}, param2={params['param2']}"
        )

        try:
            # ROI detection - get first frame from viewer layer or file
            first_frame = None

            # Check if current file is HDF5 or AVI to decide source
            is_hdf5 = current_file.lower().endswith((".h5", ".hdf5"))
            is_avi = current_file.lower().endswith(".avi")

            # For HDF5 files, always read from file (not from viewer layer)
            # For AVI batch, try to use existing layer first
            if is_hdf5:
                self._log_message(
                    "HDF5 file detected - reading first frame from file..."
                )
                first_frame = get_first_frame(current_file)
            elif is_avi or (
                hasattr(self, "avi_batch_loaded") and self.avi_batch_loaded
            ):
                # Try to get frame from existing napari layer (for AVI batch)
                if len(self.viewer.layers) > 0:
                    layer = self.viewer.layers[0]
                    if hasattr(layer, "data"):
                        first_frame = layer.data
                        if len(first_frame.shape) == 3 and first_frame.shape[0] > 1:
                            # Multi-frame layer, take first frame
                            first_frame = first_frame[0]
                        self._log_message(
                            f"Using frame from viewer layer: {layer.name} (shape: {first_frame.shape})"
                        )

            # Final fallback: read from file
            if first_frame is None:
                self._log_message("Fallback: Reading first frame from file...")
                first_frame = get_first_frame(current_file)

            if first_frame is None:
                self._log_message("ERROR: Could not read first frame")
                return

            # Convert to grayscale and enhance
            if len(first_frame.shape) == 3:
                gray_frame = cv2.cvtColor(first_frame, cv2.COLOR_RGB2GRAY)
            else:
                gray_frame = first_frame.copy()

            # Ensure uint8 for CLAHE (required by OpenCV)
            if gray_frame.dtype != np.uint8:
                # Normalize to 0-255 range and convert to uint8
                gray_float = gray_frame.astype(np.float32)
                gray_min, gray_max = gray_float.min(), gray_float.max()
                if gray_max > gray_min:
                    gray_norm = (gray_float - gray_min) / (gray_max - gray_min)
                else:
                    gray_norm = gray_float
                gray_frame = (gray_norm * 255).astype(np.uint8)
                self._log_message(
                    f"Converted frame from {first_frame.dtype} to uint8 for CLAHE"
                )

            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            enhanced_frame = clahe.apply(gray_frame)

            # Detect circles
            circles = cv2.HoughCircles(
                enhanced_frame,
                cv2.HOUGH_GRADIENT,
                dp=params["dp"],
                minDist=params["min_dist"],
                param1=params["param1"],
                param2=params["param2"],
                minRadius=params["min_radius"],
                maxRadius=params["max_radius"],
            )

            # Create masks and labeled frame
            masks = []
            if circles is not None:
                circles = np.uint16(np.around(circles))

                # Remove extra dimension from HoughCircles if present
                if len(circles.shape) == 3:
                    circles = circles[0]

                # Robust row-based sorting for multi-well plates
                # Detect number of rows (assuming 18 wells = 3 rows × 6 cols)
                num_circles = len(circles)
                if num_circles == 18:
                    expected_rows = 3
                elif num_circles == 12:
                    expected_rows = 3
                elif num_circles == 24:
                    expected_rows = 4
                elif num_circles == 6:
                    expected_rows = 2
                else:
                    expected_rows = int(np.sqrt(num_circles))

                # Sort all circles by Y coordinate
                y_sorted_indices = np.argsort(circles[:, 1])
                y_sorted = circles[y_sorted_indices]

                # Group into rows
                circles_per_row = num_circles // expected_rows
                sorted_circles = []

                for row_idx in range(expected_rows):
                    start_idx = row_idx * circles_per_row
                    end_idx = start_idx + circles_per_row
                    if row_idx == expected_rows - 1:
                        end_idx = num_circles  # Include remaining circles in last row

                    row_circles = y_sorted[start_idx:end_idx]

                    # Sort this row by X coordinate (left to right)
                    x_sorted_indices = np.argsort(row_circles[:, 0])
                    row_sorted = row_circles[x_sorted_indices]

                    # Reverse every odd row (0-indexed) for meandering pattern
                    # Row 0: L→R, Row 1: R→L, Row 2: L→R, etc.
                    if row_idx % 2 == 1:
                        row_sorted = row_sorted[::-1]

                    sorted_circles.extend(row_sorted)

                sorted_circles = np.array(sorted_circles, dtype=np.uint16)

                # Store original circles for later scaling
                self._original_circles = sorted_circles.copy()
                self._original_frame_shape = gray_frame.shape
                self._original_first_frame = first_frame.copy()
                self.btn_apply_scale.setEnabled(True)

                # Apply current scale factor
                scale = self.roi_scale.value()
                scaled_circles = sorted_circles.copy().astype(np.float32)
                scaled_circles[:, 2] = scaled_circles[:, 2] * scale  # Scale radius only
                scaled_circles = np.round(scaled_circles).astype(np.uint16)

                for idx, circle in enumerate(scaled_circles):
                    mask = np.zeros(gray_frame.shape, dtype=np.uint8)
                    cv2.circle(
                        mask, (circle[0], circle[1]), circle[2], 255, thickness=-1
                    )
                    masks.append(mask)

                # Create labeled frame with scaled circles
                if len(first_frame.shape) == 3:
                    labeled_frame = first_frame.copy()
                else:
                    labeled_frame = cv2.cvtColor(gray_frame, cv2.COLOR_GRAY2RGB)

                for idx, circle in enumerate(scaled_circles):
                    color = (
                        (255, 165, 0) if dataset_type == "CALIBRATION" else (0, 255, 0)
                    )
                    cv2.circle(
                        labeled_frame, (circle[0], circle[1]), circle[2], color, 2
                    )
                    cv2.putText(
                        labeled_frame,
                        f"{idx + 1}",
                        (circle[0] - 10, circle[1] + 10),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        2.5,
                        (255, 0, 0),
                        3,
                    )

            # Store results based on dataset type
            if dataset_type == "CALIBRATION":
                # Store calibration results (for baseline processing only)
                self.calibration_masks = masks.copy()
                self.calibration_labeled_frame = labeled_frame.copy()

                # CRITICAL: Set masks for immediate calibration processing but preserve main
                self.masks = masks  # Temporary for calibration baseline processing
                self.labeled_frame = labeled_frame

                # Enable calibration baseline processing
                if hasattr(self, "btn_process_calibration_baseline"):
                    self.btn_process_calibration_baseline.setEnabled(True)

                # Update status
                if hasattr(self, "calibration_status_label"):
                    self.calibration_status_label.setText(
                        "✅ 1. Calibration file selected\n"
                        "✅ 2. Calibration first frame loaded\n"
                        "✅ 3. Calibration ROIs detected\n"
                        "4. Process baseline (Analysis tab)\n"
                        "5. Return to main dataset for analysis"
                    )

                self._log_message(
                    "Next: Go to Analysis tab and click 'Process Calibration Baseline'"
                )
                self._log_message(
                    f"Applied automatic meandering sort to {len(sorted_circles)} ROIs"
                )
            else:  # MAIN dataset
                # Store main results permanently
                self.main_masks = masks.copy()
                self.main_labeled_frame = labeled_frame.copy()
                self.masks = masks
                self.labeled_frame = labeled_frame

                # Also update stored main dataset if we're in main mode
                if hasattr(self, "main_dataset_stored") and self.main_dataset_stored:
                    self.main_masks = masks.copy()

            # Add layers to viewer
            self._add_roi_layers_to_viewer(labeled_frame, masks, dataset_type)

            result_msg = f"{dataset_type}: Detected {len(masks)} ROIs"
            self.lbl_file_info.setText(result_msg)
            self._log_message(result_msg)

            # Persist the parameters that led to a successful detection
            self._save_roi_settings()

            # Populate ROI selection checkboxes
            if dataset_type == "MAIN":
                self._populate_roi_checkboxes(len(masks))

        except Exception as e:
            self._log_message(f"ERROR in ROI detection: {e}")

    # ------------------------------------------------------------------
    # ROI detection settings persistence (QSettings → Windows registry)
    # ------------------------------------------------------------------
    _ROI_SETTINGS_ORG = "napari-hdf5-activity"
    _ROI_SETTINGS_APP = "HDF5AnalysisWidget"

    def _save_roi_settings(self):
        """Persist current ROI detection parameters via QSettings."""
        s = QSettings(self._ROI_SETTINGS_ORG, self._ROI_SETTINGS_APP)
        s.setValue("roi/min_radius", self.min_radius.value())
        s.setValue("roi/max_radius", self.max_radius.value())
        s.setValue("roi/dp_param", self.dp_param.value())
        s.setValue("roi/min_dist", self.min_dist.value())
        s.setValue("roi/param1", self.param1.value())
        s.setValue("roi/param2", self.param2.value())
        s.setValue("roi/roi_scale", self.roi_scale.value())
        preset = "6well" if self.chk_6well.isChecked() else (
            "12well" if self.chk_12well.isChecked() else "none"
        )
        s.setValue("roi/preset", preset)
        s.sync()

    def _load_roi_settings(self):
        """Restore ROI detection parameters saved by a previous session."""
        s = QSettings(self._ROI_SETTINGS_ORG, self._ROI_SETTINGS_APP)
        if not s.contains("roi/min_radius"):
            return  # No saved settings yet → keep widget defaults
        self.min_radius.setValue(int(s.value("roi/min_radius", 100)))
        self.max_radius.setValue(int(s.value("roi/max_radius", 145)))
        self.dp_param.setValue(float(s.value("roi/dp_param", 1.0)))
        self.min_dist.setValue(int(s.value("roi/min_dist", 300)))
        self.param1.setValue(int(s.value("roi/param1", 30)))
        self.param2.setValue(int(s.value("roi/param2", 60)))
        self.roi_scale.setValue(float(s.value("roi/roi_scale", 1.0)))
        preset = s.value("roi/preset", "6well")
        self.chk_6well.setChecked(preset == "6well")
        self.chk_12well.setChecked(preset == "12well")

    def reset_for_new_analysis(self):
        """Reset all analysis data and UI state for a new analysis."""
        try:
            # Clear analysis results
            self.merged_results = {}
            self.roi_baseline_means = {}
            self.roi_upper_thresholds = {}
            self.roi_lower_thresholds = {}
            self.roi_statistics = {}
            self.movement_data = {}
            self.fraction_data = {}
            self.sleep_data = {}
            self.quiescence_data = {}
            self.roi_colors = {}
            self.roi_band_widths = {}

            # Clear ROI detection
            self.masks = []
            self.labeled_frame = None

            # Clear calibration state
            self.current_dataset_type = "main"
            self.calibration_file_path_stored = None
            self.calibration_masks = []
            self.calibration_labeled_frame = None
            self.calibration_baseline_processed = False
            self.calibration_baseline_statistics = {}

            # Reset file paths
            self.file_path = None
            self.directory = None
            self.main_dataset_path = None

            # Clear viewer layers
            self.viewer.layers.clear()

            # Reset UI elements
            self.lbl_file_info.setText("No file loaded")
            self.results_label.setText("Results will be displayed here.")
            self.status_label.setText("Ready to start analysis")
            self.progress_bar.setValue(0)

            # Reset calibration UI
            if hasattr(self, "calibration_file_path"):
                self.calibration_file_path.setText("No calibration file selected")
                self.btn_load_calibration_dataset.setEnabled(False)
                self.btn_process_calibration_baseline.setEnabled(False)

            if hasattr(self, "calibration_status_label"):
                self.calibration_status_label.setText(
                    "1. Select calibration file\n"
                    "2. Load calibration dataset\n"
                    "3. Detect ROIs (Input tab)\n"
                    "4. Process baseline"
                )

            # Clear log
            self.log_text.clear()

            # Clear matplotlib figure
            if hasattr(self, "figure"):
                self.figure.clear()
                self.canvas.draw()

            self._log_message("Analysis reset complete - ready for new analysis")

        except Exception as e:
            self._log_message(f"Error during reset: {e}")

    def process_calibration_baseline(self):
        """Process calibration dataset to create baseline statistics with progress bar."""
        if not self.calibration_masks:
            self._log_message("No calibration ROIs detected")
            return

        if not self.calibration_file_path_stored:
            self._log_message("No calibration file selected")
            return

        # Start progress monitoring and disable button
        self.btn_process_calibration_baseline.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_label.setText("Processing calibration baseline...")

        @thread_worker(start_thread=False)
        def _calibration_worker():
            return self._process_calibration_baseline_worker()

        worker = _calibration_worker()
        worker.returned.connect(self._calibration_finished)
        worker.errored.connect(self._calibration_errored)
        worker.finished.connect(self._calibration_done)
        worker.start()

    def _process_calibration_baseline_worker(self):
        """Worker function for calibration baseline processing."""

        def progress_callback(percent, message):
            self.progress_updated.emit(int(percent))
            self.status_updated.emit(message)

        try:
            self._log_message("Processing calibration baseline...")
            self._log_message(
                f"File: {os.path.basename(self.calibration_file_path_stored)}"
            )
            self._log_message(f"ROIs: {len(self.calibration_masks)}")

            progress_callback(5, "Initializing calibration processing...")

            # Try to process calibration file with enhanced error handling
            calibration_roi_changes = None
            calibration_duration = 0

            try:
                from ._reader import process_hdf5_file

                progress_callback(10, "Processing calibration file (method 1)...")
                self._log_message("Attempting to process calibration file...")

                _, calibration_roi_changes, calibration_duration = process_hdf5_file(
                    file_path=self.calibration_file_path_stored,
                    masks=self.calibration_masks,
                    chunk_size=self.chunk_size.value(),
                    progress_callback=lambda p, m: progress_callback(
                        10 + (p * 0.4), f"Calibration: {m}"
                    ),
                    frame_interval=self.frame_interval.value(),
                )
                self._log_message("Successfully processed calibration file")

            except Exception as reader_error:
                self._log_message(f"Reader error encountered: {reader_error}")

                # Try alternative processing method
                try:
                    from ._reader import process_single_file_in_parallel_dual_structure

                    progress_callback(20, "Trying alternative processing method...")
                    self._log_message("Trying alternative processing method...")

                    _, calibration_roi_changes, calibration_duration = (
                        process_single_file_in_parallel_dual_structure(
                            self.calibration_file_path_stored,
                            self.calibration_masks,
                            chunk_size=self.chunk_size.value(),
                            progress_callback=lambda p, m: progress_callback(
                                20 + (p * 0.3), f"Alt method: {m}"
                            ),
                            frame_interval=self.frame_interval.value(),
                            num_processes=1,  # Use single process to avoid issues
                        )
                    )
                    self._log_message("Alternative processing successful")

                except Exception as alt_error:
                    self._log_message(
                        f"Alternative processing also failed: {alt_error}"
                    )

                    # Final fallback - try the basic reader functions
                    try:
                        from ._reader import process_hdf5_files

                        progress_callback(30, "Trying basic processing method...")
                        self._log_message("Trying basic processing method...")

                        # Use directory processing as fallback
                        cal_dir = os.path.dirname(self.calibration_file_path_stored)
                        results, durations, _, _ = process_hdf5_files(
                            cal_dir,
                            masks=self.calibration_masks,
                            num_processes=1,
                            chunk_size=self.chunk_size.value(),
                            progress_callback=lambda p, m: progress_callback(
                                30 + (p * 0.2), f"Basic: {m}"
                            ),
                            frame_interval=self.frame_interval.value(),
                        )

                        # Extract results for our specific file
                        cal_filename = os.path.basename(
                            self.calibration_file_path_stored
                        )
                        calibration_roi_changes = None
                        calibration_duration = 0

                        for file_path, roi_data in results.items():
                            if cal_filename in file_path:
                                calibration_roi_changes = roi_data
                                calibration_duration = durations.get(file_path, 0)
                                break

                        if calibration_roi_changes is None:
                            raise Exception(
                                "Could not find calibration data in results"
                            )

                        self._log_message("Basic processing successful")

                    except Exception as final_error:
                        return {
                            "success": False,
                            "error": f"All processing methods failed: {final_error}",
                        }

            # Continue with the rest of the processing if we got valid data
            if not calibration_roi_changes:
                return {
                    "success": False,
                    "error": "No calibration data obtained - processing failed",
                }

            progress_callback(60, "Applying preprocessing...")
            self._log_message(
                f"Calibration data processed: {len(calibration_roi_changes)} ROIs"
            )

            # Apply same preprocessing as main dataset
            from ._calc import (
                apply_matlab_normalization_to_merged_results,
                improved_full_dataset_detrending,
            )

            # MATLAB normalization
            progress_callback(70, "Applying MATLAB normalization...")
            self._log_message("Applying MATLAB normalization to calibration data...")
            normalized_calibration = apply_matlab_normalization_to_merged_results(
                calibration_roi_changes, enable_matlab_norm=True
            )

            # Detrending (if enabled)
            progress_callback(80, "Applying detrending...")
            if (
                hasattr(self, "enable_detrending")
                and self.enable_detrending.isChecked()
            ):
                self._log_message("Applying detrending to calibration data...")
                processed_calibration = improved_full_dataset_detrending(
                    normalized_calibration
                )
            else:
                self._log_message("Skipping detrending (disabled)")
                processed_calibration = normalized_calibration

            # Calculate baseline statistics for each ROI
            progress_callback(90, "Calculating baseline statistics...")
            self._log_message(
                "Calculating baseline statistics from COMPLETE calibration dataset..."
            )
            self._log_message(
                f"Calibration duration: {calibration_duration/60:.1f} minutes"
            )
            calibration_baseline_statistics = {}

            for roi, data in processed_calibration.items():
                if not data:
                    self._log_message(f"No data for ROI {roi}, skipping")
                    continue

                # Extract all values from complete calibration dataset
                values = np.array([val for _, val in data])

                # Calculate comprehensive statistics
                cal_mean = np.mean(values)
                cal_std = np.std(values)
                cal_multiplier = self.calibration_multiplier.value()

                # Calculate hysteresis thresholds
                threshold_band = cal_multiplier * cal_std
                upper_threshold = cal_mean + threshold_band
                lower_threshold = max(0, cal_mean - threshold_band)  # Don't go negative

                calibration_baseline_statistics[roi] = {
                    "baseline_mean": cal_mean,
                    "baseline_std": cal_std,
                    "upper_threshold": upper_threshold,
                    "lower_threshold": lower_threshold,
                    "threshold_band": threshold_band,
                    "multiplier": cal_multiplier,
                    "data_points": len(values),
                    "duration_minutes": calibration_duration / 60,
                    "data_range": (float(np.min(values)), float(np.max(values))),
                }

                self._log_message(
                    f"ROI {roi}: mean={cal_mean:.1f}, std={cal_std:.1f}, thresholds=[{lower_threshold:.1f}, {upper_threshold:.1f}], frames={len(values)}"
                )

            progress_callback(100, "Calibration baseline complete")

            return {
                "success": True,
                "statistics": calibration_baseline_statistics,
                "duration": calibration_duration,
                "roi_count": len(calibration_baseline_statistics),
            }

        except Exception as e:
            return {"success": False, "error": str(e)}

    def _calibration_finished(self, result):
        """Handle successful calibration completion."""
        if result["success"]:
            self.calibration_baseline_statistics = result["statistics"]
            self.calibration_baseline_processed = True

            # Update status
            if hasattr(self, "calibration_status_label"):
                self.calibration_status_label.setText(
                    "✅ 1. Calibration file selected\n"
                    "✅ 2. Calibration dataset loaded\n"
                    "✅ 3. Calibration ROIs detected\n"
                    "✅ 4. Calibration baseline processed\n"
                    "Ready for analysis!"
                )

            self.status_label.setText("Calibration baseline processing complete")

            # Log results
            successful_rois = result["roi_count"]
            self._log_message("Calibration baseline processing complete:")
            self._log_message(f"  ROIs processed: {successful_rois}")
            self._log_message(f"  Duration: {result['duration']/60:.1f} minutes")
            self._log_message(
                "Calibration baseline ready! Switch to main dataset and run analysis."
            )

        else:
            self._log_message(f"Calibration failed: {result['error']}")
            self.status_label.setText(f"Calibration failed: {result['error']}")

    def _calibration_errored(self, exc):
        """Handle calibration errors."""
        self.status_label.setText(f"Calibration error: {exc}")
        self._log_message(f"Calibration error: {exc}")

    def _calibration_done(self):
        """Cleanup after calibration completion."""
        self.btn_process_calibration_baseline.setEnabled(True)
        self.progress_bar.setValue(0)

    def add_calibration_layers_to_viewer(self, labeled_frame, masks):
        """Add calibration dataset layers with clear naming and organization."""
        try:
            basename = os.path.basename(self.calibration_file_path_stored)

            # Add calibration raw frame
            cal_raw_layer = self.viewer.add_image(
                labeled_frame,
                name=f"CALIBRATION - {basename} - ROI Detection",
                colormap="gray",
                visible=True,
                opacity=0.8,
            )

            # Store calibration info in metadata
            cal_raw_layer.metadata.update(
                {
                    "dataset_type": "calibration",
                    "file_path": self.calibration_file_path_stored,
                    "roi_count": len(masks),
                    "workflow_step": "roi_detection",
                }
            )

            self._log_message(
                f"Added calibration ROI detection layer: {len(masks)} ROIs"
            )

        except Exception as e:
            self._log_message(f"Error adding calibration layers: {e}")

    def add_main_dataset_layers_to_viewer(self, labeled_frame, masks):
        """Add main dataset layers with clear naming and organization."""
        try:
            basename = (
                os.path.basename(self.file_path) if self.file_path else "main_dataset"
            )

            # Add main dataset frame
            main_raw_layer = self.viewer.add_image(
                labeled_frame,
                name=f"MAIN - {basename} - ROI Detection",
                colormap="gray",
                visible=True,
                opacity=0.8,
            )

            # Store main dataset info in metadata
            main_raw_layer.metadata.update(
                {
                    "dataset_type": "main",
                    "file_path": self.file_path,
                    "roi_count": len(masks),
                    "workflow_step": "roi_detection",
                }
            )

            self._log_message(f"Added main ROI detection layer: {len(masks)} ROIs")

        except Exception as e:
            self._log_message(f"Error adding main dataset layers: {e}")

    def manage_workflow_layers(self, workflow_step):
        """
        Manage layer visibility based on workflow step.

        Args:
            workflow_step: 'main_dataset', 'calibration_setup', 'comparison', 'final_analysis'
        """
        try:
            if workflow_step == "main_dataset":
                # Show only main dataset layers
                self._set_layer_visibility_by_type("main", True)
                self._set_layer_visibility_by_type("calibration", False)
                self._set_layer_visibility_by_type("comparison", False)
                self._log_message("Switched view: Main dataset only")

            elif workflow_step == "calibration_setup":
                # Show only calibration layers
                self._set_layer_visibility_by_type("main", False)
                self._set_layer_visibility_by_type("calibration", True)
                self._set_layer_visibility_by_type("comparison", False)
                self._log_message("Switched view: Calibration dataset only")

            elif workflow_step == "comparison":
                # Show comparison view
                self.switch_to_comparison_view()

            elif workflow_step == "final_analysis":
                # Show main dataset for final analysis
                self._set_layer_visibility_by_type("main", True)
                self._set_layer_visibility_by_type("calibration", False)
                self._set_layer_visibility_by_type("comparison", False)
                self._log_message("Switched view: Main dataset for analysis")

        except Exception as e:
            self._log_message(f"Error managing workflow layers: {e}")

    def _set_layer_visibility_by_type(self, dataset_type, visible):
        """Set visibility for all layers of a specific dataset type."""
        try:
            count = 0
            for layer in self.viewer.layers:
                if (
                    hasattr(layer, "metadata")
                    and layer.metadata.get("dataset_type") == dataset_type
                ):
                    layer.visible = visible
                    count += 1

            if count > 0:
                status = "visible" if visible else "hidden"
                self._log_message(f"Set {count} {dataset_type} layers to {status}")

        except Exception as e:
            self._log_message(f"Error setting layer visibility: {e}")

    def _create_roi_comparison_image(self, cal_frame, main_frame):
        """Create side-by-side comparison of ROI detection between calibration and main datasets."""
        try:
            # Convert frames to RGB if needed
            if len(cal_frame.shape) == 3:
                cal_rgb = cal_frame.copy()
            else:
                cal_rgb = cv2.cvtColor(cal_frame, cv2.COLOR_GRAY2RGB)

            if len(main_frame.shape) == 3:
                main_rgb = main_frame.copy()
            else:
                main_rgb = cv2.cvtColor(main_frame, cv2.COLOR_GRAY2RGB)

            # Resize frames to same height for comparison
            target_height = min(
                cal_rgb.shape[0], main_rgb.shape[0], 800
            )  # Max height 800px

            # Calculate new widths maintaining aspect ratio
            cal_ratio = cal_rgb.shape[1] / cal_rgb.shape[0]
            main_ratio = main_rgb.shape[1] / main_rgb.shape[0]

            cal_width = int(target_height * cal_ratio)
            main_width = int(target_height * main_ratio)

            cal_resized = cv2.resize(cal_rgb, (cal_width, target_height))
            main_resized = cv2.resize(main_rgb, (main_width, target_height))

            # Draw ROIs on both images
            cal_with_rois = self._draw_rois_on_image(
                cal_resized, self.calibration_masks, "CAL", (255, 165, 0)
            )  # Orange
            main_with_rois = self._draw_rois_on_image(
                main_resized, self.main_masks, "MAIN", (0, 255, 0)
            )  # Green

            # Add labels at the top
            cv2.putText(
                cal_with_rois,
                "CALIBRATION DATASET",
                (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX,
                1,
                (255, 255, 255),
                2,
            )
            cv2.putText(
                main_with_rois,
                "MAIN DATASET",
                (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX,
                1,
                (255, 255, 255),
                2,
            )

            # Create separator
            separator = (
                np.ones((target_height, 20, 3), dtype=np.uint8) * 128
            )  # Gray separator

            # Combine side by side with separator
            comparison = np.hstack([cal_with_rois, separator, main_with_rois])

            self._log_message(f"Created ROI comparison image: {comparison.shape}")
            return comparison

        except Exception as e:
            self._log_message(f"Error creating ROI comparison image: {e}")
            # Return a simple concatenation as fallback
            try:
                if cal_frame.shape == main_frame.shape:
                    return np.hstack([cal_frame, main_frame])
                else:
                    # If shapes don't match, return the calibration frame
                    return cal_frame
            except:
                return cal_frame

    def _draw_rois_on_image(self, image, masks, prefix, color):
        """Draw ROI circles and labels on image with specified color."""
        import cv2

        if not masks:
            return image.copy()

        result = image.copy()

        try:
            for i, mask in enumerate(masks):
                # Get ROI center and radius from mask
                center = self._get_roi_center(mask)
                radius = int(self._get_roi_radius(mask))

                if center[0] > 0 and center[1] > 0 and radius > 0:
                    # Scale coordinates if image was resized
                    scale_x = (
                        result.shape[1] / mask.shape[1] if mask.shape[1] > 0 else 1
                    )
                    scale_y = (
                        result.shape[0] / mask.shape[0] if mask.shape[0] > 0 else 1
                    )

                    scaled_center = (int(center[0] * scale_x), int(center[1] * scale_y))
                    scaled_radius = int(radius * min(scale_x, scale_y))

                    # Draw circle
                    cv2.circle(result, scaled_center, scaled_radius, color, 2)

                    # Draw ROI label
                    label = f"{prefix} {i+1}"
                    label_pos = (
                        scaled_center[0] - 20,
                        scaled_center[1] - scaled_radius - 10,
                    )

                    # Ensure label position is within image bounds
                    label_pos = (max(5, label_pos[0]), max(20, label_pos[1]))

                    cv2.putText(
                        result,
                        label,
                        label_pos,
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.6,
                        color,
                        2,
                    )

        except Exception as e:
            self._log_message(f"Error drawing ROIs on image: {e}")

        return result

    def switch_to_comparison_view(self):
        """Switch viewer to show both calibration and main datasets for comparison."""
        try:
            # Show both dataset layers
            for layer in self.viewer.layers:
                if hasattr(layer, "metadata"):
                    dataset_type = layer.metadata.get("dataset_type", "")
                    if dataset_type in ["calibration", "main"]:
                        layer.visible = True

            # Create and add comparison layer if both datasets have ROIs
            if (
                hasattr(self, "calibration_labeled_frame")
                and self.calibration_labeled_frame is not None
                and hasattr(self, "main_labeled_frame")
                and self.main_labeled_frame is not None
            ):

                self._log_message("Creating ROI correspondence comparison...")

                # Create side-by-side comparison
                comparison_image = self._create_roi_comparison_image(
                    self.calibration_labeled_frame, self.main_labeled_frame
                )

                # Remove existing comparison layer if it exists
                layers_to_remove = []
                for layer in self.viewer.layers:
                    if "Comparison" in layer.name:
                        layers_to_remove.append(layer)

                for layer in layers_to_remove:
                    self.viewer.layers.remove(layer)

                # Add new comparison layer
                comparison_layer = self.viewer.add_image(
                    comparison_image,
                    name="ROI Correspondence Comparison",
                    colormap="gray",
                    visible=True,
                )

                comparison_layer.metadata.update(
                    {"dataset_type": "comparison", "workflow_step": "roi_comparison"}
                )

                self._log_message("Added ROI comparison view")
                self._log_message(
                    "Orange circles = Calibration ROIs, Green circles = Main ROIs"
                )
                self._log_message(
                    "Verify that ROI numbers correspond to same physical locations"
                )
            else:
                self._log_message(
                    "Cannot create comparison - missing ROI data from one or both datasets"
                )

        except Exception as e:
            self._log_message(f"Error creating comparison view: {e}")

    def switch_to_main_dataset(self):
        """Switch back to main dataset for analysis."""
        if self.current_dataset_type == "calibration" and self.main_dataset_path:
            try:
                # Store calibration state
                self.calibration_masks = self.masks.copy()
                self.calibration_labeled_frame = self.labeled_frame

                # Restore main dataset
                self.current_dataset_type = "main"
                self.file_path = self.main_dataset_path
                self.masks = self.main_masks.copy()
                self.labeled_frame = self.main_labeled_frame

                # Reload main dataset in viewer
                reader = napari_get_reader(self.main_dataset_path)
                if reader:
                    self.viewer.layers.clear()
                    layers = reader(self.main_dataset_path)
                    for data, meta, layer_type in layers:
                        name = meta.get(
                            "name", os.path.basename(self.main_dataset_path)
                        )
                        kwargs = {k: v for k, v in meta.items() if k not in ("name",)}

                        if layer_type == "image":
                            self.viewer.add_image(data, name=name, **kwargs)
                        elif layer_type == "labels":
                            self.viewer.add_labels(data, name=name, **kwargs)

                    # Re-add main dataset ROIs if they exist
                    if self.masks:
                        self._add_roi_layers_to_viewer(self.labeled_frame, self.masks)

                self.lbl_file_info.setText(
                    f"MAIN DATASET: {os.path.basename(self.main_dataset_path)}"
                )
                self._log_message("Switched back to main dataset")

            except Exception as e:
                self._log_message(f"Error switching to main dataset: {e}")

    def _on_6well_preset_changed(self, state):
        """Handle 6-well preset checkbox change - make mutually exclusive."""
        if state == 2:  # Checked
            self.chk_12well.blockSignals(True)
            self.chk_12well.setChecked(False)
            self.chk_12well.blockSignals(False)
            # Update spinbox values to show preset (tested optimal for 6-well)
            self.min_radius.setValue(100)
            self.max_radius.setValue(145)
            self.dp_param.setValue(1.0)
            self.min_dist.setValue(300)
            self.param1.setValue(30)
            self.param2.setValue(60)

    def _on_12well_preset_changed(self, state):
        """Handle 12-well preset checkbox change - make mutually exclusive."""
        if state == 2:  # Checked
            self.chk_6well.blockSignals(True)
            self.chk_6well.setChecked(False)
            self.chk_6well.blockSignals(False)
            # Update spinbox values to show preset
            self.min_radius.setValue(70)
            self.max_radius.setValue(120)
            self.dp_param.setValue(1.0)
            self.min_dist.setValue(150)
            self.param1.setValue(50)
            self.param2.setValue(25)

    def _apply_roi_scale(self):
        """Apply scale factor to detected ROIs without re-detecting."""
        if not hasattr(self, "_original_circles") or self._original_circles is None:
            self._log_message("ERROR: No ROIs detected yet. Run detection first.")
            return

        try:
            scale = self.roi_scale.value()
            self._log_message(f"Applying ROI scale: {scale:.2f}")

            # Get original data
            original_circles = self._original_circles
            frame_shape = self._original_frame_shape
            first_frame = self._original_first_frame

            # Scale the radii (keep centers unchanged)
            scaled_circles = original_circles.copy().astype(np.float32)
            scaled_circles[:, 2] = scaled_circles[:, 2] * scale
            scaled_circles = np.round(scaled_circles).astype(np.uint16)

            # Recreate masks with scaled radii
            masks = []
            for circle in scaled_circles:
                mask = np.zeros(frame_shape, dtype=np.uint8)
                cv2.circle(mask, (circle[0], circle[1]), circle[2], 255, thickness=-1)
                masks.append(mask)

            # Recreate labeled frame
            if len(first_frame.shape) == 3:
                labeled_frame = first_frame.copy()
            else:
                labeled_frame = cv2.cvtColor(first_frame, cv2.COLOR_GRAY2RGB)

            current_type = getattr(self, "current_dataset_type", "main")
            dataset_type = "CALIBRATION" if current_type == "calibration" else "MAIN"

            for idx, circle in enumerate(scaled_circles):
                color = (255, 165, 0) if dataset_type == "CALIBRATION" else (0, 255, 0)
                cv2.circle(labeled_frame, (circle[0], circle[1]), circle[2], color, 2)
                cv2.putText(
                    labeled_frame,
                    f"{idx + 1}",
                    (circle[0] - 10, circle[1] + 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    2.5,
                    (255, 0, 0),
                    3,
                )

            # Update stored masks
            if dataset_type == "CALIBRATION":
                self.calibration_masks = masks.copy()
                self.calibration_labeled_frame = labeled_frame.copy()
                self.masks = masks
                self.labeled_frame = labeled_frame
            else:
                self.main_masks = masks.copy()
                self.main_labeled_frame = labeled_frame.copy()
                self.masks = masks
                self.labeled_frame = labeled_frame

            # Update viewer layer
            self._add_roi_layers_to_viewer(labeled_frame, masks, dataset_type)

            avg_radius = np.mean(scaled_circles[:, 2])
            self._log_message(
                f"Scaled {len(masks)} ROIs to {scale:.0%} (avg radius: {avg_radius:.0f}px)"
            )
            self.lbl_file_info.setText(
                f"{dataset_type}: {len(masks)} ROIs (scale: {scale:.0%})"
            )

        except Exception as e:
            self._log_message(f"ERROR applying ROI scale: {e}")

    def clear_roi_detection(self):
        """Enhanced ROI detection clearing with proper event disconnection."""
        try:
            self._log_message("Clear ROI Detection button clicked")

            layers_to_remove = []
            for layer in self.viewer.layers:
                # Check if this is an ROI layer
                is_roi_layer = (
                    "ROI" in layer.name
                    or "Detected" in layer.name
                    or (
                        hasattr(layer, "metadata")
                        and layer.metadata.get("roi_type") == "circular_detection"
                    )
                )

                if is_roi_layer:
                    self._log_message(f"  Marking layer for removal: {layer.name}")
                    layers_to_remove.append(layer)

            if len(layers_to_remove) == 0:
                self._log_message("No ROI layers found to remove")
                # Still clear variables in case they exist
                self.masks = []
                self.labeled_frame = None
                return

            # Disconnect any connected events before removing layers
            for layer in layers_to_remove:
                try:
                    # Disconnect contrast events if they exist
                    if hasattr(layer, "events"):
                        layer.events.contrast_limits.disconnect()
                except Exception:
                    pass  # Event might not be connected

                self.viewer.layers.remove(layer)
                self._log_message(f"  Removed layer: {layer.name}")

            # Clear all ROI-related variables
            self.masks = []
            self.labeled_frame = None
            self.main_masks = []
            self.main_labeled_frame = None
            self.calibration_masks = []
            self.calibration_labeled_frame = None

            self._log_message(
                f"✓ Successfully removed {len(layers_to_remove)} ROI layers and cleaned up all ROI data"
            )

        except Exception as e:
            self._log_message(f"ERROR clearing ROI detection: {e}")
            import traceback

            self._log_message(traceback.format_exc())

    def _add_roi_layers_to_viewer(self, labeled_frame, masks, dataset_type):
        """Add ROI layers with clear dataset identification."""
        try:
            if dataset_type == "CALIBRATION":
                file_name = os.path.basename(self.calibration_file_path_stored)
                layer_name = f"CALIBRATION - {file_name} - ROIs ({len(masks)})"
                colormap = "gray"
            else:
                file_name = os.path.basename(self.file_path)
                layer_name = f"MAIN - {file_name} - ROIs ({len(masks)})"
                colormap = "gray"

            # Add ROI detection layer
            roi_layer = self.viewer.add_image(
                labeled_frame,
                name=layer_name,
                colormap=colormap,
                visible=True,
                opacity=0.8,
            )

            # Store metadata
            roi_layer.metadata.update(
                {
                    "dataset_type": dataset_type.lower(),
                    "file_path": (
                        self.calibration_file_path_stored
                        if dataset_type == "CALIBRATION"
                        else self.file_path
                    ),
                    "roi_count": len(masks),
                    "workflow_step": "roi_detection",
                    "analysis_ready": True,
                }
            )

            self._log_message(f"Added {dataset_type} ROI layer: {len(masks)} ROIs")

        except Exception as e:
            self._log_message(f"Error adding ROI layer: {e}")

    def _get_roi_center(self, mask):
        """Calculate the center of an ROI mask."""
        try:
            y_coords, x_coords = np.where(mask > 0)
            if len(x_coords) > 0 and len(y_coords) > 0:
                center_x = np.mean(x_coords)
                center_y = np.mean(y_coords)
                return (int(center_x), int(center_y))
        except Exception:
            pass
        return (0, 0)

    def _get_roi_radius(self, mask):
        """Calculate the radius of an ROI mask."""
        try:
            y_coords, x_coords = np.where(mask > 0)
            if len(x_coords) > 0 and len(y_coords) > 0:
                center_x = np.mean(x_coords)
                center_y = np.mean(y_coords)
                distances = np.sqrt(
                    (x_coords - center_x) ** 2 + (y_coords - center_y) ** 2
                )
                radius = np.mean(distances)
                return radius
        except Exception:
            pass
        return 0.0

    # ===================================================================
    # SIMPLIFIED ANALYSIS EXECUTION - NOW USING _calc.py
    # ===================================================================

    def run_analysis(self):
        """Start analysis using the separated calculation module."""
        # Check if analysis is already running
        if hasattr(self, "current_worker") and self.current_worker is not None:
            self._log_message(
                "⚠️ Analysis already running! Please wait or stop current analysis first."
            )
            self.status_label.setText("Analysis already running!")
            return

        if not self.masks:
            self.status_label.setText(
                "Error: No ROIs detected. Please run ROI detection first."
            )
            self._log_message(
                "ERROR: No ROIs detected. Please run ROI detection first."
            )
            return

        # Quick validation using calc module
        is_valid, error_msg = validate_analysis_parameters(
            self.frame_interval.value(),
            self.chunk_size.value(),
            self.baseline_duration_minutes.value(),
        )

        if not is_valid:
            self.status_label.setText(f"Error: {error_msg}")
            self._log_message(f"ERROR: {error_msg}")
            return

        # Stop any previously running worker before starting new one
        if hasattr(self, "current_worker") and self.current_worker is not None:
            self._cancel_requested = True
            try:
                self.current_worker.returned.disconnect()
                self.current_worker.errored.disconnect()
                self.current_worker.finished.disconnect()
            except Exception:
                pass
            self.current_worker = None

        # UI state management
        self._cancel_requested = False
        self.analysis_start_time = time.time()
        self.btn_analyze.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_plot.setEnabled(False)  # Disable until analysis fully completes
        self.progress_bar.setValue(0)
        self.status_label.setText("Initializing analysis...")
        self.performance_timer.start()

        # Log analysis start
        self._log_analysis_parameters()

        @thread_worker(start_thread=False)
        def _analysis_worker():
            return self._run_analysis_with_calc_module()

        worker_instance = _analysis_worker()
        worker_instance.returned.connect(self._analysis_finished)
        worker_instance.errored.connect(self._analysis_errored)
        worker_instance.finished.connect(self._analysis_done)
        worker_instance.start()
        self.current_worker = worker_instance

    def load_calibration_file(self):
        """Enhanced calibration file loading with workflow state management."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Calibration File",
            "",
            "Video Files (*.h5 *.hdf5 *.avi);;HDF5 Files (*.h5 *.hdf5);;AVI Files (*.avi);;All Files (*.*)",
        )
        if file_path:
            self.calibration_file_path.setText(os.path.basename(file_path))
            self.calibration_file_path.setProperty("full_path", file_path)
            self.calibration_file_path_stored = file_path

            # Enable the load calibration dataset button
            self.btn_load_calibration_dataset.setEnabled(True)

            # Update status
            self.calibration_status_label.setText(
                "✅ 1. Calibration file selected\n"
                "2. Click 'Load Calibration Dataset'\n"
                "3. Detect ROIs (Input tab)\n"
                "4. Process baseline"
            )

            # Reset calibration state
            self.calibration_baseline_processed = False
            self.calibration_baseline_statistics = {}
            self.btn_process_calibration_baseline.setEnabled(False)

            self._log_message(
                f"Calibration file selected: {os.path.basename(file_path)}"
            )
        else:
            self.calibration_file_path.setText("No calibration file selected")
            self.calibration_file_path.setProperty("full_path", None)
            self.calibration_file_path_stored = None
            self.btn_load_calibration_dataset.setEnabled(False)

            self.calibration_status_label.setText(
                "1. Select calibration file\n"
                "2. Load calibration dataset\n"
                "3. Detect ROIs (Input tab)\n"
                "4. Process baseline"
            )

    # def load_calibration_dataset(self):
    #     """Load calibration dataset into viewer for ROI detection - DEBUG VERSION."""

    #     # IMMEDIATE DEBUG OUTPUT
    #     self._log_message("=== LOAD_CALIBRATION_DATASET METHOD CALLED ===")
    #     print("LOAD_CALIBRATION_DATASET METHOD CALLED")  # Also print to console

    #     # Check prerequisites
    #     self._log_message(f"calibration_file_path_stored: {getattr(self, 'calibration_file_path_stored', 'NOT_SET')}")

    #     if not hasattr(self, 'calibration_file_path_stored') or not self.calibration_file_path_stored:
    #         self._log_message("ERROR: No calibration file selected")
    #         return

    #     self._log_message(f"Calibration file path: {self.calibration_file_path_stored}")

    #     if not os.path.exists(self.calibration_file_path_stored):
    #         self._log_message(f"ERROR: Calibration file not found: {self.calibration_file_path_stored}")
    #         return

    #     self._log_message("File exists, proceeding with calibration dataset loading...")

    #     try:
    #         # Check current state
    #         current_type = getattr(self, 'current_dataset_type', 'NOT_SET')
    #         current_file = getattr(self, 'file_path', 'NOT_SET')

    #         self._log_message(f"Before switch - current_dataset_type: {current_type}")
    #         self._log_message(f"Before switch - file_path: {current_file}")

    #         # Store main dataset state FIRST
    #         if current_type == "main" or current_type == 'NOT_SET':
    #             self.main_dataset_path = current_file
    #             self.main_masks = getattr(self, 'masks', []).copy()
    #             self.main_labeled_frame = getattr(self, 'labeled_frame', None)
    #             self._log_message("Stored main dataset state")

    #         # Switch to calibration dataset
    #         self.current_dataset_type = "calibration"
    #         self.file_path = self.calibration_file_path_stored
    #         self.directory = None

    #         self._log_message(f"After switch - current_dataset_type: {self.current_dataset_type}")
    #         self._log_message(f"After switch - file_path: {self.file_path}")

    #         # Clear current ROI detection
    #         self.masks = []
    #         self.labeled_frame = None
    #         self._log_message("Cleared ROI detection state")

    #         # Check if reader exists
    #         try:
    #             from ._reader import napari_get_reader
    #             reader = napari_get_reader(self.calibration_file_path_stored)
    #             self._log_message(f"Reader obtained: {reader is not None}")
    #         except Exception as reader_error:
    #             self._log_message(f"ERROR getting reader: {reader_error}")
    #             return

    #         if reader is None:
    #             self._log_message("ERROR: Cannot read calibration file - no valid reader")
    #             return

    #         # Clear viewer
    #         self._log_message(f"Clearing viewer - current layers: {len(self.viewer.layers)}")
    #         self.viewer.layers.clear()
    #         self._log_message("Viewer cleared")

    #         # Load calibration data
    #         self._log_message("Loading calibration layers...")
    #         try:
    #             layers = reader(self.calibration_file_path_stored)
    #             self._log_message(f"Reader returned {len(layers)} layers")

    #             for i, (data, meta, layer_type) in enumerate(layers):
    #                 name = f"CALIBRATION - {meta.get('name', os.path.basename(self.calibration_file_path_stored))}"
    #                 kwargs = {k: v for k, v in meta.items() if k not in ("name",)}

    #                 self._log_message(f"Adding layer {i}: {layer_type} - {name}")

    #                 if layer_type == "image":
    #                     layer = self.viewer.add_image(data, name=name, **kwargs)
    #                     self._log_message(f"Added image layer: {layer.name}")
    #                 elif layer_type == "labels":
    #                     layer = self.viewer.add_labels(data, name=name, **kwargs)
    #                     self._log_message(f"Added labels layer: {layer.name}")
    #                 else:
    #                     self._log_message(f"Unknown layer type: {layer_type}")

    #         except Exception as layer_error:
    #             self._log_message(f"ERROR loading layers: {layer_error}")
    #             import traceback
    #             self._log_message(f"Layer loading traceback: {traceback.format_exc()}")
    #             return

    #         # Update file info
    #         basename = os.path.basename(self.calibration_file_path_stored)
    #         self.lbl_file_info.setText(f"CALIBRATION DATASET: {basename}")
    #         self._log_message(f"Updated file info to: CALIBRATION DATASET: {basename}")

    #         # Update status
    #         if hasattr(self, 'calibration_status_label'):
    #             self.calibration_status_label.setText(
    #                 "✅ 1. Calibration file selected\n"
    #                 "✅ 2. Calibration dataset loaded\n"
    #                 "3. Detect ROIs (Input tab)\n"
    #                 "4. Process baseline"
    #             )
    #             self._log_message("Updated calibration status label")
    #         else:
    #             self._log_message("WARNING: No calibration_status_label found")

    #         self._log_message("=== CALIBRATION DATASET LOADING COMPLETE ===")
    #         self._log_message("Next step: Go to Input tab and click 'Detect ROIs'")

    #     except Exception as e:
    #         self._log_message(f"CRITICAL ERROR in load_calibration_dataset: {e}")
    #         import traceback
    #         self._log_message(f"Full traceback: {traceback.format_exc()}")
    #         # Reset to main dataset on error
    #         self.current_dataset_type = "main"
    def enhanced_load_calibration_dataset(self):
        """Load calibration dataset while preserving main dataset."""
        self._log_message("=== LOADING CALIBRATION DATASET ===")

        if (
            not hasattr(self, "calibration_file_path_stored")
            or not self.calibration_file_path_stored
        ):
            self._log_message("ERROR: No calibration file selected")
            return

        try:
            # CRITICAL: Store main dataset BEFORE any calibration operations
            if not getattr(self, "main_dataset_stored", False):
                if not self.store_main_dataset_state():
                    self._log_message("ERROR: Failed to store main dataset state")
                    self._log_message(
                        "Cannot proceed with calibration without preserving main dataset"
                    )
                    return

            # Switch to calibration mode
            self.current_dataset_type = "calibration"

            # Load calibration first frame (don't change self.file_path)
            if DUAL_STRUCTURE_AVAILABLE:
                from ._reader import get_first_frame_enhanced

                first_frame, structure_info = get_first_frame_enhanced(
                    self.calibration_file_path_stored
                )
            else:
                from ._reader import get_first_frame

                first_frame = get_first_frame(self.calibration_file_path_stored)

            if first_frame is None:
                self._log_message("ERROR: Could not read calibration first frame")
                return

            self._log_message(f"Loaded calibration first frame: {first_frame.shape}")

            # Add calibration layer
            basename = os.path.basename(self.calibration_file_path_stored)
            layer_name = f"CALIBRATION - {basename} (First Frame)"

            cal_layer = self.viewer.add_image(
                first_frame, name=layer_name, colormap="plasma", visible=True
            )

            cal_layer.metadata.update(
                {
                    "dataset_type": "calibration",
                    "file_path": self.calibration_file_path_stored,
                    "workflow_step": "first_frame_loaded",
                }
            )

            # Update UI
            self.lbl_file_info.setText(
                f"CALIBRATION: {basename} (Main dataset preserved)"
            )

            # Update status
            if hasattr(self, "calibration_status_label"):
                self.calibration_status_label.setText(
                    "✅ 1. Calibration file selected\n"
                    "✅ 2. Calibration first frame loaded\n"
                    "3. Detect ROIs (Input tab)\n"
                    "4. Process baseline\n"
                    "5. Run analysis on main dataset"
                )

            self._log_message("Calibration first frame loaded (main dataset preserved)")

        except Exception as e:
            self._log_message(f"ERROR loading calibration dataset: {e}")
            self.current_dataset_type = "main"

    def restore_main_dataset_for_analysis(self):
        """
        CRITICAL: Restore main dataset before running analysis.
        This ensures analysis runs on the correct (main) dataset.
        """
        self._log_message("=== RESTORING MAIN DATASET FOR ANALYSIS ===")

        if not hasattr(self, "main_dataset_stored") or not self.main_dataset_stored:
            self._log_message("WARNING: No main dataset stored - analysis may fail")
            return False

        try:
            # Restore main dataset state
            self.file_path = self.main_dataset_path
            self.merged_results = self.main_merged_results.copy()
            self.masks = self.main_masks.copy()
            self.labeled_frame = self.main_labeled_frame
            self.current_dataset_type = "main"

            # Verify restoration
            if self.merged_results:
                sample_roi = list(self.merged_results.keys())[0]
                sample_data = self.merged_results[sample_roi]
                if sample_data:
                    restored_duration = (sample_data[-1][0] - sample_data[0][0]) / 60
                    self._log_message("✅ Main dataset restored:")
                    self._log_message(f"   Path: {os.path.basename(self.file_path)}")
                    self._log_message(f"   ROIs: {len(self.merged_results)}")
                    self._log_message(f"   Duration: {restored_duration:.1f} minutes")
                    self._log_message(f"   Data points: {len(sample_data)}")
                    return True

            self._log_message("ERROR: Restored dataset appears empty")
            return False

        except Exception as e:
            self._log_message(f"ERROR restoring main dataset: {e}")
            return False

    def _apply_automatic_timing_fix(
        self, merged_results: Dict[int, List[Tuple[float, float]]]
    ) -> Tuple[Dict, bool]:
        """
        Automatically detect and fix timing issues in HDF5 data.

        Returns:
            (corrected_merged_results, was_corrected)
        """
        if not merged_results:
            return merged_results, False

        # Get sample data to analyze timing
        sample_roi = list(merged_results.keys())[0]
        sample_data = merged_results[sample_roi]

        if len(sample_data) < 3:
            self._log_message("Insufficient data for timing analysis")
            return merged_results, False

        # Calculate actual intervals from timestamps
        times = [t for t, _ in sample_data[:20]]  # Use first 20 points
        intervals = [times[i + 1] - times[i] for i in range(len(times) - 1)]

        actual_interval = np.median(intervals)
        expected_interval = self.frame_interval.value()
        interval_std = np.std(intervals)

        self._log_message("🔍 AUTOMATIC TIMING ANALYSIS:")
        self._log_message(f"  Expected interval: {expected_interval:.1f}s")
        self._log_message(f"  Detected interval: {actual_interval:.1f}s")
        self._log_message(f"  Interval std: {interval_std:.2f}s")

        # Check if correction is needed (tolerance: 1 second or 10% of expected)
        tolerance = max(1.0, expected_interval * 0.1)
        needs_correction = abs(actual_interval - expected_interval) > tolerance

        if not needs_correction:
            self._log_message("✅ Timing is correct - no correction needed")
            return merged_results, False

        # AUTOMATIC CORRECTION
        correction_factor = actual_interval / expected_interval
        self._log_message("⚠️  TIMING MISMATCH DETECTED - Applying automatic fix")
        self._log_message(f"  Correction factor: {correction_factor:.2f}x")

        # 1. Update frame interval
        self.frame_interval.setValue(actual_interval)
        self._log_message(
            f"✅ Updated frame interval: {expected_interval:.1f}s → {actual_interval:.1f}s"
        )

        # 2. Adjust baseline duration (keep same number of frames)
        original_baseline_min = self.baseline_duration_minutes.value()
        baseline_frames = int((original_baseline_min * 60) / expected_interval)
        corrected_baseline_min = (baseline_frames * actual_interval) / 60
        self.baseline_duration_minutes.setValue(corrected_baseline_min)
        self._log_message(
            f"✅ Adjusted baseline: {original_baseline_min:.1f}min → {corrected_baseline_min:.1f}min ({baseline_frames} frames)"
        )

        # 3. Adjust bin size for fraction movement
        original_bin = self.bin_size_seconds.value()
        frames_per_bin = max(1, round(original_bin / actual_interval))
        corrected_bin = frames_per_bin * actual_interval
        self.bin_size_seconds.setValue(int(corrected_bin))
        self._log_message(
            f"✅ Adjusted bin size: {original_bin}s → {corrected_bin:.0f}s ({frames_per_bin} frames/bin)"
        )

        # 4. Update plot time ranges if they exist
        if hasattr(self, "plot_end_time"):
            self.update_end_time()  # This will recalculate based on new frame interval

        self._log_message("🎉 AUTOMATIC TIMING CORRECTION COMPLETE!")

        return merged_results, True

    def _extract_analysis_parameters(self) -> Dict[str, Any]:
        """Enhanced parameter extraction that handles calibration method specifics."""
        # Determine threshold method from active tab
        threshold_method = self._get_current_threshold_method()

        # Basic parameters
        params = {
            "threshold_method": threshold_method,
            "enable_matlab_norm": True,
            "enable_detrending": self.enable_detrending.isChecked(),
            "use_improved_detrending": True,
            "frame_interval": self.frame_interval.value(),
            "apply_hdf5_timing_correction_flag": True,
            "bin_size_seconds": self.bin_size_seconds.value(),
            "quiescence_threshold": self.quiescence_threshold.value(),
            "sleep_threshold_minutes": self.sleep_threshold_minutes.value(),
            "num_processes": self.num_processes.value(),
        }

        # Add method-specific parameters
        if threshold_method == "baseline":
            params.update(
                {
                    "baseline_duration_minutes": self.baseline_duration_minutes.value(),
                    "multiplier": self.threshold_multiplier.value(),
                    "enable_jump_correction": self.enable_jump_correction.isChecked(),
                }
            )
        elif threshold_method == "calibration":
            # For the new pre-computed workflow, we don't need file path/masks
            # because they're already processed and stored in calibration_baseline_statistics
            params.update(
                {
                    "calibration_multiplier": self.calibration_multiplier.value(),
                    # Note: calibration_file_path and masks are handled separately
                    # in the new workflow since baseline is pre-computed
                }
            )

            # Only add these for legacy workflow
            if not (
                hasattr(self, "calibration_baseline_processed")
                and self.calibration_baseline_processed
            ):
                cal_file_path = self.calibration_file_path.property("full_path")
                params.update(
                    {
                        "calibration_file_path": cal_file_path,
                        "masks": self.masks,
                    }
                )
        elif threshold_method == "adaptive":
            params.update(
                {
                    "adaptive_duration_minutes": self.adaptive_duration_minutes.value(),
                    "adaptive_multiplier": self.adaptive_base_multiplier.value(),
                }
            )

        return params

    def update_calibration_workflow_status(
        self, step: str, success: bool = True, message: str = ""
    ):
        """Update the calibration workflow status display."""
        if not hasattr(self, "calibration_status_label"):
            return

        steps = {
            "file_selected": (
                "✅ 1. Calibration file selected"
                if success
                else "❌ 1. Select calibration file"
            ),
            "dataset_loaded": (
                "✅ 2. Calibration dataset loaded"
                if success
                else "2. Load calibration dataset"
            ),
            "rois_detected": (
                "✅ 3. Calibration ROIs detected"
                if success
                else "3. Detect ROIs (Input tab)"
            ),
            "baseline_processed": (
                "✅ 4. Calibration baseline processed"
                if success
                else "4. Process baseline"
            ),
        }

        current_status = []
        for step_key, step_text in steps.items():
            if step_key == step:
                current_status.append(step_text)
                if success and step != "baseline_processed":
                    # Add next step
                    next_steps = list(steps.keys())
                    current_idx = next_steps.index(step_key)
                    if current_idx + 1 < len(next_steps):
                        next_step_key = next_steps[current_idx + 1]
                        current_status.append(steps[next_step_key])
                break
            else:
                current_status.append(step_text)

        if message:
            current_status.append(f"\n{message}")

        self.calibration_status_label.setText("\n".join(current_status))

    def _run_analysis_with_calc_module(self):
        """Simplified analysis using integration system."""
        try:
            # Determine method
            method_text = self._get_current_threshold_method_display()
            method = (
                "calibration"
                if "Calibration" in method_text
                else "baseline" if "Baseline" in method_text else "adaptive"
            )

            # Determine file and masks to use
            if (
                method == "calibration"
                and hasattr(self, "main_dataset_path")
                and self.main_dataset_path
            ):
                file_to_process = self.main_dataset_path
                masks_to_use = getattr(self, "main_masks", [])
            else:
                file_to_process = self.file_path
                masks_to_use = self._get_active_masks()

            # Log excluded ROIs
            excluded = self._get_excluded_roi_indices()
            if excluded:
                excluded_names = [f"ROI {i+1}" for i in excluded]
                self._log_message(f"Excluding {len(excluded)} ROI(s): {', '.join(excluded_names)}")

            if not file_to_process or not masks_to_use:
                raise RuntimeError("No file or masks available for processing")

            # Progress callback
            def progress_callback(percent, msg):
                if self._cancel_requested:
                    raise RuntimeError("Analysis canceled")
                self.progress_updated.emit(int(percent))
                self.status_updated.emit(msg)

            # Check if this is an AVI batch
            if hasattr(self, "avi_batch_paths") and self.avi_batch_paths:
                self._log_message("Processing AVI batch - loading all frames...")
                _, merged_results, _ = self._process_avi_batch_for_analysis(
                    self.avi_batch_paths,
                    masks_to_use,
                    self.chunk_size.value(),
                    progress_callback,
                    self.avi_batch_interval,
                )
            else:
                # Process complete dataset using reader (HDF5)
                _, merged_results, _ = process_single_file_in_parallel_dual_structure(
                    file_to_process,
                    masks_to_use,
                    self.chunk_size.value(),
                    progress_callback,
                    self.frame_interval.value(),
                    self.num_processes.value(),
                )

            # Remap ROI indices to preserve original numbering when ROIs are excluded
            roi_mapping = self._get_roi_index_mapping()
            if any(k != v for k, v in roi_mapping.items()):
                merged_results = {roi_mapping.get(k, k): v for k, v in merged_results.items()}
                self._log_message(f"ROI index mapping applied: {roi_mapping}")

            # Store raw data as fallback (will be overwritten with normalized
            # data in _analysis_finished if the calc step succeeds)
            self.merged_results = merged_results

            # Signal that reader is done but calc is still running
            self.status_updated.emit("Computing thresholds & movement detection...")
            self.progress_updated.emit(95)  # Not 100% yet - calc step still pending

            # Apply timing correction
            merged_results, _ = self._apply_automatic_timing_fix(merged_results)

            # Get analysis parameters
            analysis_params = self._extract_analysis_parameters()
            if method == "calibration":
                analysis_params["calibration_baseline_statistics"] = (
                    self.calibration_baseline_statistics
                )

            # Use integration system
            return run_analysis_with_method(merged_results, method, **analysis_params)

        except Exception as e:
            self._log_message(f"Analysis error: {e}")
            raise

    def _log_analysis_parameters(self):
        """Log analysis parameters for debugging."""
        self._log_message("=" * 50)
        self._log_message("STARTING ANALYSIS WITH _calc.py MODULE")
        self._log_message(f"ROIs: {len(self.masks)}")
        self._log_message(f"Processes: {self.num_processes.value()}")
        self._log_message(f"Chunk size: {self.chunk_size.value()}")
        self._log_message(f"Method: {self._get_current_threshold_method_display()}")
        self._log_message(f"Frame interval: {self.frame_interval.value()}s")
        self._log_message(
            f"Baseline duration: {self.baseline_duration_minutes.value()} min"
        )
        self._log_message(f"Threshold multiplier: {self.threshold_multiplier.value()}")
        self._log_message("MATLAB normalization: Enabled")
        self._log_message("HDF5 timing correction: Enabled")
        self._log_message("=" * 50)

    def _log_timing_analysis_with_units(self, timing_info):
        """Log timing analysis with clear unit display."""
        self._log_message("TIMING ANALYSIS (units clarified):")

        if "mean_frame_drift" in timing_info:
            drift_s = timing_info["mean_frame_drift"]
            drift_ms = drift_s * 1000
            self._log_message(f"  Mean Frame Drift: {drift_s:.3f}s ({drift_ms:.1f}ms)")

        if "max_frame_drift" in timing_info:
            max_drift_s = timing_info["max_frame_drift"]
            max_drift_ms = max_drift_s * 1000
            self._log_message(
                f"  Max Frame Drift: {max_drift_s:.3f}s ({max_drift_ms:.1f}ms)"
            )

        if "std_frame_drift" in timing_info:
            std_drift_s = timing_info["std_frame_drift"]
            std_drift_ms = std_drift_s * 1000
            self._log_message(
                f"  Drift Std Dev: {std_drift_s:.3f}s ({std_drift_ms:.1f}ms)"
            )

    def stop_analysis(self):
        """Stop analysis with proper cleanup."""
        self._cancel_requested = True
        self.status_label.setText("Stopping analysis...")
        self._log_message("STOP requested by user")

        # Stop performance monitoring
        self.performance_timer.stop()

        # Try to gracefully stop the worker
        if hasattr(self, "current_worker") and self.current_worker is not None:
            try:
                # Disconnect signals to prevent callbacks during shutdown
                self.current_worker.returned.disconnect()
                self.current_worker.errored.disconnect()
                self.current_worker.finished.disconnect()
            except Exception as e:
                self._log_message(f"Warning during worker cleanup: {e}")

            # Clear worker reference
            self.current_worker = None

        # Reset UI state
        self.btn_analyze.setEnabled(True)
        self.btn_stop.setEnabled(False)

    def _auto_adjust_time_range(self):
        """Set plot_start_time / plot_end_time from the actual recording duration."""
        data = getattr(self, "merged_results", None) or getattr(self, "movement_data", None)
        if not data:
            return

        max_time_seconds = 0.0
        for roi_data in data.values():
            if roi_data:
                times = [t for t, _ in roi_data]
                if times:
                    max_time_seconds = max(max_time_seconds, max(times))

        max_time_minutes = max_time_seconds / 60.0
        if max_time_minutes <= 0:
            return

        if hasattr(self, "plot_end_time"):
            self.plot_end_time.setRange(0.0, max_time_minutes * 1.05)
            self.plot_end_time.setValue(max_time_minutes)
            self.plot_start_time.setRange(0.0, max_time_minutes * 1.05)
            self.plot_start_time.setValue(0.0)
            self._log_message(
                f"Plot time range set: 0 – {max_time_minutes:.1f} min "
                f"({max_time_minutes / 60:.2f} h)"
            )

        # Cap baseline duration spinbox to recording length
        if hasattr(self, "baseline_duration_minutes"):
            self.baseline_duration_minutes.setMaximum(max_time_minutes)
            if self.baseline_duration_minutes.value() > max_time_minutes:
                self.baseline_duration_minutes.setValue(max_time_minutes)

        # Update Extended Analysis time range (hours axis)
        max_time_hours = max_time_seconds / 3600.0
        if hasattr(self, "cycle_end_time") and max_time_hours > 0:
            self.cycle_end_time.setRange(0.0, max_time_hours * 1.05)
            self.cycle_end_time.setValue(max_time_hours)
            self.cycle_start_time.setRange(0.0, max_time_hours * 1.05)
            self.cycle_start_time.setValue(0.0)
            self._log_message(
                f"Extended Analysis time range set: 0 – {max_time_hours:.2f} h"
            )

    def _analysis_finished(self, result: Dict[str, Any]):
        """Handle successful analysis completion using results from _calc.py."""
        try:
            # Capture start time immediately before it can be cleared by _analysis_done()
            start_time = self.analysis_start_time

            # Store normalized [0,1] data (default for display)
            self.merged_results = result.get("processed_data", {})
            self.roi_baseline_means = result.get("baseline_means", {})
            self.roi_upper_thresholds = result.get("upper_thresholds", {})
            self.roi_lower_thresholds = result.get("lower_thresholds", {})
            # Store raw amplitude data (pre-MinMax, for real amplitude view)
            self.merged_results_raw = result.get("processed_data_raw", {})
            self.roi_baseline_means_raw = result.get("baseline_means_raw", {})
            self.roi_upper_thresholds_raw = result.get("upper_thresholds_raw", {})
            self.roi_lower_thresholds_raw = result.get("lower_thresholds_raw", {})
            self.roi_statistics = result.get("roi_statistics", {})
            self.movement_data = result.get("movement_data", {})
            self.fraction_data = result.get("fraction_data", {})
            self.sleep_data = result.get("sleep_data", {})
            # Cache LED data from HDF5 for lighting overlay in plots
            self.led_data = self._extract_led_data_from_hdf5() or {}
            # Pixel count per ROI for sum-mode amplitude display (sum = mean × n_pixels)
            masks = getattr(self, "masks", [])
            self.roi_pixel_counts = {
                i + 1: int(np.sum(m > 0)) for i, m in enumerate(masks)
            }
            # Normalization factor: 255 for uint8, 65535 for uint16 cameras
            # Used to convert back to raw pixel units (MATLAB-equivalent values)
            if DUAL_STRUCTURE_AVAILABLE and getattr(self, "file_path", None):
                self.frame_norm_factor = get_frame_norm_factor(self.file_path)
            else:
                self.frame_norm_factor = 1.0
            self._log_message(f"  Frame normalization factor: {self.frame_norm_factor:.0f} (bit depth)")

            # Log inactive ROIs (MinMax normalization is now applied in _calc.py)
            roi_active = result.get("roi_active", {})
            inactive_rois = [roi for roi, active in roi_active.items() if not active]
            if inactive_rois:
                self._log_message(f"⚠️ {len(inactive_rois)} inactive ROI(s) detected (low amplitude): {inactive_rois}")
            self.roi_active = roi_active

            # Get ROI colors from calc results
            self.roi_colors = result.get("roi_colors", {})

            # Fallback if no ROI colors provided
            if not self.roi_colors and self.merged_results:
                self.roi_colors = {
                    roi: f"C{(roi - 1) % 10}"
                    for roi in sorted(self.merged_results.keys())
                }

            self._log_message(f"✅ ROI colors set: {self.roi_colors}")

            # Calculate quiescence data using _calc.py functions
            if self.fraction_data:
                self.quiescence_data = bin_quiescence(
                    self.fraction_data, self.quiescence_threshold.value()
                )

            # Calculate sleep quality metrics (MATLAB-compatible hourly analysis)
            if self.sleep_data:
                from ._calc import calculate_sleep_quality_hourly
                self.sleep_quality_data = calculate_sleep_quality_hourly(
                    self.sleep_data
                )
                self._log_message("✅ Sleep quality metrics calculated (min/h, transitions/h, bout length/h)")

            # Calculate band widths for plotting compatibility (normalized)
            self.roi_band_widths = {}
            for roi in self.roi_baseline_means:
                if (
                    roi in self.roi_upper_thresholds
                    and roi in self.roi_lower_thresholds
                ):
                    upper = self.roi_upper_thresholds[roi]
                    lower = self.roi_lower_thresholds[roi]
                    self.roi_band_widths[roi] = (upper - lower) / 2
            # Also for raw amplitude mode
            self.roi_band_widths_raw = {}
            for roi in self.roi_baseline_means_raw:
                if (
                    roi in self.roi_upper_thresholds_raw
                    and roi in self.roi_lower_thresholds_raw
                ):
                    upper = self.roi_upper_thresholds_raw[roi]
                    lower = self.roi_lower_thresholds_raw[roi]
                    self.roi_band_widths_raw[roi] = (upper - lower) / 2

            # Update plot time range based on actual data duration
            self._auto_adjust_time_range()

            # Calculate performance metrics using _calc.py
            total_frames = (
                sum(len(data) for data in self.merged_results.values())
                if self.merged_results
                else 0
            )
            performance_metrics = get_performance_metrics(start_time, total_frames)

            # Generate summary using _calc.py
            summary = get_analysis_summary(result)
            self.status_label.setText(
                f"Analysis completed: {performance_metrics['fps']:.1f} fps"
            )
            self.results_label.setText("Analysis completed successfully")

            # Log completion with summary
            self._log_message("=" * 60)
            self._log_message("ANALYSIS COMPLETED SUCCESSFULLY")
            self._log_message(f"Processing rate: {performance_metrics['fps']:.1f} fps")
            self._log_message(f"ROIs processed: {len(self.merged_results)}")
            self._log_message(f"Total data points: {total_frames}")
            self._log_message("ANALYSIS SUMMARY:")
            for line in summary.split("\n"):
                if line.strip():
                    self._log_message(line)

            # Log timing diagnostics if available
            timing_info = result.get("timing_diagnostics", {})
            if timing_info:
                self._log_message(
                    f"HDF5 Timing: {timing_info.get('timing_type', 'unknown')}"
                )
                self._log_message(
                    f"Timing correction: {timing_info.get('needs_hdf5_correction', False)}"
                )

            # Re-enable Generate Plot button and set progress to 100%
            self.btn_plot.setEnabled(True)
            self.progress_bar.setValue(100)

            # Auto-generate plot so user doesn't have to click manually
            self.generate_plot()

        except Exception as e:
            self._log_message(f"Error in analysis completion: {str(e)}")
            import traceback

            self._log_message(f"Full traceback: {traceback.format_exc()}")
        finally:
            # Always re-enable Generate Plot button
            self.btn_plot.setEnabled(True)
            self.progress_bar.setValue(100)

    def _analysis_errored(self, exc):
        """Handle analysis errors."""
        self.performance_timer.stop()
        import traceback
        self._log_message(f"ANALYSIS ERROR: {exc}")
        self._log_message(f"Traceback: {''.join(traceback.format_exception(type(exc), exc, exc.__traceback__))}")
        if "canceled" in str(exc).lower():
            self.status_label.setText("Analysis canceled by user.")
            self._log_message("Analysis CANCELED by user")
        else:
            self.status_label.setText(f"Analysis error: {exc}")
            self._log_message(f"ERROR: {exc}")

        # Reset UI state
        self.btn_analyze.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.btn_plot.setEnabled(True)  # Re-enable even on error
        self.progress_bar.setValue(0)

    def _analysis_done(self):
        """Cleanup after analysis completion or cancellation."""
        self.performance_timer.stop()
        self._cancel_requested = False
        self.analysis_start_time = None
        self.current_worker = None  # Clear worker reference

        # Reset UI state
        self.btn_analyze.setEnabled(True)
        self.btn_stop.setEnabled(False)

    # ===================================================================
    # TESTING AND DIAGNOSTICS USING _calc.py
    # ===================================================================

    def run_quick_analysis_test(self):
        """Run quick analysis test using _calc.py module."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("No data available for testing")
            return

        try:
            test_summary = quick_method_test(self.merged_results)
            self._log_message("QUICK ANALYSIS TEST RESULTS:")
            for line in test_summary.split("\n"):
                if line.strip():
                    self._log_message(line)
        except Exception as e:
            self._log_message(f"Quick test failed: {e}")

    def validate_hdf5_timing(self):
        """Validate HDF5 timing using _calc.py module."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("No data available for timing validation")
            return

        try:
            timing_diagnostics = validate_hdf5_timing_in_data(
                self.merged_results, self.frame_interval.value()
            )

            self._log_message("HDF5 TIMING DIAGNOSTICS:")
            self._log_message(f"Timing type: {timing_diagnostics['timing_type']}")
            self._log_message(
                f"First timestamp: {timing_diagnostics['first_time']:.1f}s"
            )
            self._log_message(
                f"Average interval: {timing_diagnostics['avg_interval']:.1f}s"
            )
            self._log_message(
                f"Expected interval: {timing_diagnostics['expected_interval']:.1f}s"
            )
            self._log_message(
                f"Interval consistent: {timing_diagnostics['interval_consistent']}"
            )
            self._log_message(
                f"Needs correction: {timing_diagnostics['needs_hdf5_correction']}"
            )
            self._log_message(
                f"Recommendation: {timing_diagnostics['recommended_action']}"
            )

        except Exception as e:
            self._log_message(f"Timing validation failed: {e}")

    # ===================================================================
    # PLOTTING METHODS - NOW USING _plot.py MODULE
    # ===================================================================

    def _on_plot_bin_changed(self):
        """Refresh plot when plot bin size changes."""
        if hasattr(self, "merged_results") and self.merged_results:
            self.generate_plot()

    def _on_plot_type_changed(self):
        """Handle plot type dropdown change - show/hide sub-selectors."""
        plot_type = self.plot_type_combo.currentText()
        # Show sleep quality metric selector only for Sleep Quality
        if hasattr(self, "sleep_quality_metric_combo"):
            self.sleep_quality_metric_combo.setVisible(plot_type == "Sleep Quality")
        # Generate plot
        self.generate_plot()

    def generate_plot(self):
        """Generate plot using PlotGenerator."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.results_label.setText("No analysis results to plot.")
            return

        # ROI colors should now come from _calc.py, but add safety check
        if not hasattr(self, "roi_colors") or not self.roi_colors:
            self._log_message("No ROI colors from analysis, creating fallback")
            self.roi_colors = {
                roi: f"C{(roi - 1) % 10}" for roi in sorted(self.merged_results.keys())
            }

        # Force clear the canvas to prevent artifacts
        try:
            if hasattr(self, "canvas") and self.canvas:
                self.canvas.clear()

                # Clear any cached renderers
                if hasattr(self.canvas.figure, "_cachedRenderer"):
                    self.canvas.figure._cachedRenderer = None

                # Force immediate refresh
                self.canvas.draw_idle()
                self.canvas.flush_events()
        except Exception as e:
            self._log_message(f"Canvas cleanup warning: {e}")

        # Initialize plot generator if needed
        if self.plot_generator is None:
            try:
                from ._plot import PlotGenerator

                self.plot_generator = PlotGenerator(self.figure)
            except Exception as e:
                self._log_message(f"Plot generator init failed: {e}")
                self.results_label.setText("Plot system initialization failed.")
                return

        plot_type = self.plot_type_combo.currentText()

        try:
            # Check amplitude mode toggle
            use_real_amplitude = (
                hasattr(self, "show_real_amplitude")
                and self.show_real_amplitude.isChecked()
            )

            # ZT mode and lighting overlay settings
            zt_mode = self.chk_zt_axis.isChecked() if hasattr(self, "chk_zt_axis") else False
            show_lighting = self.chk_show_lighting.isChecked() if hasattr(self, "chk_show_lighting") else False
            led_data_for_plot = getattr(self, "led_data", None) if show_lighting else None

            # Get data based on plot type
            if plot_type == "Raw Intensity Changes":
                pixel_sum_scales = {}  # {roi: scale_factor} — used to scale thresholds too
                if use_real_amplitude and getattr(self, "merged_results_raw", {}):
                    divide_by_pixels = (
                        hasattr(self, "chk_divide_by_pixels")
                        and self.chk_divide_by_pixels.isChecked()
                    )
                    pixel_counts = getattr(self, "roi_pixel_counts", {})
                    norm_factor = getattr(self, "frame_norm_factor", 1.0)
                    if not divide_by_pixels and pixel_counts:
                        # MATLAB-equivalent pixel sum:
                        # mean × n_pixels × norm_factor = Σ|ΔPixel_raw| per ROI
                        pixel_sum_scales = {
                            roi: pixel_counts.get(roi, 1) * norm_factor
                            for roi in self.merged_results_raw
                        }
                        data_dict = {
                            roi: [(t, v * pixel_sum_scales[roi]) for t, v in data]
                            for roi, data in self.merged_results_raw.items()
                        }
                        self._log_message(f"Plot mode: Pixel Sum MATLAB (×{norm_factor:.0f})")
                    else:
                        # Per-pixel mean mode: sum(|Δpixel|) ÷ n_pixels
                        data_dict = self.merged_results_raw
                        self._log_message("Plot mode: Per-pixel mean")
                else:
                    data_dict = self.merged_results
                    self._log_message("Plot mode: Normalized [0-1]")

                from ._plot import create_hysteresis_kwargs

                kwargs = create_hysteresis_kwargs(
                    widget_instance=self, use_real_amplitude=use_real_amplitude
                )
                # Remove merged_results from kwargs to avoid duplicate argument
                kwargs.pop("merged_results", None)

                # Scale threshold lines to match pixel sum mode
                if pixel_sum_scales:
                    for thresh_key in ("roi_baseline_means", "roi_upper_thresholds",
                                       "roi_lower_thresholds", "roi_band_widths"):
                        if thresh_key in kwargs and kwargs[thresh_key]:
                            kwargs[thresh_key] = {
                                roi: val * pixel_sum_scales.get(roi, 1.0)
                                for roi, val in kwargs[thresh_key].items()
                            }

                # ZT mode and lighting overlay
                kwargs["zt_mode"] = zt_mode
                kwargs["led_data"] = led_data_for_plot

            elif plot_type == "Movement":
                data_dict = getattr(self, "movement_data", {})
                kwargs = {"zt_mode": zt_mode, "led_data": led_data_for_plot}

            elif plot_type == "Fraction Movement":
                plot_bin_value = getattr(self, "plot_bin_minutes", None)
                bin_minutes = plot_bin_value.value() if plot_bin_value else 0
                data_dict, _, _ = self._get_rebinned_behavioral_data(bin_minutes)
                if bin_minutes > 0:
                    original_bin = self.bin_size_seconds.value() if hasattr(self, "bin_size_seconds") else 60
                    new_bin_seconds = bin_minutes * 60
                    if new_bin_seconds > original_bin:
                        self._log_message(f"Fraction Movement re-binned: {original_bin}s → {new_bin_seconds}s")
                kwargs = {"zt_mode": zt_mode, "led_data": led_data_for_plot}

            elif plot_type == "Quiescence":
                plot_bin_value = getattr(self, "plot_bin_minutes", None)
                bin_minutes = plot_bin_value.value() if plot_bin_value else 0
                _, data_dict, _ = self._get_rebinned_behavioral_data(bin_minutes)
                if bin_minutes > 0:
                    original_bin = self.bin_size_seconds.value() if hasattr(self, "bin_size_seconds") else 60
                    new_bin_seconds = bin_minutes * 60
                    if new_bin_seconds > original_bin:
                        self._log_message(f"Quiescence re-derived from {new_bin_seconds}s rebinned fraction data")
                kwargs = {"zt_mode": zt_mode, "led_data": led_data_for_plot}

            elif plot_type == "Sleep":
                plot_bin_value = getattr(self, "plot_bin_minutes", None)
                bin_minutes = plot_bin_value.value() if plot_bin_value else 0
                _, _, data_dict = self._get_rebinned_behavioral_data(bin_minutes)
                if bin_minutes > 0:
                    original_bin = self.bin_size_seconds.value() if hasattr(self, "bin_size_seconds") else 60
                    new_bin_seconds = bin_minutes * 60
                    if new_bin_seconds > original_bin:
                        self._log_message(f"Sleep re-derived from {new_bin_seconds}s rebinned fraction data")
                kwargs = {"zt_mode": zt_mode, "led_data": led_data_for_plot}

            elif plot_type == "Sleep Quality":
                data_dict = getattr(self, "sleep_quality_data", {})
                # Get metric from sub-selector if available
                sleep_metric = "sleep_minutes"
                if hasattr(self, "sleep_quality_metric_combo"):
                    metric_text = self.sleep_quality_metric_combo.currentText()
                    if "Transitions" in metric_text:
                        sleep_metric = "transitions"
                    elif "Bout" in metric_text:
                        sleep_metric = "bout_length"
                kwargs = {"sleep_metric": sleep_metric}

            elif plot_type == "Lighting Conditions (dark IR)":
                data_dict = getattr(self, "fraction_data", {})
                from ._plot import create_hysteresis_kwargs

                kwargs = create_hysteresis_kwargs(
                    widget_instance=self
                )  # Keep merged_results for lighting

                # Use plot_bin_minutes from GUI (default 10 if not set)
                plot_bin_value = getattr(self, "plot_bin_minutes", None)
                bin_minutes = plot_bin_value.value() if plot_bin_value else 10
                kwargs.update({"bin_minutes": bin_minutes})
                self._log_message(f"Using plot binning: {bin_minutes} minutes")

                # Extract LED data from HDF5 if available
                led_data = self._extract_led_data_from_hdf5()
                if led_data:
                    kwargs["led_data"] = led_data
                    self._log_message(
                        f"Using LED data from HDF5: {len(led_data.get('times', []))} data points"
                    )

            else:
                self.results_label.setText(f"Unknown plot type: {plot_type}")
                return

            if not data_dict:
                self.results_label.setText(f"No {plot_type.lower()} data available.")
                return

            # Create plot config
            from ._plot import create_plot_config

            plot_config = create_plot_config(self)

            # Generate plot
            success = self.plot_generator.generate_plot(
                plot_type, data_dict, self.roi_colors, plot_config, **kwargs
            )

            if success:
                # Force complete canvas refresh
                self.canvas.draw()
                self.canvas.flush_events()
                self.results_label.setText(f"Generated {plot_type} plot successfully.")
                self._log_message(f"Generated {plot_type} plot")
            else:
                self.results_label.setText(f"Failed to generate {plot_type} plot.")
                self._log_message("Plot generation returned False")
                # Add debugging info
                self._log_message(f"Debug - Plot type: {plot_type}")
                self._log_message(
                    f"Debug - Data dict keys: {list(data_dict.keys()) if data_dict else 'None'}"
                )
                self._log_message(
                    f"Debug - ROI colors: {len(self.roi_colors) if self.roi_colors else 0} colors"
                )

        except Exception as e:
            self._log_message(f"Plot error: {e}")
            self.results_label.setText(f"Plot generation failed: {str(e)}")
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def _extract_led_data_from_hdf5(self):
        """Extract LED power timeseries from HDF5 or AVI file.

        Light phase = white LED ON (alone or with IR LED)
        Dark phase = only IR LED ON (white LED OFF)

        Returns:
            dict or None: Dictionary with 'times', 'white_powers', and 'ir_powers' keys if LED data exists,
                         None if no LED data is available
        """
        try:
            if not hasattr(self, "file_path") or not self.file_path:
                return None

            # Check if this is an AVI file
            if self.file_path.lower().endswith(".avi"):
                return self._extract_led_data_from_avi()

            # HDF5 file processing
            with h5py.File(self.file_path, "r") as f:
                if "timeseries" not in f:
                    return None

                timeseries = f["timeseries"]

                # Try to find white LED data (various possible names)
                white_led = None
                white_led_names = [
                    "led_white_power_percent",
                    "white_led_power",
                    "led_white_power",
                    "white_led_power_percent",
                ]
                for name in white_led_names:
                    if name in timeseries:
                        white_led = timeseries[name][:]
                        self._log_message(f"Found white LED data: {name}")
                        break

                # Special case: "led_power_percent" without specific white/IR separation
                # This is typically IR-only systems (old recordings) - don't use for lighting detection
                if white_led is None and "led_power_percent" in timeseries:
                    self._log_message(
                        "Found generic 'led_power_percent' but no white LED channel - likely IR-only system"
                    )
                    self._log_message(
                        "→ Using legacy 12h light/dark cycles for visualization"
                    )
                    return None  # Will trigger legacy 12h cycle visualization

                # Try to find IR LED data (various possible names)
                ir_led = None
                ir_led_names = [
                    "led_ir_power_percent",
                    "ir_led_power",
                    "led_ir_power",
                    "ir_led_power_percent",
                ]
                for name in ir_led_names:
                    if name in timeseries:
                        ir_led = timeseries[name][:]
                        self._log_message(f"Found IR LED data: {name}")
                        break

                # If no separate LED channels found, return None
                if white_led is None:
                    self._log_message("No white LED data found in HDF5 timeseries")
                    return None

                # Get timestamps (try capture_timestamps first, fallback to calculated times)
                if "capture_timestamps" in timeseries:
                    times = timeseries["capture_timestamps"][:]
                else:
                    # Fallback: use frame interval to calculate times
                    frame_interval = self.frame_interval.value()
                    times = np.arange(len(white_led)) * frame_interval

                result = {"times": times.tolist(), "white_powers": white_led.tolist()}

                if ir_led is not None:
                    result["ir_powers"] = ir_led.tolist()

                return result

        except Exception as e:
            self._log_message(f"Could not extract LED data from HDF5: {e}")
            return None

    def _extract_led_data_from_avi(self):
        """AVI files don't contain LED data.

        LED data is only available in HDF5 files.
        For AVIs, plots will not show lighting conditions.

        Returns:
            None (AVIs don't have LED data)
        """
        self._log_message(
            "AVI files don't contain LED data - lighting conditions will not be shown in plots"
        )
        return None

    def apply_time_range(self):
        """Apply time range and regenerate plot."""
        self.generate_plot()

    def save_current_plot(self):
        """Save the current plot using _plot.py module."""
        # Build a meaningful default filename
        try:
            base = os.path.splitext(os.path.basename(self.file_path))[0]
        except Exception:
            base = "analysis"
        plot_type = (
            self.plot_type_combo.currentText().replace(" ", "_").lower()
            if hasattr(self, "plot_type_combo")
            else "plot"
        )
        default_name = f"{base}_{plot_type}.png"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Plot",
            default_name,
            "PNG Files (*.png);;PDF Files (*.pdf);;SVG Files (*.svg);;All Files (*)",
        )

        if file_path:
            dpi = self.plot_dpi_spin.value()
            success = save_plot(self.figure, file_path, dpi)

            if success:
                self._log_message(f"Plot saved: {os.path.basename(file_path)}")
                self.results_label.setText(f"Plot saved: {os.path.basename(file_path)}")
            else:
                error_msg = "Failed to save plot"
                self.results_label.setText(error_msg)
                self._log_message(f"ERROR: {error_msg}")

    def save_all_plots(self):
        """Save all plot types using _plot.py module."""
        directory = QFileDialog.getExistingDirectory(
            self, "Select Directory to Save All Plots"
        )

        if not directory:
            return

        try:
            # Extract current display settings so saved plots match what is shown
            zt_mode = self.chk_zt_axis.isChecked() if hasattr(self, "chk_zt_axis") else False
            show_lighting = self.chk_show_lighting.isChecked() if hasattr(self, "chk_show_lighting") else False
            led_data_for_plot = getattr(self, "led_data", None) if show_lighting else None

            # Apply pixel-sum mode to raw intensity data if the toggle is off (= sum mode)
            divide_by_pixels = hasattr(self, "chk_divide_by_pixels") and self.chk_divide_by_pixels.isChecked()
            pixel_counts = getattr(self, "roi_pixel_counts", {})
            norm_factor = getattr(self, "frame_norm_factor", 1.0)
            raw_results = getattr(self, "merged_results_raw", {})
            if not divide_by_pixels and pixel_counts and raw_results:
                merged_for_plot = {
                    roi: [(t, v * pixel_counts.get(roi, 1) * norm_factor) for t, v in data]
                    for roi, data in raw_results.items()
                }
            else:
                merged_for_plot = getattr(self, "merged_results", {})

            # Prepare all data sets
            data_sets = {
                "merged_results": merged_for_plot,
                "movement_data": getattr(self, "movement_data", {}),
                "fraction_data": getattr(self, "fraction_data", {}),
                "quiescence_data": getattr(self, "quiescence_data", {}),
                "sleep_data": getattr(self, "sleep_data", {}),
            }

            # Create plot configuration
            plot_config = create_plot_config(self)

            # Generate timestamp
            timestamp = str(int(time.time()))

            # Save all plots using the separated module
            saved_files = save_all_plot_types(
                self.plot_generator,
                data_sets,
                self.roi_colors,
                plot_config,
                directory,
                timestamp,
                zt_mode=zt_mode,
                led_data=led_data_for_plot,
            )

            if saved_files:
                self.results_label.setText(
                    f"Saved {len(saved_files)} plots to {directory}"
                )
                self._log_message(f"Saved {len(saved_files)} plots successfully")

                # Restore original plot type
                self.generate_plot()
            else:
                self.results_label.setText("No plots were saved")
                self._log_message("WARNING: No plots were saved")

        except Exception as e:
            error_msg = f"Error saving plots: {str(e)}"
            self.results_label.setText(error_msg)
            self._log_message(f"ERROR: {error_msg}")

    # ===================================================================
    # EXPORT AND RESULTS METHODS
    # ===================================================================

    def check_hdf5_structure(self):
        """Enhanced HDF5 structure checking with dual structure support."""
        if not self.file_path:
            self._log_message("No file loaded")
            return

        # Skip structure check for AVI files
        if self.file_path.lower().endswith(".avi") or (
            hasattr(self, "avi_batch_paths") and self.avi_batch_paths
        ):
            self._log_message("AVI file(s) loaded - skipping HDF5 structure check")
            return

        import h5py

        try:
            if DUAL_STRUCTURE_AVAILABLE:
                # Use enhanced structure detection
                structure_info = detect_hdf5_structure_type(self.file_path)

                self._log_message("=== ENHANCED HDF5 FILE STRUCTURE ANALYSIS ===")
                self._log_message(f"Structure type: {structure_info['type']}")

                if structure_info["type"] == "stacked_frames":
                    self._log_message("✅ Stacked frames structure")
                    self._log_message(f"   Dataset: {structure_info['dataset_name']}")
                    self._log_message(
                        f"   Frame count: {structure_info['frame_count']}"
                    )
                    self._log_message(
                        f"   Frame shape: {structure_info['frame_shape']}"
                    )

                elif structure_info["type"] == "individual_frames":
                    self._log_message("✅ Individual frames structure")
                    self._log_message(f"   Group: {structure_info['group_name']}")
                    self._log_message(
                        f"   Frame count: {structure_info['frame_count']}"
                    )
                    self._log_message(
                        f"   Frame shape: {structure_info['frame_shape']}"
                    )
                    self._log_message(
                        f"   Sample keys: {structure_info['frame_keys'][:5]}..."
                    )

                elif structure_info["type"] == "alternative_dataset":
                    self._log_message("✅ Alternative dataset structure")
                    self._log_message(f"   Dataset: {structure_info['dataset_name']}")
                    self._log_message(
                        f"   Frame count: {structure_info['frame_count']}"
                    )

                elif structure_info["type"] == "error":
                    self._log_message(
                        f"❌ Structure detection failed: {structure_info['error']}"
                    )

            else:
                # Fallback to original method
                with h5py.File(self.file_path, "r") as f:
                    self._log_message("=== BASIC HDF5 FILE STRUCTURE ===")
                    self._log_message(f"Root keys: {list(f.keys())}")

                    def print_structure(name, obj):
                        if isinstance(obj, h5py.Group):
                            self._log_message(
                                f"Group: {name} - keys: {list(obj.keys())}"
                            )
                        elif isinstance(obj, h5py.Dataset):
                            self._log_message(f"Dataset: {name} - shape: {obj.shape}")

                    f.visititems(print_structure)

        except Exception as e:
            self._log_message(f"HDF5 structure check failed: {e}")

    def _create_basic_matlab_export(
        self, analysis_results: Dict[str, Any], output_dir: str
    ) -> List[str]:
        """Fallback MATLAB export when modern system not available."""
        import csv
        import json
        from datetime import datetime

        created_files = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            # Export basic CSV for MATLAB
            csv_file = os.path.join(output_dir, f"matlab_export_{timestamp}.csv")

            with open(csv_file, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)

                # Header
                writer.writerow(["# MATLAB Export"])
                writer.writerow(
                    [f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                )
                writer.writerow([])

                # ROI summary
                writer.writerow(
                    ["ROI_ID", "Baseline_Mean", "Upper_Threshold", "Lower_Threshold"]
                )

                baseline_means = analysis_results.get("baseline_means", {})
                upper_thresholds = analysis_results.get("upper_thresholds", {})
                lower_thresholds = analysis_results.get("lower_thresholds", {})

                for roi in sorted(baseline_means.keys()):
                    writer.writerow(
                        [
                            roi,
                            baseline_means.get(roi, 0),
                            upper_thresholds.get(roi, 0),
                            lower_thresholds.get(roi, 0),
                        ]
                    )

            created_files.append(csv_file)

            # Export parameters as JSON
            json_file = os.path.join(output_dir, f"matlab_parameters_{timestamp}.json")
            with open(json_file, "w") as f:
                json.dump(
                    analysis_results.get("parameters", {}), f, indent=2, default=str
                )

            created_files.append(json_file)

        except Exception as e:
            print(f"Error in basic MATLAB export: {e}")

        return created_files

    def export_results_for_matlab_compatibility(self):
        """Export analysis results for MATLAB compatibility."""
        # 1) Check preconditions
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.results_label.setText("No results to export.")
            return

        directory = QFileDialog.getExistingDirectory(
            self, "Select Directory for MATLAB Export"
        )
        if not directory:
            return

        try:
            # 2) Prepare analysis object
            analysis_results = {
                "merged_results": getattr(self, "merged_results", {}),
                "baseline_means": getattr(self, "roi_baseline_means", {}),
                "upper_thresholds": getattr(self, "roi_upper_thresholds", {}),
                "lower_thresholds": getattr(self, "roi_lower_thresholds", {}),
                "movement_data": getattr(self, "movement_data", {}),
                "fraction_data": getattr(self, "fraction_data", {}),
                "sleep_data": getattr(self, "sleep_data", {}),
                "parameters": {
                    "threshold_method": self._get_current_threshold_method_display(),
                    "frame_interval": self.frame_interval.value(),
                    "enable_matlab_norm": True,
                    "enable_detrending": self.enable_detrending.isChecked(),
                },
            }

            # Add method-specific parameters
            if hasattr(self, "baseline_duration_minutes"):
                analysis_results["parameters"][
                    "baseline_duration_minutes"
                ] = self.baseline_duration_minutes.value()
            if hasattr(self, "threshold_multiplier"):
                analysis_results["parameters"][
                    "threshold_multiplier"
                ] = self.threshold_multiplier.value()
            if hasattr(self, "calibration_multiplier"):
                analysis_results["parameters"][
                    "calibration_multiplier"
                ] = self.calibration_multiplier.value()

            # 3) Try modern export
            created_files = []
            try:
                from ._calc_integration import export_results_for_matlab as _export_fn

                created_files = _export_fn(analysis_results, directory)
            except Exception:
                _export_fn = None

            # 4) Fallback export if modern system not available
            if not created_files:
                created_files = self._create_basic_matlab_export(
                    analysis_results, directory
                )

            # 5) UI feedback
            if created_files:
                self.results_label.setText(
                    f"Exported {len(created_files)} files for MATLAB"
                )
                self._log_message(
                    f"MATLAB export completed: {len(created_files)} files"
                )
                for p in created_files:
                    self._log_message(f"  Created: {os.path.basename(p)}")
            else:
                self.results_label.setText("Export failed")
                self._log_message("MATLAB export failed")

        except Exception as e:
            err = f"Error exporting for MATLAB: {e}"
            self.results_label.setText(err)
            self._log_message(f"ERROR: {err}")

    def save_results_consolidated_complete(self):
        """
        UPDATED VERSION: Complete consolidated save function with all sheets.
        This replaces the previous save_results_consolidated method.
        """
        import os
        import time

        # Check if analysis results are available
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.results_label.setText(
                "❌ No analysis results to save. Run analysis first."
            )
            self._log_message("Save failed: No analysis results available")
            return

        # Check if we have behavioral analysis data
        has_behavioral_data = (
            hasattr(self, "movement_data")
            and self.movement_data
            and hasattr(self, "fraction_data")
            and self.fraction_data
        )

        if not has_behavioral_data:
            self.results_label.setText("⚠️ Incomplete analysis data detected.")
            self._log_message("Warning: Saving with incomplete behavioral analysis")

        # Get base filename from user
        from qtpy.QtWidgets import QFileDialog

        base_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Analysis Results",
            f"analysis_results_{int(time.time())}",  # No extension - we'll add them
            "All Files (*)",
        )

        if not base_path:
            self._log_message("Save cancelled by user")
            return

        # Remove any extension from base_path to ensure clean naming
        base_path = os.path.splitext(base_path)[0]

        saved_files = []
        sheets_created = []

        try:
            # === SAVE CSV VERSION ===
            csv_path = f"{base_path}.csv"
            self._log_message(f"Saving CSV version: {csv_path}")

            try:
                self._save_results_csv(csv_path)
                saved_files.append(("CSV", csv_path))
                self._log_message("✅ CSV saved successfully")
            except Exception as e:
                self._log_message(f"❌ CSV save failed: {e}")

            # === SAVE COMPLETE EXCEL VERSION (if possible) ===
            try:
                import pandas as pd
                import openpyxl

                excel_path = f"{base_path}.xlsx"
                self._log_message(
                    f"Saving complete Excel version with all sheets: {excel_path}"
                )

                # Use the complete Excel save method
                self._save_results_excel_to_path(excel_path)
                saved_files.append(("Excel", excel_path))

                # Count sheets created
                wb = openpyxl.load_workbook(excel_path)
                sheets_created = wb.sheetnames
                wb.close()

                self._log_message(
                    f"✅ Excel saved with {len(sheets_created)} sheets: {', '.join(sheets_created)}"
                )

            except ImportError:
                self._log_message(
                    "⚠️ Excel export not available (missing pandas/openpyxl)"
                )
                self._log_message("   Install with: pip install pandas openpyxl")
            except Exception as e:
                self._log_message(f"❌ Excel save failed: {e}")
                import traceback

                self._log_message(f"Traceback: {traceback.format_exc()}")

            # === SHOW THRESHOLD STATS IN LOG ===
            if hasattr(self, "roi_baseline_means") and self.roi_baseline_means:
                self._log_message("\n" + "=" * 50)
                self._log_message("THRESHOLD STATISTICS (included with save)")
                self._log_message("=" * 50)
                self._show_threshold_stats_in_log()

            # === UPDATE UI WITH RESULTS ===
            if saved_files:
                if sheets_created:
                    file_list = f"CSV + Excel ({len(sheets_created)} sheets: {', '.join(sheets_created)})"
                else:
                    file_list = ", ".join(
                        [
                            f"{fmt} ({os.path.basename(path)})"
                            for fmt, path in saved_files
                        ]
                    )

                self.results_label.setText(f"✅ Saved: {file_list}")
                self._log_message(
                    f"\n🎉 SAVE COMPLETE: {len(saved_files)} files created"
                )

                # Show success dialog with file details
                self._show_save_success_dialog_complete(saved_files, sheets_created)
            else:
                self.results_label.setText(
                    "❌ All save attempts failed - check log for details"
                )
                self._log_message("❌ No files were saved successfully")

        except Exception as e:
            error_msg = f"Save operation failed: {e}"
            self.results_label.setText(error_msg)
            self._log_message(f"❌ {error_msg}")
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def _show_save_success_dialog_complete(self, saved_files, sheets_created):
        """Show success dialog with complete file and sheet details."""
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)
        msg.setWindowTitle("Save Complete")
        msg.setText(
            f"Successfully saved analysis results in {len(saved_files)} format(s):"
        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                file_size = "Unknown size"

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        if sheets_created:
            file_details.append("")
            file_details.append("Excel Sheets Created:")
            for sheet in sheets_created:
                file_details.append(f"  - {sheet}")

        msg.setDetailedText("\n".join(file_details))
        msg.setInformativeText("Analysis results saved successfully")
        msg.exec_()

    def add_nematostella_analysis_to_widget(widget_instance):
        """
        Add Nematostella-specific analysis capabilities to the existing widget.
        This function can be called from the widget to enable enhanced analysis.

        Args:
            widget_instance: Instance of HDF5AnalysisWidget
        """
        # Add new button to widget if it doesn't exist
        if not hasattr(widget_instance, "btn_nematostella_analysis"):
            from qtpy.QtWidgets import QPushButton

            widget_instance.btn_nematostella_analysis = QPushButton(
                "Nematostella Timeseries Analysis"
            )
            widget_instance.btn_nematostella_analysis.setToolTip(
                "Run specialized Nematostella timeseries analysis"
            )
            widget_instance.btn_nematostella_analysis.setStyleSheet(
                "QPushButton { background-color: #9C27B0; color: white; font-weight: bold; }"
            )

            # Add to the existing results tab layout
            if hasattr(widget_instance, "tab_results"):
                layout = widget_instance.tab_results.layout()
                if layout:
                    # Find the plot buttons group and add after it
                    for i in range(layout.count()):
                        item = layout.itemAt(i)
                        if (
                            item
                            and hasattr(item.widget(), "title")
                            and "Controls" in item.widget().title()
                        ):
                            # Insert after plot controls group
                            layout.insertWidget(
                                i + 1, widget_instance.btn_nematostella_analysis
                            )
                            break
                    else:
                        # Fallback: add at the end
                        layout.addWidget(widget_instance.btn_nematostella_analysis)

            # Connect the button
            widget_instance.btn_nematostella_analysis.clicked.connect(
                lambda: run_nematostella_analysis_from_widget(widget_instance)
            )

    def run_nematostella_analysis_from_widget(widget_instance):
        """
        Run Nematostella analysis from within the napari widget.

        Args:
            widget_instance: Instance of HDF5AnalysisWidget
        """
        if not hasattr(widget_instance, "file_path") or not widget_instance.file_path:
            widget_instance._log_message(
                "No HDF5 file loaded for Nematostella analysis"
            )
            widget_instance.results_label.setText("Error: No HDF5 file loaded")
            return

        try:
            widget_instance._log_message("Starting Nematostella timeseries analysis...")
            widget_instance.results_label.setText("Running Nematostella analysis...")

            # Get quick summary first
            summary = get_nematostella_timeseries_summary(widget_instance.file_path)
            widget_instance._log_message("Timeseries summary:")
            for line in summary.split("\n"):
                if line.strip():
                    widget_instance._log_message(f"  {line}")

            # Run full analysis
            results = analyze_nematostella_hdf5_file(widget_instance.file_path)

            if results["success"]:
                widget_instance._log_message(
                    "Nematostella analysis completed successfully!"
                )
                widget_instance._log_message(
                    f"Excel file created: {results['excel_file']}"
                )
                widget_instance._log_message(
                    f"Report file created: {results['report_file']}"
                )
                widget_instance._log_message(
                    f"Sheets created: {', '.join(results['sheets_created'])}"
                )

                # Update results display
                widget_instance.results_label.setText(
                    f"Nematostella analysis complete: {len(results['sheets_created'])} Excel sheets created"
                )

                # Log key findings from report
                widget_instance._log_message("Key Analysis Results:")
                report_lines = results["report"].split("\n")
                in_important_section = False
                for line in report_lines:
                    if any(
                        section in line
                        for section in [
                            "## Timing Analysis",
                            "## LED System Analysis",
                            "## Environmental Conditions",
                        ]
                    ):
                        in_important_section = True
                        widget_instance._log_message(line)
                    elif line.startswith("##") and in_important_section:
                        in_important_section = False
                    elif in_important_section and line.strip().startswith("-"):
                        widget_instance._log_message(f"  {line.strip()}")

            else:
                widget_instance._log_message(
                    f"Nematostella analysis failed: {results['error']}"
                )
                widget_instance.results_label.setText(
                    f"Analysis failed: {results['error']}"
                )

        except Exception as e:
            error_msg = f"Nematostella analysis error: {e}"
            widget_instance._log_message(error_msg)
            widget_instance.results_label.setText(error_msg)

    def _show_save_success_dialog_with_metadata(
        self, saved_files, metadata_dict, nematostella_results=None
    ):
        """Show success dialog with metadata details and optional Nematostella analysis."""
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)

        # Adjust title and message based on whether Nematostella analysis was performed
        if nematostella_results and nematostella_results["success"]:
            msg.setWindowTitle("Save Complete: Metadata + Nematostella Analysis")
            msg.setText(
                f"Successfully saved metadata with specialized Nematostella timeseries analysis in {len(saved_files)} format(s):"
            )
        else:
            msg.setWindowTitle("Save with Metadata Complete")
            msg.setText(
                f"Successfully saved analysis results with metadata in {len(saved_files)} format(s):"
            )

        # Count metadata statistics
        total_static_params = 0
        total_timeseries_params = 0
        total_timeseries_points = 0

        for source_name, metadata in metadata_dict.items():
            # Skip Nematostella analysis entry for counting (it's not traditional metadata)
            if source_name == "nematostella_analysis":
                continue

            # Count static parameters
            static_data = {k: v for k, v in metadata.items() if k != "timeseries_data"}
            total_static_params += len(static_data)

            # Count time-series parameters
            if "timeseries_data" in metadata and metadata["timeseries_data"]:
                ts_data = metadata["timeseries_data"]
                total_timeseries_params += len(ts_data)
                for param_data in ts_data.values():
                    if hasattr(param_data, "__len__"):
                        total_timeseries_points = max(
                            total_timeseries_points, len(param_data)
                        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            file_size = "Unknown size"
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:  # > 1MB
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:  # > 1KB
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                pass

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        # Add Nematostella analysis files if available
        if nematostella_results and nematostella_results["success"]:
            file_details.append("")
            file_details.append("Nematostella Analysis Files:")

            # Add Excel analysis file
            excel_file = nematostella_results["excel_file"]
            try:
                excel_size_bytes = os.path.getsize(excel_file)
                if excel_size_bytes > 1024 * 1024:
                    excel_size = f"{excel_size_bytes/(1024*1024):.1f} MB"
                elif excel_size_bytes > 1024:
                    excel_size = f"{excel_size_bytes/1024:.1f} KB"
                else:
                    excel_size = f"{excel_size_bytes} bytes"
            except:
                excel_size = "Unknown size"

            file_details.append(
                f"• Excel Analysis: {os.path.basename(excel_file)} ({excel_size})"
            )

            # Add text report file
            report_file = nematostella_results["report_file"]
            try:
                report_size_bytes = os.path.getsize(report_file)
                if report_size_bytes > 1024:
                    report_size = f"{report_size_bytes/1024:.1f} KB"
                else:
                    report_size = f"{report_size_bytes} bytes"
            except:
                report_size = "Unknown size"

            file_details.append(
                f"• Text Report: {os.path.basename(report_file)} ({report_size})"
            )
            file_details.append(
                f"• Analysis Sheets: {len(nematostella_results['sheets_created'])}"
            )

        # Add metadata summary
        file_details.append("")
        file_details.append("Metadata Summary:")
        file_details.append(f"• Static parameters: {total_static_params}")
        file_details.append(f"• Time-series parameters: {total_timeseries_params}")
        if total_timeseries_points > 0:
            file_details.append(
                f"• Time-series length: {total_timeseries_points} time points"
            )
            duration_min = (
                total_timeseries_points * self.frame_interval.value()
            ) / 60.0
            file_details.append(f"• Total duration: {duration_min:.1f} minutes")

        # Add Nematostella analysis summary if available
        if nematostella_results and nematostella_results["success"]:
            file_details.append("")
            file_details.append("Nematostella Analysis Summary:")

            # Extract key metrics from the analysis results
            if (
                "analysis_results" in nematostella_results
                and nematostella_results["analysis_results"]
            ):
                analysis_results = nematostella_results["analysis_results"]

                # Timing analysis summary
                if "timing_analysis" in analysis_results:
                    timing = analysis_results["timing_analysis"]
                    if "timing" in timing:
                        accuracy = timing["timing"]["timing_accuracy"]
                        file_details.append(f"• Timing Accuracy: {accuracy:.1%}")

                # Environmental stability
                if (
                    "env_analysis" in analysis_results
                    and analysis_results["env_analysis"]
                ):
                    env = analysis_results["env_analysis"]["environment"]
                    if "temperature" in env:
                        temp_range = env["temperature"]["range"]
                        file_details.append(
                            f"• Temperature Stability: ±{temp_range/2:.1f}°C"
                        )

                # LED system performance
                if (
                    "led_analysis" in analysis_results
                    and analysis_results["led_analysis"]
                ):
                    led = analysis_results["led_analysis"]
                    if "led_sync" in led:
                        sync_rate = led["led_sync"]["success_rate"]
                        file_details.append(f"• LED Sync Success: {sync_rate:.1%}")

        msg.setDetailedText("\n".join(file_details))

        # Adjust informative text based on analysis type
        if nematostella_results and nematostella_results["success"]:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata AND specialized Nematostella timeseries analysis with timing, environmental, and LED system evaluation."
            )
        else:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata in time-series format matching analysis data structure."
            )

        msg.exec_()
        # Adjust informative text based on analysis type
        if nematostella_results and nematostella_results["success"]:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata AND specialized Nematostella timeseries analysis with timing, environmental, and LED system evaluation."
            )
        else:
            msg.setInformativeText(
                "Files include comprehensive HDF5 metadata in time-series format matching analysis data structure."
            )

        msg.exec_()

    def _show_threshold_stats_in_log(self):
        """
        HELPER METHOD: Show threshold statistics in the log.
        This replaces the separate "Show Threshold Stats" button functionality.
        """
        if not hasattr(self, "roi_baseline_means") or not self.roi_baseline_means:
            self._log_message("No threshold statistics available")
            return

        # Generate threshold statistics for log
        method = self._get_current_threshold_method_display()
        self._log_message(f"Method: {method}")

        if hasattr(self, "baseline_duration_minutes"):
            self._log_message(
                f"Baseline Duration: {self.baseline_duration_minutes.value():.1f} minutes"
            )
        if hasattr(self, "threshold_multiplier"):
            self._log_message(
                f"Hysteresis Multiplier: {self.threshold_multiplier.value():.2f}"
            )
        elif hasattr(self, "calibration_multiplier"):
            self._log_message(
                f"Calibration Multiplier: {self.calibration_multiplier.value():.2f}"
            )

        self._log_message(
            f"Detrending: {'Enabled' if getattr(self, 'enable_detrending', False) and self.enable_detrending.isChecked() else 'Disabled'}"
        )

        roi_band_widths = getattr(self, "roi_band_widths", {})
        roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
        roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})

        # Show statistics for first 5 ROIs to avoid log spam
        rois_to_show = sorted(self.roi_baseline_means.keys())[:5]

        for roi in rois_to_show:
            baseline_mean = self.roi_baseline_means[roi]
            band_width = roi_band_widths.get(roi, 0)
            upper_threshold = roi_upper_thresholds.get(roi, baseline_mean + band_width)
            lower_threshold = roi_lower_thresholds.get(roi, baseline_mean - band_width)

            self._log_message(f"\nROI {roi} HYSTERESIS SYSTEM:")
            self._log_message(f"  Baseline Mean: {baseline_mean:.3f}")
            self._log_message(f"  Band Width: ±{band_width:.3f}")
            self._log_message(
                f"  Upper Threshold: {upper_threshold:.3f} (Movement = TRUE when above)"
            )
            self._log_message(
                f"  Lower Threshold: {lower_threshold:.3f} (Movement = FALSE when below)"
            )
            self._log_message(
                f"  Hysteresis Zone: {lower_threshold:.3f} to {upper_threshold:.3f} (State unchanged)"
            )

        if len(self.roi_baseline_means) > 5:
            self._log_message(f"\n... and {len(self.roi_baseline_means) - 5} more ROIs")

    def _show_save_success_dialog(self, saved_files):
        """
        HELPER METHOD: Show success dialog with list of saved files.
        """
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)
        msg.setWindowTitle("Save Results Complete")
        msg.setText(
            f"Successfully saved analysis results in {len(saved_files)} format(s):"
        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            file_size = "Unknown size"
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:  # > 1MB
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:  # > 1KB
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                pass

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        msg.setDetailedText("\n".join(file_details))
        msg.setInformativeText(
            "Files can be opened in Excel, analyzed further, or imported into other analysis software."
        )
        msg.exec_()

    def _save_results_excel_to_path(self, excel_path: str):
        """
        COMPLETE METHOD: Save Excel results with ALL sheets to a specific file path.
        Creates the same multi-sheet structure as shown in the screenshot.
        """
        try:
            import pandas as pd

            # Create Excel writer
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

                # === SHEET 1: SUMMARY ===
                summary_data = []
                sorted_rois = sorted(self.merged_results.keys())

                for roi in sorted_rois:
                    row_data = {
                        "ROI": roi,
                        "Baseline Mean": getattr(self, "roi_baseline_means", {}).get(
                            roi, 0
                        ),
                        "Upper Threshold": getattr(
                            self, "roi_upper_thresholds", {}
                        ).get(roi, 0),
                        "Lower Threshold": getattr(
                            self, "roi_lower_thresholds", {}
                        ).get(roi, 0),
                        "Threshold Band Width": getattr(
                            self, "roi_band_widths", {}
                        ).get(roi, 0),
                    }

                    # Calculate movement statistics
                    movement_data = getattr(self, "movement_data", {})
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        row_data["Total Movement Events"] = sum(movement_values)
                        row_data["Movement Percentage"] = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                    # Calculate sleep statistics
                    sleep_data = getattr(self, "sleep_data", {})
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        row_data["Total Sleep Bins"] = total_sleep_bins
                        row_data["Sleep Time (min)"] = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                    summary_data.append(row_data)

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=1)
                writer.sheets["Summary"].cell(row=1, column=1).value = (
                    "Summary of ROI statistics: movement events, sleep bins, and detection "
                    "thresholds per ROI. One row per ROI."
                )

                # === SHEET 2: RAW INTENSITY ===
                if hasattr(self, "merged_results") and self.merged_results:
                    intensity_df = self._create_time_series_dataframe(
                        self.merged_results,
                        sorted_rois,
                        "Intensity",
                        convert_to_minutes=True,
                    )
                    intensity_df.to_excel(
                        writer, sheet_name="Raw_Intensity", index=False, startrow=1
                    )
                    writer.sheets["Raw_Intensity"].cell(row=1, column=1).value = (
                        "MinMax-normalized [0–1] raw intensity per ROI (per-pixel mean, "
                        "after MinMax scaling). See 'Real_Amplitude_MATLAB' sheet for "
                        "MATLAB-equivalent pixel sum values. Time column in minutes."
                    )

                # === SHEET 3: MOVEMENT ===
                if hasattr(self, "movement_data") and self.movement_data:
                    movement_df = self._create_time_series_dataframe(
                        self.movement_data,
                        sorted_rois,
                        "Movement",
                        convert_to_minutes=True,
                    )
                    movement_df.to_excel(writer, sheet_name="Movement", index=False, startrow=1)
                    writer.sheets["Movement"].cell(row=1, column=1).value = (
                        "Binary movement detection per ROI (1=moving, 0=stationary). "
                        "Based on hysteresis threshold crossing of raw intensity signal. "
                        "Time column in minutes."
                    )

                # === SHEET 4: FRACTION MOVEMENT ===
                if hasattr(self, "fraction_data") and self.fraction_data:
                    fraction_df = self._create_time_series_dataframe(
                        self.fraction_data,
                        sorted_rois,
                        "Fraction",
                        convert_to_minutes=True,
                    )
                    fraction_df.to_excel(
                        writer, sheet_name="Fraction_Movement", index=False, startrow=1
                    )
                    writer.sheets["Fraction_Movement"].cell(row=1, column=1).value = (
                        "Fraction of time bins with movement per ROI (0–1 scale). "
                        "Derived from binary movement data, binned over the analysis bin size. "
                        "Time column in minutes."
                    )

                # === SHEET 5: SLEEP ===
                if hasattr(self, "sleep_data") and self.sleep_data:
                    sleep_df = self._create_time_series_dataframe(
                        self.sleep_data, sorted_rois, "Sleep", convert_to_minutes=True
                    )
                    sleep_df.to_excel(writer, sheet_name="Sleep", index=False, startrow=1)
                    writer.sheets["Sleep"].cell(row=1, column=1).value = (
                        "Binary sleep state per ROI (1=sleeping, 0=not sleeping). "
                        "Sleep = continuous quiescence exceeding the minimum sleep duration threshold. "
                        "Time column in minutes."
                    )

                # === SHEET 6: QUIESCENCE ===
                if hasattr(self, "quiescence_data") and self.quiescence_data:
                    quiescence_df = self._create_time_series_dataframe(
                        self.quiescence_data,
                        sorted_rois,
                        "Quiescence",
                        convert_to_minutes=True,
                    )
                    quiescence_df.to_excel(
                        writer, sheet_name="Quiescence", index=False, startrow=1
                    )
                    writer.sheets["Quiescence"].cell(row=1, column=1).value = (
                        "Binary quiescence state per ROI (1=quiescent, 0=active). "
                        "Quiescence = fraction movement below the quiescence threshold. "
                        "Time column in minutes."
                    )

                # === SHEET 7: LIGHTING CONDITIONS ===
                if hasattr(self, "fraction_data") and self.fraction_data:
                    # Create lighting conditions data (binned activity for circadian analysis)
                    try:
                        # Use 30-minute bins for lighting analysis
                        from ._calc import bin_activity_data_for_lighting

                        lighting_data = bin_activity_data_for_lighting(
                            self.fraction_data, bin_minutes=30
                        )

                        if lighting_data:
                            lighting_df = self._create_time_series_dataframe(
                                lighting_data,
                                sorted_rois,
                                "Activity_30min_bins",
                                convert_to_minutes=True,
                            )
                            lighting_df.to_excel(
                                writer, sheet_name="Lighting_Conditions", index=False, startrow=1
                            )
                            writer.sheets["Lighting_Conditions"].cell(row=1, column=1).value = (
                                "Fraction movement binned in 30-minute intervals per ROI. "
                                "Suitable for circadian rhythm and lighting condition analysis. "
                                "Time column in minutes."
                            )
                    except Exception as e:
                        self._log_message(
                            f"Warning: Could not create lighting conditions sheet: {e}"
                        )

                # === SHEET 8: REAL AMPLITUDE (MATLAB-EQUIVALENT) ===
                if hasattr(self, "merged_results_raw") and self.merged_results_raw:
                    _norm_factor = getattr(self, "frame_norm_factor", 1.0)
                    _pixel_counts = getattr(self, "roi_pixel_counts", {})
                    _dtype_label = "uint8" if _norm_factor == 255.0 else "uint16"
                    matlab_data = {}
                    for roi, data in self.merged_results_raw.items():
                        _scale = _pixel_counts.get(roi, 1) * _norm_factor
                        matlab_data[roi] = [(t, v * _scale) for t, v in data]
                    matlab_df = self._create_time_series_dataframe(
                        matlab_data, sorted_rois, "Amplitude_raw", convert_to_minutes=True
                    )
                    matlab_df.to_excel(
                        writer, sheet_name="Real_Amplitude_MATLAB", index=False, startrow=1
                    )
                    writer.sheets["Real_Amplitude_MATLAB"].cell(row=1, column=1).value = (
                        f"MATLAB-equivalent pixel sum amplitude: per-pixel-mean x n_pixels x "
                        f"norm_factor (norm_factor={_norm_factor:.0f}, {_dtype_label}). "
                        "Units: sum of absolute pixel differences per frame per ROI "
                        "(equivalent to MATLAB diffs.sum()). Time column in minutes."
                    )

                # === SHEET 9: PARAMETERS ===
                # Determine source type (HDF5 or AVI)
                is_avi = hasattr(self, "avi_batch_paths") and self.avi_batch_paths
                source_label = "AVI Batch" if is_avi else "HDF5"

                params_data = {
                    "Parameter": [
                        "Data Source Type",
                        "Analysis Method",
                        "Frame Interval (s)",
                        "Baseline Duration (min)",
                        "Threshold Multiplier",
                        "Detrending Enabled",
                        "Jump Correction Enabled",
                        "Bin Size (s)",
                        "Quiescence Threshold",
                        "Sleep Threshold (min)",
                        "Number of ROIs",
                        "Total Analysis Time (min)",
                        "Generated Date",
                        "Software Version",
                    ],
                    "Value": [
                        source_label,
                        self._get_current_threshold_method_display(),
                        self.frame_interval.value(),
                        (
                            self.baseline_duration_minutes.value()
                            if hasattr(self, "baseline_duration_minutes")
                            else "N/A"
                        ),
                        (
                            self.threshold_multiplier.value()
                            if hasattr(self, "threshold_multiplier")
                            else "N/A"
                        ),
                        (
                            self.enable_detrending.isChecked()
                            if hasattr(self, "enable_detrending")
                            else "N/A"
                        ),
                        (
                            self.enable_jump_correction.isChecked()
                            if hasattr(self, "enable_jump_correction")
                            else "N/A"
                        ),
                        self.bin_size_seconds.value(),
                        self.quiescence_threshold.value(),
                        self.sleep_threshold_minutes.value(),
                        len(sorted_rois),
                        (
                            self.plot_end_time.value()
                            if hasattr(self, "plot_end_time")
                            else "N/A"
                        ),
                        pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "HDF5 Analysis Widget v1.0",
                    ],
                    "Description": [
                        "Type of data source (HDF5 file or AVI batch)",
                        "Threshold calculation method used",
                        "Time interval between frames",
                        "Duration of baseline period for threshold calculation",
                        "Multiplier for hysteresis band width",
                        "Whether detrending was applied to remove drift",
                        "Whether jump correction was applied",
                        "Time bin size for fraction movement calculation",
                        "Threshold below which animal is considered quiescent",
                        "Minimum continuous quiescence duration for sleep",
                        "Total number of ROIs analyzed",
                        "Total duration of analysis",
                        "When this file was generated",
                        "Software version and name",
                    ],
                }

                # Add AVI-specific parameters if applicable
                if is_avi and hasattr(self, "avi_batch_paths"):
                    params_data["Parameter"].extend(
                        [
                            "Number of AVI Files",
                            "AVI Start Time (s)",
                            "AVI End Time (s)",
                            "Total Duration (min)",
                        ]
                    )

                    # Calculate total duration from merged_results
                    total_duration_s = 0
                    start_time_s = 0
                    end_time_s = 0
                    if self.merged_results:
                        first_roi = next(iter(self.merged_results.values()))
                        if first_roi:
                            start_time_s = first_roi[0][0]
                            end_time_s = first_roi[-1][0]
                            total_duration_s = end_time_s - start_time_s

                    params_data["Value"].extend(
                        [
                            len(self.avi_batch_paths),
                            f"{start_time_s:.1f}",
                            f"{end_time_s:.1f}",
                            f"{total_duration_s / 60:.1f}",
                        ]
                    )

                    params_data["Description"].extend(
                        [
                            "Number of AVI video files processed",
                            "Start time of the analysis (in seconds)",
                            "End time of the analysis (in seconds)",
                            "Total duration of the video sequence (in minutes)",
                        ]
                    )

                # Add normalization / pixel-count parameters
                _p_norm = getattr(self, "frame_norm_factor", 1.0)
                _p_pixels = getattr(self, "roi_pixel_counts", {})
                _p_pixel_str = (
                    "; ".join(f"{roi}: {px}" for roi, px in sorted(_p_pixels.items()))
                    if _p_pixels else "N/A"
                )
                params_data["Parameter"].extend(
                    ["Frame Norm Factor", "ROI Pixel Counts"]
                )
                params_data["Value"].extend(
                    [f"{_p_norm:.0f}", _p_pixel_str]
                )
                params_data["Description"].extend(
                    [
                        "Pixel intensity normalization factor (255 for uint8 / 65535 for uint16). "
                        "Used to compute MATLAB-equivalent pixel sum in 'Real_Amplitude_MATLAB' sheet.",
                        "Number of pixels per ROI used for MATLAB-equivalent pixel sum scaling.",
                    ]
                )

                params_df = pd.DataFrame(params_data)
                params_df.to_excel(writer, sheet_name="Parameters", index=False, startrow=1)
                writer.sheets["Parameters"].cell(row=1, column=1).value = (
                    "Analysis parameters and settings used to generate this dataset. "
                    "Refer to these values to reproduce the analysis."
                )

        except Exception as e:
            raise Exception(f"Complete Excel save error: {e}")

    def _save_results_csv(self, file_path: str):
        """Save results in clear CSV format."""
        try:
            sorted_rois = sorted(self.merged_results.keys())

            # Create main data CSV
            with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)

                # === HEADER SECTION ===
                # Determine source type (HDF5 or AVI)
                is_avi = hasattr(self, "avi_batch_paths") and self.avi_batch_paths
                source_label = "AVI" if is_avi else "HDF5"
                writer.writerow([f"{source_label} Analysis Results"])
                writer.writerow(
                    [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                )
                writer.writerow(
                    [f"Analysis Method: {self._get_current_threshold_method_display()}"]
                )
                writer.writerow(
                    [f"Frame Interval: {self.frame_interval.value()} seconds"]
                )
                writer.writerow([f"Number of ROIs: {len(sorted_rois)}"])
                writer.writerow([])  # Empty row

                # === ROI SUMMARY TABLE ===
                writer.writerow(["ROI SUMMARY"])
                summary_headers = [
                    "ROI",
                    "Baseline Mean",
                    "Upper Threshold",
                    "Lower Threshold",
                    "Movement %",
                    "Sleep Time (min)",
                ]
                writer.writerow(summary_headers)

                roi_baseline_means = getattr(self, "roi_baseline_means", {})
                roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
                roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})
                movement_data = getattr(self, "movement_data", {})
                sleep_data = getattr(self, "sleep_data", {})

                for roi in sorted_rois:
                    # Calculate statistics
                    movement_pct = 0
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        movement_pct = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                    sleep_minutes = 0
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        sleep_minutes = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                    writer.writerow(
                        [
                            roi,
                            f"{roi_baseline_means.get(roi, 0):.3f}",
                            f"{roi_upper_thresholds.get(roi, 0):.3f}",
                            f"{roi_lower_thresholds.get(roi, 0):.3f}",
                            f"{movement_pct:.1f}",
                            f"{sleep_minutes:.1f}",
                        ]
                    )

                writer.writerow([])  # Empty row
                writer.writerow([])  # Empty row

                # === TIME SERIES DATA ===
                writer.writerow(["RAW INTENSITY DATA (Time in minutes)"])

                # Create time-aligned data
                all_times = set()
                for roi_data in self.merged_results.values():
                    for time_point, _ in roi_data:
                        all_times.add(round(time_point / 60.0, 2))  # Convert to minutes

                sorted_times = sorted(all_times)

                # Header row: Time, ROI1, ROI2, ROI3, ...
                header = ["Time (min)"] + [f"ROI_{roi}" for roi in sorted_rois]
                writer.writerow(header)

                # Create data rows
                for time_min in sorted_times:
                    row = [f"{time_min:.2f}"]

                    for roi in sorted_rois:
                        # Find value at this time point
                        value = None
                        if roi in self.merged_results:
                            for t, v in self.merged_results[roi]:
                                if abs(t / 60.0 - time_min) < 0.01:  # Within tolerance
                                    value = v
                                    break

                        row.append(f"{value:.6f}" if value is not None else "")

                    writer.writerow(row)

            self._log_message(f"CSV saved: {file_path}")

        except Exception as e:
            self._log_message(f"Error in CSV export: {e}")
            raise

    def _create_time_series_dataframe(
        self,
        data_dict: Dict,
        sorted_rois: List[int],
        data_type: str,
        convert_to_minutes: bool = True,
    ):
        """
        Create a pandas DataFrame with clear time-series structure for Excel export.
        This method already exists in your code but here's the complete version to ensure compatibility.
        """

        # Collect all unique time points
        all_times = set()
        for roi_data in data_dict.values():
            for time_point, _ in roi_data:
                if convert_to_minutes:
                    all_times.add(round(time_point / 60.0, 2))  # Round to 2 decimals
                else:
                    all_times.add(round(time_point, 2))

        sorted_times = sorted(all_times)

        # Create DataFrame structure
        df_data = {"Time (min)" if convert_to_minutes else "Time (s)": sorted_times}

        # Add data for each ROI
        for roi in sorted_rois:
            roi_values = []

            # Create time->value mapping for this ROI
            time_value_map = {}
            if roi in data_dict:
                for time_point, value in data_dict[roi]:
                    if convert_to_minutes:
                        time_key = round(time_point / 60.0, 2)
                    else:
                        time_key = round(time_point, 2)
                    time_value_map[time_key] = value

            # Fill values for all time points
            for time_point in sorted_times:
                if time_point in time_value_map:
                    roi_values.append(time_value_map[time_point])
                else:
                    roi_values.append(None)  # Missing data as None

            # Column name format: ROI_1, ROI_2, etc.
            df_data[f"ROI_{roi}"] = roi_values

        return pd.DataFrame(df_data)

    def save_results(self):
        """Save analysis results - let user choose format."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.results_label.setText("No results to save.")
            return

        # Ask user for format
        file_path, file_type = QFileDialog.getSaveFileName(
            self,
            "Save Results",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
        )

        if not file_path:
            return

        try:
            if file_path.endswith(".xlsx") or "Excel" in file_type:
                # Ensure .xlsx extension
                if not file_path.endswith(".xlsx"):
                    file_path += ".xlsx"
                self._save_results_excel_to_path(file_path)
                self.results_label.setText(
                    f"Results saved to {os.path.basename(file_path)}"
                )
                self._log_message(f"Excel results saved: {file_path}")
            else:
                # Default to CSV
                if not file_path.endswith(".csv"):
                    file_path += ".csv"
                self._save_results_csv(file_path)
                self.results_label.setText(
                    f"Results saved to {os.path.basename(file_path)}"
                )
                self._log_message(f"CSV results saved: {file_path}")

        except Exception as e:
            self.results_label.setText(f"Error saving results: {str(e)}")
            self._log_message(f"ERROR saving results: {str(e)}")
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def show_threshold_statistics(self):
        """Show detailed hysteresis statistics."""
        if not hasattr(self, "roi_baseline_means") or not self.roi_baseline_means:
            self._log_message("No hysteresis statistics available")
            return

        # Create statistics summary
        stats_text = "HYSTERESIS DETECTION SYSTEM STATISTICS\n"
        stats_text += "=" * 50 + "\n\n"

        # Add method parameters
        stats_text += f"Method: {self._get_current_threshold_method_display()}\n"
        if hasattr(self, "baseline_duration_minutes"):
            stats_text += f"Baseline Duration: {self.baseline_duration_minutes.value():.1f} minutes\n"
        if hasattr(self, "threshold_multiplier"):
            stats_text += (
                f"Hysteresis Multiplier: {self.threshold_multiplier.value():.2f}\n"
            )
        elif hasattr(self, "calibration_multiplier"):
            stats_text += (
                f"Calibration Multiplier: {self.calibration_multiplier.value():.2f}\n"
            )
        stats_text += f"Detrending: {'Enabled' if self.enable_detrending.isChecked() else 'Disabled'}\n\n"

        roi_band_widths = getattr(self, "roi_band_widths", {})
        roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
        roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})
        roi_statistics = getattr(self, "roi_statistics", {})

        for roi in sorted(self.roi_baseline_means.keys()):
            baseline_mean = self.roi_baseline_means[roi]
            band_width = roi_band_widths.get(roi, 0)
            stats = roi_statistics.get(roi, {})

            upper_threshold = roi_upper_thresholds.get(roi, baseline_mean + band_width)
            lower_threshold = roi_lower_thresholds.get(roi, baseline_mean - band_width)

            stats_text += f"ROI {roi} - HYSTERESIS SYSTEM:\n"
            stats_text += f"  Baseline Mean: {baseline_mean:.3f}\n"
            stats_text += f"  Band Width: ±{band_width:.3f}\n"
            stats_text += f"  Upper Threshold: {upper_threshold:.3f} (Movement = TRUE when above)\n"
            stats_text += f"  Lower Threshold: {lower_threshold:.3f} (Movement = FALSE when below)\n"
            stats_text += f"  Hysteresis Zone: {lower_threshold:.3f} to {upper_threshold:.3f} (State unchanged)\n"

            if stats.get("was_detrended", False):
                stats_text += "  Detrending: Applied\n"

            # Add method-specific information
            method = stats.get("method", "unknown")
            if "calibration" in method:
                snr = stats.get("signal_to_noise_ratio", 0)
                quality = stats.get("calibration_quality", "unknown")
                stats_text += f"  Calibration Quality: {quality}\n"
                stats_text += f"  Signal-to-Noise Ratio: {snr:.2f}\n"

            stats_text += "\n"

        # Display in log
        self._log_message("DETAILED HYSTERESIS STATISTICS:")
        for line in stats_text.split("\n"):
            if line.strip():
                self._log_message(line)

    # def save_results_with_metadata(self):
    #     """
    #     Save HDF5 metadata with optional Nematostella timeseries analysis.
    #     Enhanced to automatically detect and analyze Nematostella experiments.
    #     """

    #     # Check if we have a file loaded
    #     if not hasattr(self, 'file_path') or not self.file_path:
    #         self.results_label.setText("No HDF5 file loaded. Load a file first.")
    #         self._log_message("Save failed: No HDF5 file loaded")
    #         return

    #     # Analysis results are optional for metadata extraction
    #     has_analysis_results = (hasattr(self, "merged_results") and self.merged_results)

    #     if has_analysis_results:
    #         self._log_message("Saving analysis results with HDF5 metadata...")
    #     else:
    #         self._log_message("Saving HDF5 metadata only (no analysis results available)...")

    #     # NEW: Check for Nematostella timeseries data
    #     nematostella_results = None
    #     # Direkte Prüfung statt globaler Variable
    #     try:
    #         from ._metadata import analyze_nematostella_hdf5_file
    #         nematostella_available = True
    #     except ImportError:
    #         nematostella_available = False

    #     if nematostella_available:
    #         try:
    #             self._log_message("Checking for Nematostella timeseries data...")

    #             # Quick check if this is a Nematostella experiment
    #             with h5py.File(self.file_path, 'r') as h5_file:
    #                 if 'timeseries' in h5_file:
    #                     ts_group = h5_file['timeseries']
    #                     # Check for typical Nematostella parameters
    #                     nematostella_indicators = [
    #                         'actual_intervals', 'expected_intervals', 'frame_drift',
    #                         'temperature', 'humidity', 'led_power_percent'
    #                     ]

    #                     found_indicators = [key for key in ts_group.keys() if key in nematostella_indicators]

    #                     if len(found_indicators) >= 2:  # At least 2 indicators found
    #                         self._log_message(f"Nematostella experiment detected! Found: {', '.join(found_indicators)}")
    #                         self._log_message("Running specialized Nematostella timeseries analysis...")

    #                         # Run Nematostella analysis
    #                         nematostella_results = analyze_nematostella_hdf5_file(self.file_path)

    #                         if nematostella_results['success']:
    #                             self._log_message(f"Nematostella analysis completed: {len(nematostella_results['sheets_created'])} sheets")
    #                         else:
    #                             self._log_message(f"Nematostella analysis failed: {nematostella_results['error']}")
    #                     else:
    #                         self._log_message("No Nematostella-specific timeseries detected")
    #         except Exception as e:
    #             self._log_message(f"Nematostella detection failed: {e}")

    #     # Get base filename from user
    #     from qtpy.QtWidgets import QFileDialog

    #     if nematostella_results and nematostella_results['success']:
    #         dialog_title = "Save HDF5 Metadata with Nematostella Analysis"
    #         default_name = f"nematostella_metadata_{int(time.time())}"
    #     else:
    #         dialog_title = "Save HDF5 Metadata" + (" with Analysis Results" if has_analysis_results else "")
    #         default_name = f"hdf5_metadata_{int(time.time())}"

    #     base_path, _ = QFileDialog.getSaveFileName(
    #         self, dialog_title, default_name, "All Files (*)"
    #     )

    #     if not base_path:
    #         self._log_message("Save cancelled by user")
    #         return

    #     base_path = os.path.splitext(base_path)[0]
    #     saved_files = []

    #     try:
    #         # Extract regular metadata
    #         self._log_message("Extracting HDF5 metadata with time-series support...")
    #         metadata_dict = {}

    #         # Extract from main file with time-series capability
    #         if hasattr(self, 'file_path') and self.file_path:
    #             self._log_message(f"   Extracting from main file: {os.path.basename(self.file_path)}")
    #             try:
    #                 main_metadata = extract_hdf5_metadata_timeseries(self.file_path)
    #                 metadata_dict['main_file'] = main_metadata

    #                 if 'timeseries_data' in main_metadata and main_metadata['timeseries_data']:
    #                     ts_data = main_metadata['timeseries_data']
    #                     self._log_message(f"     Found {len(ts_data)} time-series parameters")

    #             except Exception as e:
    #                 self._log_message(f"     Main file metadata extraction failed: {e}")
    #                 metadata_dict['main_file'] = {'error': str(e), 'timeseries_data': {}}

    #         # Add analysis metadata (only if we have analysis results)
    #         if has_analysis_results:
    #             metadata_dict['analysis_info'] = {
    #                 'analysis_method': self._get_current_threshold_method_display(),
    #                 'analysis_timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
    #                 'frame_interval': self.frame_interval.value(),
    #                 'rois_analyzed': len(self.merged_results),
    #                 'software_version': 'HDF5 Activity Analysis Widget v1.0',
    #                 'parameters': self._get_analysis_parameters_for_metadata(),
    #                 'timeseries_data': {}
    #             }
    #         else:
    #             metadata_dict['file_info_only'] = {
    #                 'extraction_type': 'HDF5 metadata only',
    #                 'extraction_timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
    #                 'source_file': os.path.basename(self.file_path),
    #                 'software_version': 'HDF5 Activity Analysis Widget v1.0',
    #                 'timeseries_data': {}
    #             }

    #         # NEW: Add Nematostella analysis results if available
    #         if nematostella_results and nematostella_results['success']:
    #             metadata_dict['nematostella_analysis'] = {
    #                 'analysis_type': 'Nematostella Timeseries Analysis',
    #                 'excel_file': nematostella_results['excel_file'],
    #                 'report_file': nematostella_results['report_file'],
    #                 'sheets_created': nematostella_results['sheets_created'],
    #                 'analysis_timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
    #                 'timeseries_data': {}
    #             }

    #         self._log_message("Metadata extraction completed")

    #         # Save CSV with metadata
    #         csv_path = f"{base_path}_metadata.csv"
    #         self._log_message(f"Saving CSV with metadata: {os.path.basename(csv_path)}")

    #         try:
    #             self._save_results_csv_with_metadata(csv_path, metadata_dict, has_analysis_results)
    #             saved_files.append(("CSV with Metadata", csv_path))
    #             self._log_message("CSV with metadata saved successfully")
    #         except Exception as e:
    #             self._log_message(f"CSV save failed: {e}")

    #         # Save Excel with metadata (if pandas available)
    #         try:
    #             import pandas as pd
    #             excel_path = f"{base_path}_metadata.xlsx"
    #             self._log_message(f"Saving Excel with metadata: {os.path.basename(excel_path)}")

    #             self._save_results_excel_with_metadata(excel_path, metadata_dict, has_analysis_results)
    #             saved_files.append(("Excel with Metadata", excel_path))
    #             self._log_message("Excel with metadata saved successfully")

    #         except ImportError:
    #             self._log_message("Excel export not available (missing pandas/openpyxl)")
    #         except Exception as e:
    #             self._log_message(f"Excel save failed: {e}")

    #         # Update UI
    #         if saved_files:
    #             file_list = ", ".join([f"{fmt} ({os.path.basename(path)})" for fmt, path in saved_files])

    #             if nematostella_results and nematostella_results['success']:
    #                 result_msg = f"Saved metadata + Nematostella analysis: {file_list}"
    #                 result_msg += f" + {nematostella_results['excel_file']}"
    #             else:
    #                 result_msg = f"Saved metadata: {file_list}"

    #             self.results_label.setText(result_msg)
    #             self._log_message(f"Save with metadata complete: {len(saved_files)} files created")

    #             # Log Nematostella results if available
    #             if nematostella_results and nematostella_results['success']:
    #                 self._log_message("Nematostella Analysis Summary:")
    #                 report_lines = nematostella_results['report'].split('\n')
    #                 for line in report_lines:
    #                     if any(section in line for section in ['## Timing Analysis', '## Environmental Conditions', '## LED System']):
    #                         self._log_message(line)
    #                     elif line.strip().startswith('-') and any(keyword in line for keyword in ['Mean', 'Accuracy', 'Success Rate']):
    #                         self._log_message(f"  {line.strip()}")

    #             # Check if method supports nematostella_results parameter
    #             try:
    #                 self._show_save_success_dialog_with_metadata(saved_files, metadata_dict, nematostella_results)
    #             except TypeError:
    #                 # Fallback to old method signature
    #                 self._show_save_success_dialog_with_metadata(saved_files, metadata_dict)
    #         else:
    #             self.results_label.setText("All save attempts failed - check log")

    #     except Exception as e:
    #         error_msg = f"Save with metadata failed: {e}"
    #         self.results_label.setText(error_msg)
    #         self._log_message(error_msg)
    #         import traceback
    #         self._log_message(f"Traceback: {traceback.format_exc()}")
    def save_results_with_metadata(self):
        """
        Save HDF5 metadata with automatic legacy enhancement and optional Nematostella analysis.
        Enhanced to automatically detect legacy files and add unit documentation.
        """

        # Check if we have a file loaded
        if not hasattr(self, "file_path") or not self.file_path:
            self.results_label.setText("No HDF5 file loaded. Load a file first.")
            self._log_message("Save failed: No HDF5 file loaded")
            return

        # Analysis results are optional for metadata extraction
        has_analysis_results = hasattr(self, "merged_results") and self.merged_results

        if has_analysis_results:
            self._log_message("Saving analysis results with HDF5 metadata...")
        else:
            self._log_message(
                "Saving HDF5 metadata only (no analysis results available)..."
            )

        # NEW: Check for Nematostella timeseries data
        nematostella_results = None
        # Direkte Prüfung statt globaler Variable
        try:
            from ._metadata import analyze_nematostella_hdf5_file

            nematostella_available = True
        except ImportError:
            nematostella_available = False

        if nematostella_available:
            try:
                self._log_message("Checking for Nematostella timeseries data...")

                # Quick check if this is a Nematostella experiment
                with h5py.File(self.file_path, "r") as h5_file:
                    if "timeseries" in h5_file:
                        ts_group = h5_file["timeseries"]
                        # Check for typical Nematostella parameters
                        nematostella_indicators = [
                            "actual_intervals",
                            "expected_intervals",
                            "frame_drift",
                            "temperature",
                            "humidity",
                            "led_power_percent",
                        ]

                        found_indicators = [
                            key
                            for key in ts_group.keys()
                            if key in nematostella_indicators
                        ]

                        if len(found_indicators) >= 2:  # At least 2 indicators found
                            self._log_message(
                                f"Nematostella experiment detected! Found: {', '.join(found_indicators)}"
                            )
                            self._log_message(
                                "Running specialized Nematostella timeseries analysis..."
                            )

                            # Run Nematostella analysis
                            nematostella_results = analyze_nematostella_hdf5_file(
                                self.file_path
                            )

                            if nematostella_results["success"]:
                                self._log_message(
                                    f"Nematostella analysis completed: {len(nematostella_results['sheets_created'])} sheets"
                                )
                            else:
                                self._log_message(
                                    f"Nematostella analysis failed: {nematostella_results['error']}"
                                )
                        else:
                            self._log_message(
                                "No Nematostella-specific timeseries detected"
                            )
            except Exception as e:
                self._log_message(f"Nematostella detection failed: {e}")

        # Get base filename from user
        from qtpy.QtWidgets import QFileDialog

        if nematostella_results and nematostella_results["success"]:
            dialog_title = "Save HDF5 Metadata with Nematostella Analysis"
            default_name = f"nematostella_metadata_{int(time.time())}"
        else:
            dialog_title = "Save HDF5 Metadata" + (
                " with Analysis Results" if has_analysis_results else ""
            )
            default_name = f"hdf5_metadata_{int(time.time())}"

        base_path, _ = QFileDialog.getSaveFileName(
            self, dialog_title, default_name, "All Files (*)"
        )

        if not base_path:
            self._log_message("Save cancelled by user")
            return

        base_path = os.path.splitext(base_path)[0]
        saved_files = []

        try:
            # === AUTOMATIC LEGACY ENHANCEMENT INTEGRATION ===
            self._log_message(
                "Extracting HDF5 metadata with automatic legacy enhancement..."
            )
            metadata_dict = {}

            # Extract from main file with automatic legacy enhancement
            if hasattr(self, "file_path") and self.file_path:
                self._log_message(
                    f"   Extracting from main file: {os.path.basename(self.file_path)}"
                )
                try:
                    # This function now automatically enhances legacy files
                    main_metadata = extract_hdf5_metadata_timeseries(self.file_path)
                    metadata_dict["main_file"] = main_metadata

                    # Log automatic legacy enhancement results
                    if main_metadata.get("legacy_enhanced", False):
                        enhancement_info = main_metadata.get("_enhancement_summary", {})
                        enhanced_params = enhancement_info.get("parameters_enhanced", 0)
                        self._log_message("     ✅ Legacy file automatically enhanced!")
                        self._log_message(
                            f"     📏 Unit documentation added for {enhanced_params} parameters"
                        )
                        self._log_message(
                            f"     🕒 Enhancement timestamp: {main_metadata.get('enhancement_timestamp', 'unknown')}"
                        )
                        self._log_message(
                            "     📊 Unit standard: seconds for timing, celsius for temp, percent for humidity"
                        )
                    elif main_metadata.get("modern_file", False):
                        self._log_message(
                            "     ✅ Modern file with existing unit documentation detected"
                        )
                    else:
                        self._log_message("     ⚠️ File type could not be determined")

                    # Log timeseries data found
                    if (
                        "timeseries_data" in main_metadata
                        and main_metadata["timeseries_data"]
                    ):
                        ts_data = main_metadata["timeseries_data"]
                        # Count non-metadata parameters
                        param_count = len(
                            [k for k in ts_data.keys() if not k.startswith("_")]
                        )
                        self._log_message(
                            f"     📈 Found {param_count} time-series parameters"
                        )

                        # Log unit enhancement details if available
                        unit_info = ts_data.get("_unit_info", {})
                        if unit_info:
                            timing_params = [
                                k
                                for k, v in unit_info.items()
                                if v.get("units") == "seconds"
                            ]
                            environmental_params = [
                                k
                                for k, v in unit_info.items()
                                if v.get("units") in ["celsius", "percent"]
                            ]
                            if timing_params:
                                self._log_message(
                                    f"       ⏱️ Timing parameters: {', '.join(timing_params[:3])}{'...' if len(timing_params) > 3 else ''}"
                                )
                            if environmental_params:
                                self._log_message(
                                    f"       🌡️ Environmental parameters: {', '.join(environmental_params)}"
                                )

                except Exception as e:  # <-- JETZT KORREKT EINGERÜCKT
                    self._log_message(f"     Main file metadata extraction failed: {e}")
                    metadata_dict["main_file"] = {
                        "error": str(e),
                        "timeseries_data": {},
                    }

            # Add analysis metadata (only if we have analysis results)
            if has_analysis_results:
                metadata_dict["analysis_info"] = {
                    "analysis_method": self._get_current_threshold_method_display(),
                    "analysis_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "frame_interval": self.frame_interval.value(),
                    "rois_analyzed": len(self.merged_results),
                    "software_version": "HDF5 Activity Analysis Widget v1.0 (Legacy Enhanced)",
                    "parameters": self._get_analysis_parameters_for_metadata(),
                    "timeseries_data": {},
                    "legacy_compatibility": True,  # Mark as legacy-compatible
                }
            else:
                metadata_dict["file_info_only"] = {
                    "extraction_type": "HDF5 metadata only (Legacy Enhanced)",
                    "extraction_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "source_file": os.path.basename(self.file_path),
                    "software_version": "HDF5 Activity Analysis Widget v1.0 (Legacy Enhanced)",
                    "timeseries_data": {},
                    "legacy_compatibility": True,
                }

            # NEW: Add Nematostella analysis results if available
            if nematostella_results and nematostella_results["success"]:
                metadata_dict["nematostella_analysis"] = {
                    "analysis_type": "Nematostella Timeseries Analysis (Legacy Enhanced)",
                    "excel_file": nematostella_results["excel_file"],
                    "report_file": nematostella_results["report_file"],
                    "sheets_created": nematostella_results["sheets_created"],
                    "analysis_timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "timeseries_data": {},
                    "legacy_enhanced": metadata_dict["main_file"].get(
                        "legacy_enhanced", False
                    ),
                }

            self._log_message("Metadata extraction with legacy enhancement completed")

            # Save CSV with enhanced metadata
            csv_path = f"{base_path}_metadata.csv"
            self._log_message(
                f"Saving enhanced CSV with metadata: {os.path.basename(csv_path)}"
            )

            try:
                self._save_results_csv_with_metadata(
                    csv_path, metadata_dict, has_analysis_results
                )
                saved_files.append(("Enhanced CSV with Metadata", csv_path))
                self._log_message("Enhanced CSV with metadata saved successfully")
            except Exception as e:
                self._log_message(f"CSV save failed: {e}")

            # Save Excel with enhanced metadata (if pandas available)
            try:
                import pandas as pd

                excel_path = f"{base_path}_metadata.xlsx"
                self._log_message(
                    f"Saving enhanced Excel with metadata: {os.path.basename(excel_path)}"
                )

                if has_analysis_results:
                    # Step 1: Write all analysis sheets (Movement, Sleep, etc.)
                    self._save_results_excel_to_path(excel_path)
                    # Step 2: Append HDF5 sensor timeseries sheets to same file
                    self._append_hdf5_sheets_to_excel(excel_path, metadata_dict)
                else:
                    # No analysis data — write HDF5 metadata only
                    self._save_results_excel_with_metadata(
                        excel_path, metadata_dict, False
                    )
                saved_files.append(("Enhanced Excel with Metadata", excel_path))
                self._log_message("Enhanced Excel with metadata saved successfully")

            except ImportError:
                self._log_message(
                    "Excel export not available (missing pandas/openpyxl)"
                )
            except Exception as e:
                self._log_message(f"Excel save failed: {e}")

            # Update UI with legacy enhancement information
            if saved_files:
                file_list = ", ".join(
                    [f"{fmt} ({os.path.basename(path)})" for fmt, path in saved_files]
                )

                # Enhanced result message
                is_legacy = metadata_dict.get("main_file", {}).get(
                    "legacy_enhanced", False
                )
                legacy_suffix = " (Legacy Enhanced)" if is_legacy else ""

                if nematostella_results and nematostella_results["success"]:
                    result_msg = f"Saved metadata + Nematostella analysis{legacy_suffix}: {file_list}"
                    result_msg += (
                        f" + {os.path.basename(nematostella_results['excel_file'])}"
                    )
                else:
                    result_msg = f"Saved metadata{legacy_suffix}: {file_list}"

                self.results_label.setText(result_msg)
                self._log_message(
                    f"Save with metadata complete: {len(saved_files)} files created"
                )

                # Log enhancement summary
                if is_legacy:
                    enhancement_summary = metadata_dict["main_file"].get(
                        "_enhancement_summary", {}
                    )
                    enhanced_count = enhancement_summary.get("parameters_enhanced", 0)
                    self._log_message("📋 Legacy Enhancement Summary:")
                    self._log_message(f"   Parameters enhanced: {enhanced_count}")
                    self._log_message(
                        f"   Unit standard applied: {enhancement_summary.get('unit_standard', 'Unknown')}"
                    )
                    self._log_message(
                        "   Files include comprehensive unit documentation"
                    )

                # Log Nematostella results if available
                if nematostella_results and nematostella_results["success"]:
                    self._log_message("Nematostella Analysis Summary:")
                    report_lines = nematostella_results["report"].split("\n")
                    for line in report_lines:
                        if any(
                            section in line
                            for section in [
                                "## Timing Analysis",
                                "## Environmental Conditions",
                                "## LED System",
                            ]
                        ):
                            self._log_message(line)
                        elif line.strip().startswith("-") and any(
                            keyword in line
                            for keyword in ["Mean", "Accuracy", "Success Rate"]
                        ):
                            self._log_message(f"  {line.strip()}")

                # Show enhanced success dialog
                try:
                    self._show_save_success_dialog_with_metadata(
                        saved_files, metadata_dict, nematostella_results
                    )
                except TypeError:
                    # Fallback to old method signature
                    self._show_save_success_dialog_with_metadata(
                        saved_files, metadata_dict
                    )
            else:
                self.results_label.setText("All save attempts failed - check log")

        except Exception as e:
            error_msg = f"Save with metadata failed: {e}"
            self.results_label.setText(error_msg)
            self._log_message(error_msg)
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def _create_legacy_enhanced_sheets(
        self, writer, ts_data: dict, unit_info: dict, source_name: str
    ):
        """Create enhanced sheets for legacy data with automatic unit documentation."""

        # Erstelle DataFrame mit Unit-erweiterten Spalten-Namen
        enhanced_columns = {}

        for param_name, param_data in ts_data.items():
            if param_name.startswith("_"):
                continue  # Skip metadata

            unit = unit_info.get(param_name, {}).get("units", "unknown")

            if unit == "seconds" and "drift" in param_name.lower():
                # Für Timing-Daten: beide Einheiten
                enhanced_columns[f"{param_name}_sec"] = param_data
                enhanced_columns[f"{param_name}_ms"] = [
                    d * 1000 if d else 0 for d in param_data
                ]
            else:
                # Standard Parameter mit Unit-Suffix
                enhanced_columns[f"{param_name}_{unit}"] = param_data

        if enhanced_columns:
            # Frame index hinzufügen
            max_length = max(
                len(data)
                for data in enhanced_columns.values()
                if isinstance(data, (list, tuple))
            )
            enhanced_columns["frame_index"] = list(range(max_length))

            df = pd.DataFrame(enhanced_columns)
            sheet_name = f"Enhanced_{source_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _create_automatic_units_reference_sheet(self, writer, metadata_dict: dict):
        """Automatically create units reference sheet for legacy files."""

        units_found = set()

        # Sammle alle gefundenen Parameter und ihre Units
        for metadata in metadata_dict.values():
            if "timeseries_data" in metadata:
                unit_info = metadata["timeseries_data"].get("_unit_info", {})
                for param, info in unit_info.items():
                    units_found.add(
                        (param, info["units"], info.get("display_hint", ""))
                    )

        if units_found:
            units_data = []
            for param, unit, hint in sorted(units_found):
                units_data.append(
                    {
                        "Parameter": param,
                        "Units": unit,
                        "Display_Hint": hint,
                        "Enhancement": "Automatically added for legacy compatibility",
                    }
                )

            units_df = pd.DataFrame(units_data)
            units_df.to_excel(writer, sheet_name="Auto_Units_Reference", index=False)

    def _save_results_csv_with_metadata(
        self, file_path: str, metadata_dict: dict, has_analysis_results: bool = True
    ):
        """
        Save CSV with HDF5 metadata time-series (with optional analysis results).
        """
        import csv
        from datetime import datetime

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            # === HEADER SECTION ===
            if has_analysis_results:
                writer.writerow(["HDF5 Analysis Results with Time-Series Metadata"])
                sorted_rois = sorted(self.merged_results.keys())
                writer.writerow([f"Number of ROIs: {len(sorted_rois)}"])
            else:
                writer.writerow(["HDF5 Time-Series Metadata Only"])
                writer.writerow(
                    ["No analysis results available - metadata extraction only"]
                )

            writer.writerow(
                [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            )
            writer.writerow([f"Source file: {os.path.basename(self.file_path)}"])
            writer.writerow([])

            # === ANALYSIS RESULTS SUMMARY (only if available) ===
            if has_analysis_results:
                writer.writerow(["=== ANALYSIS RESULTS SUMMARY ==="])
                writer.writerow(
                    [
                        "ROI",
                        "Baseline Mean",
                        "Upper Threshold",
                        "Lower Threshold",
                        "Movement %",
                        "Sleep Time (min)",
                    ]
                )

                # Get analysis data
                roi_baseline_means = getattr(self, "roi_baseline_means", {})
                roi_upper_thresholds = getattr(self, "roi_upper_thresholds", {})
                roi_lower_thresholds = getattr(self, "roi_lower_thresholds", {})
                movement_data = getattr(self, "movement_data", {})
                sleep_data = getattr(self, "sleep_data", {})

                for roi in sorted_rois:
                    # Calculate statistics
                    movement_pct = 0
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        movement_pct = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                    sleep_minutes = 0
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        sleep_minutes = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                    writer.writerow(
                        [
                            roi,
                            f"{roi_baseline_means.get(roi, 0):.3f}",
                            f"{roi_upper_thresholds.get(roi, 0):.3f}",
                            f"{roi_lower_thresholds.get(roi, 0):.3f}",
                            f"{movement_pct:.1f}",
                            f"{sleep_minutes:.1f}",
                        ]
                    )

                writer.writerow([])
                writer.writerow([])

            # === STATIC HDF5 METADATA SECTIONS ===
            for source_name, metadata in metadata_dict.items():
                if source_name in ["analysis_info", "file_info_only"]:
                    continue  # Handle separately

                writer.writerow(
                    [f"=== {source_name.upper().replace('_', ' ')} STATIC METADATA ==="]
                )

                # Write static metadata (excluding timeseries_data)
                static_metadata = {
                    k: v for k, v in metadata.items() if k != "timeseries_data"
                }
                if static_metadata:
                    from ._metadata import write_metadata_to_csv

                    write_metadata_to_csv(writer, static_metadata, source_name.upper())
                else:
                    writer.writerow(["No static metadata available"])

                writer.writerow([])

            # === HDF5 TIME-SERIES METADATA SECTIONS ===
            has_hdf5_timeseries = False
            for source_name, metadata in metadata_dict.items():
                if "timeseries_data" in metadata and metadata["timeseries_data"]:
                    has_hdf5_timeseries = True
                    writer.writerow(
                        [
                            f"=== {source_name.upper().replace('_', ' ')} HDF5 TIME-SERIES METADATA ==="
                        ]
                    )

                    ts_data = metadata["timeseries_data"]

                    # Filter out analysis-related data - only keep actual HDF5 metadata
                    from ._metadata import filter_hdf5_metadata_only

                    hdf5_metadata_only = filter_hdf5_metadata_only(ts_data)

                    if hdf5_metadata_only:
                        max_length = max(
                            len(data) for data in hdf5_metadata_only.values()
                        )

                        # Log what we're including
                        param_names = list(hdf5_metadata_only.keys())
                        self._log_message(
                            f"   Including HDF5 time-series: {param_names}"
                        )

                        # Align time with analysis data (or use generic timing)
                        frame_interval = (
                            self.frame_interval.value() if has_analysis_results else 5.0
                        )

                        # Header: Time (min), parameters...
                        header = ["Time (min)"] + param_names
                        writer.writerow(header)

                        # Data rows
                        for i in range(max_length):
                            time_min = (i * frame_interval) / 60.0
                            row = [f"{time_min:.2f}"]

                            for param_name in param_names:
                                param_data = hdf5_metadata_only[param_name]
                                if i < len(param_data):
                                    value = param_data[i]
                                    if isinstance(value, (int, float)):
                                        row.append(f"{value:.6f}")
                                    else:
                                        row.append(str(value))
                                else:
                                    row.append("")  # Missing data

                            writer.writerow(row)

                        writer.writerow([])
                        writer.writerow(
                            [
                                f"HDF5 time-series metadata: {len(hdf5_metadata_only)} parameters, {max_length} time points"
                            ]
                        )

                        # List all parameters
                        writer.writerow(["Parameters included:"])
                        for param in param_names:
                            writer.writerow([f"  - {param}"])

                    else:
                        writer.writerow(["No HDF5 time-series metadata found"])

                    writer.writerow([])
                    writer.writerow([])

            if not has_hdf5_timeseries:
                writer.writerow(["=== NO HDF5 TIME-SERIES METADATA FOUND ==="])
                writer.writerow(
                    ["Your HDF5 file may not contain time-series metadata."]
                )
                writer.writerow([])

            # === ANALYSIS/FILE INFO PARAMETERS ===
            info_section = metadata_dict.get("analysis_info") or metadata_dict.get(
                "file_info_only"
            )
            if info_section:
                section_name = (
                    "ANALYSIS PARAMETERS"
                    if has_analysis_results
                    else "FILE INFORMATION"
                )
                writer.writerow([f"=== {section_name} ==="])

                # Write parameters
                writer.writerow(["Parameter", "Value", "Description"])

                param_descriptions = {
                    "analysis_method": "Threshold calculation method used",
                    "analysis_timestamp": "When analysis was performed",
                    "extraction_timestamp": "When metadata was extracted",
                    "frame_interval": "Time interval between frames (seconds)",
                    "rois_analyzed": "Total number of ROIs analyzed",
                    "extraction_type": "Type of extraction performed",
                    "source_file": "Source HDF5 file name",
                    "software_version": "Analysis software version",
                }

                for param, value in info_section.items():
                    if param != "parameters":
                        description = param_descriptions.get(param, "Parameter")
                        writer.writerow([param, str(value), description])

                # Write nested parameters (if analysis results available)
                if "parameters" in info_section:
                    writer.writerow([])
                    writer.writerow(["Detailed Parameters:"])
                    writer.writerow(["Parameter", "Value"])

                    for param, value in info_section["parameters"].items():
                        writer.writerow([param, str(value)])

            # === FOOTER ===
            writer.writerow([])
            writer.writerow(["=== END OF FILE ==="])
            writer.writerow(
                [f"Generation time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            )

    def _show_save_success_dialog_with_metadata(self, saved_files, metadata_dict):
        """Show success dialog with metadata details."""
        from qtpy.QtWidgets import QMessageBox

        msg = QMessageBox(self)
        msg.setWindowTitle("Save with Metadata Complete")
        msg.setText(
            f"Successfully saved analysis results with metadata in {len(saved_files)} format(s):"
        )

        # Count metadata statistics
        total_static_params = 0
        total_timeseries_params = 0
        total_timeseries_points = 0

        for source_name, metadata in metadata_dict.items():
            # Count static parameters
            static_data = {k: v for k, v in metadata.items() if k != "timeseries_data"}
            total_static_params += len(static_data)

            # Count time-series parameters
            if "timeseries_data" in metadata and metadata["timeseries_data"]:
                ts_data = metadata["timeseries_data"]
                total_timeseries_params += len(ts_data)
                for param_data in ts_data.values():
                    if hasattr(param_data, "__len__"):
                        total_timeseries_points = max(
                            total_timeseries_points, len(param_data)
                        )

        file_details = []
        for file_format, file_path in saved_files:
            filename = os.path.basename(file_path)
            file_size = "Unknown size"
            try:
                size_bytes = os.path.getsize(file_path)
                if size_bytes > 1024 * 1024:  # > 1MB
                    file_size = f"{size_bytes/(1024*1024):.1f} MB"
                elif size_bytes > 1024:  # > 1KB
                    file_size = f"{size_bytes/1024:.1f} KB"
                else:
                    file_size = f"{size_bytes} bytes"
            except:
                pass

            file_details.append(f"• {file_format}: {filename} ({file_size})")

        # Add metadata summary
        file_details.append("")
        file_details.append("Metadata Summary:")
        file_details.append(f"• Static parameters: {total_static_params}")
        file_details.append(f"• Time-series parameters: {total_timeseries_params}")
        if total_timeseries_points > 0:
            file_details.append(
                f"• Time-series length: {total_timeseries_points} time points"
            )
            duration_min = (
                total_timeseries_points * self.frame_interval.value()
            ) / 60.0
            file_details.append(f"• Total duration: {duration_min:.1f} minutes")

        msg.setDetailedText("\n".join(file_details))
        msg.setInformativeText(
            "Files include comprehensive HDF5 metadata in time-series format matching analysis data structure."
        )
        msg.exec_()

    def _append_hdf5_sheets_to_excel(self, excel_path: str, metadata_dict: dict):
        """
        Append HDF5 sensor timeseries sheets (temperature, humidity, LED, frame drift, etc.)
        to an existing Excel file without overwriting the analysis sheets already in it.
        """
        import pandas as pd

        is_legacy_enhanced = metadata_dict.get("main_file", {}).get(
            "legacy_enhanced", False
        )
        appended_sheets = []

        try:
            with pd.ExcelWriter(
                excel_path, engine="openpyxl", mode="a", if_sheet_exists="new"
            ) as writer:
                for source_name, metadata in metadata_dict.items():
                    if source_name in [
                        "analysis_info",
                        "file_info_only",
                        "nematostella_analysis",
                    ]:
                        continue

                    # HDF5 timeseries sheets
                    if "timeseries_data" in metadata and metadata["timeseries_data"]:
                        ts_data = metadata["timeseries_data"]
                        unit_info = ts_data.get("_unit_info", {})
                        hdf5_metadata_only = self._filter_hdf5_metadata_only(ts_data)

                        if hdf5_metadata_only:
                            frame_interval = self.frame_interval.value()

                            if is_legacy_enhanced and unit_info:
                                created = self._create_unit_enhanced_timeseries_sheets(
                                    writer,
                                    hdf5_metadata_only,
                                    unit_info,
                                    frame_interval,
                                    source_name,
                                )
                            else:
                                try:
                                    from ._metadata import (
                                        create_individual_timeseries_sheets,
                                    )

                                    created = create_individual_timeseries_sheets(
                                        writer, hdf5_metadata_only, frame_interval
                                    )
                                except Exception:
                                    created = self._create_timeseries_sheets_manually(
                                        writer, hdf5_metadata_only, source_name
                                    )

                            appended_sheets.extend(created)
                            for s in created:
                                self._log_message(f"   ✓ HDF5 sheet appended: '{s}'")

                    # Static HDF5 metadata sheet
                    static_metadata = {
                        k: v for k, v in metadata.items() if k != "timeseries_data"
                    }
                    if static_metadata:
                        try:
                            try:
                                from ._metadata import create_metadata_dataframe

                                meta_df = create_metadata_dataframe(
                                    static_metadata, source_name
                                )
                            except ImportError:
                                meta_df = self._create_metadata_dataframe_manually(
                                    static_metadata, source_name
                                )

                            sheet_name = f"Static_{source_name}"[:31]
                            meta_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            appended_sheets.append(sheet_name)
                            self._log_message(
                                f"   ✓ Static metadata sheet appended: '{sheet_name}'"
                            )
                        except Exception as e:
                            self._log_message(
                                f"   Warning: Could not create static sheet: {e}"
                            )

            self._log_message(
                f"HDF5 sensor sheets appended ({len(appended_sheets)} sheets): "
                + ", ".join(appended_sheets)
            )
        except Exception as e:
            self._log_message(f"Warning: Could not append HDF5 sheets: {e}")

    def _save_results_excel_with_metadata(
        self, excel_path: str, metadata_dict: dict, has_analysis_results: bool = True
    ):
        """
        Save Excel with unit-enhanced headers for legacy files and individual HDF5 sheets.
        """
        import pandas as pd

        # Check if this is a legacy enhanced file
        is_legacy_enhanced = metadata_dict.get("main_file", {}).get(
            "legacy_enhanced", False
        )

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

            # === SUMMARY SHEET (with unit-aware columns if legacy enhanced) ===
            if has_analysis_results:
                sorted_rois = sorted(self.merged_results.keys())

                summary_data = []
                for roi in sorted_rois:
                    if is_legacy_enhanced:
                        # Enhanced column names with units
                        row_data = {
                            "ROI": roi,
                            "Baseline_Mean_intensity": getattr(
                                self, "roi_baseline_means", {}
                            ).get(roi, 0),
                            "Upper_Threshold_intensity": getattr(
                                self, "roi_upper_thresholds", {}
                            ).get(roi, 0),
                            "Lower_Threshold_intensity": getattr(
                                self, "roi_lower_thresholds", {}
                            ).get(roi, 0),
                        }
                    else:
                        # Traditional column names
                        row_data = {
                            "ROI": roi,
                            "Baseline Mean": getattr(
                                self, "roi_baseline_means", {}
                            ).get(roi, 0),
                            "Upper Threshold": getattr(
                                self, "roi_upper_thresholds", {}
                            ).get(roi, 0),
                            "Lower Threshold": getattr(
                                self, "roi_lower_thresholds", {}
                            ).get(roi, 0),
                        }

                    # Add movement and sleep statistics with unit-aware names
                    movement_data = getattr(self, "movement_data", {})
                    if roi in movement_data and movement_data[roi]:
                        movement_values = [m for _, m in movement_data[roi]]
                        movement_pct = (
                            (sum(movement_values) / len(movement_values) * 100)
                            if movement_values
                            else 0
                        )

                        if is_legacy_enhanced:
                            row_data["Movement_Percentage_0to100"] = movement_pct
                        else:
                            row_data["Movement Percentage"] = movement_pct

                    sleep_data = getattr(self, "sleep_data", {})
                    if roi in sleep_data and sleep_data[roi]:
                        sleep_values = [s for _, s in sleep_data[roi]]
                        total_sleep_bins = sum(sleep_values)
                        sleep_minutes = (
                            total_sleep_bins * self.bin_size_seconds.value()
                        ) / 60

                        if is_legacy_enhanced:
                            row_data["Sleep_Time_minutes"] = sleep_minutes
                        else:
                            row_data["Sleep Time (min)"] = sleep_minutes

                    summary_data.append(row_data)

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                all_sheets_created = ["Summary"]
            else:
                # Create info sheet with legacy enhancement info
                info_data = [
                    {
                        "Property": "Extraction Type",
                        "Value": (
                            "HDF5 Metadata Only (Legacy Enhanced)"
                            if is_legacy_enhanced
                            else "HDF5 Metadata Only"
                        ),
                        "Description": "No analysis results available",
                    },
                    {
                        "Property": "Source File",
                        "Value": os.path.basename(self.file_path),
                        "Description": "HDF5 file analyzed",
                    },
                    {
                        "Property": "Extraction Date",
                        "Value": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Description": "When metadata was extracted",
                    },
                ]

                if is_legacy_enhanced:
                    info_data.append(
                        {
                            "Property": "Legacy Enhancement",
                            "Value": "Applied",
                            "Description": "Unit documentation added automatically",
                        }
                    )

                info_df = pd.DataFrame(info_data)
                info_df.to_excel(writer, sheet_name="File_Info", index=False)
                all_sheets_created = ["File_Info"]

            # === PROCESS HDF5 TIME-SERIES METADATA WITH UNIT ENHANCEMENT ===
            for source_name, metadata in metadata_dict.items():
                if source_name in [
                    "analysis_info",
                    "file_info_only",
                    "nematostella_analysis",
                ]:
                    continue

                # Process HDF5 time-series metadata
                if "timeseries_data" in metadata and metadata["timeseries_data"]:
                    ts_data = metadata["timeseries_data"]

                    # Get unit information if available
                    unit_info = ts_data.get("_unit_info", {})

                    # Filter to get only HDF5 metadata
                    hdf5_metadata_only = self._filter_hdf5_metadata_only(ts_data)

                    if hdf5_metadata_only:
                        self._log_message(
                            f"Creating unit-enhanced Excel sheets for {len(hdf5_metadata_only)} HDF5 parameters from {source_name}"
                        )

                        # Use frame interval from analysis if available, otherwise default
                        frame_interval = (
                            self.frame_interval.value() if has_analysis_results else 5.0
                        )

                        # CREATE UNIT-ENHANCED SHEETS
                        if is_legacy_enhanced and unit_info:
                            # Use enhanced unit-aware sheet creation
                            created_sheets = (
                                self._create_unit_enhanced_timeseries_sheets(
                                    writer,
                                    hdf5_metadata_only,
                                    unit_info,
                                    frame_interval,
                                    source_name,
                                )
                            )
                            all_sheets_created.extend(created_sheets)

                            self._log_message(
                                f"   ✅ Created {len(created_sheets)} unit-enhanced sheets"
                            )
                            for sheet_name in created_sheets:
                                self._log_message(f"   - {sheet_name}")

                        else:
                            # Fallback to regular sheet creation
                            try:
                                from ._metadata import (
                                    create_individual_timeseries_sheets,
                                    create_combined_timeseries_sheet,
                                )

                                individual_sheets = create_individual_timeseries_sheets(
                                    writer, hdf5_metadata_only, frame_interval
                                )
                                all_sheets_created.extend(individual_sheets)

                                for sheet_name in individual_sheets:
                                    self._log_message(
                                        f"   ✓ Created sheet '{sheet_name}'"
                                    )

                            except ImportError:
                                created_sheets = (
                                    self._create_timeseries_sheets_manually(
                                        writer, hdf5_metadata_only, source_name
                                    )
                                )
                                all_sheets_created.extend(created_sheets)
                            except Exception as e:
                                self._log_message(
                                    f"   Error with metadata functions: {e}"
                                )
                                created_sheets = (
                                    self._create_timeseries_sheets_manually(
                                        writer, hdf5_metadata_only, source_name
                                    )
                                )
                                all_sheets_created.extend(created_sheets)

                # Static HDF5 metadata sheet
                static_metadata = {
                    k: v for k, v in metadata.items() if k != "timeseries_data"
                }
                if static_metadata:
                    try:
                        try:
                            from ._metadata import create_metadata_dataframe

                            meta_df = create_metadata_dataframe(
                                static_metadata, source_name
                            )
                        except ImportError:
                            meta_df = self._create_metadata_dataframe_manually(
                                static_metadata, source_name
                            )

                        sheet_name = f"Static_{source_name}"[:31]
                        meta_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        all_sheets_created.append(sheet_name)
                        self._log_message(
                            f"   ✓ Created static metadata sheet '{sheet_name}'"
                        )
                    except Exception as e:
                        self._log_message(
                            f"   Warning: Could not create static sheet: {e}"
                        )

            # === PARAMETERS SHEET (only if analysis available) ===
            if has_analysis_results and "analysis_info" in metadata_dict:
                try:
                    params_data = []
                    analysis_info = metadata_dict["analysis_info"]

                    for key, value in analysis_info.items():
                        if key != "parameters":
                            params_data.append(
                                {
                                    "Parameter": key,
                                    "Value": str(value),
                                    "Category": "Analysis Info",
                                }
                            )

                    if "parameters" in analysis_info:
                        for key, value in analysis_info["parameters"].items():
                            params_data.append(
                                {
                                    "Parameter": key,
                                    "Value": str(value),
                                    "Category": "Analysis Parameters",
                                }
                            )

                    if params_data:
                        params_df = pd.DataFrame(params_data)
                        params_df.to_excel(
                            writer, sheet_name="Analysis_Parameters", index=False
                        )
                        all_sheets_created.append("Analysis_Parameters")
                except Exception as e:
                    self._log_message(
                        f"   Warning: Could not create parameters sheet: {e}"
                    )

            # Log final summary with enhancement info
            enhancement_info = " (with unit enhancement)" if is_legacy_enhanced else ""
            self._log_message(
                f"Excel file created{enhancement_info} with {len(all_sheets_created)} sheets:"
            )
            for sheet in all_sheets_created:
                self._log_message(f"   - {sheet}")

    def _create_unit_enhanced_timeseries_sheets(
        self,
        writer,
        hdf5_metadata: dict,
        unit_info: dict,
        frame_interval: float,
        source_name: str,
    ):
        """Create individual timeseries sheets with unit-enhanced column headers."""
        sheets_created = []

        # Create individual sheet for each parameter with unit-enhanced names
        for param_name, param_data in hdf5_metadata.items():
            if (
                not isinstance(param_data, (list, tuple, np.ndarray))
                or len(param_data) == 0
            ):
                continue

            try:
                max_length = len(param_data)
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                # Get unit information
                unit = unit_info.get(param_name, {}).get("units", "unknown")

                # Create unit-enhanced column names
                if unit == "seconds" and "drift" in param_name.lower():
                    # For timing parameters: create both seconds and milliseconds columns
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_seconds": param_data,
                        f"{param_name}_milliseconds": [
                            d * 1000 if d else 0 for d in param_data
                        ],
                    }
                elif unit == "celsius":
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_celsius": param_data,
                    }
                elif unit == "percent":
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_percent": param_data,
                    }
                elif unit == "milliseconds":
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_milliseconds": param_data,
                    }
                else:
                    df_data = {
                        "Time_minutes": time_minutes,
                        f"{param_name}_{unit}": param_data,
                    }

                # Create DataFrame and sheet
                param_df = pd.DataFrame(df_data)
                clean_name = self._clean_sheet_name(f"{param_name}_Enhanced")

                # Ensure unique sheet name
                original_clean_name = clean_name
                counter = 1
                while clean_name in sheets_created:
                    clean_name = f"{original_clean_name[:28]}_{counter}"
                    counter += 1

                param_df.to_excel(writer, sheet_name=clean_name, index=False)
                sheets_created.append(clean_name)

            except Exception as e:
                self._log_message(
                    f"   Warning: Could not create enhanced sheet for {param_name}: {e}"
                )
                continue

        # Create combined sheet with all enhanced parameters
        if len(hdf5_metadata) > 1:
            try:
                max_length = max(len(data) for data in hdf5_metadata.values())
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                combined_data = {"Time_minutes": time_minutes}

                for param_name, param_data in hdf5_metadata.items():
                    unit = unit_info.get(param_name, {}).get("units", "unknown")

                    # Pad data if necessary
                    if len(param_data) < max_length:
                        padded_data = list(param_data) + [np.nan] * (
                            max_length - len(param_data)
                        )
                    else:
                        padded_data = param_data

                    # Add with unit-enhanced name
                    if unit == "seconds" and "drift" in param_name.lower():
                        combined_data[f"{param_name}_seconds"] = padded_data
                        combined_data[f"{param_name}_ms"] = [
                            d * 1000 if d and not np.isnan(d) else np.nan
                            for d in padded_data
                        ]
                    else:
                        combined_data[f"{param_name}_{unit}"] = padded_data

                combined_df = pd.DataFrame(combined_data)
                combined_name = f"Enhanced_All_{source_name}"[:31]
                combined_df.to_excel(writer, sheet_name=combined_name, index=False)
                sheets_created.append(combined_name)

            except Exception as e:
                self._log_message(
                    f"   Warning: Could not create enhanced combined sheet: {e}"
                )

        return sheets_created

    def _filter_hdf5_metadata_only(self, ts_data: dict) -> dict:
        """Filter to keep only actual HDF5 metadata, excluding analysis results."""
        hdf5_metadata_only = {}

        # Only exclude specific analysis result patterns
        analysis_result_patterns = [
            "roi_",
            "baseline_",
            "threshold_",
            "upper_threshold",
            "lower_threshold",
            "movement_data",
            "fraction_data",
            "sleep_data",
            "quiescence_data",
            "intensity_roi_",
            "analysis_",
            "processed_",
            "calculated_",
        ]

        for param_name, param_data in ts_data.items():
            param_lower = param_name.lower()

            # Keep the parameter unless it matches specific analysis result patterns
            is_analysis_result = any(
                pattern in param_lower for pattern in analysis_result_patterns
            )

            if not is_analysis_result:
                hdf5_metadata_only[param_name] = param_data

        return hdf5_metadata_only

    def _create_timeseries_sheets_manually(
        self, writer, hdf5_metadata: dict, source_name: str
    ):
        """Manual fallback for creating time-series sheets."""
        import pandas as pd
        import numpy as np

        sheets_created = []

        for param_name, param_data in hdf5_metadata.items():
            try:
                if not hasattr(param_data, "__len__") or len(param_data) == 0:
                    continue

                # Create DataFrame with this parameter
                max_length = len(param_data)
                frame_interval = self.frame_interval.value()

                # Create time column aligned with analysis
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                # Create DataFrame
                df_data = {"Time (min)": time_minutes, param_name: param_data}

                param_df = pd.DataFrame(df_data)

                # Clean parameter name for Excel sheet (max 31 chars)
                clean_name = self._clean_sheet_name(param_name)

                # Ensure unique sheet name
                original_clean_name = clean_name
                counter = 1
                while clean_name in sheets_created:
                    clean_name = f"{original_clean_name[:28]}_{counter}"
                    counter += 1

                # Create the sheet
                param_df.to_excel(writer, sheet_name=clean_name, index=False)
                sheets_created.append(clean_name)
                self._log_message(
                    f"   ✓ Created manual sheet '{clean_name}' for {param_name}"
                )

            except Exception as e:
                self._log_message(
                    f"   Warning: Could not create sheet for {param_name}: {e}"
                )
                continue

        # Also create a combined sheet if we have multiple parameters
        if len(hdf5_metadata) > 1:
            try:
                # Find the maximum length across all parameters
                max_length = max(len(data) for data in hdf5_metadata.values())
                frame_interval = self.frame_interval.value()
                time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

                # Create combined DataFrame
                df_data = {"Time (min)": time_minutes}

                for param_name, param_data in hdf5_metadata.items():
                    # Pad shorter series with NaN
                    if len(param_data) < max_length:
                        padded_data = list(param_data) + [np.nan] * (
                            max_length - len(param_data)
                        )
                    else:
                        padded_data = param_data
                    df_data[param_name] = padded_data

                combined_df = pd.DataFrame(df_data)
                combined_name = f"All_HDF5_{source_name}"[:31]
                combined_df.to_excel(writer, sheet_name=combined_name, index=False)
                sheets_created.append(combined_name)
                self._log_message(
                    f"   ✓ Created combined manual sheet '{combined_name}'"
                )

            except Exception as e:
                self._log_message(f"   Warning: Could not create combined sheet: {e}")

        return sheets_created

    def _clean_sheet_name(self, param_name: str) -> str:
        """Clean parameter name to be valid Excel sheet name."""
        # Remove or replace invalid characters
        clean = param_name.replace("/", "_").replace("\\", "_").replace(":", "_")
        clean = (
            clean.replace("*", "star")
            .replace("?", "q")
            .replace("[", "")
            .replace("]", "")
        )
        clean = clean.replace("<", "lt").replace(">", "gt").replace("|", "_")

        # Truncate to 31 characters max
        if len(clean) > 31:
            # Try to keep meaningful parts
            if "_" in clean:
                parts = clean.split("_")
                if len(parts[0]) <= 25:
                    clean = parts[0] + "_" + "".join(p[0] for p in parts[1:] if p)[:5]
                else:
                    clean = clean[:31]
            else:
                clean = clean[:31]

        # Remove trailing underscores and ensure not empty
        clean = clean.rstrip("_")
        if not clean:
            clean = "unnamed_param"

        return clean

    def _create_metadata_dataframe_manually(self, metadata: dict, source_name: str):
        """Manual fallback for creating metadata DataFrame."""
        import pandas as pd

        rows = []

        def flatten_metadata(d, prefix=""):
            for key, value in d.items():
                full_key = f"{prefix}.{key}" if prefix else key

                if isinstance(value, dict):
                    flatten_metadata(value, full_key)
                else:
                    try:
                        str_value = str(value)
                        data_type = type(value).__name__
                    except:
                        str_value = f"<{type(value).__name__}>"
                        data_type = "Complex"

                    rows.append(
                        {
                            "Category": prefix if prefix else "Root",
                            "Parameter": key,
                            "Value": str_value,
                            "Data_Type": data_type,
                            "Source": source_name,
                        }
                    )

        flatten_metadata(metadata)
        return pd.DataFrame(rows)

    def _create_individual_sheets_fallback(
        self, writer, hdf5_metadata: dict, sheets_created: list
    ):
        """Fallback method using new module functions if available."""
        try:
            from ._metadata import create_individual_timeseries_sheets

            new_sheets = create_individual_timeseries_sheets(
                writer, hdf5_metadata, self.frame_interval.value()
            )
            sheets_created.extend(new_sheets)
            self._log_message(f"   ✓ Created {len(new_sheets)} fallback sheets")
        except Exception as e:
            self._log_message(f"   ✗ Fallback method failed: {e}")

    def _create_metadata_sheet_fallback(
        self, writer, hdf5_metadata: dict, source_name: str
    ):
        """
        Fallback method to create HDF5 metadata sheet when import fails.
        """
        import pandas as pd
        import numpy as np

        try:
            if not hdf5_metadata:
                return

            # Find the maximum length
            max_length = max(
                len(data) if hasattr(data, "__len__") else 1
                for data in hdf5_metadata.values()
            )

            # Create time column
            frame_interval = self.frame_interval.value()
            time_minutes = [(i * frame_interval) / 60.0 for i in range(max_length)]

            # Build DataFrame
            df_data = {"Time (min)": time_minutes}

            for param_name, param_data in hdf5_metadata.items():
                if hasattr(param_data, "__len__") and len(param_data) > 0:
                    # Pad shorter series with NaN
                    padded_data = list(param_data) + [np.nan] * (
                        max_length - len(param_data)
                    )
                    df_data[param_name] = padded_data
                else:
                    # Single value or empty data
                    df_data[param_name] = (
                        [param_data] * max_length if max_length > 0 else []
                    )

            df = pd.DataFrame(df_data)
            sheet_name = f"HDF5_{source_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            self._log_message(f"   Created fallback metadata sheet '{sheet_name}'")

        except Exception as e:
            self._log_message(f"   Error in fallback metadata sheet creation: {e}")

    def _create_static_metadata_sheet_fallback(
        self, writer, static_metadata: dict, source_name: str
    ):
        """
        Fallback method to create static metadata sheet when import fails.
        """
        import pandas as pd

        try:
            rows = []

            def flatten_metadata(d, prefix=""):
                for key, value in d.items():
                    full_key = f"{prefix}.{key}" if prefix else key

                    if isinstance(value, dict):
                        flatten_metadata(value, full_key)
                    else:
                        rows.append(
                            {
                                "Category": prefix if prefix else "Root",
                                "Parameter": key,
                                "Value": str(value),
                                "Data_Type": type(value).__name__,
                                "Source": source_name,
                            }
                        )

            flatten_metadata(static_metadata)

            df = pd.DataFrame(rows)
            sheet_name = f"Static_{source_name}"[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            self._log_message(f"   Created fallback static sheet '{sheet_name}'")

        except Exception as e:
            self._log_message(f"   Error in fallback static sheet creation: {e}")

    # def _create_metadata_dataframe(self, metadata: dict, source_name: str):
    #     """Helper to create DataFrame from metadata."""

    #     rows = []

    #     def flatten_metadata(d, prefix=""):
    #         for key, value in d.items():
    #             full_key = f"{prefix}.{key}" if prefix else key

    #             if isinstance(value, dict):
    #                 flatten_metadata(value, full_key)
    #             else:
    #                 try:
    #                     str_value = str(value)
    #                     data_type = type(value).__name__
    #                 except:
    #                     str_value = f"<{type(value).__name__}>"
    #                     data_type = 'Complex'

    #                 rows.append({
    #                     'Category': prefix if prefix else 'Root',
    #                     'Parameter': key,
    #                     'Value': str_value,
    #                     'Data_Type': data_type,
    #                     'Source': source_name
    #                 })

    #     flatten_metadata(metadata)
    #     return pd.DataFrame(rows)

    def _get_analysis_parameters_for_metadata(self) -> dict:
        """Extract analysis parameters for metadata."""
        params = {
            "method": self._get_current_threshold_method_display(),
            "frame_interval_seconds": self.frame_interval.value(),
            "bin_size_seconds": self.bin_size_seconds.value(),
            "quiescence_threshold": self.quiescence_threshold.value(),
            "sleep_threshold_minutes": self.sleep_threshold_minutes.value(),
        }

        # Method-specific parameters
        if hasattr(self, "baseline_duration_minutes"):
            params["baseline_duration_minutes"] = self.baseline_duration_minutes.value()
        if hasattr(self, "threshold_multiplier"):
            params["threshold_multiplier"] = self.threshold_multiplier.value()
        if hasattr(self, "calibration_multiplier"):
            params["calibration_multiplier"] = self.calibration_multiplier.value()
        if hasattr(self, "enable_detrending"):
            params["detrending_enabled"] = self.enable_detrending.isChecked()

        return params

    # ===================================================================
    # UI EVENT HANDLERS
    # ===================================================================

    def _on_progress_update(self, percent: int):
        """Update progress bar."""
        self.progress_bar.setValue(percent)

    def _on_status_update(self, message: str):
        """Update status label."""
        self.status_label.setText(message)
        self._log_message(message)

    def _on_performance_update(self, metrics: str):
        """Update performance metrics."""
        self.performance_label.setText(metrics)

    def _update_performance_metrics(self):
        """Update real-time performance metrics during analysis."""
        if self.analysis_start_time and not self._cancel_requested:
            elapsed = time.time() - self.analysis_start_time
            cpu_percent = psutil.cpu_percent(interval=None)
            memory_info = psutil.virtual_memory()

            metrics = (
                f"Elapsed: {elapsed:.1f}s | "
                f"CPU: {cpu_percent:.1f}% | "
                f"Memory: {memory_info.percent:.1f}% | "
                f"Processes: {self.num_processes.value()}"
            )

            self.performance_updated.emit(metrics)

    def _on_auto_scale_toggled(self, checked: bool):
        """Enable/disable manual Y-axis controls and advanced options based on auto scale setting."""
        # Manual controls
        self.y_min_spin.setEnabled(not checked)
        self.y_max_spin.setEnabled(not checked)
        self.btn_apply_y_range.setEnabled(not checked)

        # Advanced auto-scaling options
        self.robust_scaling.setEnabled(checked)
        self.adaptive_scaling.setEnabled(checked)
        self.center_around_zero.setEnabled(checked)
        self.lower_percentile_spin.setEnabled(
            checked and self.robust_scaling.isChecked()
        )
        self.upper_percentile_spin.setEnabled(
            checked and self.robust_scaling.isChecked()
        )

        # Regenerate plot with new scaling
        if hasattr(self, "merged_results") and self.merged_results:
            self.generate_plot()

    def _on_threshold_tab_changed(self, tab_index: int):
        """Handle threshold method tab changes."""
        # Tab index directly determines the method
        # 0 = Baseline, 1 = Calibration, 2 = Adaptive
        method_names = ["Baseline", "Calibration", "Adaptive"]
        if 0 <= tab_index < len(method_names):
            self._log_message(f"Threshold method changed to: {method_names[tab_index]}")

    def _get_current_threshold_method(self) -> str:
        """Get current threshold method based on active tab."""
        tab_index = self.threshold_params_stack.currentIndex()
        method_map = {0: "baseline", 1: "calibration", 2: "adaptive"}
        return method_map.get(tab_index, "baseline")

    def _get_current_threshold_method_display(self) -> str:
        """Get current threshold method display name based on active tab."""
        tab_index = self.threshold_params_stack.currentIndex()
        method_map = {
            0: "Baseline (First Frames)",
            1: "Calibration (Sedated Animals)",
            2: "Adaptive (Smart Detection)",
        }
        return method_map.get(tab_index, "Baseline (First Frames)")

    def load_calibration_file(self):
        """Enhanced calibration file loading with robust UI updates."""

        self._log_message("=== LOAD_CALIBRATION_FILE METHOD CALLED ===")

        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Select Calibration File",
                "",
                "Video Files (*.h5 *.hdf5 *.avi);;HDF5 Files (*.h5 *.hdf5);;AVI Files (*.avi);;All Files (*.*)",
            )

            self._log_message(f"File dialog returned: '{file_path}'")

            if file_path and os.path.exists(file_path):
                basename = os.path.basename(file_path)

                # Store the path first
                self.calibration_file_path_stored = file_path
                self._log_message(f"Stored calibration file path: {file_path}")

                # Force UI update with multiple methods
                self.calibration_file_path.setText(basename)
                self.calibration_file_path.setProperty("full_path", file_path)

                # Force the widget to process events and update
                from qtpy.QtCore import QCoreApplication

                QCoreApplication.processEvents()

                # Verify the text was set
                current_text = self.calibration_file_path.text()
                self._log_message(f"UI text field now shows: '{current_text}'")

                if current_text != basename:
                    self._log_message(
                        f"WARNING: UI text mismatch! Expected '{basename}', got '{current_text}'"
                    )
                    # Try setting it again
                    self.calibration_file_path.setText(basename)
                    QCoreApplication.processEvents()
                    self._log_message(
                        f"After retry, UI shows: '{self.calibration_file_path.text()}'"
                    )

                # Enable the load dataset button
                if hasattr(self, "btn_load_calibration_dataset"):
                    self.btn_load_calibration_dataset.setEnabled(True)
                    self._log_message("Enabled 'Load Calibration Dataset' button")

                # Update status
                if hasattr(self, "calibration_status_label"):
                    self.calibration_status_label.setText(
                        "✅ 1. Calibration file selected\n"
                        "2. Click 'Load Calibration Dataset'\n"
                        "3. Detect ROIs (Input tab)\n"
                        "4. Process baseline"
                    )

                # Reset calibration processing state
                self.calibration_baseline_processed = False
                self.calibration_baseline_statistics = {}
                if hasattr(self, "btn_process_calibration_baseline"):
                    self.btn_process_calibration_baseline.setEnabled(False)

                self._log_message(f"Calibration file selection complete: {basename}")

            else:
                self._log_message("No valid file selected or file doesn't exist")

                # Reset UI
                self.calibration_file_path.setText("No calibration file selected")
                self.calibration_file_path_stored = None

                if hasattr(self, "btn_load_calibration_dataset"):
                    self.btn_load_calibration_dataset.setEnabled(False)

        except Exception as e:
            self._log_message(f"ERROR in load_calibration_file: {e}")
            import traceback

            self._log_message(f"Traceback: {traceback.format_exc()}")

    def _on_6well_toggled(self, checked: bool):
        """Enable/disable ROI spinboxes when 6-well preset is toggled."""
        if checked:
            self.min_radius.setValue(100)
            self.max_radius.setValue(145)
            self.dp_param.setValue(1.0)
            self.min_dist.setValue(300)
            self.param1.setValue(30)
            self.param2.setValue(60)

        for widget in (
            self.min_radius,
            self.max_radius,
            self.dp_param,
            self.min_dist,
            self.param1,
            self.param2,
        ):
            widget.setEnabled(not checked)

    def _on_12well_toggled(self, checked: bool):
        """Enable/disable and populate ROI controls when preset is toggled."""
        # 12-well plate preset values
        preset = {
            "min_radius": 40,
            "max_radius": 75,
            "dp": 1.0,
            "min_dist": 75,
            "param1": 50,
            "param2": 30,
        }

        if checked:
            # Push the preset into the spin-boxes
            self.min_radius.setValue(preset["min_radius"])
            self.max_radius.setValue(preset["max_radius"])
            self.dp_param.setValue(preset["dp"])
            self.min_dist.setValue(preset["min_dist"])
            self.param1.setValue(preset["param1"])
            self.param2.setValue(preset["param2"])

        # Disable or re-enable editing
        for widget in (
            self.min_radius,
            self.max_radius,
            self.dp_param,
            self.min_dist,
            self.param1,
            self.param2,
        ):
            widget.setEnabled(not checked)

    def on_tab_changed(self, index: int):
        """Handle tab changes."""
        # Add specific logic for when tabs are changed if needed
        pass

    # ===================================================================
    # ROI MANAGEMENT METHODS
    # ===================================================================

    def _toggle_all_roi_visibility(self):
        """Toggle visibility of all ROI mask layers."""
        roi_layers = [
            layer
            for layer in self.viewer.layers
            if hasattr(layer, "metadata")
            and layer.metadata.get("roi_type") == "circular_detection"
        ]

        if not roi_layers:
            self._log_message("No ROI layers found")
            return

        # Check current state and toggle
        any_visible = any(layer.visible for layer in roi_layers)
        new_visibility = not any_visible

        for layer in roi_layers:
            layer.visible = new_visibility

        status = "visible" if new_visibility else "hidden"
        self._log_message(f"All ROI layers are now {status}")

    def _show_only_selected_roi(self):
        """Show only the currently selected ROI layer."""
        selected_layers = list(self.viewer.layers.selection)
        roi_layers = [
            layer
            for layer in self.viewer.layers
            if hasattr(layer, "metadata")
            and layer.metadata.get("roi_type") == "circular_detection"
        ]

        if not roi_layers:
            self._log_message("No ROI layers found")
            return

        selected_roi_layers = [
            layer for layer in selected_layers if layer in roi_layers
        ]

        if not selected_roi_layers:
            self._log_message("No ROI layer selected")
            return

        # Hide all ROI layers first
        for layer in roi_layers:
            layer.visible = False

        # Show only selected ROI layers
        for layer in selected_roi_layers:
            layer.visible = True
            roi_id = layer.metadata.get("roi_id", "unknown")
            self._log_message(f"Showing only ROI {roi_id}")

    def _reset_roi_visibility(self):
        """Reset ROI layer visibility to default state."""
        roi_layers = [
            layer
            for layer in self.viewer.layers
            if hasattr(layer, "metadata")
            and layer.metadata.get("roi_type") == "circular_detection"
        ]

        for layer in roi_layers:
            layer.visible = False  # Default state
            layer.opacity = 0.6  # Reset opacity
            layer.blending = "additive"  # Reset blending

        self._log_message(f"Reset visibility for {len(roi_layers)} ROI layers")

    # ===================================================================
    # EXTENDED ANALYSIS METHODS (FISCHER Z-TRANSFORMATION)
    # ===================================================================

    def _set_period_preset(self, min_period, max_period):
        """Set period range from preset."""
        self.fisher_min_period.setValue(min_period)
        self.fisher_max_period.setValue(max_period)
        self._log_message(
            f"Period range set to: {min_period:.1f} - {max_period:.1f} hours"
        )

    def _auto_detect_period_range(self):
        """Automatically detect optimal period range based on recording duration."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("⚠️ No data loaded. Please run analysis first.")
            return

        # Get recording duration from first ROI
        first_roi_data = next(iter(self.merged_results.values()))
        if not first_roi_data:
            return

        times = [t for t, _ in first_roi_data]
        duration_seconds = max(times) - min(times)
        duration_hours = duration_seconds / 3600.0

        # Determine optimal range based on duration
        # Rule: max_period should be at most 1/2 of recording duration
        if duration_hours < 12:
            # Short recording: 0.5h - duration/2
            min_period = 0.5
            max_period = max(3.0, duration_hours / 2)
            preset_name = "Short Cycles"
        elif duration_hours < 48:
            # Medium recording: 1h - duration/2
            min_period = 1.0
            max_period = min(18.0, duration_hours / 2)
            preset_name = "Ultradian/Semi-Daily"
        else:
            # Long recording: circadian analysis
            min_period = 12.0
            max_period = min(36.0, duration_hours / 2)
            preset_name = "Circadian"

        self.fisher_min_period.setValue(min_period)
        self.fisher_max_period.setValue(max_period)

        self._log_message(
            f"Auto-detected: Recording duration = {duration_hours:.1f}h\n"
            f"  → Using '{preset_name}' preset: {min_period:.1f} - {max_period:.1f} hours"
        )

    def _on_fisher_method_changed(self, index):
        """Handle change in analysis method selection."""
        methods = [
            "Chi² Periodogram",
            "FFT Power Spectrum",
            "Cosinor Analysis",
            "ROI Similarity Matrix",
            "Coherence Analysis",
            "Phase Clustering",
        ]
        self._log_message(f"Analysis method changed to: {methods[index]}")

        # Update data source dropdown based on method
        self._update_data_source_for_method(index)

    def _update_data_source_for_method(self, method_index):
        """Update data source dropdown based on selected analysis method.

        All methods (Fisher, FFT, Cosinor, etc.) can operate on either
        Fraction Movement or Raw Intensity data.
        """
        if not hasattr(self, "data_source_combo"):
            return

        self.data_source_combo.blockSignals(True)

        current_index = self.data_source_combo.currentIndex()
        self.data_source_combo.clear()
        self.data_source_combo.addItems(
            ["Fraction Movement (0-1)", "Raw Intensity (continuous)"]
        )
        if current_index < self.data_source_combo.count():
            self.data_source_combo.setCurrentIndex(current_index)
        else:
            self.data_source_combo.setCurrentIndex(0)
        self.data_source_combo.setEnabled(True)
        self.data_source_combo.setToolTip(
            "Choose data source for analysis:\n"
            "• Fraction Movement: Binned activity ratio (0-1), smoother signal\n"
            "• Raw Intensity: Continuous per-pixel intensity changes (MinMax 0-1)"
        )

        self.data_source_combo.blockSignals(False)

    def _on_cycle_selection_toggled(self, state):
        """Handle toggling of cycle selection checkbox."""
        enabled = state == 2  # Qt.Checked = 2

        # Enable/disable cycle selection controls
        self.cycle_start_time.setEnabled(enabled)
        self.cycle_end_time.setEnabled(enabled)
        self.btn_cycle_first24.setEnabled(enabled)
        self.btn_cycle_second24.setEnabled(enabled)
        self.btn_cycle_last24.setEnabled(enabled)
        self.btn_cycle_reset.setEnabled(enabled)

        if enabled:
            self._log_message(
                "Time range selection enabled - analysis will use specified window"
            )
        else:
            self._log_message(
                "Time range selection disabled - analysis will use full recording"
            )

    def _set_cycle_range(self, start_hours, end_hours):
        """Set the cycle selection range to specific hours."""
        self.cycle_start_time.setValue(start_hours)
        self.cycle_end_time.setValue(end_hours)
        self._log_message(f"Time range set to {start_hours:.1f}h - {end_hours:.1f}h")

    def _set_cycle_last_24h(self):
        """Set the cycle selection to the last 24 hours of the recording."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("⚠️ No data loaded - cannot determine recording duration")
            return

        # Get recording duration from first ROI
        first_roi = list(self.merged_results.keys())[0]
        data = self.merged_results[first_roi]

        if not data:
            self._log_message("⚠️ No data available")
            return

        # Calculate duration in hours
        duration_seconds = data[-1][0] - data[0][0]
        duration_hours = duration_seconds / 3600.0

        if duration_hours < 24:
            self._log_message(
                f"⚠️ Recording is only {duration_hours:.1f}h - using full recording"
            )
            self._set_cycle_range(0, duration_hours)
        else:
            start_hours = duration_hours - 24
            self._set_cycle_range(start_hours, duration_hours)

    def _reset_cycle_range(self):
        """Reset cycle selection to full recording."""
        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("⚠️ No data loaded")
            return

        # Get recording duration from first ROI
        first_roi = list(self.merged_results.keys())[0]
        data = self.merged_results[first_roi]

        if not data:
            self._log_message("⚠️ No data available")
            return

        # Calculate duration in hours
        duration_seconds = data[-1][0] - data[0][0]
        duration_hours = duration_seconds / 3600.0

        self._set_cycle_range(0, duration_hours)

    def _adjust_analysis_bin_size(self, direction):
        """Adjust analysis bin size by one step in the given direction."""
        current = self.analysis_bin_size.value()
        step = self.analysis_bin_size.singleStep()
        new_value = current + (direction * step)

        # Clamp to valid range
        new_value = max(
            self.analysis_bin_size.minimum(),
            min(self.analysis_bin_size.maximum(), new_value),
        )
        self.analysis_bin_size.setValue(new_value)

    def _set_analysis_bin_preset(self, preset):
        """Set analysis bin size to a preset value."""
        if preset == "original":
            # Use original bin size from main analysis
            original_bin = (
                self.bin_size_seconds.value()
                if hasattr(self, "bin_size_seconds")
                else 60
            )
            self.analysis_bin_size.setValue(original_bin)
            self._log_message(f"Analysis bin size set to original: {original_bin}s")
        else:
            # Numeric preset
            self.analysis_bin_size.setValue(preset)
            self._log_message(f"Analysis bin size set to: {preset}s")

    def _update_bin_size_info(self):
        """Update the info label showing bin size comparison."""
        analysis_bin = self.analysis_bin_size.value()
        original_bin = (
            self.bin_size_seconds.value() if hasattr(self, "bin_size_seconds") else 60
        )

        if analysis_bin == original_bin:
            self.bin_size_info_label.setText(
                f"✓ Using original bin size ({original_bin}s) - no re-binning needed"
            )
            self.bin_size_info_label.setStyleSheet(
                "color: #27ae60; font-size: 10px; font-style: italic;"
            )
        elif analysis_bin > original_bin:
            factor = analysis_bin / original_bin
            self.bin_size_info_label.setText(
                f"⚠ Data will be re-binned: {original_bin}s → {analysis_bin}s ({factor:.1f}x larger bins)"
            )
            self.bin_size_info_label.setStyleSheet(
                "color: #f39c12; font-size: 10px; font-style: italic;"
            )
        else:
            self.bin_size_info_label.setText(
                f"⚠ Cannot re-bin to smaller size: {analysis_bin}s < {original_bin}s (original). Using original."
            )
            self.bin_size_info_label.setStyleSheet(
                "color: #e74c3c; font-size: 10px; font-style: italic;"
            )

    def _on_data_source_changed(self, index):
        """Handle change in data source selection."""
        source_names = [
            "Fraction Movement (0-1)",
            "Raw Intensity (continuous)",
            "Normalized Movement (0-1)",
        ]
        if index < len(source_names):
            self._log_message(f"Data source changed to: {source_names[index]}")

    def _rebin_timeseries_data(
        self,
        data_dict: Dict[int, List[Tuple[float, float]]],
        new_bin_size: int,
        original_bin_size: int,
    ) -> Dict[int, List[Tuple[float, float]]]:
        """
        Re-bin timeseries data to a larger bin size.

        Args:
            data_dict: Dictionary of {roi_id: [(time, value), ...]}
            new_bin_size: New bin size in seconds
            original_bin_size: Original bin size in seconds

        Returns:
            Dictionary with re-binned data
        """
        if new_bin_size <= original_bin_size:
            # Cannot re-bin to smaller size, return original
            return data_dict

        # Calculate binning factor (must be integer for proper binning)
        factor = int(round(new_bin_size / original_bin_size))

        rebinned_data = {}

        for roi_id, timeseries in data_dict.items():
            if not timeseries:
                rebinned_data[roi_id] = []
                continue

            # Group data points into larger bins
            rebinned = []
            for i in range(0, len(timeseries), factor):
                bin_data = timeseries[i : i + factor]

                if not bin_data:
                    continue

                # Use middle timepoint of the bin
                times = [t for t, _ in bin_data]
                values = [v for _, v in bin_data]

                avg_time = sum(times) / len(times)
                avg_value = sum(values) / len(values)

                rebinned.append((avg_time, avg_value))

            rebinned_data[roi_id] = rebinned

        return rebinned_data

    def _get_rebinned_behavioral_data(self, bin_minutes: int):
        """Rebin fraction data and recompute quiescence and sleep consistently.

        Returns (fraction_data, quiescence_data, sleep_data) all at the same
        bin resolution. When bin_minutes=0 the stored originals are returned.
        """
        from ._calc import bin_quiescence, define_sleep_periods

        fraction = getattr(self, "fraction_data", {})
        original_bin = self.bin_size_seconds.value() if hasattr(self, "bin_size_seconds") else 60

        if bin_minutes > 0:
            new_bin_seconds = bin_minutes * 60
            if new_bin_seconds > original_bin:
                fraction = self._rebin_timeseries_data(fraction, new_bin_seconds, original_bin)
                bin_seconds = new_bin_seconds
            else:
                bin_seconds = original_bin
        else:
            bin_seconds = original_bin

        quiescence = bin_quiescence(fraction, self.quiescence_threshold.value())
        sleep = define_sleep_periods(
            quiescence,
            self.sleep_threshold_minutes.value(),
            bin_seconds,
        )
        return fraction, quiescence, sleep

    def _load_results_from_hdf5(self):
        """Load ALL analysis results from HDF5 file (core + extended analysis)."""
        from qtpy.QtWidgets import QFileDialog
        from ._results_io import load_comprehensive_results

        # Open file dialog to select results file
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Results from HDF5",
            "",
            "HDF5 Files (*.h5 *.hdf5);;All Files (*)",
        )

        if not file_path:
            self._log_message("Load cancelled by user")
            return

        try:
            self._log_message(f"Loading comprehensive results from: {file_path}")

            # Load all results using comprehensive load function
            loaded_data = load_comprehensive_results(file_path)

            # Restore core analysis results
            if "core_analysis" in loaded_data:
                core = loaded_data["core_analysis"]

                # Restore raw signal (merged_results) — new key name
                if "merged_results" in core:
                    self.merged_results = core["merged_results"]
                    self._log_message(
                        f"  ✓ Loaded raw signal for {len(self.merged_results)} ROIs"
                    )
                elif "movement_data" in core:  # backward compat with old saves
                    self.merged_results = core["movement_data"]
                    self._log_message(
                        f"  ✓ Loaded raw signal (legacy key) for {len(self.merged_results)} ROIs"
                    )

                # Restore binary movement data
                if "movement_data" in core:
                    self.movement_data = core["movement_data"]
                    self._log_message(
                        f"  ✓ Loaded binary movement data for {len(self.movement_data)} ROIs"
                    )

                # Restore fraction data
                if "fraction_data" in core:
                    self.fraction_data = core["fraction_data"]
                    self._log_message(
                        f"  ✓ Loaded fraction data for {len(self.fraction_data)} ROIs"
                    )

                # Restore quiescence data
                if "quiescence_data" in core:
                    self.quiescence_data = core["quiescence_data"]
                    self._log_message(
                        f"  ✓ Loaded quiescence data for {len(self.quiescence_data)} ROIs"
                    )

                # Restore sleep data
                if "sleep_data" in core:
                    self.sleep_data = core["sleep_data"]
                    self._log_message(
                        f"  ✓ Loaded sleep data for {len(self.sleep_data)} ROIs"
                    )
                    # Recompute sleep quality metrics from restored sleep data
                    try:
                        from ._calc import calculate_sleep_quality_hourly
                        self.sleep_quality_data = calculate_sleep_quality_hourly(
                            self.sleep_data
                        )
                        self._log_message("  ✓ Sleep quality metrics recomputed")
                    except Exception:
                        pass

                # Restore ROI colors
                if "roi_colors" in core:
                    self.roi_colors = core["roi_colors"]
                    self._log_message(
                        f"  ✓ Loaded ROI colors for {len(self.roi_colors)} ROIs"
                    )

                # Restore ROI statistics
                if "roi_statistics" in core:
                    self.roi_statistics = core["roi_statistics"]
                    self._log_message(
                        f"  ✓ Loaded ROI statistics for {len(self.roi_statistics)} ROIs"
                    )

                # Restore thresholds
                if "thresholds" in core:
                    self.roi_baseline_means = {}
                    self.roi_upper_thresholds = {}
                    self.roi_lower_thresholds = {}

                    for roi_id, thresh_data in core["thresholds"].items():
                        self.roi_baseline_means[roi_id] = thresh_data.get(
                            "baseline_mean", 0.0
                        )
                        self.roi_upper_thresholds[roi_id] = thresh_data.get(
                            "upper_threshold", 0.0
                        )
                        self.roi_lower_thresholds[roi_id] = thresh_data.get(
                            "lower_threshold", 0.0
                        )

                    self._log_message(
                        f"  ✓ Loaded thresholds for {len(self.roi_baseline_means)} ROIs"
                    )

            # Restore extended analysis results
            if "extended_analysis" in loaded_data and loaded_data["extended_analysis"]:
                extended = loaded_data["extended_analysis"]

                # Restore fisher analysis results
                if "results" in extended:
                    self.fisher_analysis_results = extended["results"]
                    self._log_message("  ✓ Loaded extended analysis results")

                # Restore method selection
                if "method_index" in extended:
                    method_idx = extended["method_index"]
                    self.current_fisher_method = method_idx
                    if hasattr(self, "fisher_method_combo"):
                        self.fisher_method_combo.setCurrentIndex(method_idx)
                    self._log_message(
                        f"  ✓ Restored analysis method: {extended.get('method_name', 'Unknown')}"
                    )

                    # Re-create plot for extended analysis
                    if hasattr(self, "fisher_analysis_results"):
                        self._create_circadian_plot(
                            self.fisher_analysis_results, method_idx
                        )

                    # Generate and display summary
                    if method_idx == 0:  # Chi² Periodogram
                        from ._fisher_analysis import generate_circadian_summary

                        summary = generate_circadian_summary(
                            self.fisher_analysis_results
                        )
                    elif method_idx == 1:  # FFT Power Spectrum
                        from ._circadian_fft import generate_fft_summary

                        summary = generate_fft_summary(self.fisher_analysis_results)
                    elif method_idx == 2:  # Cosinor Analysis
                        summary = self._generate_cosinor_summary(
                            self.fisher_analysis_results
                        )
                    elif method_idx == 3:  # ROI Similarity Matrix
                        summary = self._generate_similarity_summary(
                            self.fisher_analysis_results
                        )
                    elif method_idx == 4:  # Coherence Analysis
                        summary = self._generate_coherence_summary(
                            self.fisher_analysis_results
                        )
                    elif method_idx == 5:  # Phase Clustering
                        summary = self._generate_phase_clustering_summary(
                            self.fisher_analysis_results
                        )
                    else:
                        summary = "Extended analysis results loaded successfully."

                    self.fisher_results_text.setPlainText(summary)

            # Restore analysis parameters
            if "analysis_parameters" in loaded_data:
                params = loaded_data["analysis_parameters"]

                # Restore core parameters
                if "core" in params:
                    core_params = params["core"]
                    if (
                        "frame_interval" in core_params
                        and core_params["frame_interval"] is not None
                    ):
                        if hasattr(self, "frame_interval"):
                            self.frame_interval.setValue(core_params["frame_interval"])
                    if (
                        "bin_size_seconds" in core_params
                        and core_params["bin_size_seconds"] is not None
                    ):
                        if hasattr(self, "bin_size_seconds"):
                            self.bin_size_seconds.setValue(
                                core_params["bin_size_seconds"]
                            )

                # Restore extended parameters
                if "extended" in params:
                    ext_params = params["extended"]
                    if (
                        "min_period_hours" in ext_params
                        and ext_params["min_period_hours"] is not None
                    ):
                        if hasattr(self, "fisher_min_period"):
                            self.fisher_min_period.setValue(
                                ext_params["min_period_hours"]
                            )
                    if (
                        "max_period_hours" in ext_params
                        and ext_params["max_period_hours"] is not None
                    ):
                        if hasattr(self, "fisher_max_period"):
                            self.fisher_max_period.setValue(
                                ext_params["max_period_hours"]
                            )
                    if (
                        "significance_level" in ext_params
                        and ext_params["significance_level"] is not None
                    ):
                        if hasattr(self, "fisher_significance"):
                            self.fisher_significance.setValue(
                                ext_params["significance_level"]
                            )

                self._log_message("  ✓ Restored analysis parameters")

            # Build summary message
            summary_lines = [
                "Successfully loaded comprehensive results from HDF5!",
                "",
                f"File: {file_path}",
            ]

            if hasattr(self, "merged_results") and self.merged_results:
                summary_lines.append(f"ROIs loaded: {len(self.merged_results)}")
                summary_lines.append(f"ROI IDs: {sorted(self.merged_results.keys())}")

            summary_lines.append("")
            summary_lines.append("Loaded Data:")

            if hasattr(self, "merged_results") and self.merged_results:
                summary_lines.append(
                    f"  • Movement data: {len(self.merged_results)} ROIs"
                )

            if hasattr(self, "fraction_data") and self.fraction_data:
                summary_lines.append(
                    f"  • Fraction data: {len(self.fraction_data)} ROIs"
                )

            if hasattr(self, "roi_baseline_means") and self.roi_baseline_means:
                summary_lines.append(
                    f"  • Thresholds: {len(self.roi_baseline_means)} ROIs"
                )

            if (
                hasattr(self, "fisher_analysis_results")
                and self.fisher_analysis_results
            ):
                method_name = loaded_data.get("extended_analysis", {}).get(
                    "method_name", "Unknown"
                )
                summary_lines.append(f"  • Extended analysis: {method_name}")

            summary_lines.extend(
                [
                    "",
                    "All analysis results have been restored.",
                    "You can now perform post-hoc cycle/period analysis or export results.",
                ]
            )

            self._log_message("✓ Successfully loaded comprehensive results")

            # Auto-adjust time range spinboxes to match loaded recording duration
            self._auto_adjust_time_range()

            # Compute pixel counts per ROI for Real Amplitude (sum) display mode.
            # Uses current self.masks if they are loaded (from ROI detection step).
            masks = getattr(self, "masks", [])
            if masks:
                self.roi_pixel_counts = {
                    i + 1: int(np.sum(m > 0)) for i, m in enumerate(masks)
                }
                self._log_message(
                    f"  ✓ Pixel counts computed for {len(self.roi_pixel_counts)} ROIs"
                )
            else:
                self.roi_pixel_counts = {}
                self._log_message(
                    "  ⚠ No masks available — Real Amplitude (pixel sum) mode not available"
                )

            # Only show summary in text widget if no extended analysis was loaded
            # (if extended analysis was loaded, its summary is already shown above)
            if not (
                "extended_analysis" in loaded_data and loaded_data["extended_analysis"]
            ):
                self.fisher_results_text.setPlainText("\n".join(summary_lines))

            # Enable export button if we have extended results
            if (
                hasattr(self, "fisher_analysis_results")
                and self.fisher_analysis_results
            ):
                if hasattr(self, "btn_export_fisher"):
                    self.btn_export_fisher.setEnabled(True)

        except Exception as e:
            self.fisher_results_text.setPlainText(
                f"ERROR loading results from HDF5:\n\n{str(e)}"
            )
            self._log_message(f"ERROR loading results from HDF5: {e}")
            import traceback

            traceback.print_exc()

    def _save_results_to_hdf5(self):
        """Save ALL analysis results to HDF5 file (core + extended analysis)."""
        from qtpy.QtWidgets import QFileDialog
        import os
        from ._results_io import save_comprehensive_results

        # Check if we have results to save
        if not hasattr(self, "merged_results") or not self.merged_results:
            self.fisher_results_text.setPlainText(
                "ERROR: No analysis results to save.\n\n"
                "Please run analysis first or load data before saving."
            )
            self._log_message("⚠️ Cannot save: No analysis results available")
            return

        # Open file dialog to select save location
        default_name = "comprehensive_analysis_results.h5"
        if hasattr(self, "file_path") and self.file_path:
            # Suggest name based on current file
            base_name = os.path.splitext(os.path.basename(self.file_path))[0]
            default_name = f"{base_name}_comprehensive_results.h5"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Results to HDF5",
            default_name,
            "HDF5 Files (*.h5 *.hdf5);;All Files (*)",
        )

        if not file_path:
            self._log_message("Save cancelled by user")
            return

        try:
            self._log_message(f"Saving comprehensive results to: {file_path}")

            # Collect core analysis results
            core_results = {
                "merged_results": self.merged_results,        # raw frame differences
                "movement_data_binary": getattr(self, "movement_data", {}),  # binary per bin
                "fraction_data": getattr(self, "fraction_data", {}),
                "quiescence_data": getattr(self, "quiescence_data", {}),
                "sleep_data": getattr(self, "sleep_data", {}),
                "roi_colors": getattr(self, "roi_colors", {}),
                "roi_statistics": getattr(self, "roi_statistics", {}),
                "roi_summary": {},
                "thresholds": {},
            }

            # Add threshold information (core values + extra stats from roi_statistics)
            if hasattr(self, "roi_baseline_means"):
                for roi_id in self.roi_baseline_means:
                    thresh_dict = {
                        "baseline_mean": self.roi_baseline_means.get(roi_id, 0.0),
                        "upper_threshold": self.roi_upper_thresholds.get(roi_id, 0.0),
                        "lower_threshold": self.roi_lower_thresholds.get(roi_id, 0.0),
                    }
                    # Append extra stats if available
                    if hasattr(self, "roi_statistics") and roi_id in self.roi_statistics:
                        stats = self.roi_statistics[roi_id]
                        for key in ("std", "threshold_band", "min_baseline",
                                    "multiplier", "mean"):
                            if key in stats:
                                thresh_dict[key] = float(stats[key])
                    core_results["thresholds"][roi_id] = thresh_dict

            # Collect extended analysis results if available
            extended_results = None
            if (
                hasattr(self, "fisher_analysis_results")
                and self.fisher_analysis_results
            ):
                method_index = getattr(self, "current_fisher_method", 0)
                method_name = (
                    self.fisher_method_combo.currentText()
                    if hasattr(self, "fisher_method_combo")
                    else "Unknown"
                )

                # Format extended results to match expected structure
                # The save function expects keys like 'fisher', 'fft', 'cosinor', etc.
                extended_results = {
                    "method_index": method_index,
                    "method_name": method_name,
                }

                # Map method index to expected key names
                results_data = self.fisher_analysis_results
                if method_index == 0:  # Chi² Periodogram
                    # Filter to ROI data only (integers), format for save
                    fisher_data = {}
                    for roi_id, roi_result in results_data.items():
                        if isinstance(roi_id, int):
                            periodogram = roi_result.get("periodogram", {})
                            fisher_data[roi_id] = {
                                "periods": periodogram.get("periods", []),
                                "z_scores": periodogram.get("z_scores", []),
                                "dominant_period": periodogram.get(
                                    "dominant_period", 0
                                ),
                                "dominant_z_score": periodogram.get(
                                    "dominant_z_score", 0
                                ),
                                "is_significant": periodogram.get(
                                    "is_significant", False
                                ),
                            }
                    extended_results["fisher"] = fisher_data
                    # Save sleep phase results if available
                    if "sleep_phase_results" in results_data:
                        extended_results["sleep_phase_results"] = results_data[
                            "sleep_phase_results"
                        ]

                elif method_index == 1:  # FFT Power Spectrum
                    fft_data = {}
                    for roi_id, roi_result in results_data.items():
                        if isinstance(roi_id, int):
                            fft_data[roi_id] = {
                                "periods": roi_result.get("relevant_periods", []),
                                "power": roi_result.get("relevant_power", []),
                                "dominant_period": roi_result.get("dominant_period", 0),
                                "dominant_power": roi_result.get("dominant_power", 0),
                            }
                    extended_results["fft"] = fft_data
                    if "sleep_phase_results" in results_data:
                        extended_results["sleep_phase_results"] = results_data[
                            "sleep_phase_results"
                        ]

                elif method_index == 2:  # Cosinor Analysis
                    cosinor_data = {}
                    roi_results = results_data.get("roi_results", {})
                    for roi_id, roi_result in roi_results.items():
                        if isinstance(roi_id, int):
                            best = roi_result.get("best_result", {})
                            period = roi_result.get("best_period", 24)
                            cosinor_data[roi_id] = {
                                period: {
                                    "mesor": best.get("mesor", 0),
                                    "amplitude": best.get("amplitude", 0),
                                    "peak_time": best.get("peak_time", 0),
                                    "r_squared": best.get("r_squared", 0),
                                    "p_value": best.get("p_value", 1),
                                    "significant": best.get("significant", False),
                                }
                            }
                    extended_results["cosinor"] = cosinor_data
                    if "sleep_phase_results" in results_data:
                        extended_results["sleep_phase_results"] = results_data[
                            "sleep_phase_results"
                        ]

                elif method_index == 3:  # ROI Similarity
                    extended_results["similarity"] = {
                        "matrix": results_data.get("correlation_matrix", []),
                        "roi_ids": results_data.get("roi_ids", []),
                    }

                elif method_index == 4:  # Coherence
                    extended_results["coherence_matrix"] = results_data.get(
                        "coherence_matrix", []
                    )
                    extended_results["coherence_roi_ids"] = results_data.get(
                        "roi_ids", []
                    )

                elif method_index == 5:  # Phase Clustering
                    phase_data = {}
                    roi_phases = results_data.get("roi_phases", {})
                    for roi_id, phase_info in roi_phases.items():
                        if isinstance(roi_id, int):
                            phase_data[roi_id] = {
                                "phase": phase_info.get("phase_radians", 0),
                                "amplitude": phase_info.get("amplitude", 0),
                            }
                    extended_results["phase"] = phase_data

            # Collect analysis parameters
            analysis_params = {
                "core": {
                    "frame_interval": (
                        self.frame_interval.value()
                        if hasattr(self, "frame_interval")
                        else None
                    ),
                    "bin_size_seconds": (
                        self.bin_size_seconds.value()
                        if hasattr(self, "bin_size_seconds")
                        else None
                    ),
                },
                "extended": {
                    "min_period_hours": (
                        self.fisher_min_period.value()
                        if hasattr(self, "fisher_min_period")
                        else None
                    ),
                    "max_period_hours": (
                        self.fisher_max_period.value()
                        if hasattr(self, "fisher_max_period")
                        else None
                    ),
                    "significance_level": (
                        self.fisher_significance.value()
                        if hasattr(self, "fisher_significance")
                        else None
                    ),
                    "analysis_method": (
                        self.fisher_method_combo.currentText()
                        if hasattr(self, "fisher_method_combo")
                        else None
                    ),
                },
            }

            # Collect metadata
            metadata = {
                "saved_from": "napari-hdf5-activity comprehensive analysis",
                "save_timestamp": datetime.now().isoformat(),
                "n_rois": len(self.merged_results),
                "source_file": (
                    os.path.basename(self.file_path)
                    if hasattr(self, "file_path") and self.file_path
                    else None
                ),
            }

            # Save using comprehensive save function
            success = save_comprehensive_results(
                file_path=file_path,
                core_results=core_results,
                extended_results=extended_results,
                analysis_params=analysis_params,
                metadata=metadata,
            )

            if success:
                # Build summary message
                summary_lines = [
                    "Successfully saved comprehensive results to HDF5!",
                    "",
                    f"File: {file_path}",
                    f"ROIs saved: {len(self.merged_results)}",
                    f"ROI IDs: {sorted(self.merged_results.keys())}",
                    "",
                    "Saved Data:",
                    f"  • Movement data: {len(self.merged_results)} ROIs",
                ]

                if hasattr(self, "fraction_data") and self.fraction_data:
                    summary_lines.append(
                        f"  • Fraction data: {len(self.fraction_data)} ROIs"
                    )

                if hasattr(self, "roi_baseline_means") and self.roi_baseline_means:
                    summary_lines.append(
                        f"  • Thresholds: {len(self.roi_baseline_means)} ROIs"
                    )

                if extended_results:
                    summary_lines.append(
                        f"  • Extended analysis: {extended_results['method_name']}"
                    )

                summary_lines.extend(
                    [
                        "",
                        "You can reload these results anytime using 'Load Results from HDF5'.",
                        "All data is available for post-hoc cycle/period analysis.",
                    ]
                )

                self._log_message(
                    f"✓ Successfully saved comprehensive results for {len(self.merged_results)} ROIs"
                )
                self.fisher_results_text.setPlainText("\n".join(summary_lines))
            else:
                raise Exception("Save operation returned failure status")

        except Exception as e:
            self.fisher_results_text.setPlainText(
                f"ERROR saving results to HDF5:\n\n{str(e)}"
            )
            self._log_message(f"ERROR saving results: {e}")
            import traceback

            traceback.print_exc()

    def run_fisher_analysis(self):
        """Run rhythmic pattern analysis on movement data using selected method."""
        # Check if we have analysis results
        if not hasattr(self, "fraction_data") or not self.fraction_data:
            self.fisher_results_text.setPlainText(
                "ERROR: No analysis results available.\n\n"
                "Please run the main analysis first (Analysis tab) before "
                "attempting rhythmic pattern detection."
            )
            self._log_message(
                "⚠️ Rhythmic pattern analysis requires movement data from main analysis"
            )
            return

        # Get selected method
        method_index = self.fisher_method_combo.currentIndex()
        method_name = self.fisher_method_combo.currentText()

        self._log_message(f"Starting {method_name} analysis...")
        self.fisher_results_text.setPlainText("Running analysis...\n")

        try:
            # Get parameters
            min_period = self.fisher_min_period.value()
            max_period = self.fisher_max_period.value()
            significance = self.fisher_significance.value()
            sampling_interval = self.frame_interval.value()

            # Select data source based on user choice
            data_source_index = self.data_source_combo.currentIndex()
            if data_source_index == 0:
                # Fraction movement (0-1)
                if not hasattr(self, "fraction_data") or not self.fraction_data:
                    self.fisher_results_text.setPlainText(
                        "ERROR: No fraction movement data available.\n\n"
                        "Please run the main analysis first."
                    )
                    self._log_message(
                        "⚠️ No fraction movement data available for rhythmic pattern analysis"
                    )
                    return
                source_data = self.fraction_data
                data_type_name = "Fraction Movement (0-1)"
            elif data_source_index == 1:
                # Raw movement data
                if not hasattr(self, "merged_results") or not self.merged_results:
                    self.fisher_results_text.setPlainText(
                        "ERROR: No movement data available.\n\n"
                        "Please run the main analysis first."
                    )
                    self._log_message(
                        "⚠️ No movement data available for rhythmic pattern analysis"
                    )
                    return
                source_data = self.merged_results
                data_type_name = "Raw Intensity (continuous)"
            else:
                # Normalized movement (min/max per ROI, comparable to literature)
                if not hasattr(self, "merged_results") or not self.merged_results:
                    self.fisher_results_text.setPlainText(
                        "ERROR: No movement data available.\n\n"
                        "Please run the main analysis first."
                    )
                    self._log_message(
                        "⚠️ No movement data available for rhythmic pattern analysis"
                    )
                    return
                from ._calc import bin_and_normalize_movement

                norm_bin_size = (
                    self.analysis_bin_size.value()
                    if hasattr(self, "analysis_bin_size")
                    else self.bin_size_seconds.value()
                )
                source_data = bin_and_normalize_movement(
                    self.merged_results, norm_bin_size
                )
                data_type_name = "Normalized Movement (min/max, 0-1)"

            # Get bin sizes
            original_bin_size = self.bin_size_seconds.value()
            analysis_bin_size = self.analysis_bin_size.value()

            self._log_message(f"  Data source: {data_type_name}")
            self._log_message(f"  Original bin size: {original_bin_size}s")
            self._log_message(f"  Analysis bin size: {analysis_bin_size}s")
            self._log_message(
                f"  Detecting periods: {min_period:.1f} - {max_period:.1f} hours"
            )

            # Apply re-binning if needed
            # (Normalized Movement already binned at analysis_bin_size)
            if data_source_index == 2:
                bin_size = norm_bin_size
                self._log_message(
                    f"  Normalized Movement binned at {bin_size}s"
                )
            elif analysis_bin_size > original_bin_size:
                self._log_message(
                    f"  Re-binning data: {original_bin_size}s → {analysis_bin_size}s"
                )
                source_data = self._rebin_timeseries_data(
                    source_data, analysis_bin_size, original_bin_size
                )
                bin_size = analysis_bin_size
            else:
                bin_size = original_bin_size

            # Check if cycle/time range selection is enabled
            analysis_data = source_data
            if (
                hasattr(self, "enable_cycle_selection")
                and self.enable_cycle_selection.isChecked()
            ):
                start_hours = self.cycle_start_time.value()
                end_hours = self.cycle_end_time.value()

                # Convert hours to seconds for filtering
                start_time = start_hours * 3600.0
                end_time = end_hours * 3600.0

                # Filter data to time range
                filtered_data = {}
                for roi_id, data in source_data.items():
                    filtered = [
                        (t, v) for t, v in data
                        if start_time <= t <= end_time
                    ]
                    if filtered:
                        filtered_data[roi_id] = filtered
                analysis_data = filtered_data

                self._log_message(
                    f"  ✓ Time range filter applied: {start_hours:.1f}h - {end_hours:.1f}h"
                )
                self._log_message(
                    f"  Analyzing {len(analysis_data)} ROIs in selected time window"
                )
            else:
                self._log_message("  Using full recording for analysis")

            # Route to appropriate analysis method
            if method_index == 0:  # Chi² Periodogram
                results, summary = self._run_fisher_method(
                    min_period,
                    max_period,
                    significance,
                    sampling_interval,
                    bin_size,
                    analysis_data,
                )
            elif method_index == 1:  # FFT Power Spectrum
                results, summary = self._run_fft_method(
                    min_period,
                    max_period,
                    significance,
                    sampling_interval,
                    bin_size,
                    analysis_data,
                )
            elif method_index == 2:  # Cosinor Analysis
                self._log_message(
                    f"  Using {data_type_name} (with rebinning/time range if set)"
                )
                results, summary = self._run_cosinor_method(
                    min_period,
                    max_period,
                    significance,
                    sampling_interval,
                    bin_size,
                    analysis_data,
                )
            elif method_index == 3:  # ROI Similarity Matrix
                results, summary = self._run_similarity_method(
                    sampling_interval, bin_size, analysis_data
                )
            elif method_index == 4:  # Coherence Analysis
                results, summary = self._run_coherence_method(
                    sampling_interval, bin_size, analysis_data
                )
            elif method_index == 5:  # Phase Clustering
                results, summary = self._run_phase_clustering_method(
                    sampling_interval, bin_size, analysis_data
                )
            else:
                raise ValueError(f"Unknown method index: {method_index}")

            # Calculate sleep phase if enabled (for methods that support it)
            # Note: Cosinor (method_index=2) excluded because it requires continuous
            # sinusoidal data, not binary sleep states
            sleep_results = None
            sleep_summary = ""
            sleep_source_name = ""
            if self.chk_calculate_sleep_phase.isChecked() and method_index in [
                0,
                1,
            ]:  # Fisher, FFT only (not Cosinor)
                # Determine sleep data source from user selection
                sleep_source_index = (
                    self.sleep_source_combo.currentIndex()
                    if hasattr(self, "sleep_source_combo")
                    else 1
                )

                sleep_analysis_data = None
                if sleep_source_index == 0:
                    # Quiescence data (same temporal resolution as activity)
                    if hasattr(self, "quiescence_data") and self.quiescence_data:
                        total_points = sum(
                            sum(1 for t, v in data if v == 1)
                            for data in self.quiescence_data.values()
                        )
                        sleep_source_name = "quiescence_data (binary rest state)"
                        self._log_message(
                            f"  Analyzing sleep rhythms from quiescence_data "
                            f"({total_points} quiescence points, same resolution as activity)..."
                        )
                        sleep_analysis_data = self.quiescence_data
                    else:
                        self._log_message(
                            "  WARNING: No quiescence data available. "
                            "Run main analysis first."
                        )
                else:
                    # Sleep data (≥8min sustained quiescence)
                    if hasattr(self, "sleep_data") and self.sleep_data:
                        sleep_threshold = self.sleep_threshold_minutes.value()
                        total_points = sum(
                            sum(1 for t, v in data if v == 1)
                            for data in self.sleep_data.values()
                        )
                        sleep_source_name = (
                            f"sleep_data (≥{sleep_threshold}min sustained quiescence)"
                        )
                        self._log_message(
                            f"  Analyzing sleep rhythms from sleep_data "
                            f"(≥{sleep_threshold}min, {total_points} sleep points, "
                            f"low-pass filtered)..."
                        )
                        sleep_analysis_data = self.sleep_data
                    elif hasattr(self, "quiescence_data") and self.quiescence_data:
                        # Fallback to quiescence data
                        total_points = sum(
                            sum(1 for t, v in data if v == 1)
                            for data in self.quiescence_data.values()
                        )
                        sleep_source_name = "quiescence_data (fallback)"
                        self._log_message(
                            f"  No sleep_data available, falling back to quiescence_data "
                            f"({total_points} quiescence points)..."
                        )
                        sleep_analysis_data = self.quiescence_data
                    else:
                        self._log_message(
                            "  WARNING: No sleep/quiescence data available. "
                            "Run main analysis first."
                        )

                if sleep_analysis_data:
                    # Apply same time range filter if enabled
                    if (
                        hasattr(self, "enable_cycle_selection")
                        and self.enable_cycle_selection.isChecked()
                    ):
                        start_hours = self.cycle_start_time.value()
                        end_hours = self.cycle_end_time.value()
                        start_time = start_hours * 3600.0
                        end_time = end_hours * 3600.0
                        filtered_sleep = {}
                        for roi_id, data in sleep_analysis_data.items():
                            filtered = [
                                (t, v) for t, v in data
                                if start_time <= t <= end_time
                            ]
                            if filtered:
                                filtered_sleep[roi_id] = filtered
                        sleep_analysis_data = filtered_sleep

                    # Run the same analysis on sleep data
                    if method_index == 0:  # Chi² Periodogram
                        sleep_results, sleep_summary = self._run_fisher_method(
                            min_period,
                            max_period,
                            significance,
                            sampling_interval,
                            bin_size,
                            sleep_analysis_data,
                        )
                    elif method_index == 1:  # FFT Power Spectrum
                        sleep_results, sleep_summary = self._run_fft_method(
                            min_period,
                            max_period,
                            significance,
                            sampling_interval,
                            bin_size,
                            sleep_analysis_data,
                        )

                    # Log sleep analysis results
                    if sleep_results:
                        # Handle both structures: direct ROI keys (Fisher/FFT) or nested roi_results (Cosinor)
                        if "roi_results" in sleep_results:
                            sleep_roi_count = len(
                                [
                                    k
                                    for k in sleep_results["roi_results"].keys()
                                    if isinstance(k, int)
                                ]
                            )
                        else:
                            sleep_roi_count = len(
                                [k for k in sleep_results.keys() if isinstance(k, int)]
                            )
                        self._log_message(
                            f"  ✓ Sleep rhythm analysis complete: {sleep_roi_count} ROIs"
                        )
                    else:
                        self._log_message(
                            "  ⚠️ Sleep rhythm analysis returned no results"
                        )

                    # Combine results
                    results["sleep_phase_results"] = sleep_results
                    summary = self._combine_activity_sleep_summary(
                        summary, sleep_summary, results, sleep_results,
                        sleep_source_name=sleep_source_name,
                    )

            # Store results + data context for plot functions
            self.fisher_analysis_results = results
            self.current_fisher_method = method_index
            self.fisher_data_type_name = data_type_name    # e.g. "Fraction Movement (0-1)"
            self.fisher_analysis_data = analysis_data      # the actual data that was analysed

            # Warn if period range was capped by recording duration
            if method_index in [0, 1]:
                for roi_id, roi_result in results.items():
                    if not isinstance(roi_id, int):
                        continue
                    peri = roi_result.get("periodogram", roi_result)
                    if peri.get("period_capped", False):
                        actual_max = peri.get("actual_max_period", max_period)
                        self._log_message(
                            f"  ⚠️ Max period capped at {actual_max:.1f}h "
                            f"(recording duration / 2). Your setting of {max_period:.1f}h "
                            f"requires a recording ≥ {max_period * 2:.0f}h."
                        )
                        break  # one warning is enough

            # Display results
            self.fisher_results_text.setPlainText(summary)

            # Create and display plot
            self._create_circadian_plot(results, method_index)

            # Enable export button
            self.btn_export_fisher.setEnabled(True)

            self._log_message(f"✓ Rhythmic pattern analysis complete ({method_name})")

        except Exception as e:
            error_msg = f"ERROR during rhythmic pattern analysis:\n\n{str(e)}\n\nPlease check the console for details."
            self.fisher_results_text.setPlainText(error_msg)
            self._log_message(f"❌ Rhythmic pattern analysis failed: {e}")
            import traceback

            traceback.print_exc()

    def _run_fisher_method(
        self,
        min_period,
        max_period,
        significance,
        sampling_interval,
        bin_size,
        fraction_data=None,
    ):
        """Run Chi² Periodogram analysis."""
        from ._fisher_analysis import (
            analyze_roi_circadian_patterns,
            generate_circadian_summary,
        )

        # Use provided fraction_data or default to self.fraction_data
        data_to_analyze = (
            fraction_data if fraction_data is not None else self.fraction_data
        )

        results = analyze_roi_circadian_patterns(
            data_to_analyze,  # Use fraction_data (proportion 0-1)
            sampling_interval=sampling_interval,
            min_period_hours=min_period,
            max_period_hours=max_period,
            significance_level=significance,
            phase_threshold=0.5,  # Keep for backward compatibility, unused
            bin_size_seconds=bin_size,
        )

        summary = generate_circadian_summary(results)
        return results, summary

    def _create_inverted_activity_data(self, activity_data):
        """
        Create inverted activity data for sleep phase analysis.
        For fraction movement (0-1): inverted = 1 - value
        For raw movement: inverted = max - value
        """
        inverted_data = {}

        for roi, data_list in activity_data.items():
            if not data_list:
                inverted_data[roi] = []
                continue

            # Extract values to determine if this is fraction (0-1) or raw data
            values = [v for t, v in data_list]
            max_val = max(values) if values else 1.0

            inverted_list = []
            for timestamp, value in data_list:
                if max_val <= 1.0:
                    # Fraction movement (0-1): invert as 1 - value
                    inverted_value = 1.0 - value
                else:
                    # Raw movement: invert as max - value
                    inverted_value = max_val - value
                inverted_list.append((timestamp, inverted_value))

            inverted_data[roi] = inverted_list

        return inverted_data

    def _combine_activity_sleep_summary(
        self, activity_summary, sleep_summary, activity_results, sleep_results,
        sleep_source_name="",
    ):
        """Combine activity and sleep phase summaries into one comprehensive report."""
        combined = []

        # Get data source name from combo box
        data_source_index = self.data_source_combo.currentIndex()
        if data_source_index == 0:
            data_source_name = "Fraction Movement"
        elif data_source_index == 1:
            data_source_name = "Raw Intensity"
        else:
            data_source_name = "Normalized Movement (min/max)"

        combined.append("=" * 60)
        combined.append("ACTIVITY & SLEEP RHYTHM ANALYSIS RESULTS")
        combined.append("=" * 60)
        combined.append("")

        # Activity Phase Section
        combined.append("─" * 40)
        combined.append(f"ACTIVITY RHYTHMS (from {data_source_name})")
        combined.append("─" * 40)

        # Extract key info from activity results
        # Handle both structures: direct ROI keys (Fisher) or nested "roi_results" (Cosinor)
        if "roi_results" in activity_results:
            act_roi_data = activity_results["roi_results"]
        else:
            # Fisher results have ROI IDs as direct keys
            act_roi_data = {
                k: v for k, v in activity_results.items() if isinstance(k, int)
            }

        if act_roi_data:
            roi_items = {k: v for k, v in act_roi_data.items() if isinstance(k, int)}
            for roi_id, roi_data in sorted(roi_items.items()):
                # Handle nested structure for different analysis methods
                if "periodogram" in roi_data:
                    # Fisher/FFT structure
                    periodogram = roi_data.get("periodogram", {})
                    period = periodogram.get("dominant_period", "N/A")
                    significant = periodogram.get("is_significant", False)
                    z_score = periodogram.get(
                        "dominant_z_score", periodogram.get("max_power", "N/A")
                    )
                    peak_time = (
                        "N/A (use Cosinor)"  # Fisher doesn't calculate peak time
                    )
                elif "best_result" in roi_data:
                    # Cosinor structure: best_period at top, details in best_result
                    period = roi_data.get("best_period", "N/A")
                    best_result = roi_data.get("best_result", {})
                    peak_time = best_result.get("peak_time", "N/A")
                    significant = best_result.get("significant", False)
                    z_score = None
                elif "dominant_period" in roi_data:
                    # FFT direct structure
                    period = roi_data.get("dominant_period", "N/A")
                    peak_time = "N/A"
                    significant = roi_data.get("is_significant", False)
                    z_score = roi_data.get("dominant_power", None)
                else:
                    continue

                sig_str = "✓" if significant else "✗"

                if isinstance(period, (int, float)):
                    period_str = f"{period:.1f}h"
                else:
                    period_str = str(period)

                if isinstance(peak_time, (int, float)):
                    peak_time_str = f"{peak_time:.1f}h"
                else:
                    peak_time_str = str(peak_time)

                if z_score is not None and isinstance(z_score, (int, float)):
                    combined.append(
                        f"  ROI {roi_id}: Period={period_str}, Z={z_score:.1f} {sig_str}"
                    )
                else:
                    combined.append(
                        f"  ROI {roi_id}: Period={period_str}, Peak Time={peak_time_str} {sig_str}"
                    )

        combined.append("")

        # Sleep Phase Section
        if not sleep_source_name:
            sleep_threshold = (
                self.sleep_threshold_minutes.value()
                if hasattr(self, "sleep_threshold_minutes")
                else 8
            )
            sleep_source_name = f"sleep_data (≥{sleep_threshold}min quiescence)"
        combined.append("─" * 40)
        combined.append(f"SLEEP RHYTHMS (from {sleep_source_name})")
        combined.append("─" * 40)

        # Handle both structures for sleep results
        if sleep_results:
            if "roi_results" in sleep_results:
                slp_roi_data = sleep_results["roi_results"]
            else:
                slp_roi_data = {
                    k: v for k, v in sleep_results.items() if isinstance(k, int)
                }

            sleep_roi_items = {
                k: v for k, v in slp_roi_data.items() if isinstance(k, int)
            }
            for roi_id, roi_data in sorted(sleep_roi_items.items()):
                # Handle nested structure for different analysis methods
                if "periodogram" in roi_data:
                    # Fisher/FFT structure
                    periodogram = roi_data.get("periodogram", {})
                    period = periodogram.get("dominant_period", "N/A")
                    significant = periodogram.get("is_significant", False)
                    z_score = periodogram.get(
                        "dominant_z_score", periodogram.get("max_power", "N/A")
                    )
                    sleep_phase = "N/A"
                elif "best_result" in roi_data:
                    # Cosinor structure: best_period at top, details in best_result
                    period = roi_data.get("best_period", "N/A")
                    best_result = roi_data.get("best_result", {})
                    sleep_phase = best_result.get("peak_time", "N/A")
                    significant = best_result.get("significant", False)
                    z_score = None
                elif "dominant_period" in roi_data:
                    # FFT direct structure
                    period = roi_data.get("dominant_period", "N/A")
                    sleep_phase = "N/A"
                    significant = roi_data.get("is_significant", False)
                    z_score = roi_data.get("dominant_power", None)
                else:
                    continue

                sig_str = "✓" if significant else "✗"

                if isinstance(period, (int, float)):
                    period_str = f"{period:.1f}h"
                else:
                    period_str = str(period)

                if z_score is not None and isinstance(z_score, (int, float)):
                    combined.append(
                        f"  ROI {roi_id}: Period={period_str}, Z={z_score:.1f} {sig_str}"
                    )
                else:
                    if isinstance(sleep_phase, (int, float)):
                        sleep_str = f"{sleep_phase:.1f}h"
                    else:
                        sleep_str = str(sleep_phase)
                    combined.append(
                        f"  ROI {roi_id}: Period={period_str}, Sleep Phase={sleep_str} {sig_str}"
                    )

        combined.append("")

        # Summary comparison
        combined.append("─" * 40)
        combined.append("PERIOD COMPARISON (Activity vs Sleep)")
        combined.append("─" * 40)

        # Use previously extracted ROI data
        if act_roi_data and sleep_results:
            # Get sleep ROI data (handle both structures)
            if "roi_results" in sleep_results:
                slp_compare_data = sleep_results["roi_results"]
            else:
                slp_compare_data = {
                    k: v for k, v in sleep_results.items() if isinstance(k, int)
                }

            act_roi_keys = [k for k in act_roi_data.keys() if isinstance(k, int)]
            for roi_id in sorted(act_roi_keys):
                act_data = act_roi_data.get(roi_id, {})
                slp_data = slp_compare_data.get(roi_id, {})

                # Handle nested structure for Fisher/FFT
                if "periodogram" in act_data:
                    act_period = act_data.get("periodogram", {}).get("dominant_period")
                else:
                    act_period = act_data.get("dominant_period")

                if "periodogram" in slp_data:
                    slp_period = slp_data.get("periodogram", {}).get("dominant_period")
                else:
                    slp_period = slp_data.get("dominant_period")

                # Also try to get peak_time for Cosinor
                act_phase = act_data.get("peak_time")
                slp_phase = slp_data.get("peak_time")

                if act_period is not None and slp_period is not None:
                    if act_phase is not None and slp_phase is not None:
                        # Cosinor: show fitted peak time comparison
                        diff = abs(act_phase - slp_phase)
                        if diff > 12:
                            diff = 24 - diff
                        combined.append(
                            f"  ROI {roi_id}: Act Peak={act_phase:.1f}h, "
                            f"Sleep Peak={slp_phase:.1f}h, Δ={diff:.1f}h"
                        )
                    else:
                        # Fisher/FFT: show period comparison
                        combined.append(
                            f"  ROI {roi_id}: Act Period={act_period:.1f}h, "
                            f"Sleep Period={slp_period:.1f}h"
                        )

        combined.append("")
        combined.append("=" * 60)
        combined.append("Legend: ✓ = significant rhythm, ✗ = not significant")
        combined.append("=" * 60)

        return "\n".join(combined)

    def _run_fft_method(
        self,
        min_period,
        max_period,
        significance,
        sampling_interval,
        bin_size,
        fraction_data=None,
    ):
        """Run FFT Power Spectrum analysis with significance testing."""
        from ._circadian_fft import analyze_roi_fft_patterns, generate_fft_summary

        # Use provided fraction_data or default to self.fraction_data
        data_to_analyze = (
            fraction_data if fraction_data is not None else self.fraction_data
        )

        results = analyze_roi_fft_patterns(
            data_to_analyze,  # Use fraction_data (proportion 0-1)
            sampling_interval=sampling_interval,
            min_period_hours=min_period,
            max_period_hours=max_period,
            bin_size_seconds=bin_size,
            window="hann",
            significance_level=significance,
            n_permutations=1000,  # Use 1000 permutations for good statistical power
        )

        summary = generate_fft_summary(results)
        return results, summary

    def _run_cosinor_method(
        self,
        min_period,
        max_period,
        significance,
        sampling_interval,
        bin_size,
        data=None,
    ):
        """Run Cosinor analysis on any time-series data (fraction movement or raw intensity)."""
        from ._cosinor_analysis import (
            multi_period_cosinor,
            population_cosinor,
        )

        # Prepare test periods - test common periods in the specified range
        test_periods = []
        if min_period <= 12 and max_period >= 12:
            test_periods.append(12.0)  # 12h ultradian
        if min_period <= 24 and max_period >= 24:
            test_periods.append(24.0)  # 24h circadian
        if min_period <= 30 and max_period >= 30:
            test_periods.append(30.0)  # 30h infradian

        # Also add min, max, and midpoint
        midpoint = (min_period + max_period) / 2
        for p in [min_period, midpoint, max_period]:
            if p not in test_periods:
                test_periods.append(p)

        test_periods = sorted(test_periods)

        # Use provided data or fall back to fraction_data
        data_to_analyze = (
            data if data is not None else self.fraction_data
        )

        # Compute recording duration for cycle-count warnings / plot annotation
        recording_duration_h = 0.0
        for _dl in data_to_analyze.values():
            if _dl:
                _ts = [t for t, _ in _dl]
                recording_duration_h = max(
                    recording_duration_h, (max(_ts) - min(_ts)) / 3600.0
                )

        # Warn for every test period that yields < 2 complete cycles
        for _tp in test_periods:
            _n_cyc = recording_duration_h / _tp if _tp > 0 else 0
            if _n_cyc < 2.0:
                self._log_message(
                    f"⚠️ Cosinor: period {_tp:.1f} h → only {_n_cyc:.1f} complete cycle(s) "
                    f"in {recording_duration_h:.1f} h recording — result unreliable (need ≥ 2 cycles)"
                )

        # Analyze each ROI with cosinor
        roi_results = {}
        population_time_series = []
        population_timestamps = []

        for roi_id, data_list in data_to_analyze.items():
            if not data_list:
                continue

            # Use actual timestamps from the data — critical for correct period detection.
            # Passing timestamps=None would cause the cosinor to invent synthetic timestamps
            # using bin_size, which is wrong for raw intensity data (5s intervals ≠ 60s bins).
            timestamps_s = np.array([t for t, _ in data_list])  # seconds
            values = np.array([v for _, v in data_list])

            multi_result = multi_period_cosinor(
                time_series=values,
                timestamps=timestamps_s,
                test_periods=test_periods,
                sampling_interval=bin_size,
                alpha=significance,
            )

            roi_results[roi_id] = multi_result
            population_time_series.append(values)
            population_timestamps.append(timestamps_s)

        # Population-level cosinor for best period across all ROIs
        if population_time_series:
            # Find the most common best period
            best_periods = [
                r["best_period"] for r in roi_results.values() if "best_period" in r
            ]
            if best_periods:
                # Use median of best periods as population period
                population_period = float(np.median(best_periods))

                pop_result = population_cosinor(
                    time_series_list=population_time_series,
                    timestamps_list=population_timestamps,
                    period_hours=population_period,
                    sampling_interval=bin_size,
                    alpha=significance,
                )
            else:
                pop_result = {"error": "No valid ROI results"}
        else:
            pop_result = {"error": "No time series data"}

        # Combine results
        results = {
            "roi_results": roi_results,
            "population_result": pop_result,
            "test_periods": test_periods,
            "sampling_interval": sampling_interval,
            "recording_duration_h": recording_duration_h,
        }

        # Generate summary
        summary = self._generate_cosinor_summary(results)

        return results, summary

    def _generate_cosinor_summary(self, results):
        """Generate text summary for cosinor analysis results."""
        data_type_name = getattr(self, "fisher_data_type_name", "Fraction Movement (0-1)")
        lines = [
            "=" * 70,
            "COSINOR ANALYSIS - Circadian Rhythm Quantification",
            "=" * 70,
            "",
            f"Data source: {data_type_name}",
            "",
            "Cosinor analysis fits a cosine curve to the signal:",
            "  y(t) = MESOR + Amplitude × cos(2πt/τ + φ)",
            "",
            "Parameters:",
            "  • MESOR: Midline Estimating Statistic of Rhythm (mean level)",
            "  • Amplitude: Half the difference between peak and trough",
            "  • φ (phase angle): phase offset of the fitted cosine (radians)",
            "  • Peak Time: time from recording start to first fitted cosine peak",
            "    NOTE: not the biological acrophase without a known ZT reference",
            "  • Period (τ): Duration of one complete cycle",
            "",
        ]

        roi_results = results.get("roi_results", {})
        test_periods = results.get("test_periods", [])

        # Diagnostic checks for period range issues
        warnings = []
        boundary_count = 0
        best_periods = []

        for roi_id, roi_data in roi_results.items():
            best_result = roi_data.get("best_result", {})
            if "error" in best_result or not best_result.get("significant", False):
                continue

            best_period = best_result.get("period", 0)
            if best_period > 0:
                best_periods.append(best_period)

            # Check if best period is at boundary of test_periods
            if len(test_periods) > 0:
                min_test = min(test_periods)
                max_test = max(test_periods)

                # Check if best period matches boundary (exactly or very close)
                if best_period == min_test or best_period == max_test:
                    boundary_count += 1

        # Generate warnings
        n_significant = len(best_periods)
        if (
            boundary_count > n_significant * 0.3 and n_significant > 0
        ):  # More than 30% at boundaries
            warnings.append(
                f"⚠️  WARNING: {boundary_count}/{n_significant} ROIs have best-fit periods at test range boundaries.\n"
                f"   This suggests the period range may be too narrow.\n"
                f"   Consider expanding the test period range to capture true rhythms."
            )

        # Check if best periods cluster at extremes
        if len(best_periods) >= 3:
            best_periods_array = np.array(best_periods)
            min_best = best_periods_array.min()
            max_best = best_periods_array.max()

            if len(test_periods) > 0:
                test_min = min(test_periods)
                test_max = max(test_periods)

                # Suggest expanding range if periods cluster near boundaries
                if max_best > 12.0 and test_max < 24.0:
                    warnings.append(
                        f"ℹ️  INFO: Some best-fit periods exceed 12h (max: {max_best:.1f}h).\n"
                        f"   Consider testing 24h period for circadian analysis."
                    )
                elif min_best < 2.0 and test_min > 1.0:
                    warnings.append(
                        f"ℹ️  INFO: Some best-fit periods below 2h (min: {min_best:.1f}h).\n"
                        f"   Consider testing shorter periods (0.5-1h) for ultradian analysis."
                    )

        if warnings:
            lines.extend(warnings)
            lines.append("")

        lines.extend(
            [
                "=" * 70,
                "INDIVIDUAL ROI RESULTS",
                "=" * 70,
                "",
            ]
        )

        for roi_id in sorted(roi_results.keys()):
            roi_data = roi_results[roi_id]
            best_result = roi_data.get("best_result", {})

            if "error" in best_result:
                lines.extend([f"ROI {roi_id}:", f"  Error: {best_result['error']}", ""])
                continue

            # Check if period is at boundary
            best_period = best_result.get("period", 0)
            boundary_marker = ""
            if len(test_periods) > 0:
                min_test = min(test_periods)
                max_test = max(test_periods)
                if best_period == max_test:
                    boundary_marker = f" ⚠️ (at upper test boundary {max_test:.2f}h)"
                elif best_period == min_test:
                    boundary_marker = f" ⚠️ (at lower test boundary {min_test:.2f}h)"

            lines.extend(
                [
                    f"ROI {roi_id}:",
                    f"  Best-fit period: {best_period:.2f} hours{boundary_marker}",
                    f"  MESOR (mean level): {best_result.get('mesor', 0):.4f}",
                    f"  Amplitude: {best_result.get('amplitude', 0):.4f}",
                    f"  Peak Time (fitted cosine peak): {best_result.get('peak_time', 0):.2f} h from start",
                    f"  R²: {best_result.get('r_squared', 0):.3f}",
                    f"  p-value: {_fmt_p(best_result.get('p_value', 1))} {'***' if best_result.get('p_value', 1) < 0.001 else '**' if best_result.get('p_value', 1) < 0.01 else '*' if best_result.get('p_value', 1) < 0.05 else 'ns'}",
                    f"  Significant: {'YES' if best_result.get('significant', False) else 'NO'}",
                ]
            )

            # Add confidence intervals if available
            if "ci_amplitude" in best_result:
                ci_amp = best_result["ci_amplitude"]
                ci_pt = best_result.get("ci_peak_time", (float("nan"), float("nan")))
                lines.extend(
                    [
                        f"  95% CI Amplitude: [{ci_amp[0]:.4f}, {ci_amp[1]:.4f}]",
                        f"  95% CI Peak Time: [{ci_pt[0]:.2f}h, {ci_pt[1]:.2f}h]",
                    ]
                )

            lines.append("")

            # Show all tested periods
            all_results = roi_data.get("all_results", [])
            if all_results:
                lines.append("  Tested periods:")
                for res in all_results:
                    sig_marker = "✓" if res.get("significant", False) else " "
                    lines.append(
                        f"    [{sig_marker}] {res.get('test_period', 0):.1f}h: "
                        f"R²={res.get('r_squared', 0):.3f}, "
                        f"Amp={res.get('amplitude', 0):.4f}, "
                        f"p={_fmt_p(res.get('p_value', 1))}"
                    )
                lines.append("")

        # Population-level results
        pop_result = results.get("population_result", {})
        if "error" not in pop_result:
            lines.extend(
                [
                    "=" * 70,
                    "POPULATION-LEVEL COSINOR",
                    "=" * 70,
                    "",
                    f"Population MESOR: {pop_result.get('population_mesor', 0):.4f}",
                    f"Population Amplitude: {pop_result.get('population_amplitude', 0):.4f}",
                    f"Population Peak Time (circular mean): {pop_result.get('population_peak_time', 0):.2f} h from start",
                    f"Test period: {pop_result.get('period', 0):.2f} hours",
                    f"p-value: {_fmt_p(pop_result.get('p_value', 1))}",
                    f"Significant rhythm: {'YES' if pop_result.get('significant', False) else 'NO'}",
                    "",
                    f"Individual ROIs analyzed: {pop_result.get('n_individuals', 0)}",
                    f"ROIs with significant rhythm: {pop_result.get('n_significant', 0)} "
                    f"({pop_result.get('proportion_significant', 0)*100:.1f}%)",
                    "",
                ]
            )

        lines.extend(
            [
                "=" * 70,
                "INTERPRETATION GUIDE",
                "=" * 70,
                "",
                "MESOR:",
                "  The rhythm-adjusted mean activity level (baseline activity)",
                "",
                "Amplitude:",
                "  Larger amplitude = stronger rhythm (higher variation from mean)",
                "  Small amplitude suggests weak or no rhythmic pattern",
                "",
                "Acrophase:",
                "  Time of peak activity in the cycle",
                "  For 24h rhythm: 0h=start of recording, 12h=halfway through",
                "",
                "Significance (p-value):",
                "  p < 0.05: Significant rhythm detected",
                "  p < 0.01: Highly significant rhythm (**)",
                "  p < 0.001: Very highly significant rhythm (***)",
                "",
                "R² (goodness of fit):",
                "  >0.3: Strong rhythmic pattern",
                "  0.1-0.3: Moderate rhythm",
                "  <0.1: Weak or no rhythm",
                "",
            ]
        )

        return "\n".join(lines)

    def _run_similarity_method(self, sampling_interval, bin_size, fraction_data=None):
        """Run ROI Similarity analysis."""
        from ._circadian_similarity import (
            calculate_roi_correlation_matrix,
            hierarchical_clustering,
            generate_similarity_summary,
        )

        # Use provided fraction_data or default to self.fraction_data
        data_to_analyze = (
            fraction_data if fraction_data is not None else self.fraction_data
        )

        # Use half of max period as max lag (adaptive to period range)
        max_lag = self.fisher_max_period.value() / 2

        # Get significance level from UI
        significance = self.fisher_significance.value()

        correlation_results = calculate_roi_correlation_matrix(
            data_to_analyze,  # Use fraction_data (proportion 0-1)
            sampling_interval=sampling_interval,
            bin_size_seconds=bin_size,
            max_lag_hours=max_lag,
            significance_level=significance,
        )

        clustering_results = hierarchical_clustering(
            correlation_results["correlation_matrix"],
            correlation_results["roi_ids"],
            method="average",
        )

        correlation_results["clustering"] = clustering_results
        summary = generate_similarity_summary(correlation_results, clustering_results)

        return correlation_results, summary

    def _run_coherence_method(self, sampling_interval, bin_size, fraction_data=None):
        """Run Coherence analysis."""
        from ._circadian_coherence import (
            calculate_coherence_matrix,
            generate_coherence_summary,
        )

        # Use provided fraction_data or default to self.fraction_data
        data_to_analyze = (
            fraction_data if fraction_data is not None else self.fraction_data
        )

        # Use midpoint of period range from GUI instead of hardcoded 24h
        midpoint_period = (
            self.fisher_min_period.value() + self.fisher_max_period.value()
        ) / 2

        # Get significance level from UI
        significance = self.fisher_significance.value()

        results = calculate_coherence_matrix(
            data_to_analyze,  # Use fraction_data (proportion 0-1)
            sampling_interval=sampling_interval,
            bin_size_seconds=bin_size,
            target_period_hours=midpoint_period,
            significance_level=significance,
        )

        summary = generate_coherence_summary(results)
        return results, summary

    def _run_phase_clustering_method(
        self, sampling_interval, bin_size, fraction_data=None
    ):
        """Run Phase Clustering analysis."""
        from ._circadian_coherence import detect_phase_clusters

        # Use provided fraction_data or default to self.fraction_data
        data_to_analyze = (
            fraction_data if fraction_data is not None else self.fraction_data
        )

        # Use midpoint of period range from GUI instead of hardcoded 24h
        midpoint_period = (
            self.fisher_min_period.value() + self.fisher_max_period.value()
        ) / 2

        results = detect_phase_clusters(
            data_to_analyze,  # Use fraction_data (proportion 0-1)
            dominant_period_hours=midpoint_period,
            sampling_interval=sampling_interval,
            bin_size_seconds=bin_size,
        )

        # Generate summary
        summary_lines = [
            "=" * 60,
            "PHASE CLUSTERING ANALYSIS",
            "=" * 60,
            "",
            f"Total ROIs analyzed: {results['n_rois']}",
            f"Dominant period: {results['dominant_period_hours']:.1f} hours",
            "",
            "ROI Phase Clusters:",
            "",
        ]

        for cluster_name, roi_list in results["phase_clusters"].items():
            if roi_list:
                summary_lines.append(
                    f"  {cluster_name.replace('_', ' ').title()}: {len(roi_list)} ROIs"
                )
                summary_lines.append(f"    ROIs: {', '.join(map(str, roi_list))}")
                summary_lines.append("")

        summary_lines.append("")
        summary_lines.append("Individual ROI Phases:")
        summary_lines.append("")

        for roi_id, phase_info in sorted(results["roi_phases"].items()):
            summary_lines.append(
                f"  ROI {roi_id}: Peak activity at {phase_info['phase_hours']:.1f}h "
                f"(amplitude: {phase_info['amplitude']:.2f})"
            )

        summary_lines.append("")
        summary_lines.append("=" * 60)

        summary = "\n".join(summary_lines)
        return results, summary

    def _create_circadian_plot(self, results: Dict, method_index: int):
        """Create and display plot based on selected analysis method."""
        if method_index == 0:  # Chi² Periodogram
            self._create_fisher_plot(results)
        elif method_index == 1:  # FFT Power Spectrum
            self._create_fft_plot(results)
        elif method_index == 2:  # Cosinor Analysis
            self._create_cosinor_plot(results)
        elif method_index == 3:  # ROI Similarity Matrix
            self._create_similarity_plot(results)
        elif method_index == 4:  # Coherence Analysis
            self._create_coherence_plot(results)
        elif method_index == 5:  # Phase Clustering
            self._create_phase_cluster_plot(results)

    def _create_fft_plot(self, fft_results: Dict[int, Dict]):
        """Create FFT power spectrum plots."""
        try:
            import matplotlib.pyplot as plt
            from qtpy.QtGui import QPixmap
            import io

            if hasattr(self, "fisher_plot_figure") and self.fisher_plot_figure:
                plt.close(self.fisher_plot_figure)

            # Get data source for plot title
            data_source_index = self.data_source_combo.currentIndex()
            data_source = (
                "Fraction Movement" if data_source_index == 0 else "Raw Intensity"
            )

            # Check if sleep phase results are available
            sleep_results = fft_results.get("sleep_phase_results", None)
            has_sleep = sleep_results is not None and len(sleep_results) > 0

            # Filter to only ROI keys (integers)
            roi_only_results = {
                k: v for k, v in fft_results.items() if isinstance(k, int)
            }
            n_rois = len(roi_only_results)
            n_cols = min(3, n_rois)
            n_rows_per_section = (n_rois + n_cols - 1) // n_cols

            # If we have sleep data, double the rows
            if has_sleep:
                total_rows = n_rows_per_section * 2
                fig_height = 3.5 * total_rows + 1.5
            else:
                total_rows = n_rows_per_section
                fig_height = 3.5 * total_rows + 0.5

            fig_width = 4 * n_cols
            fig, axes = plt.subplots(
                total_rows, n_cols, figsize=(fig_width, fig_height), squeeze=False
            )

            # Check if max period was capped for any ROI
            cap_note = ""
            requested_max = self.fisher_max_period.value()
            for _roi_id, _roi_res in roi_only_results.items():
                _periods = _roi_res.get("relevant_periods", [])
                if len(_periods) > 0:
                    _actual_max = max(_periods)
                    if _actual_max < requested_max * 0.95:  # capped by >5%
                        cap_note = f"  (max period capped at {_actual_max:.1f}h = recording / 2)"
                    break

            if has_sleep:
                fig.suptitle(
                    f"FFT Power Spectrum  —  Data source: {data_source}{cap_note}{self._get_recording_start_str()}",
                    fontsize=11,
                    fontweight="bold",
                    y=0.99,
                )
            else:
                fig.suptitle(
                    f"FFT Power Spectrum  —  Activity from {data_source}{cap_note}{self._get_recording_start_str()}",
                    fontsize=11,
                    fontweight="bold",
                    y=0.99,
                )

            # Helper function to plot a section
            def plot_section(results_dict, start_row, section_label):
                roi_items = {
                    k: v for k, v in results_dict.items() if isinstance(k, int)
                }

                for idx, (roi_id, result) in enumerate(sorted(roi_items.items())):
                    row = start_row + (idx // n_cols)
                    col = idx % n_cols
                    ax = axes[row, col]

                    roi_color = (
                        self.roi_colors.get(roi_id, f"C{idx}")
                        if hasattr(self, "roi_colors")
                        else f"C{idx}"
                    )

                    if "error" in result:
                        ax.text(
                            0.5,
                            0.5,
                            f"ROI {roi_id}\n{result['error']}",
                            ha="center",
                            va="center",
                            transform=ax.transAxes,
                        )
                        ax.set_xticks([])
                        ax.set_yticks([])
                    else:
                        periods = result.get("relevant_periods", [])
                        power = result.get("relevant_power", [])

                        ax.plot(periods, power, color=roi_color, linewidth=1.5)
                        ax.set_xlabel("Period (h)", fontsize=9)
                        ax.set_ylabel("Power (a.u.)", fontsize=9)
                        ax.set_title(
                            f"ROI {roi_id} - {section_label}",
                            fontsize=9,
                            color=roi_color,
                            fontweight="bold",
                            pad=4,
                            loc="left",
                        )
                        ax.tick_params(axis="both", labelsize=8)
                        ax.grid(True, alpha=0.3)

                        # Per-ROI Y-axis scaling
                        if len(power) > 0:
                            roi_y_max = max(power) * 1.1
                            ax.set_ylim(0, roi_y_max)

                        if "dominant_period" in result:
                            dominant_period = result["dominant_period"]
                            dominant_power = result.get("dominant_power", 0)
                            ax.axvline(
                                x=dominant_period,
                                color=roi_color,
                                linestyle="--",
                                linewidth=1.5,
                                alpha=0.5,
                            )
                            ax.plot(
                                dominant_period,
                                dominant_power,
                                "o",
                                color=roi_color,
                                markersize=8,
                                markeredgecolor="black",
                                markeredgewidth=1,
                                label=f"Peak: {dominant_period:.1f}h",
                            )
                            ax.legend(fontsize=7)

                            # Stats box: period, peak time (phase), p-value
                            peak_t = result.get("dominant_peak_time_hours", None)
                            p_val  = result.get("p_value", 1.0)
                            sig_marker = " *" if result.get("is_significant", False) else ""
                            if peak_t is not None and not np.isnan(peak_t):
                                stats_text = (
                                    f"Period: {dominant_period:.2f}h\n"
                                    f"Peak time: {peak_t:.1f}h\n"
                                    f"p: {_fmt_p(p_val)}{sig_marker}"
                                )
                            else:
                                stats_text = (
                                    f"Period: {dominant_period:.2f}h\n"
                                    f"p: {_fmt_p(p_val)}{sig_marker}"
                                )
                            ax.text(
                                0.97, 0.97, stats_text,
                                transform=ax.transAxes, fontsize=8,
                                verticalalignment="top",
                                horizontalalignment="right",
                                bbox=dict(boxstyle="round", facecolor="wheat",
                                          alpha=0.5, pad=0.5),
                            )

                        # Set x-axis to actual data range (analysis caps max at duration/2)
                        if len(periods) > 0:
                            ax.set_xlim(min(periods), max(periods))

                # Hide unused subplots in this section
                for idx in range(n_rois, n_rows_per_section * n_cols):
                    row = start_row + (idx // n_cols)
                    col = idx % n_cols
                    if row < total_rows:
                        axes[row, col].axis("off")

            # Plot Activity section
            plot_section(roi_only_results, 0, "Activity")

            # Plot Sleep section if available
            if has_sleep:
                plot_section(sleep_results, n_rows_per_section, "Sleep")

            plt.tight_layout(rect=[0, 0, 1, 0.96], h_pad=2.0, w_pad=1.0)
            self.fisher_plot_figure = fig

            buf = io.BytesIO()
            fig.savefig(
                buf,
                format="png",
                dpi=150,
                bbox_inches="tight",
                facecolor="white",
                edgecolor="none",
            )
            buf.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buf.read())

            # Scale pixmap to fit canvas while maintaining aspect ratio
            scaled_pixmap = pixmap.scaled(
                self.fisher_plot_canvas.size(),
                1,  # Qt.KeepAspectRatio
                1,  # Qt.SmoothTransformation
            )
            self.fisher_plot_canvas.setPixmap(scaled_pixmap)

            # Enable pop-out button after successful plot creation
            if hasattr(self, "btn_popout_plot"):
                self.btn_popout_plot.setEnabled(True)
                if hasattr(self, "btn_save_fisher_plot"):
                    self.btn_save_fisher_plot.setEnabled(True)

        except Exception as e:
            self._log_message(f"⚠️ Could not create FFT plot: {e}")
            import traceback

            traceback.print_exc()

    def _create_cosinor_plot(self, cosinor_results: Dict):
        """Create cosinor analysis plots showing data and fitted curves.

        Cosinor analysis only uses raw movement data (continuous values),
        not fraction movement or binary sleep data.
        """
        try:
            import matplotlib.pyplot as plt
            from qtpy.QtGui import QPixmap
            import io

            # Close the previous figure and disconnect all Qt event callbacks
            # before creating a new one to prevent NavigationToolbar2QT dangling refs.
            if hasattr(self, "fisher_plot_figure") and self.fisher_plot_figure:
                try:
                    plt.close(self.fisher_plot_figure)
                except Exception:
                    pass
                self.fisher_plot_figure = None

            roi_results = cosinor_results.get("roi_results", {})
            # Filter to only integer ROI keys
            roi_results = {k: v for k, v in roi_results.items() if isinstance(k, int)}
            n_rois = len(roi_results)

            # Handle case with no ROIs
            if n_rois == 0:
                self._log_message("⚠️ No ROI results found for Cosinor plot")
                return

            n_cols = min(3, n_rois)
            n_rows_per_section = (n_rois + n_cols - 1) // n_cols

            show_population = (
                hasattr(self, "chk_cosinor_population")
                and self.chk_cosinor_population.isChecked()
            )
            pop_result = cosinor_results.get("population_result", {})
            has_population = show_population and "error" not in pop_result

            fig_height = 4.5 * n_rows_per_section + 1 + (3.5 if has_population else 0)
            fig = plt.figure(figsize=(13, fig_height))

            if has_population:
                gs = fig.add_gridspec(
                    n_rows_per_section + 1, n_cols,
                    hspace=0.45,
                    height_ratios=[1] * n_rows_per_section + [0.8],
                )
                axes = np.array(
                    [[fig.add_subplot(gs[r, c]) for c in range(n_cols)]
                     for r in range(n_rows_per_section)]
                )
                ax_pop = fig.add_subplot(gs[n_rows_per_section, :])
            else:
                gs = fig.add_gridspec(n_rows_per_section, n_cols, hspace=0.45)
                axes = np.array(
                    [[fig.add_subplot(gs[r, c]) for c in range(n_cols)]
                     for r in range(n_rows_per_section)]
                )
                ax_pop = None

            data_type_name = getattr(self, "fisher_data_type_name", "Fraction Movement (0-1)")
            fig.suptitle(
                f"Cosinor Analysis  —  {data_type_name}{self._get_recording_start_str()}",
                fontsize=13,
                fontweight="bold",
                y=0.99,
            )

            # Recording duration (hours) — needed for figure-level warning and per-ROI annotation
            _recording_dur_h = cosinor_results.get("recording_duration_h", 0.0)

            # Figure-level warning if recording < 2× any test period
            test_periods_plot = cosinor_results.get("test_periods", [])
            short_periods = [
                tp for tp in test_periods_plot
                if _recording_dur_h > 0 and _recording_dur_h / tp < 2.0
            ]
            if short_periods:
                warn_txt = (
                    f"⚠  Recording ({_recording_dur_h:.1f} h) < 2× period — "
                    f"result(s) unreliable for: "
                    + ", ".join(f"{p:.1f} h" for p in short_periods)
                )
                fig.text(
                    0.5, 0.005, warn_txt,
                    ha="center", va="bottom", fontsize=8,
                    color="#B71C1C",
                    bbox=dict(boxstyle="round", fc="#FFEBEE", ec="#EF9A9A", alpha=0.9),
                )

            # Use the data that was actually analysed (stored by run_fisher_analysis)
            activity_data_dict = getattr(self, "fisher_analysis_data", None)
            if not activity_data_dict:
                activity_data_dict = self.fraction_data if hasattr(self, "fraction_data") else {}
            activity_y_label = data_type_name

            # Helper function to plot ROI results
            def plot_section(
                results_dict, raw_data_dict, start_row, section_label, y_label
            ):
                roi_items = {
                    k: v for k, v in results_dict.items() if isinstance(k, int)
                }

                for idx, (roi_id, roi_data) in enumerate(sorted(roi_items.items())):
                    row = start_row + (idx // n_cols)
                    col = idx % n_cols
                    ax = axes[row, col]

                    roi_color = (
                        self.roi_colors.get(roi_id, f"C{idx}")
                        if hasattr(self, "roi_colors")
                        else f"C{idx}"
                    )

                    best_result = roi_data.get("best_result", {})

                    if "error" in best_result:
                        ax.text(
                            0.5,
                            0.5,
                            f"ROI {roi_id}\n{best_result['error']}",
                            ha="center",
                            va="center",
                            transform=ax.transAxes,
                        )
                        ax.set_xticks([])
                        ax.set_yticks([])
                        continue

                    # Get data for this ROI
                    if roi_id in raw_data_dict:
                        times_hours = np.array(
                            [t / 3600 for t, _ in raw_data_dict[roi_id]]
                        )
                        values = np.array([v for _, v in raw_data_dict[roi_id]])

                        # Plot actual data as a continuous line
                        ax.plot(
                            times_hours,
                            values,
                            alpha=0.6,
                            linewidth=0.8,
                            color=roi_color,
                            label=f"{section_label} data",
                        )

                        # Reconstruct fitted curve at raw data time points
                        # (avoids length mismatch when analysis used rebinned data)
                        period = best_result.get("period", 0)
                        amplitude = best_result.get("amplitude", 0)
                        mesor = best_result.get("mesor", 0)
                        peak_time = best_result.get("peak_time", 0)
                        phase_angle_rad = best_result.get("phase_angle_rad", 0)

                        if period > 0 and not np.isnan(period):
                            omega = 2 * np.pi / period
                            # Reconstruct using the phase angle φ:
                            # y(t) = MESOR + A·cos(ω·t + φ)
                            fitted_curve = mesor + amplitude * np.cos(
                                omega * times_hours + phase_angle_rad
                            )
                            ax.plot(
                                times_hours,
                                fitted_curve,
                                color="black",
                                linewidth=2,
                                label=f"Fitted curve ({period:.1f}h)",
                            )

                            # MESOR line
                            ax.axhline(
                                y=mesor,
                                color="gray",
                                linestyle="--",
                                linewidth=1,
                                alpha=0.5,
                                label=f"MESOR={mesor:.3f}",
                            )

                            # Peak time marker
                            peak_value = mesor + amplitude
                            if (
                                isinstance(peak_time, (int, float))
                                and peak_time < times_hours[-1]
                            ):
                                ax.plot(
                                    peak_time,
                                    peak_value,
                                    "^",
                                    color="red",
                                    markersize=10,
                                    markeredgecolor="black",
                                    markeredgewidth=0.5,
                                    label=f"Peak Time={peak_time:.1f}h",
                                )

                        textstr = f"MESOR: {best_result.get('mesor', 0):.3f}\n"
                        textstr += f"Amplitude: {best_result.get('amplitude', 0):.3f}\n"
                        textstr += f"R²: {best_result.get('r_squared', 0):.3f}\n"
                        textstr += f"p: {_fmt_p(best_result.get('p_value', 1))}"
                        if best_result.get("significant", False):
                            textstr += " *"
                        if period > 0 and _recording_dur_h > 0:
                            n_cycles = _recording_dur_h / period
                            warn_flag = "⚠ " if n_cycles < 2.0 else ""
                            textstr += f"\n{warn_flag}cycles: {n_cycles:.1f}"

                        ax.text(
                            0.95,
                            0.05,
                            textstr,
                            transform=ax.transAxes,
                            fontsize=8,
                            verticalalignment="bottom",
                            horizontalalignment="right",
                            bbox=dict(
                                boxstyle="round", facecolor="wheat", alpha=0.5, pad=0.5
                            ),
                        )

                    ax.set_xlabel("Time (h)", fontsize=9)
                    ax.set_ylabel(y_label, fontsize=9)

                    # Per-ROI Y-axis scaling - include fitted curve range
                    if roi_id in raw_data_dict and len(values) > 0:
                        data_max = max(values)
                        data_min = min(values)
                        # Also consider fitted curve range if it exists
                        if period > 0 and not np.isnan(period):
                            fitted_max = mesor + amplitude
                            fitted_min = mesor - amplitude
                            roi_y_max = max(data_max, fitted_max) * 1.1
                            roi_y_min = (
                                min(0, data_min, fitted_min) * 1.1
                                if min(0, data_min, fitted_min) < 0
                                else min(0, data_min)
                            )
                        else:
                            roi_y_max = data_max * 1.1
                            roi_y_min = min(0, data_min)
                        ax.set_ylim(roi_y_min, roi_y_max)
                    ax.tick_params(axis="both", labelsize=8)
                    is_significant = best_result.get("significant", False)
                    title_weight = "bold" if is_significant else "normal"
                    ax.set_title(
                        f"ROI {roi_id} - {section_label}",
                        fontsize=9,
                        color=roi_color,
                        fontweight=title_weight,
                        pad=4,
                        loc="left",
                    )
                    ax.legend(fontsize=7, loc="upper left")
                    ax.grid(True, alpha=0.3)

                # Hide unused subplots in this section
                for idx in range(n_rois, n_rows_per_section * n_cols):
                    row = start_row + (idx // n_cols)
                    col = idx % n_cols
                    if row < total_rows:
                        axes[row, col].axis("off")

            # Plot Activity section — use a short label derived from data_type_name
            # e.g. "Fraction Movement (0-1)" → "Fraction Movement"
            section_label = data_type_name.split("(")[0].strip()
            plot_section(
                roi_results,
                activity_data_dict,
                0,
                section_label,
                activity_y_label,
            )

            # --- Optional population mean subplot ---
            if has_population and ax_pop is not None:
                # Build common time grid from union of all ROI time ranges
                all_times_h = []
                for data_list in activity_data_dict.values():
                    if data_list:
                        all_times_h.extend(t / 3600.0 for t, _ in data_list)
                t_min = min(all_times_h) if all_times_h else 0.0
                t_max = max(all_times_h) if all_times_h else 24.0
                t_grid = np.linspace(t_min, t_max, 500)

                # Faint individual fitted curves
                for idx, (roi_id, roi_data) in enumerate(sorted(roi_results.items())):
                    best = roi_data.get("best_result", {})
                    period = best.get("period", 0)
                    amplitude = best.get("amplitude", 0)
                    mesor = best.get("mesor", 0)
                    phi = best.get("phase_angle_rad", 0)
                    if period > 0 and not np.isnan(period):
                        omega = 2 * np.pi / period
                        fitted = mesor + amplitude * np.cos(omega * t_grid + phi)
                        roi_color = (
                            self.roi_colors.get(roi_id, f"C{idx}")
                            if hasattr(self, "roi_colors") else f"C{idx}"
                        )
                        ax_pop.plot(t_grid, fitted, color=roi_color,
                                    alpha=0.25, linewidth=1)

                # Population mean fit
                pop_period = pop_result.get("period", 24.0)
                pop_amplitude = pop_result.get("population_amplitude", 0)
                pop_mesor = pop_result.get("population_mesor", 0)
                pop_peak_time = pop_result.get("population_peak_time", 0)

                if pop_period > 0:
                    omega_pop = 2 * np.pi / pop_period
                    phase_pop = -omega_pop * pop_peak_time
                    pop_curve = pop_mesor + pop_amplitude * np.cos(
                        omega_pop * t_grid + phase_pop
                    )
                    ax_pop.plot(t_grid, pop_curve, color="black", linewidth=2.5,
                                label=f"Population mean ({pop_period:.1f}h)", zorder=5)
                    ax_pop.axhline(y=pop_mesor, color="gray", linestyle="--",
                                   linewidth=1, alpha=0.5,
                                   label=f"Pop. MESOR={pop_mesor:.3f}")

                p_val = pop_result.get("p_value", 1.0)
                sig_marker = " *" if pop_result.get("significant", False) else ""
                n_sig = pop_result.get("n_significant", 0)
                n_ind = pop_result.get("n_individuals", 0)
                stats_text = (
                    f"MESOR: {pop_mesor:.3f}\n"
                    f"Amplitude: {pop_amplitude:.3f}\n"
                    f"Peak time: {pop_peak_time:.1f}h\n"
                    f"p: {_fmt_p(p_val)}{sig_marker}\n"
                    f"n={n_ind} ROIs  ({n_sig} significant)"
                )
                ax_pop.text(
                    0.97, 0.05, stats_text, transform=ax_pop.transAxes,
                    fontsize=8, verticalalignment="bottom",
                    horizontalalignment="right",
                    bbox=dict(boxstyle="round", facecolor="wheat",
                              alpha=0.5, pad=0.5),
                )
                ax_pop.set_title("Population Mean", fontsize=10,
                                 fontweight="bold", loc="left")
                ax_pop.set_xlabel("Time (h)", fontsize=9)
                ax_pop.set_ylabel(activity_y_label, fontsize=9)
                ax_pop.legend(fontsize=7, loc="upper left")
                ax_pop.grid(True, alpha=0.3)
                ax_pop.tick_params(axis="both", labelsize=8)

            # subplots_adjust instead of tight_layout: the latter does not
            # support axes that span multiple columns (population subplot).
            fig.subplots_adjust(top=0.94, left=0.07, right=0.97,
                                bottom=0.04, hspace=0.50, wspace=0.30)
            self.fisher_plot_figure = fig

            buf = io.BytesIO()
            fig.savefig(
                buf,
                format="png",
                dpi=150,
                bbox_inches="tight",
                facecolor="white",
                edgecolor="none",
            )
            buf.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buf.read())
            self.fisher_plot_canvas.setPixmap(
                pixmap.scaled(self.fisher_plot_canvas.size(), 1, 1)
            )

            # Enable pop-out button after successful plot creation
            if hasattr(self, "btn_popout_plot"):
                self.btn_popout_plot.setEnabled(True)
                if hasattr(self, "btn_save_fisher_plot"):
                    self.btn_save_fisher_plot.setEnabled(True)

        except Exception as e:
            self._log_message(f"⚠️ Could not create Cosinor plot: {e}")
            import traceback

            traceback.print_exc()

    def _create_similarity_plot(self, similarity_results: Dict):
        """Create correlation matrix heatmap and dendrogram."""
        try:
            import matplotlib.pyplot as plt
            from qtpy.QtGui import QPixmap
            import io
            from scipy.cluster import hierarchy

            if hasattr(self, "fisher_plot_figure") and self.fisher_plot_figure:
                plt.close(self.fisher_plot_figure)

            fig = plt.figure(figsize=(14, 6))

            # Get data source for plot title
            data_source_index = self.data_source_combo.currentIndex()
            data_source = (
                "Fraction Movement" if data_source_index == 0 else "Raw Intensity"
            )
            fig.suptitle(
                f"ROI Similarity Analysis\n(Activity from {data_source})",
                fontsize=14,
                fontweight="bold",
            )

            # Correlation heatmap
            ax1 = plt.subplot(1, 2, 1)
            corr_matrix = similarity_results["correlation_matrix"]
            roi_ids = similarity_results["roi_ids"]

            im = ax1.imshow(corr_matrix, cmap="RdYlGn", vmin=-1, vmax=1, aspect="auto")
            ax1.set_xticks(range(len(roi_ids)))
            ax1.set_yticks(range(len(roi_ids)))
            ax1.set_xticklabels(roi_ids, rotation=45)
            ax1.set_yticklabels(roi_ids)
            ax1.set_title("ROI Correlation Matrix")
            plt.colorbar(im, ax=ax1, label="Correlation")

            # Dendrogram
            ax2 = plt.subplot(1, 2, 2)
            if "clustering" in similarity_results:
                linkage = similarity_results["clustering"]["linkage_matrix"]

                # Calculate color threshold (30% of max distance)
                max_distance = np.max(linkage[:, 2])
                color_threshold = max_distance * 0.3

                # Create dendrogram with colors
                hierarchy.dendrogram(
                    linkage,
                    labels=[f"ROI {r}" for r in roi_ids],
                    ax=ax2,
                    color_threshold=color_threshold,
                    above_threshold_color="gray",
                )

                ax2.set_title("Hierarchical Clustering", fontsize=12, fontweight="bold")
                ax2.set_xlabel("ROI", fontsize=11)
                ax2.set_ylabel("Distance (1 - correlation)", fontsize=11)
                ax2.tick_params(axis="both", labelsize=10)

                # Add horizontal line at threshold
                ax2.axhline(
                    y=color_threshold,
                    color="red",
                    linestyle="--",
                    linewidth=1.5,
                    alpha=0.7,
                    label=f"Cluster threshold ({color_threshold:.2f})",
                )
                ax2.legend(fontsize=9)

                # Add grid for better readability
                ax2.grid(True, alpha=0.3, axis="y")

            plt.tight_layout()
            self.fisher_plot_figure = fig

            buf = io.BytesIO()
            fig.savefig(
                buf,
                format="png",
                dpi=150,
                bbox_inches="tight",
                facecolor="white",
                edgecolor="none",
            )
            buf.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buf.read())
            self.fisher_plot_canvas.setPixmap(
                pixmap.scaled(self.fisher_plot_canvas.size(), 1, 1)
            )

            # Enable pop-out button after successful plot creation
            if hasattr(self, "btn_popout_plot"):
                self.btn_popout_plot.setEnabled(True)
                if hasattr(self, "btn_save_fisher_plot"):
                    self.btn_save_fisher_plot.setEnabled(True)

        except Exception as e:
            self._log_message(f"⚠️ Could not create similarity plot: {e}")
            import traceback

            traceback.print_exc()

    def _create_coherence_plot(self, coherence_results: Dict):
        """Create coherence matrix heatmap."""
        try:
            import matplotlib.pyplot as plt
            from qtpy.QtGui import QPixmap
            import io

            if hasattr(self, "fisher_plot_figure") and self.fisher_plot_figure:
                plt.close(self.fisher_plot_figure)

            fig, ax = plt.subplots(figsize=(10, 8))

            # Get data source for plot title
            data_source_index = self.data_source_combo.currentIndex()
            data_source = (
                "Fraction Movement" if data_source_index == 0 else "Raw Intensity"
            )

            coherence_matrix = coherence_results["coherence_matrix"]
            roi_ids = coherence_results["roi_ids"]

            im = ax.imshow(coherence_matrix, cmap="hot", vmin=0, vmax=1, aspect="auto")
            ax.set_xticks(range(len(roi_ids)))
            ax.set_yticks(range(len(roi_ids)))
            ax.set_xticklabels(roi_ids, rotation=45)
            ax.set_yticklabels(roi_ids)
            ax.set_title(
                f"ROI Coherence at ~{coherence_results['target_period_hours']:.0f}h Period\n(Activity from {data_source})"
            )
            plt.colorbar(im, ax=ax, label="Coherence")

            plt.tight_layout()
            self.fisher_plot_figure = fig

            buf = io.BytesIO()
            fig.savefig(
                buf,
                format="png",
                dpi=150,
                bbox_inches="tight",
                facecolor="white",
                edgecolor="none",
            )
            buf.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buf.read())
            self.fisher_plot_canvas.setPixmap(
                pixmap.scaled(self.fisher_plot_canvas.size(), 1, 1)
            )

            # Enable pop-out button after successful plot creation
            if hasattr(self, "btn_popout_plot"):
                self.btn_popout_plot.setEnabled(True)
                if hasattr(self, "btn_save_fisher_plot"):
                    self.btn_save_fisher_plot.setEnabled(True)

        except Exception as e:
            self._log_message(f"⚠️ Could not create coherence plot: {e}")
            import traceback

            traceback.print_exc()

    def _create_phase_cluster_plot(self, phase_results: Dict):
        """Create polar plot showing ROI phases."""
        try:
            import matplotlib.pyplot as plt
            from qtpy.QtGui import QPixmap
            import io

            if hasattr(self, "fisher_plot_figure") and self.fisher_plot_figure:
                plt.close(self.fisher_plot_figure)

            fig = plt.figure(figsize=(10, 10))
            ax = fig.add_subplot(111, projection="polar")

            roi_phases = phase_results["roi_phases"]
            # Filter to only integer ROI keys
            roi_phases_filtered = {
                k: v for k, v in roi_phases.items() if isinstance(k, int)
            }

            for idx, (roi_id, phase_info) in enumerate(
                sorted(roi_phases_filtered.items())
            ):
                # Get ROI-specific color
                roi_color = (
                    self.roi_colors.get(roi_id, f"C{idx}")
                    if hasattr(self, "roi_colors")
                    else f"C{idx}"
                )

                theta = (phase_info["phase_radians"] + np.pi) % (
                    2 * np.pi
                )  # Adjust for polar plot
                r = phase_info["amplitude"]
                ax.plot(
                    [theta, theta],
                    [0, r],
                    color=roi_color,
                    linewidth=2,
                    label=f"ROI {roi_id}",
                )
                ax.scatter(
                    [theta],
                    [r],
                    color=roi_color,
                    s=100,
                    edgecolor="black",
                    linewidth=1,
                    zorder=5,
                )

            ax.set_theta_zero_location("N")
            ax.set_theta_direction(-1)
            # Get data source for plot title
            data_source_index = self.data_source_combo.currentIndex()
            data_source = (
                "Fraction Movement" if data_source_index == 0 else "Raw Intensity"
            )
            ax.set_title(
                f"ROI Activity Phases\n(Activity from {data_source})",
                fontsize=14,
                fontweight="bold",
                pad=20,
            )
            ax.legend(loc="upper left", bbox_to_anchor=(1.1, 1.0), fontsize=8)

            plt.tight_layout()
            self.fisher_plot_figure = fig

            buf = io.BytesIO()
            fig.savefig(
                buf,
                format="png",
                dpi=150,
                bbox_inches="tight",
                facecolor="white",
                edgecolor="none",
            )
            buf.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buf.read())
            self.fisher_plot_canvas.setPixmap(
                pixmap.scaled(self.fisher_plot_canvas.size(), 1, 1)
            )

            # Enable pop-out button after successful plot creation
            if hasattr(self, "btn_popout_plot"):
                self.btn_popout_plot.setEnabled(True)
                if hasattr(self, "btn_save_fisher_plot"):
                    self.btn_save_fisher_plot.setEnabled(True)

        except Exception as e:
            self._log_message(f"⚠️ Could not create phase plot: {e}")
            import traceback

            traceback.print_exc()

    def export_fisher_results(self):
        """Export rhythmic pattern analysis results to Excel/CSV."""
        if not hasattr(self, "fisher_analysis_results"):
            self._log_message("⚠️ No analysis results to export")
            return

        if not hasattr(self, "current_fisher_method"):
            self._log_message("⚠️ Cannot determine analysis method")
            return

        from qtpy.QtWidgets import QFileDialog
        import os

        # Determine method name for file naming
        method_names = {
            0: "fisher_z",
            1: "fft_spectrum",
            2: "cosinor",
            3: "roi_similarity",
            4: "coherence_analysis",
            5: "phase_clustering",
        }
        method_name = method_names.get(self.current_fisher_method, "rhythmic_analysis")

        # Get save location
        default_name = f"rhythmic_{method_name}_results"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Rhythmic Pattern Analysis Results",
            default_name,
            "Excel Files (*.xlsx);;All Files (*.*)",
        )

        if not file_path:
            self._log_message("Export cancelled by user")
            return

        # Ensure .xlsx extension
        if not file_path.endswith(".xlsx"):
            file_path = f"{file_path}.xlsx"

        # Remove extension for consistent naming
        base_path = os.path.splitext(file_path)[0]

        try:
            # Route to appropriate export method based on analysis type
            if self.current_fisher_method == 0:  # Chi² Periodogram
                self._export_fisher_method_results(base_path)
            elif self.current_fisher_method == 1:  # FFT Power Spectrum
                self._export_fft_method_results(base_path)
            elif self.current_fisher_method == 2:  # Cosinor Analysis
                self._export_cosinor_method_results(base_path)
            elif self.current_fisher_method == 3:  # ROI Similarity Matrix
                self._export_similarity_method_results(file_path)
            elif self.current_fisher_method == 4:  # Coherence Analysis
                self._export_coherence_method_results(file_path)
            elif self.current_fisher_method == 5:  # Phase Clustering
                self._export_phase_clustering_method_results(file_path)
            else:
                self._log_message(
                    f"❌ Unknown analysis method: {self.current_fisher_method}"
                )
                return

            # Export plot if available
            if hasattr(self, "fisher_plot_figure") and self.fisher_plot_figure:
                plot_path = f"{base_path}_plot.png"
                self.fisher_plot_figure.savefig(
                    plot_path, dpi=300, bbox_inches="tight", facecolor="white"
                )
                self._log_message(f"✓ Exported plot to: {plot_path}")

            self._log_message("✓ Export complete!")

        except Exception as e:
            self._log_message(f"❌ Export failed: {e}")
            import traceback

            traceback.print_exc()

    def _export_fisher_method_results(self, base_path):
        """Export Chi² Periodogram results to CSV and Excel."""
        # Export to CSV
        csv_path = f"{base_path}.csv"
        self._export_fisher_to_csv(csv_path)
        self._log_message(f"✓ Exported Fisher results to CSV: {csv_path}")

        # Export to Excel
        try:
            import pandas as pd

            excel_path = f"{base_path}.xlsx"
            self._export_fisher_to_excel(excel_path)
            self._log_message(f"✓ Exported Fisher results to Excel: {excel_path}")
        except ImportError:
            self._log_message("⚠️ Excel export not available (pandas not installed)")

    def _export_fft_method_results(self, base_path):
        """Export FFT Power Spectrum results to CSV and Excel."""
        # Export to CSV (similar to Fisher)
        csv_path = f"{base_path}.csv"
        self._export_fft_to_csv(csv_path)
        self._log_message(f"✓ Exported FFT results to CSV: {csv_path}")

        # Export to Excel
        try:
            import pandas as pd

            excel_path = f"{base_path}.xlsx"
            self._export_fft_to_excel(excel_path)
            self._log_message(f"✓ Exported FFT results to Excel: {excel_path}")
        except ImportError:
            self._log_message("⚠️ Excel export not available (pandas not installed)")

    def _export_cosinor_method_results(self, base_path):
        """Export Cosinor Analysis results to CSV and Excel."""
        csv_path = f"{base_path}.csv"
        self._export_cosinor_to_csv(csv_path)
        self._log_message(f"✓ Exported Cosinor results to CSV: {csv_path}")

        try:
            import pandas as pd

            excel_path = f"{base_path}.xlsx"
            self._export_cosinor_to_excel(excel_path)
            self._log_message(f"✓ Exported Cosinor results to Excel: {excel_path}")
        except ImportError:
            self._log_message("⚠️ Excel export not available (pandas not installed)")
        except Exception as e:
            self._log_message(f"⚠️ Cosinor export error: {e}")

    def _export_cosinor_to_excel(self, file_path: str):
        """Export Cosinor analysis results to Excel format."""
        import pandas as pd

        results = self.fisher_analysis_results
        roi_results = results.get("roi_results", {})
        sleep_results = results.get("sleep_phase_results", {})
        sleep_roi_results = (
            sleep_results.get("roi_results", {}) if sleep_results else {}
        )

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: Activity Summary
            activity_data = []
            for roi_id, roi_data in sorted(
                {k: v for k, v in roi_results.items() if isinstance(k, int)}.items()
            ):
                best_result = roi_data.get("best_result", {})
                if "error" in best_result:
                    continue

                activity_data.append(
                    {
                        "ROI": roi_id,
                        "Best Period (hours)": roi_data.get("best_period", "N/A"),
                        "Peak Time (hours from start)": best_result.get("peak_time", "N/A"),
                        "Phase Angle (rad)": best_result.get("phase_angle_rad", "N/A"),
                        "MESOR": best_result.get("mesor", "N/A"),
                        "Amplitude": best_result.get("amplitude", "N/A"),
                        "R-squared": best_result.get("r_squared", "N/A"),
                        "P-value": best_result.get("p_value", "N/A"),
                        "Significant": best_result.get("significant", False),
                    }
                )

            if activity_data:
                activity_df = pd.DataFrame(activity_data)
                activity_df.to_excel(writer, sheet_name="Activity_Summary", index=False)

            # Sheet 2: Sleep Summary (if available)
            if sleep_roi_results:
                sleep_data = []
                for roi_id, roi_data in sorted(
                    {
                        k: v for k, v in sleep_roi_results.items() if isinstance(k, int)
                    }.items()
                ):
                    best_result = roi_data.get("best_result", {})
                    if "error" in best_result:
                        continue

                    sleep_data.append(
                        {
                            "ROI": roi_id,
                            "Best Period (hours)": roi_data.get("best_period", "N/A"),
                            "Sleep Peak Time (hours from start)": best_result.get("peak_time", "N/A"),
                            "MESOR": best_result.get("mesor", "N/A"),
                            "Amplitude": best_result.get("amplitude", "N/A"),
                            "R-squared": best_result.get("r_squared", "N/A"),
                            "P-value": best_result.get("p_value", "N/A"),
                            "Significant": best_result.get("significant", False),
                        }
                    )

                if sleep_data:
                    sleep_df = pd.DataFrame(sleep_data)
                    sleep_df.to_excel(writer, sheet_name="Sleep_Summary", index=False)

            # Sheet 3: Combined comparison (Activity vs Sleep)
            if activity_data and sleep_roi_results:
                comparison_data = []
                for roi_id in sorted(
                    set([d["ROI"] for d in activity_data]).intersection(
                        [k for k in sleep_roi_results.keys() if isinstance(k, int)]
                    )
                ):
                    act_row = next((d for d in activity_data if d["ROI"] == roi_id), {})
                    slp_roi = sleep_roi_results.get(roi_id, {})
                    slp_best = slp_roi.get("best_result", {})

                    act_peak = act_row.get("Peak Time (hours from start)", 0)
                    slp_peak = slp_best.get("peak_time", 0)

                    # Calculate peak time difference
                    if isinstance(act_peak, (int, float)) and isinstance(
                        slp_peak, (int, float)
                    ):
                        phase_diff = abs(act_peak - slp_peak)
                        if phase_diff > 12:
                            phase_diff = 24 - phase_diff
                    else:
                        phase_diff = "N/A"

                    comparison_data.append(
                        {
                            "ROI": roi_id,
                            "Activity Peak Time (h)": act_peak,
                            "Sleep Peak Time (h)": slp_peak,
                            "Peak Time Difference (h)": phase_diff,
                            "Activity Period (h)": act_row.get(
                                "Best Period (hours)", "N/A"
                            ),
                            "Sleep Period (h)": slp_roi.get("best_period", "N/A"),
                        }
                    )

                if comparison_data:
                    comparison_df = pd.DataFrame(comparison_data)
                    comparison_df.to_excel(
                        writer, sheet_name="Activity_vs_Sleep", index=False
                    )

            # Sheet 4: Parameters
            params_df = pd.DataFrame(
                {
                    "Parameter": [
                        "Minimum Period",
                        "Maximum Period",
                        "Significance Level",
                        "Sampling Interval",
                        "Sleep Phase Calculated",
                    ],
                    "Value": [
                        f"{self.fisher_min_period.value():.1f} hours",
                        f"{self.fisher_max_period.value():.1f} hours",
                        f"{self.fisher_significance.value():.3f}",
                        f"{self.frame_interval.value():.1f} seconds",
                        "Yes" if sleep_roi_results else "No",
                    ],
                }
            )
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

    def _export_similarity_method_results(self, base_path):
        """Export ROI Similarity results to CSV and Excel."""
        csv_path = f"{base_path}.csv"
        self._export_similarity_to_csv(csv_path)
        self._log_message(f"✓ Exported Similarity results to CSV: {csv_path}")

        try:
            from ._circadian_similarity import export_similarity_to_excel
            excel_path = f"{base_path}.xlsx"
            export_similarity_to_excel(excel_path, self.fisher_analysis_results)
            self._log_message(f"✓ Exported Similarity results to Excel: {excel_path}")
        except Exception as e:
            self._log_message(f"⚠️ Similarity Excel export error: {e}")

    def _export_coherence_method_results(self, base_path):
        """Export Coherence Analysis results to CSV and Excel."""
        csv_path = f"{base_path}.csv"
        self._export_coherence_to_csv(csv_path)
        self._log_message(f"✓ Exported Coherence results to CSV: {csv_path}")

        try:
            from ._circadian_coherence import export_coherence_to_excel
            excel_path = f"{base_path}.xlsx"
            export_coherence_to_excel(excel_path, self.fisher_analysis_results)
            self._log_message(f"✓ Exported Coherence results to Excel: {excel_path}")
        except Exception as e:
            self._log_message(f"⚠️ Coherence Excel export error: {e}")

    def _export_phase_clustering_method_results(self, base_path):
        """Export Phase Clustering results to CSV and Excel."""
        csv_path = f"{base_path}.csv"
        self._export_phase_clustering_to_csv(csv_path)
        self._log_message(f"✓ Exported Phase Clustering results to CSV: {csv_path}")

        try:
            import pandas as pd
            excel_path = f"{base_path}.xlsx"
            self._export_phase_clustering_to_excel(excel_path)
            self._log_message(f"✓ Exported Phase Clustering results to Excel: {excel_path}")
        except Exception as e:
            self._log_message(f"⚠️ Phase Clustering Excel export error: {e}")

    def _export_phase_clustering_to_excel(self, file_path):
        """Export Phase Clustering results to Excel."""
        import pandas as pd

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: Phase Clusters
            if "phase_clusters" in self.fisher_analysis_results:
                cluster_data = []
                for cluster_name, roi_list in self.fisher_analysis_results[
                    "phase_clusters"
                ].items():
                    for roi in roi_list:
                        cluster_data.append(
                            {
                                "ROI": roi,
                                "Cluster": cluster_name.replace("_", " ").title(),
                                "Cluster_Size": len(roi_list),
                            }
                        )

                if cluster_data:
                    cluster_df = pd.DataFrame(cluster_data)
                    cluster_df = cluster_df.sort_values(["Cluster", "ROI"])
                    cluster_df.to_excel(
                        writer, sheet_name="Phase_Clusters", index=False
                    )

            # Sheet 2: Phase Values
            if "roi_phases" in self.fisher_analysis_results:
                phase_data = []
                for roi_id, phase_info in self.fisher_analysis_results[
                    "roi_phases"
                ].items():
                    phase_data.append(
                        {
                            "ROI": roi_id,
                            "Phase_Radians": phase_info["phase_radians"],
                            "Phase_Degrees": np.degrees(phase_info["phase_radians"]),
                            "Phase_Hours": phase_info["phase_hours"],
                            "Amplitude": phase_info["amplitude"],
                        }
                    )

                phase_df = pd.DataFrame(phase_data)
                phase_df = phase_df.sort_values("ROI")
                phase_df.to_excel(writer, sheet_name="Phase_Values", index=False)

            # Sheet 3: Parameters
            params_df = pd.DataFrame(
                {
                    "Parameter": [
                        "Number of Clusters",
                        "Number of ROIs",
                        "Dominant Period (hours)",
                    ],
                    "Value": [
                        str(
                            len(self.fisher_analysis_results.get("phase_clusters", {}))
                        ),
                        str(self.fisher_analysis_results.get("n_rois", 0)),
                        f"{self.fisher_analysis_results.get('dominant_period_hours', 0):.1f}",
                    ],
                }
            )
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

        self._log_message(f"✓ Exported Phase Clustering results to Excel: {file_path}")

    def _export_cosinor_to_csv(self, file_path: str):
        """Export Cosinor Analysis results to CSV format."""
        import csv

        results = self.fisher_analysis_results
        roi_results = results.get("roi_results", {})
        sleep_results = results.get("sleep_phase_results", {})
        sleep_roi_results = sleep_results.get("roi_results", {}) if sleep_results else {}

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            writer.writerow(["Cosinor Analysis Results"])
            writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])

            # Activity Summary
            writer.writerow(["Activity Summary"])
            writer.writerow(["ROI", "Best Period (h)", "Peak Time (h)", "Phase Angle (rad)",
                             "MESOR", "Amplitude", "R-squared", "P-value", "Significant"])
            for roi_id, roi_data in sorted(
                {k: v for k, v in roi_results.items() if isinstance(k, int)}.items()
            ):
                best = roi_data.get("best_result", {})
                if "error" in best:
                    writer.writerow([roi_id, "Error", best["error"]])
                    continue
                writer.writerow([
                    roi_id,
                    f"{roi_data.get('best_period', 'N/A')}",
                    f"{best.get('peak_time', 'N/A')}",
                    f"{best.get('phase_angle_rad', 'N/A')}",
                    f"{best.get('mesor', 'N/A')}",
                    f"{best.get('amplitude', 'N/A')}",
                    f"{best.get('r_squared', 'N/A')}",
                    f"{best.get('p_value', 'N/A')}",
                    str(best.get("significant", False)),
                ])

            if sleep_roi_results:
                writer.writerow([])
                writer.writerow(["Sleep Summary"])
                writer.writerow(["ROI", "Best Period (h)", "Sleep Peak Time (h)",
                                 "MESOR", "Amplitude", "R-squared", "P-value", "Significant"])
                for roi_id, roi_data in sorted(
                    {k: v for k, v in sleep_roi_results.items() if isinstance(k, int)}.items()
                ):
                    best = roi_data.get("best_result", {})
                    if "error" in best:
                        writer.writerow([roi_id, "Error", best["error"]])
                        continue
                    writer.writerow([
                        roi_id,
                        f"{roi_data.get('best_period', 'N/A')}",
                        f"{best.get('peak_time', 'N/A')}",
                        f"{best.get('mesor', 'N/A')}",
                        f"{best.get('amplitude', 'N/A')}",
                        f"{best.get('r_squared', 'N/A')}",
                        f"{best.get('p_value', 'N/A')}",
                        str(best.get("significant", False)),
                    ])

            writer.writerow([])
            writer.writerow(["Parameters"])
            writer.writerow(["Parameter", "Value"])
            writer.writerow(["Minimum Period", f"{self.fisher_min_period.value():.1f} hours"])
            writer.writerow(["Maximum Period", f"{self.fisher_max_period.value():.1f} hours"])
            writer.writerow(["Significance Level", f"{self.fisher_significance.value():.3f}"])
            writer.writerow(["Sampling Interval", f"{self.frame_interval.value():.1f} seconds"])
            writer.writerow(["Sleep Phase Calculated", "Yes" if sleep_roi_results else "No"])

    def _export_similarity_to_csv(self, file_path: str):
        """Export ROI Similarity (correlation matrix) results to CSV format."""
        import csv

        results = self.fisher_analysis_results

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            writer.writerow(["ROI Similarity Analysis Results"])
            writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])

            # Correlation matrix
            corr_matrix = results.get("correlation_matrix", None)
            roi_ids = results.get("roi_ids", [])
            if corr_matrix is not None and len(roi_ids) > 0:
                writer.writerow(["Correlation Matrix"])
                writer.writerow(["ROI"] + [f"ROI {r}" for r in roi_ids])
                for i, roi_id in enumerate(roi_ids):
                    row = [f"ROI {roi_id}"]
                    for j in range(len(roi_ids)):
                        row.append(f"{corr_matrix[i, j]:.4f}")
                    writer.writerow(row)
            else:
                writer.writerow(["No correlation matrix available"])

            # Lag matrix
            lag_matrix = results.get("lag_matrix", None)
            if lag_matrix is not None and len(roi_ids) > 0:
                writer.writerow([])
                writer.writerow(["Lag Matrix (hours)"])
                writer.writerow(["ROI"] + [f"ROI {r}" for r in roi_ids])
                for i, roi_id in enumerate(roi_ids):
                    row = [f"ROI {roi_id}"]
                    for j in range(len(roi_ids)):
                        row.append(f"{lag_matrix[i, j]:.2f}")
                    writer.writerow(row)

            # Cluster assignments
            clusters = results.get("clusters", None)
            if clusters is not None and len(roi_ids) > 0:
                writer.writerow([])
                writer.writerow(["Cluster Assignments"])
                writer.writerow(["ROI", "Cluster"])
                for roi_id, cluster in zip(roi_ids, clusters):
                    writer.writerow([roi_id, cluster])

    def _export_coherence_to_csv(self, file_path: str):
        """Export Coherence Analysis results to CSV format."""
        import csv

        results = self.fisher_analysis_results

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            writer.writerow(["Coherence Analysis Results"])
            writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])

            # Pairwise coherence summary
            coherence_data = results.get("coherence_results", {})
            if coherence_data:
                writer.writerow(["Pairwise Coherence Summary"])
                writer.writerow(["ROI Pair", "Mean Coherence", "Peak Frequency (Hz)",
                                 "Peak Period (h)"])
                for pair_key, pair_data in coherence_data.items():
                    if isinstance(pair_data, dict):
                        freqs = pair_data.get("frequencies", [])
                        coh = pair_data.get("coherence", [])
                        mean_coh = float(np.mean(coh)) if len(coh) > 0 else 0.0
                        if len(freqs) > 0 and len(coh) > 0:
                            peak_idx = int(np.argmax(coh))
                            peak_freq = freqs[peak_idx]
                            peak_period = (1.0 / peak_freq / 3600.0) if peak_freq > 0 else 0.0
                        else:
                            peak_freq = 0.0
                            peak_period = 0.0
                        writer.writerow([pair_key, f"{mean_coh:.4f}",
                                         f"{peak_freq:.6f}", f"{peak_period:.2f}"])

                # Detailed frequency-by-frequency data per pair
                writer.writerow([])
                writer.writerow(["Detailed Coherence per Pair"])
                for pair_key, pair_data in coherence_data.items():
                    if isinstance(pair_data, dict):
                        freqs = pair_data.get("frequencies", [])
                        coh = pair_data.get("coherence", [])
                        if len(freqs) > 0:
                            writer.writerow([])
                            writer.writerow([f"Pair: {pair_key}"])
                            writer.writerow(["Frequency (Hz)", "Period (h)", "Coherence"])
                            for f, c in zip(freqs, coh):
                                period_h = (1.0 / f / 3600.0) if f > 0 else 0.0
                                writer.writerow([f"{f:.6f}", f"{period_h:.2f}", f"{c:.4f}"])
            else:
                writer.writerow(["No coherence data available"])

    def _export_phase_clustering_to_csv(self, file_path: str):
        """Export Phase Clustering results to CSV format."""
        import csv

        results = self.fisher_analysis_results

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            writer.writerow(["Phase Clustering Results"])
            writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])

            # Parameters
            writer.writerow(["Parameters"])
            writer.writerow(["Number of Clusters",
                             len(results.get("phase_clusters", {}))])
            writer.writerow(["Number of ROIs", results.get("n_rois", 0)])
            writer.writerow(["Dominant Period (h)",
                             f"{results.get('dominant_period_hours', 0):.1f}"])
            writer.writerow([])

            # Cluster assignments
            phase_clusters = results.get("phase_clusters", {})
            if phase_clusters:
                writer.writerow(["Cluster Assignments"])
                writer.writerow(["ROI", "Cluster", "Cluster Size"])
                for cluster_name, roi_list in phase_clusters.items():
                    label = cluster_name.replace("_", " ").title()
                    for roi in sorted(roi_list):
                        writer.writerow([roi, label, len(roi_list)])

            # Phase values per ROI
            roi_phases = results.get("roi_phases", {})
            if roi_phases:
                writer.writerow([])
                writer.writerow(["Phase Values per ROI"])
                writer.writerow(["ROI", "Phase (rad)", "Phase (deg)", "Phase (h)", "Amplitude"])
                for roi_id, phase_info in sorted(roi_phases.items()):
                    writer.writerow([
                        roi_id,
                        f"{phase_info.get('phase_radians', 0):.4f}",
                        f"{np.degrees(phase_info.get('phase_radians', 0)):.2f}",
                        f"{phase_info.get('phase_hours', 0):.2f}",
                        f"{phase_info.get('amplitude', 0):.4f}",
                    ])

    def _export_fft_to_csv(self, file_path: str):
        """Export FFT Power Spectrum results to CSV format."""
        import csv

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            # Header
            writer.writerow(["FFT Power Spectrum Analysis Results"])
            writer.writerow(
                [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            )
            writer.writerow([])

            # Summary table
            writer.writerow(["ROI Summary"])
            writer.writerow(
                [
                    "ROI",
                    "Dominant Period (hours)",
                    "Dominant Frequency (Hz)",
                    "Spectral Power",
                    "Number of Peaks",
                    "Mean Activity",
                    "Std Activity",
                ]
            )

            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    writer.writerow([roi_id, "Error", result["error"], "", "", "", ""])
                    continue

                dominant_period = result.get("dominant_period", 0)
                dominant_freq = result.get("dominant_frequency", 0)
                dominant_power = result.get("dominant_power", 0)
                n_peaks = len(result.get("frequency_peaks", []))
                mean_activity = result.get("mean_activity", 0)
                std_activity = result.get("std_activity", 0)

                writer.writerow(
                    [
                        roi_id,
                        f"{dominant_period:.2f}",
                        f"{dominant_freq:.6f}",
                        f"{dominant_power:.2e}",
                        n_peaks,
                        f"{mean_activity:.4f}",
                        f"{std_activity:.4f}",
                    ]
                )

            # Detailed peak information
            writer.writerow([])
            writer.writerow(["Detailed Peak Information"])
            writer.writerow(
                [
                    "ROI",
                    "Peak #",
                    "Period (hours)",
                    "Frequency (Hz)",
                    "Power",
                    "Prominence",
                ]
            )

            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    continue

                peaks = result.get("frequency_peaks", [])
                for idx, peak in enumerate(peaks[:5], 1):  # Top 5 peaks
                    writer.writerow(
                        [
                            roi_id,
                            idx,
                            f"{peak['period_hours']:.2f}",
                            f"{peak['frequency_hz']:.6f}",
                            f"{peak['power']:.2e}",
                            f"{peak['prominence']:.2e}",
                        ]
                    )

    def _export_fft_to_excel(self, file_path: str):
        """Export FFT Power Spectrum results to Excel format."""
        import pandas as pd

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: Summary
            summary_data = []
            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    summary_data.append(
                        {
                            "ROI": roi_id,
                            "Error": result["error"],
                            "Dominant_Period_h": None,
                            "Dominant_Frequency_Hz": None,
                            "Spectral_Power": None,
                            "N_Peaks": None,
                            "Mean_Activity": None,
                            "Std_Activity": None,
                        }
                    )
                    continue

                summary_data.append(
                    {
                        "ROI": roi_id,
                        "Error": None,
                        "Dominant_Period_h": result.get("dominant_period", 0),
                        "Dominant_Frequency_Hz": result.get("dominant_frequency", 0),
                        "Spectral_Power": result.get("dominant_power", 0),
                        "N_Peaks": len(result.get("frequency_peaks", [])),
                        "Mean_Activity": result.get("mean_activity", 0),
                        "Std_Activity": result.get("std_activity", 0),
                    }
                )

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Sheet 2: All Peaks
            all_peaks_data = []
            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    continue

                peaks = result.get("frequency_peaks", [])
                for idx, peak in enumerate(peaks[:10], 1):  # Top 10 peaks
                    all_peaks_data.append(
                        {
                            "ROI": roi_id,
                            "Peak_Number": idx,
                            "Period_hours": peak["period_hours"],
                            "Frequency_Hz": peak["frequency_hz"],
                            "Power": peak["power"],
                            "Prominence": peak["prominence"],
                        }
                    )

            if all_peaks_data:
                peaks_df = pd.DataFrame(all_peaks_data)
                peaks_df.to_excel(writer, sheet_name="Frequency_Peaks", index=False)

            # Sheet 3: Full Power Spectra (for plotting)
            # Export complete power spectrum data for each ROI
            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    continue

                # Get relevant period range data
                relevant_periods = result.get("relevant_periods", [])
                relevant_power = result.get("relevant_power", [])

                if len(relevant_periods) > 0 and len(relevant_power) > 0:
                    spectrum_data = pd.DataFrame(
                        {
                            "Period_hours": relevant_periods,
                            "Power": relevant_power,
                        }
                    )
                    # Limit to reasonable number of points for Excel (max 10000)
                    if len(spectrum_data) > 10000:
                        # Downsample by taking every nth point
                        step = len(spectrum_data) // 10000 + 1
                        spectrum_data = spectrum_data.iloc[::step]

                    spectrum_data.to_excel(
                        writer, sheet_name=f"ROI_{roi_id}_Spectrum", index=False
                    )

            # Sheet 4: Sleep Phase Results (if available)
            sleep_results = self.fisher_analysis_results.get("sleep_phase_results", {})
            if sleep_results:
                sleep_roi_data = {
                    k: v for k, v in sleep_results.items() if isinstance(k, int)
                }
                if sleep_roi_data:
                    sleep_summary_data = []
                    for roi_id, result in sorted(sleep_roi_data.items()):
                        if "error" in result:
                            continue
                        sleep_summary_data.append(
                            {
                                "ROI": roi_id,
                                "Dominant_Period_h": result.get("dominant_period", 0),
                                "Dominant_Power": result.get("dominant_power", 0),
                                "Significant": result.get("is_significant", False),
                            }
                        )
                    if sleep_summary_data:
                        sleep_df = pd.DataFrame(sleep_summary_data)
                        sleep_df.to_excel(
                            writer, sheet_name="Sleep_Phase_Summary", index=False
                        )

            # Sheet 5: Parameters
            has_sleep = bool(sleep_results)
            params_df = pd.DataFrame(
                {
                    "Parameter": [
                        "Analysis Method",
                        "Generated",
                        "Sleep Phase Calculated",
                    ],
                    "Value": [
                        "FFT Power Spectrum",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Yes" if has_sleep else "No",
                    ],
                }
            )
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

    def _export_fisher_to_csv(self, file_path: str):
        """Export Fisher results to CSV format."""
        import csv

        with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            # Header
            writer.writerow(
                ["Circadian Rhythm Analysis Results (Fischer Z-transformation)"]
            )
            writer.writerow(
                [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            )
            writer.writerow([])

            # Summary table
            writer.writerow(["ROI Summary"])
            writer.writerow(
                [
                    "ROI",
                    "Significant Rhythm",
                    "Dominant Period (hours)",
                    "Z-Score",
                    "P-Value",
                    "Wake Phases",
                    "Sleep Phases",
                    "Wake Fraction (%)",
                ]
            )

            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    writer.writerow(
                        [roi_id, "Error", result["error"], "", "", "", "", ""]
                    )
                    continue

                periodogram = result.get("periodogram", {})
                phase_analysis = result.get("phase_analysis", {})

                is_sig = periodogram.get("is_significant", False)
                period = periodogram.get("dominant_period", 0)
                z_score = periodogram.get("dominant_z_score", 0)
                p_value = periodogram.get("p_value", 1.0)

                n_wake = len(phase_analysis.get("wake_phases", []))
                n_sleep = len(phase_analysis.get("sleep_phases", []))
                wake_frac = phase_analysis.get("wake_fraction", 0) * 100

                writer.writerow(
                    [
                        roi_id,
                        "Yes" if is_sig else "No",
                        f"{period:.2f}",
                        f"{z_score:.2f}",
                        f"{p_value:.4f}",
                        n_wake,
                        n_sleep,
                        f"{wake_frac:.1f}",
                    ]
                )

    def _export_fisher_to_excel(self, file_path: str):
        """Export Fisher results to Excel format with multiple sheets."""
        import pandas as pd

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: Summary
            summary_data = []
            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    continue

                periodogram = result.get("periodogram", {})
                phase_analysis = result.get("phase_analysis", {})

                summary_data.append(
                    {
                        "ROI": roi_id,
                        "Significant Rhythm": periodogram.get("is_significant", False),
                        "Dominant Period (hours)": periodogram.get(
                            "dominant_period", 0
                        ),
                        "Z-Score": periodogram.get("dominant_z_score", 0),
                        "P-Value": periodogram.get("p_value", 1.0),
                        "Wake Phases": len(phase_analysis.get("wake_phases", [])),
                        "Sleep Phases": len(phase_analysis.get("sleep_phases", [])),
                        "Wake Fraction (%)": phase_analysis.get("wake_fraction", 0)
                        * 100,
                    }
                )

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # Sheet 2: Full Periodograms (for plotting)
            # Export complete periodogram data for each ROI
            for roi_id, result in sorted(
                {
                    k: v
                    for k, v in self.fisher_analysis_results.items()
                    if isinstance(k, int)
                }.items()
            ):
                if "error" in result:
                    continue

                periodogram = result.get("periodogram", {})
                periods = periodogram.get("periods", [])
                z_scores = periodogram.get("z_scores", [])

                if len(periods) > 0 and len(z_scores) > 0:
                    periodogram_data = pd.DataFrame(
                        {
                            "Period_hours": periods,
                            "Z_Score": z_scores,
                        }
                    )
                    # Limit to reasonable number of points for Excel (max 10000)
                    if len(periodogram_data) > 10000:
                        # Downsample by taking every nth point
                        step = len(periodogram_data) // 10000 + 1
                        periodogram_data = periodogram_data.iloc[::step]

                    periodogram_data.to_excel(
                        writer, sheet_name=f"ROI_{roi_id}_Periodogram", index=False
                    )

            # Sheet 3: Sleep Phase Results (if available)
            sleep_results = self.fisher_analysis_results.get("sleep_phase_results", {})
            if sleep_results:
                sleep_roi_data = {
                    k: v for k, v in sleep_results.items() if isinstance(k, int)
                }
                if sleep_roi_data:
                    sleep_summary_data = []
                    for roi_id, result in sorted(sleep_roi_data.items()):
                        if "error" in result:
                            continue
                        periodogram = result.get("periodogram", {})
                        sleep_summary_data.append(
                            {
                                "ROI": roi_id,
                                "Significant Rhythm": periodogram.get(
                                    "is_significant", False
                                ),
                                "Dominant Period (hours)": periodogram.get(
                                    "dominant_period", 0
                                ),
                                "Z-Score": periodogram.get("dominant_z_score", 0),
                                "P-Value": periodogram.get("p_value", 1.0),
                            }
                        )
                    if sleep_summary_data:
                        sleep_df = pd.DataFrame(sleep_summary_data)
                        sleep_df.to_excel(
                            writer, sheet_name="Sleep_Phase_Summary", index=False
                        )

            # Sheet 4: Parameters
            has_sleep = bool(sleep_results)
            params_df = pd.DataFrame(
                {
                    "Parameter": [
                        "Minimum Period",
                        "Maximum Period",
                        "Significance Level",
                        "Sampling Interval",
                        "Sleep Phase Calculated",
                    ],
                    "Value": [
                        f"{self.fisher_min_period.value():.1f} hours",
                        f"{self.fisher_max_period.value():.1f} hours",
                        f"{self.fisher_significance.value():.3f}",
                        f"{self.frame_interval.value():.1f} seconds",
                        "Yes" if has_sleep else "No",
                    ],
                }
            )
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

    def _export_similarity_to_excel(self, file_path: str, similarity_results: Dict):
        """Export ROI Similarity results to Excel format."""
        import pandas as pd
        import numpy as np

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: Correlation Matrix
            corr_matrix = similarity_results.get("correlation_matrix", np.array([]))
            roi_ids = similarity_results.get("roi_ids", [])

            if len(corr_matrix) > 0:
                corr_df = pd.DataFrame(
                    corr_matrix,
                    index=[f"ROI {r}" for r in roi_ids],
                    columns=[f"ROI {r}" for r in roi_ids],
                )
                corr_df.to_excel(writer, sheet_name="Correlation_Matrix")

            # Sheet 2: Pairwise Correlations
            pairwise_data = []
            for i, roi1 in enumerate(roi_ids):
                for j, roi2 in enumerate(roi_ids):
                    if i < j:  # Upper triangle only
                        correlation = corr_matrix[i, j]
                        pairwise_data.append(
                            {
                                "ROI_1": roi1,
                                "ROI_2": roi2,
                                "Correlation": correlation,
                                "Status": (
                                    "Synchronized"
                                    if correlation > 0.7
                                    else ("Moderate" if correlation > 0.3 else "Low")
                                ),
                            }
                        )

            pairwise_df = pd.DataFrame(pairwise_data)
            pairwise_df = pairwise_df.sort_values("Correlation", ascending=False)
            pairwise_df.to_excel(
                writer, sheet_name="Pairwise_Correlations", index=False
            )

            # Sheet 3: Clustering
            if "clustering" in similarity_results:
                cluster_info = similarity_results["clustering"]
                cluster_data = []
                for cluster_id, roi_list in cluster_info.get("clusters", {}).items():
                    for roi in roi_list:
                        cluster_data.append(
                            {
                                "ROI": roi,
                                "Cluster": cluster_id,
                                "Cluster_Size": len(roi_list),
                            }
                        )

                cluster_df = pd.DataFrame(cluster_data)
                cluster_df = cluster_df.sort_values(["Cluster", "ROI"])
                cluster_df.to_excel(writer, sheet_name="Clusters", index=False)

    def _export_coherence_to_excel(self, file_path: str, coherence_results: Dict):
        """Export Coherence Analysis results to Excel format."""
        import pandas as pd
        import numpy as np

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: Coherence Matrix
            coherence_matrix = coherence_results.get("coherence_matrix", np.array([]))
            roi_ids = coherence_results.get("roi_ids", [])
            target_period = coherence_results.get("target_period_hours", 24.0)

            if len(coherence_matrix) > 0:
                coherence_df = pd.DataFrame(
                    coherence_matrix,
                    index=[f"ROI {r}" for r in roi_ids],
                    columns=[f"ROI {r}" for r in roi_ids],
                )
                coherence_df.to_excel(writer, sheet_name="Coherence_Matrix")

            # Sheet 2: Pairwise Coherence
            pairwise_coherence = coherence_results.get("pairwise_coherence", {})
            pairwise_data = []

            for (roi1, roi2), result in pairwise_coherence.items():
                pairwise_data.append(
                    {
                        "ROI_1": roi1,
                        "ROI_2": roi2,
                        "Coherence": result.get("circadian_coherence", 0),
                        "Coherence_Period": result.get(
                            "circadian_period", target_period
                        ),
                        "Max_Coherence": result.get("max_coherence", 0),
                        "Max_Coherence_Period": result.get("max_coherence_period", 0),
                        "Synchronized": (
                            "Yes" if result.get("is_synchronized", False) else "No"
                        ),
                    }
                )

            pairwise_df = pd.DataFrame(pairwise_data)
            pairwise_df = pairwise_df.sort_values("Coherence", ascending=False)
            pairwise_df.to_excel(writer, sheet_name="Pairwise_Coherence", index=False)

            # Sheet 3: Parameters
            params_df = pd.DataFrame(
                {
                    "Parameter": ["Target Period", "Number of ROIs"],
                    "Value": [f"{target_period:.1f} hours", str(len(roi_ids))],
                }
            )
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

    def _export_phase_clustering_to_excel(self, file_path: str, phase_results: Dict):
        """Export Phase Clustering results to Excel format."""
        import pandas as pd

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # Sheet 1: Phase Clusters
            clusters = phase_results.get("clusters", {})
            cluster_data = []

            for cluster_name, cluster_info in clusters.items():
                roi_list = cluster_info.get("rois", [])
                for roi in roi_list:
                    cluster_data.append(
                        {
                            "Cluster": cluster_name,
                            "ROI": roi,
                            "Cluster_Size": len(roi_list),
                        }
                    )

            cluster_df = pd.DataFrame(cluster_data)
            cluster_df = cluster_df.sort_values(["Cluster", "ROI"])
            cluster_df.to_excel(writer, sheet_name="Phase_Clusters", index=False)

            # Sheet 2: Individual ROI Phases
            roi_phases = phase_results.get("roi_phases", {})
            phase_data = []

            for roi_id, phase_info in sorted(roi_phases.items()):
                phase_data.append(
                    {
                        "ROI": roi_id,
                        "Peak_Time_Hours": phase_info.get("peak_time_hours", 0),
                        "Phase_Radians": phase_info.get("phase_radians", 0),
                        "Amplitude": phase_info.get("amplitude", 0),
                        "Mean_Activity": phase_info.get("mean_activity", 0),
                    }
                )

            phase_df = pd.DataFrame(phase_data)
            phase_df.to_excel(writer, sheet_name="ROI_Phases", index=False)

            # Sheet 3: Parameters
            params_df = pd.DataFrame(
                {
                    "Parameter": [
                        "Dominant Period",
                        "Total ROIs",
                        "Number of Clusters",
                    ],
                    "Value": [
                        f"{phase_results.get('dominant_period_hours', 0):.1f} hours",
                        str(phase_results.get("n_rois", 0)),
                        str(len(clusters)),
                    ],
                }
            )
            params_df.to_excel(writer, sheet_name="Parameters", index=False)

    def _create_fisher_plot(self, fisher_results: Dict[int, Dict]):
        """Create and display periodogram plot for Fisher analysis results."""
        try:
            import matplotlib.pyplot as plt
            from qtpy.QtGui import QPixmap
            import io

            # Close old figure if it exists to free memory
            if hasattr(self, "fisher_plot_figure") and self.fisher_plot_figure:
                plt.close(self.fisher_plot_figure)
                self.fisher_plot_figure = None

            # Get data source name for titles
            data_source_index = self.data_source_combo.currentIndex()
            data_source = (
                "Fraction Movement" if data_source_index == 0 else "Raw Intensity"
            )

            # Check if sleep phase results are available
            sleep_results = fisher_results.get("sleep_phase_results", None)
            has_sleep = sleep_results is not None and len(sleep_results) > 0

            # Create figure with subplots (only count integer ROI keys)
            roi_only_results = {
                k: v for k, v in fisher_results.items() if isinstance(k, int)
            }
            n_rois = len(roi_only_results)

            # Determine layout - max 3 columns
            n_cols = min(3, n_rois)
            n_rows_per_section = (n_rois + n_cols - 1) // n_cols

            # If we have sleep data, double the rows (Activity on top, Sleep on bottom)
            if has_sleep:
                total_rows = n_rows_per_section * 2
                fig_height = 3.5 * total_rows + 1.5
            else:
                total_rows = n_rows_per_section
                fig_height = 3.5 * total_rows + 0.5

            fig_width = 4 * n_cols
            fig, axes = plt.subplots(
                total_rows, n_cols, figsize=(fig_width, fig_height), squeeze=False
            )

            # Check if max period was capped for any ROI
            cap_note = ""
            for _roi_id, _roi_res in roi_only_results.items():
                _peri = _roi_res.get("periodogram", {})
                if _peri.get("period_capped", False):
                    _actual = _peri.get("actual_max_period", 0)
                    cap_note = f"  (max period capped at {_actual:.1f}h = recording / 2)"
                    break

            if has_sleep:
                fig.suptitle(
                    f"Chi² Periodogram  —  {data_source}{cap_note}{self._get_recording_start_str()}",
                    fontsize=11,
                    fontweight="bold",
                    y=0.99,
                )
            else:
                fig.suptitle(
                    f"Chi² Periodogram  —  Activity from {data_source}{cap_note}{self._get_recording_start_str()}",
                    fontsize=11,
                    fontweight="bold",
                    y=0.99,
                )

            # Helper function to plot a section (Activity or Sleep)
            def plot_section(results_dict, start_row, section_label):
                # Filter to integer keys only
                roi_items = {
                    k: v for k, v in results_dict.items() if isinstance(k, int)
                }

                for idx, (roi_id, result) in enumerate(sorted(roi_items.items())):
                    row = start_row + (idx // n_cols)
                    col = idx % n_cols
                    ax = axes[row, col]

                    periodogram = result.get("periodogram", {})

                    # Get ROI-specific color
                    roi_color = (
                        self.roi_colors.get(roi_id, f"C{idx}")
                        if hasattr(self, "roi_colors")
                        else f"C{idx}"
                    )

                    if "error" in result or "error" in periodogram:
                        ax.text(
                            0.5,
                            0.5,
                            f"ROI {roi_id}\nInsufficient data",
                            ha="center",
                            va="center",
                            transform=ax.transAxes,
                        )
                        ax.set_xticks([])
                        ax.set_yticks([])
                    else:
                        periods = periodogram.get("periods", [])
                        z_scores = periodogram.get("z_scores", [])
                        critical_z = periodogram.get("critical_z", 0)
                        is_significant = periodogram.get("is_significant", False)

                        ax.plot(
                            periods,
                            z_scores,
                            color=roi_color,
                            linewidth=1.5,
                            label="Z-score",
                        )

                        if critical_z > 0:
                            ax.axhline(
                                y=critical_z,
                                color="gray",
                                linestyle="--",
                                linewidth=1,
                                label="Significance",
                            )

                        if is_significant:
                            dominant_period = periodogram.get("dominant_period", 0)
                            dominant_z = periodogram.get("dominant_z_score", 0)
                            ax.axvline(
                                x=dominant_period,
                                color=roi_color,
                                linestyle="--",
                                linewidth=1.5,
                                alpha=0.5,
                            )
                            ax.plot(
                                dominant_period,
                                dominant_z,
                                "o",
                                color=roi_color,
                                markersize=8,
                                markeredgecolor="black",
                                markeredgewidth=1,
                                label=f"Peak: {dominant_period:.1f}h",
                            )

                        ax.set_xlabel("Period (h)", fontsize=9)
                        ax.set_ylabel("Z-score", fontsize=9)
                        ax.tick_params(axis="both", labelsize=8)
                        title_weight = "bold" if is_significant else "normal"
                        # Include section label in ROI title
                        ax.set_title(
                            f"ROI {roi_id} - {section_label}",
                            fontsize=9,
                            color=roi_color,
                            fontweight=title_weight,
                            pad=4,
                            loc="left",
                        )
                        ax.legend(fontsize=7, loc="best")
                        ax.grid(True, alpha=0.3)

                        # Per-ROI Y-axis scaling
                        if len(z_scores) > 0:
                            roi_y_max = max(z_scores) * 1.1
                            ax.set_ylim(
                                0,
                                max(
                                    roi_y_max,
                                    critical_z * 1.2 if critical_z > 0 else roi_y_max,
                                ),
                            )

                        # Set x-axis to actual data range (analysis caps max at duration/2)
                        if len(periods) > 0:
                            ax.set_xlim(min(periods), max(periods))

                # Hide unused subplots in this section
                for idx in range(n_rois, n_rows_per_section * n_cols):
                    row = start_row + (idx // n_cols)
                    col = idx % n_cols
                    if row < total_rows:
                        axes[row, col].axis("off")

            # Plot Activity section
            plot_section(roi_only_results, 0, "Activity")

            # Plot Sleep section if available
            if has_sleep:
                plot_section(sleep_results, n_rows_per_section, "Sleep")

            plt.tight_layout(rect=[0, 0, 1, 0.96], h_pad=2.0, w_pad=1.0)

            # Store figure for later export
            self.fisher_plot_figure = fig

            # Convert to QPixmap and display with higher DPI
            buf = io.BytesIO()
            fig.savefig(
                buf,
                format="png",
                dpi=150,
                bbox_inches="tight",
                facecolor="white",
                edgecolor="none",
            )
            buf.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buf.read())

            # Scale pixmap to fit canvas while maintaining aspect ratio
            scaled_pixmap = pixmap.scaled(
                self.fisher_plot_canvas.size(),
                1,  # Qt.KeepAspectRatio
                1,  # Qt.SmoothTransformation
            )
            self.fisher_plot_canvas.setPixmap(scaled_pixmap)

            # Enable pop-out button after successful plot creation
            if hasattr(self, "btn_popout_plot"):
                self.btn_popout_plot.setEnabled(True)
                if hasattr(self, "btn_save_fisher_plot"):
                    self.btn_save_fisher_plot.setEnabled(True)

            # Note: Don't close fig - we need it for export

        except Exception as e:
            self._log_message(f"⚠️ Could not create Fisher plot: {e}")
            import traceback

            traceback.print_exc()

    def _open_plot_window(self):
        """Open the current plot in a separate, resizable window."""
        if not hasattr(self, "fisher_plot_figure") or self.fisher_plot_figure is None:
            self._log_message("⚠️ No plot available to display")
            return

        try:
            from qtpy.QtWidgets import (
                QDialog,
                QVBoxLayout,
                QHBoxLayout,
                QPushButton,
                QScrollArea,
            )
            from matplotlib.backends.backend_qt5agg import (
                FigureCanvasQTAgg as FigureCanvas,
            )
            from matplotlib.backends.backend_qt5agg import (
                NavigationToolbar2QT as NavigationToolbar,
            )

            # Create dialog window
            dialog = QDialog(self)
            dialog.setWindowTitle("Rhythmic Pattern Analysis - Plot View")
            dialog.resize(1400, 900)  # Larger default size

            layout = QVBoxLayout()
            dialog.setLayout(layout)

            # Create matplotlib canvas
            canvas = FigureCanvas(self.fisher_plot_figure)
            canvas.setMinimumSize(800, 600)

            # Add navigation toolbar for zoom/pan
            toolbar = NavigationToolbar(canvas, dialog)
            layout.addWidget(toolbar)

            # Add canvas in scroll area for very large plots
            scroll_area = QScrollArea()
            scroll_area.setWidget(canvas)
            scroll_area.setWidgetResizable(True)
            layout.addWidget(scroll_area)

            # Add close button
            button_layout = QHBoxLayout()
            button_layout.addStretch()

            btn_save = QPushButton("Save Plot...")
            btn_save.clicked.connect(lambda: self._save_plot_from_dialog(canvas))
            button_layout.addWidget(btn_save)

            btn_close = QPushButton("Close")
            btn_close.clicked.connect(dialog.close)
            button_layout.addWidget(btn_close)

            layout.addLayout(button_layout)

            # Show dialog
            dialog.exec_()

        except Exception as e:
            self._log_message(f"⚠️ Could not open plot window: {e}")
            import traceback

            traceback.print_exc()

    def _save_plot_from_dialog(self, canvas):
        """Save plot from dialog window."""
        try:
            from qtpy.QtWidgets import QFileDialog
            import os

            # Get save location
            default_name = "rhythmic_pattern_plot.png"
            if hasattr(self, "current_fisher_method"):
                method_names = [
                    "fisher",
                    "fft",
                    "cosinor",
                    "similarity",
                    "coherence",
                    "phase",
                ]
                method_name = (
                    method_names[self.current_fisher_method]
                    if self.current_fisher_method < len(method_names)
                    else "plot"
                )
                default_name = f"rhythmic_pattern_{method_name}.png"

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Plot",
                default_name,
                "PNG Files (*.png);;PDF Files (*.pdf);;SVG Files (*.svg);;All Files (*)",
            )

            if file_path:
                # Determine format from extension
                ext = os.path.splitext(file_path)[1].lower()
                if ext == ".pdf":
                    canvas.figure.savefig(
                        file_path, format="pdf", dpi=300, bbox_inches="tight"
                    )
                elif ext == ".svg":
                    canvas.figure.savefig(file_path, format="svg", bbox_inches="tight")
                else:
                    canvas.figure.savefig(
                        file_path, format="png", dpi=300, bbox_inches="tight"
                    )

                self._log_message(f"✓ Plot saved to: {file_path}")

        except Exception as e:
            self._log_message(f"⚠️ Could not save plot: {e}")
            import traceback

            traceback.print_exc()

    def _save_fisher_plot(self):
        """Save the currently displayed periodogram / analysis plot as an image file."""
        if not hasattr(self, "fisher_plot_figure") or self.fisher_plot_figure is None:
            self._log_message("⚠️ No plot to save — run an analysis first.")
            return

        # Build a default filename from the HDF5 file and current method
        try:
            base = os.path.splitext(os.path.basename(self.file_path))[0]
        except Exception:
            base = "analysis"
        method_names = {0: "fisher", 1: "fft", 2: "cosinor", 3: "similarity",
                        4: "coherence", 5: "phase_clustering"}
        method_idx = getattr(self, "current_fisher_method", 0)
        method_tag = method_names.get(method_idx, "plot")
        default_name = f"{base}_{method_tag}_plot.png"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Plot",
            default_name,
            "PNG Files (*.png);;PDF Files (*.pdf);;SVG Files (*.svg);;All Files (*)",
        )
        if not file_path:
            return

        try:
            dpi = self.plot_dpi_spin.value() if hasattr(self, "plot_dpi_spin") else 150
            self.fisher_plot_figure.savefig(
                file_path, dpi=dpi, bbox_inches="tight", facecolor="white"
            )
            self._log_message(f"✓ Plot saved: {os.path.basename(file_path)}")
        except Exception as e:
            self._log_message(f"⚠️ Could not save plot: {e}")

    # ===================================================================
    # FRAME VIEWER METHODS
    # ===================================================================

    def _update_viewer_file_combo(self):
        """Update the file selector combo box with all available files."""
        if not hasattr(self, "viewer_file_combo"):
            return

        self.viewer_file_combo.blockSignals(True)
        self.viewer_file_combo.clear()

        files_added = set()

        # Add current HDF5 file
        if hasattr(self, "file_path") and self.file_path and self.file_path not in files_added:
            self.viewer_file_combo.addItem(
                os.path.basename(self.file_path), self.file_path
            )
            files_added.add(self.file_path)

        # Add all AVI batch files
        if hasattr(self, "avi_batch_paths") and self.avi_batch_paths:
            for avi_path in self.avi_batch_paths:
                if avi_path not in files_added:
                    self.viewer_file_combo.addItem(
                        os.path.basename(avi_path), avi_path
                    )
                    files_added.add(avi_path)

        # Add HDF5 files from directory
        if hasattr(self, "directory") and self.directory:
            try:
                for f in sorted(os.listdir(self.directory)):
                    full_path = os.path.join(self.directory, f)
                    if f.lower().endswith((".h5", ".hdf5")) and full_path not in files_added:
                        self.viewer_file_combo.addItem(f, full_path)
                        files_added.add(full_path)
            except Exception:
                pass

        self.viewer_file_combo.blockSignals(False)

        if self.viewer_file_combo.count() > 0:
            self._log_message(
                f"Frame Viewer: {self.viewer_file_combo.count()} file(s) available"
            )

    def _viewer_load_data(self):
        """Load the selected file into the frame viewer."""
        # Use file from combo box if available
        selected_path = None
        if hasattr(self, "viewer_file_combo") and self.viewer_file_combo.count() > 0:
            selected_path = self.viewer_file_combo.currentData()

        # Fallback to self.file_path
        if not selected_path:
            selected_path = getattr(self, "file_path", None)

        if not selected_path:
            self.viewer_status_label.setText(
                "⚠️ No file loaded. Please load a file in the Input tab first."
            )
            self._log_message("⚠️ Frame viewer: No file loaded")
            return

        # Set file_path to the selected file for viewer loading
        self.file_path = selected_path

        try:
            # Check if HDF5 or AVI
            is_hdf5 = selected_path.lower().endswith((".h5", ".hdf5"))
            is_avi = selected_path.lower().endswith(".avi")

            if is_avi:
                # Load AVI batch
                self._viewer_load_avi_batch()
            elif is_hdf5:
                # Load HDF5
                self._viewer_load_hdf5()
            else:
                self.viewer_status_label.setText("⚠️ Unsupported file format")
                self._log_message("⚠️ Frame viewer: Unsupported file format")

        except Exception as e:
            self.viewer_status_label.setText(f"❌ Error loading data: {e}")
            self._log_message(f"❌ Frame viewer error: {e}")
            import traceback

            traceback.print_exc()

    def _viewer_load_hdf5(self):
        """Load HDF5 file frames into viewer."""
        import h5py

        self._log_message(f"Loading HDF5 file into frame viewer: {self.file_path}")

        with h5py.File(self.file_path, "r") as f:
            # Get frame interval from metadata
            if "metadata" in f.attrs:
                import json

                metadata = json.loads(f.attrs["metadata"])
                self.viewer_frame_interval = metadata.get("frame_interval", 5.0)
            else:
                # Default to 5 seconds
                self.viewer_frame_interval = 5.0

            # Find the dataset - try multiple common names
            dataset_found = False

            # Try common dataset names in order
            for dataset_name in ["frames", "images", "data"]:
                if dataset_name in f:
                    data_obj = f[dataset_name]

                    if isinstance(data_obj, h5py.Dataset):
                        # Stacked frames format: (N, H, W) or (N, H, W, C)
                        self._log_message(
                            f"Found stacked dataset: {dataset_name} with shape {data_obj.shape}"
                        )
                        self.viewer_frames = data_obj
                        self.viewer_n_frames = (
                            data_obj.shape[0] if data_obj.ndim >= 3 else 1
                        )
                        self.viewer_file_handle = h5py.File(self.file_path, "r")
                        self.viewer_dataset_name = dataset_name
                        self.viewer_is_sequence = False
                        dataset_found = True
                        break
                    elif isinstance(data_obj, h5py.Group):
                        # Individual frames format: group with frame_XXXXXX datasets
                        frame_names = sorted(
                            [k for k in data_obj.keys() if k.startswith("frame_")]
                        )
                        if frame_names:
                            self._log_message(
                                f"Found individual frames in group: {dataset_name} ({len(frame_names)} frames)"
                            )
                            self.viewer_frames = None
                            self.viewer_frame_names = frame_names
                            self.viewer_n_frames = len(frame_names)
                            self.viewer_file_handle = h5py.File(self.file_path, "r")
                            self.viewer_dataset_name = dataset_name
                            self.viewer_is_sequence = True
                            dataset_found = True
                            break

            if not dataset_found:
                # List available keys for debugging
                available_keys = list(f.keys())
                self._log_message(f"Available HDF5 keys: {available_keys}")
                raise ValueError(
                    f"No 'frames', 'images', or 'data' dataset found in HDF5 file. "
                    f"Available keys: {available_keys}"
                )

        # Update UI
        self.viewer_current_frame = 0
        self.viewer_frame_slider.setMaximum(self.viewer_n_frames - 1)
        self.viewer_frame_slider.setValue(0)
        self.viewer_frame_slider.setEnabled(True)

        # Enable controls
        self.btn_viewer_first.setEnabled(True)
        self.btn_viewer_prev.setEnabled(True)
        self.btn_viewer_play.setEnabled(True)
        self.btn_viewer_next.setEnabled(True)
        self.btn_viewer_last.setEnabled(True)

        # Enable export controls
        self.export_start_frame.setEnabled(True)
        self.export_start_frame.setMaximum(self.viewer_n_frames - 1)
        self.export_start_frame.setValue(0)

        self.export_end_frame.setEnabled(True)
        self.export_end_frame.setMaximum(self.viewer_n_frames - 1)
        self.export_end_frame.setValue(self.viewer_n_frames - 1)  # Default to all frames

        self.btn_export_video.setEnabled(True)
        self.btn_export_gif.setEnabled(True)

        self.viewer_status_label.setText(
            f"✓ Loaded HDF5: {self.viewer_n_frames} frames (Interval: {self.viewer_frame_interval}s)"
        )
        self._log_message(
            f"✓ Frame viewer: Loaded {self.viewer_n_frames} frames from HDF5 (interval: {self.viewer_frame_interval}s)"
        )

        # Pre-load frames into cache for smooth playback
        self._viewer_preload_frames()

        # Display first frame
        self._viewer_show_frame(0)

    def _viewer_load_avi_batch(self):
        """Load AVI batch as one continuous video (same sampling as analysis)."""
        import numpy as np
        import cv2
        from qtpy.QtWidgets import QProgressDialog

        avi_paths = getattr(self, "avi_batch_paths", [])
        if not avi_paths:
            raise ValueError("No AVI batch paths available.")

        target_interval = getattr(self, "avi_batch_interval", 5.0)
        self.viewer_frame_interval = target_interval

        self._log_message(
            f"Loading AVI batch into frame viewer: {len(avi_paths)} files, "
            f"sampling interval={target_interval}s"
        )

        # First pass: count total sampled frames
        total_sampled = 0
        video_infos = []
        for path in avi_paths:
            cap = cv2.VideoCapture(path)
            if not cap.isOpened():
                self._log_message(f"Cannot open: {path}")
                continue
            fps = cap.get(cv2.CAP_PROP_FPS)
            n_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            cap.release()
            frames_per_sample = max(1, int(fps * target_interval))
            sampled = len(range(0, n_frames, frames_per_sample))
            video_infos.append((path, fps, n_frames, frames_per_sample, sampled))
            total_sampled += sampled

        if total_sampled == 0:
            raise ValueError("No frames found in AVI batch.")

        self._log_message(
            f"Total sampled frames across all videos: {total_sampled}"
        )

        # Second pass: load sampled frames with progress
        progress = QProgressDialog(
            f"Loading {total_sampled} sampled frames from {len(video_infos)} videos...",
            "Cancel", 0, total_sampled, self,
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(500)

        # Pre-estimate memory: get frame size from first video
        _h, _w = 0, 0
        _probe = cv2.VideoCapture(video_infos[0][0])
        if _probe.isOpened():
            _h = int(_probe.get(cv2.CAP_PROP_FRAME_HEIGHT))
            _w = int(_probe.get(cv2.CAP_PROP_FRAME_WIDTH))
        _probe.release()
        est_bytes = total_sampled * _h * _w  # grayscale = 1 byte/pixel
        est_gb = est_bytes / 1024**3
        if est_gb > 4.0:
            self._log_message(
                f"⚠️ AVI batch too large to load ({total_sampled} frames × "
                f"{_h}×{_w} ≈ {est_gb:.1f} GB). "
                f"Use a shorter time range or increase the frame interval."
            )
            progress.close()
            return

        all_frames = []
        loaded = 0
        oom_hit = False

        for path, fps, n_frames, frames_per_sample, sampled in video_infos:
            if oom_hit:
                break
            cap = cv2.VideoCapture(path)
            if not cap.isOpened():
                continue
            for frame_idx in range(0, n_frames, frames_per_sample):
                if progress.wasCanceled():
                    cap.release()
                    self._log_message("AVI batch loading canceled.")
                    return
                cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
                try:
                    ret, frame = cap.read()
                except (SystemError, MemoryError):
                    oom_hit = True
                    break
                if ret and frame is not None:
                    try:
                        # Convert to grayscale
                        if len(frame.shape) == 3:
                            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                        all_frames.append(frame)
                    except MemoryError:
                        oom_hit = True
                        break
                loaded += 1
                if loaded % 20 == 0:
                    progress.setValue(loaded)
            cap.release()

        if oom_hit:
            self._log_message(
                f"⚠️ Out of memory after loading {len(all_frames)} frames — "
                f"showing partial batch. Use a shorter time range."
            )

        progress.setValue(total_sampled)

        if not all_frames:
            raise ValueError("No frames loaded from AVI batch.")

        try:
            self.viewer_frames = np.stack(all_frames, axis=0)
        except MemoryError:
            self._log_message(
                f"⚠️ Cannot stack {len(all_frames)} frames into memory. "
                f"Reduce the number of videos or increase frame interval."
            )
            return
        self.viewer_n_frames = len(all_frames)
        self.viewer_file_handle = None
        self.viewer_is_sequence = False

        # Update UI controls
        self.viewer_current_frame = 0
        self.viewer_frame_slider.setMaximum(self.viewer_n_frames - 1)
        self.viewer_frame_slider.setValue(0)
        self.viewer_frame_slider.setEnabled(True)

        self.btn_viewer_first.setEnabled(True)
        self.btn_viewer_prev.setEnabled(True)
        self.btn_viewer_play.setEnabled(True)
        self.btn_viewer_next.setEnabled(True)
        self.btn_viewer_last.setEnabled(True)

        self.export_start_frame.setEnabled(True)
        self.export_start_frame.setMaximum(self.viewer_n_frames - 1)
        self.export_start_frame.setValue(0)
        self.export_end_frame.setEnabled(True)
        self.export_end_frame.setMaximum(self.viewer_n_frames - 1)
        self.export_end_frame.setValue(self.viewer_n_frames - 1)
        self.btn_export_video.setEnabled(True)
        self.btn_export_gif.setEnabled(True)

        # Build per-frame video source info for status display
        self._viewer_avi_video_boundaries = []
        frame_offset = 0
        for path, fps, n_frames, frames_per_sample, sampled in video_infos:
            self._viewer_avi_video_boundaries.append(
                (frame_offset, frame_offset + sampled - 1, os.path.basename(path))
            )
            frame_offset += sampled

        total_duration = self.viewer_n_frames * target_interval
        self.viewer_status_label.setText(
            f"AVI Batch: {self.viewer_n_frames} frames from {len(video_infos)} videos "
            f"({total_duration/60:.1f} min, interval={target_interval}s)"
        )
        self._log_message(
            f"Frame viewer: Loaded {self.viewer_n_frames} sampled frames "
            f"from {len(video_infos)} AVI files as continuous video"
        )
        for start, end, name in self._viewer_avi_video_boundaries:
            self._log_message(f"  {name}: frames {start}-{end}")

        # Pre-process for display cache
        self.viewer_frame_cache = [
            self._prepare_frame_for_display(f) for f in all_frames
        ]

        # Display first frame
        self._viewer_show_frame(0)

    def _viewer_preload_frames(self):
        """Pre-load all frames into memory cache for smooth playback."""
        from qtpy.QtWidgets import QProgressDialog
        from qtpy.QtCore import Qt
        import psutil

        # Get available system memory
        available_ram_mb = psutil.virtual_memory().available / (1024 * 1024)

        # Estimate memory usage
        # Get first frame to check size
        if hasattr(self, "viewer_file_handle") and self.viewer_file_handle:
            if hasattr(self, "viewer_frame_names"):
                frame_name = self.viewer_frame_names[0]
                sample_frame = self.viewer_file_handle[
                    f"{self.viewer_dataset_name}/{frame_name}"
                ][()]
            else:
                sample_frame = self.viewer_file_handle[self.viewer_dataset_name][0]
        else:
            sample_frame = self.viewer_frames[0]

        # Calculate memory (3 channels for BGR, uint8)
        frame_size_mb = (sample_frame.shape[0] * sample_frame.shape[1] * 3) / (
            1024 * 1024
        )
        total_size_mb = frame_size_mb * self.viewer_n_frames

        # Safety limits:
        # 1. Don't use more than 50% of available RAM
        # 2. Don't use more than 4GB total
        # 3. Don't cache more than 10000 frames
        max_ram_mb = min(available_ram_mb * 0.5, 4096)  # 50% of available or 4GB max
        max_frames = min(10000, int(max_ram_mb / frame_size_mb))

        self._log_message(
            f"System RAM: {available_ram_mb:.0f} MB available, "
            f"Frame size: {frame_size_mb:.2f} MB, "
            f"Total needed: {total_size_mb:.1f} MB"
        )

        # Check if caching is feasible
        if total_size_mb > max_ram_mb:
            self._log_message(
                f"⚠ Skipping frame cache: {total_size_mb:.0f} MB needed exceeds "
                f"safe limit ({max_ram_mb:.0f} MB). Playback may be slower."
            )
            self.viewer_frame_cache = None
            return

        if self.viewer_n_frames > max_frames:
            self._log_message(
                f"⚠ Skipping frame cache: {self.viewer_n_frames} frames exceeds "
                f"limit ({max_frames}). Playback may be slower."
            )
            self.viewer_frame_cache = None
            return

        self._log_message(
            f"Pre-loading {self.viewer_n_frames} frames into cache "
            f"(~{total_size_mb:.1f} MB)..."
        )

        # Show progress dialog
        progress = QProgressDialog(
            f"Loading {self.viewer_n_frames} frames into memory...",
            "Cancel",
            0,
            self.viewer_n_frames,
            self,
        )
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(500)  # Show after 500ms

        # Pre-load and pre-process all frames
        self.viewer_frame_cache = []

        try:
            for i in range(self.viewer_n_frames):
                # Update progress
                if i % 10 == 0:  # Update every 10 frames
                    progress.setValue(i)
                    if progress.wasCanceled():
                        self._log_message("Frame cache loading canceled by user")
                        self.viewer_frame_cache = None
                        return

                # Load frame
                if hasattr(self, "viewer_file_handle") and self.viewer_file_handle:
                    if hasattr(self, "viewer_frame_names"):
                        frame_name = self.viewer_frame_names[i]
                        frame_data = self.viewer_file_handle[
                            f"{self.viewer_dataset_name}/{frame_name}"
                        ][()]
                    else:
                        frame_data = self.viewer_file_handle[self.viewer_dataset_name][
                            i
                        ]
                else:
                    frame_data = self.viewer_frames[i]

                # Pre-process frame to display format
                frame_processed = self._prepare_frame_for_display(frame_data)
                self.viewer_frame_cache.append(frame_processed)

            progress.setValue(self.viewer_n_frames)
            self._log_message(
                f"✓ Frame cache ready: {self.viewer_n_frames} frames "
                f"({total_size_mb:.1f} MB in RAM)"
            )

        except Exception as e:
            self._log_message(f"❌ Frame cache loading failed: {e}")
            self.viewer_frame_cache = None
            import traceback

            traceback.print_exc()

    def _prepare_frame_for_display(self, frame_data):
        """Pre-process a frame to display format (BGR uint8)."""
        import numpy as np
        import cv2

        # Make a writable copy
        frame = np.array(frame_data, copy=True)

        # Ensure it's uint8
        if frame.dtype != np.uint8:
            # Normalize to 0-255 range
            frame_min = frame.min()
            frame_max = frame.max()
            if frame_max > frame_min:
                frame = ((frame - frame_min) / (frame_max - frame_min) * 255).astype(
                    np.uint8
                )
            else:
                frame = np.zeros_like(frame, dtype=np.uint8)

        # Ensure 2D shape
        if frame.ndim == 3 and frame.shape[2] == 1:
            frame = frame[:, :, 0]

        # Convert to 3-channel BGR for colored text overlay
        if len(frame.shape) == 2:
            frame = cv2.cvtColor(frame, cv2.COLOR_GRAY2BGR)

        return frame

    def _draw_roi_overlay(self, frame):
        """Draw ROI circles and numbers on a frame (in-place).

        Uses the stored circle positions from ROI detection. Falls back to
        computing circle approximations from masks if circles aren't available.
        """
        import cv2
        import numpy as np

        # Get ROI exclusion info
        excluded = set()
        if hasattr(self, "roi_checkboxes") and self.roi_checkboxes:
            excluded = {
                i for i, cb in enumerate(self.roi_checkboxes) if not cb.isChecked()
            }

        # Try to use stored circle positions (most accurate)
        circles = getattr(self, "_original_circles", None)
        scale = getattr(self, "roi_scale", None)
        scale_val = scale.value() if scale is not None else 1.0

        if circles is not None:
            for idx, circle in enumerate(circles):
                cx, cy, r = int(circle[0]), int(circle[1]), int(circle[2] * scale_val)
                is_excluded = idx in excluded

                # Green for active, red for excluded
                color = (0, 0, 255) if is_excluded else (0, 255, 0)
                cv2.circle(frame, (cx, cy), r, color, 2)

                # ROI number label
                label = f"{idx + 1}"
                font = cv2.FONT_HERSHEY_SIMPLEX
                font_scale = 1.8
                thickness = 3
                (tw, th), _ = cv2.getTextSize(label, font, font_scale, thickness)
                text_x = cx - tw // 2
                text_y = cy + th // 2
                # Dark background for readability
                cv2.rectangle(
                    frame,
                    (text_x - 4, text_y - th - 4),
                    (text_x + tw + 4, text_y + 4),
                    (0, 0, 0),
                    -1,
                )
                text_color = (100, 100, 255) if is_excluded else (255, 100, 100)
                cv2.putText(
                    frame, label, (text_x, text_y),
                    font, font_scale, text_color, thickness, cv2.LINE_AA,
                )
            return

        # Fallback: derive circles from masks
        masks = getattr(self, "masks", [])
        if not masks:
            return

        for idx, mask in enumerate(masks):
            if mask is None or mask.size == 0:
                continue
            is_excluded = idx in excluded
            ys, xs = np.where(mask > 0)
            if len(xs) == 0:
                continue
            cx = int(np.mean(xs))
            cy = int(np.mean(ys))
            r = int(np.sqrt(len(xs) / np.pi))

            color = (0, 0, 255) if is_excluded else (0, 255, 0)
            cv2.circle(frame, (cx, cy), r, color, 2)

            label = f"{idx + 1}"
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 1.8
            thickness = 3
            (tw, th), _ = cv2.getTextSize(label, font, font_scale, thickness)
            text_x = cx - tw // 2
            text_y = cy + th // 2
            cv2.rectangle(
                frame,
                (text_x - 4, text_y - th - 4),
                (text_x + tw + 4, text_y + 4),
                (0, 0, 0),
                -1,
            )
            text_color = (100, 100, 255) if is_excluded else (255, 100, 100)
            cv2.putText(
                frame, label, (text_x, text_y),
                font, font_scale, text_color, thickness, cv2.LINE_AA,
            )

    def _on_viewer_roi_overlay_changed(self, enabled):
        """Re-render the current frame when the ROI overlay toggle changes."""
        if hasattr(self, "viewer_current_frame"):
            self._viewer_show_frame(self.viewer_current_frame)

    def _viewer_show_frame(self, frame_idx):
        """Display a specific frame in the napari viewer."""
        try:
            if frame_idx < 0 or frame_idx >= self.viewer_n_frames:
                return

            # Calculate time from frame index
            frame_time_seconds = frame_idx * self.viewer_frame_interval
            frame_time_minutes = frame_time_seconds / 60.0
            frame_time_hours = frame_time_minutes / 60.0

            import numpy as np
            import cv2

            # Use cached frame if available (FAST!)
            if hasattr(self, "viewer_frame_cache") and self.viewer_frame_cache:
                # Get pre-processed frame from cache (already BGR uint8)
                frame_with_text = self.viewer_frame_cache[frame_idx].copy()
                frame_data = frame_with_text  # For info display
            else:
                # Fallback: Load and process frame on-the-fly (SLOWER)
                # Get frame data
                if hasattr(self, "viewer_file_handle") and self.viewer_file_handle:
                    # HDF5 file
                    if hasattr(self, "viewer_frame_names"):
                        # Individual frame datasets
                        frame_name = self.viewer_frame_names[frame_idx]
                        frame_data = self.viewer_file_handle[
                            f"{self.viewer_dataset_name}/{frame_name}"
                        ][()]
                    else:
                        # Single dataset
                        frame_data = self.viewer_file_handle[self.viewer_dataset_name][
                            frame_idx
                        ]
                else:
                    # From napari layer (AVI or pre-loaded)
                    frame_data = self.viewer_frames[frame_idx]

                # Process frame
                frame_with_text = self._prepare_frame_for_display(frame_data)

            # Prepare time text
            time_text = f"t = {frame_time_seconds:.1f}s ({frame_time_minutes:.2f}min)"

            # Position: lower left (10 pixels from left, 30 pixels from bottom)
            height, width = frame_with_text.shape[:2]
            text_position = (10, height - 10)

            # Draw text in white (BGR: 255, 255, 255), 50% larger
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 1.2  # Increased from 0.8 (50% larger)
            color = (255, 255, 255)  # White in BGR
            thickness = 2

            cv2.putText(
                frame_with_text,
                time_text,
                text_position,
                font,
                font_scale,
                color,
                thickness,
                cv2.LINE_AA,
            )

            # Draw ROI overlay if enabled
            if (
                hasattr(self, "viewer_show_roi_overlay")
                and self.viewer_show_roi_overlay.isChecked()
            ):
                self._draw_roi_overlay(frame_with_text)

            # Create or update napari layer
            layer_name = "Frame Viewer"

            # Check if layer exists
            if layer_name in self.viewer.layers:
                # Update existing layer
                self.viewer.layers[layer_name].data = frame_with_text
            else:
                # Create new layer
                self.viewer.add_image(
                    frame_with_text,
                    name=layer_name,
                    colormap="gray",
                )

            # Update frame label with time
            self.viewer_frame_label.setText(
                f"Frame: {frame_idx + 1} / {self.viewer_n_frames} | "
                f"Time: {frame_time_seconds:.1f}s ({frame_time_minutes:.2f}min / {frame_time_hours:.3f}h)"
            )
            self.viewer_current_frame = frame_idx

            # Update info (use original frame_data, not the one with text)
            info_lines = [
                f"Frame: {frame_idx + 1} / {self.viewer_n_frames}",
                f"Time: {frame_time_seconds:.1f}s ({frame_time_minutes:.2f} min)",
                f"Hours: {frame_time_hours:.3f} h",
                f"Shape: {frame_data.shape}",
                f"Dtype: {frame_data.dtype}",
                f"Min/Max: {np.min(frame_data):.2f} / {np.max(frame_data):.2f}",
                f"Mean: {np.mean(frame_data):.2f}",
            ]
            self.viewer_info_text.setPlainText("\n".join(info_lines))

        except Exception as e:
            self._log_message(f"❌ Error showing frame {frame_idx}: {e}")

    def _on_viewer_frame_changed(self, value):
        """Handle slider value change."""
        self._viewer_show_frame(value)
        # Update synchronized plot time marker
        if hasattr(self, "sync_plots_enabled") and self.sync_plots_enabled.isChecked():
            self._update_sync_time_marker(value)

    def _toggle_sync_plots(self, enabled):
        """Toggle synchronized plots visibility."""
        self.sync_canvas.setVisible(enabled)
        if enabled:
            self._update_sync_plot()

    def _get_sync_plot_data(self):
        """Return (plot_type, data_dict) for the current sync plot settings, with rebinning applied."""
        sync_type = self.sync_plot_type.currentText()
        sync_bin = self.sync_plot_bin.value() if hasattr(self, "sync_plot_bin") else 0
        original_bin = self.bin_size_seconds.value() if hasattr(self, "bin_size_seconds") else 60
        new_bin_seconds = sync_bin * 60

        def maybe_rebin(dd):
            if dd and new_bin_seconds > original_bin:
                return self._rebin_timeseries_data(dd, new_bin_seconds, original_bin)
            return dd

        if "Raw Intensity" in sync_type:
            return "Raw Intensity Changes", self.merged_results
        elif "Movement (binary)" in sync_type:
            return "Movement", maybe_rebin(getattr(self, "movement_data", {}))
        elif "Fraction Movement" in sync_type:
            return "Fraction Movement", maybe_rebin(getattr(self, "fraction_data", {}))
        elif "Quiescence" in sync_type:
            return "Quiescence", maybe_rebin(getattr(self, "quiescence_data", {}))
        elif "Sleep Quality" in sync_type:
            return "Sleep Quality", getattr(self, "sleep_quality_data", {})
        elif "Sleep" in sync_type:
            return "Sleep", maybe_rebin(getattr(self, "sleep_data", {}))
        else:
            return "Raw Intensity Changes", self.merged_results

    def _update_sync_plot(self):
        """Update the synchronized analysis plot (same style as Analysis tab)."""
        if not hasattr(self, "sync_plots_enabled") or not self.sync_plots_enabled.isChecked():
            return

        # Check if we have analysis data
        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("⚠️ No analysis data for synchronized plots")
            return

        from ._plot import PlotGenerator, create_plot_config, create_hysteresis_kwargs

        plot_type, data_dict = self._get_sync_plot_data()

        if not data_dict:
            self._log_message(f"⚠️ No {self.sync_plot_type.currentText()} data available")
            return

        # Get ROI colors
        roi_colors = getattr(self, "roi_colors", {})
        n_rois = len(data_dict)
        if n_rois == 0:
            return

        # Clear figure
        self.sync_figure.clear()

        # Dynamically adjust figure size based on canvas size and number of ROIs
        canvas_height = self.sync_canvas.height()
        canvas_width = self.sync_canvas.width()
        dpi = self.sync_figure.get_dpi()

        # Set figure size to match canvas
        fig_width = max(10, canvas_width / dpi)
        fig_height = max(4, canvas_height / dpi)
        self.sync_figure.set_size_inches(fig_width, fig_height)

        # Create plot configuration from widget settings (same as Analysis tab)
        plot_config = create_plot_config(self)
        # Calculate optimal height per ROI based on available space
        plot_config["height_per_roi"] = max(0.6, fig_height / n_rois) if n_rois > 0 else 0.8
        plot_config["fig_width"] = fig_width

        # Get kwargs for specific plot types
        if plot_type == "Raw Intensity Changes":
            hysteresis_kwargs = create_hysteresis_kwargs(widget_instance=self)
            hysteresis_kwargs.pop("merged_results", None)
        elif plot_type == "Sleep Quality":
            hysteresis_kwargs = {"sleep_metric": "sleep_minutes"}
        else:
            hysteresis_kwargs = {}

        # Use PlotGenerator (same as Analysis tab)
        plot_generator = PlotGenerator(self.sync_figure)
        plot_generator.generate_plot(
            plot_type, data_dict, roi_colors, plot_config, **hysteresis_kwargs
        )

        # Store time markers for each subplot
        self.sync_time_markers = []
        for ax in self.sync_figure.get_axes():
            marker = ax.axvline(x=0, color="red", linewidth=2, linestyle="--", alpha=0.9, zorder=100)
            self.sync_time_markers.append(marker)

        # Adjust layout to maximize plot area (suppress warning for incompatible axes)
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                # Use subplots_adjust for better space usage
                self.sync_figure.subplots_adjust(
                    left=0.08, right=0.92, top=0.95, bottom=0.08, hspace=0.15
                )
        except Exception:
            pass

        # Draw canvas
        self.sync_canvas.draw()

        # Update marker to current position
        if hasattr(self, "viewer_current_frame"):
            self._update_sync_time_marker(self.viewer_current_frame)

    def _update_sync_time_marker(self, frame_idx):
        """Update the time marker position in synchronized plots."""
        if not hasattr(self, "sync_time_markers") or not self.sync_time_markers:
            return

        if not hasattr(self, "viewer_frame_interval"):
            return

        # Calculate time in minutes
        time_minutes = (frame_idx * self.viewer_frame_interval) / 60.0

        # Update all markers
        for marker in self.sync_time_markers:
            marker.set_xdata([time_minutes, time_minutes])

        # Redraw canvas (use blit for performance if available)
        try:
            self.sync_canvas.draw_idle()
        except Exception:
            pass

    def _export_video_with_plots(self):
        """Export video with synchronized analysis plots (same style as Analysis tab)."""
        import cv2
        import numpy as np
        from qtpy.QtWidgets import QFileDialog, QProgressDialog
        from qtpy.QtCore import Qt
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        from ._plot import PlotGenerator, create_plot_config, create_hysteresis_kwargs

        # Check prerequisites
        if not hasattr(self, "viewer_n_frames") or self.viewer_n_frames == 0:
            self._log_message("❌ No frames loaded. Load data into Frame Viewer first.")
            return

        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("❌ No analysis data. Run Main Analysis first.")
            return

        # Get frame range from export controls
        start_frame = self.export_start_frame.value()
        end_frame = self.export_end_frame.value()

        if start_frame >= end_frame:
            self._log_message("❌ Start frame must be less than end frame")
            return

        # Ask for save location
        default_name = f"video_with_plots_{start_frame}-{end_frame}.mp4"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Video with Plots",
            default_name,
            "MP4 Video (*.mp4);;All Files (*)",
        )

        if not file_path:
            return

        export_fps = self.export_fps_spin.value()
        n_frames = end_frame - start_frame

        self._log_message(f"🎬 Exporting {n_frames} frames with plots at {export_fps} FPS...")

        # Setup progress dialog
        progress = QProgressDialog("Exporting video with plots...", "Cancel", 0, n_frames, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setWindowTitle("Export Progress")
        progress.show()

        # Process events to show dialog
        from qtpy.QtWidgets import QApplication
        QApplication.processEvents()

        plot_type, data_dict = self._get_sync_plot_data()

        if not data_dict:
            self._log_message(f"❌ No {self.sync_plot_type.currentText()} data available")
            return

        roi_colors = getattr(self, "roi_colors", {})
        n_rois = len(data_dict)

        # Get first frame to determine dimensions
        first_frame = self._get_viewer_frame(start_frame)
        frame_height, frame_width = first_frame.shape[:2]

        # Create plot figure dimensions matching Analysis tab style
        plot_dpi = 100
        plot_width = frame_width
        height_per_roi = 60  # pixels per ROI
        plot_height = max(200, height_per_roi * n_rois)
        fig_width = plot_width / plot_dpi
        fig_height = plot_height / plot_dpi

        # Total output dimensions: frame on top, plots below
        output_width = frame_width
        output_height = frame_height + plot_height

        # Initialize video writer
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        out = cv2.VideoWriter(file_path, fourcc, export_fps, (output_width, output_height))

        # Create plot configuration from widget settings (same as Analysis tab)
        base_plot_config = create_plot_config(self)
        base_plot_config["fig_width"] = fig_width
        base_plot_config["height_per_roi"] = fig_height / n_rois if n_rois > 0 else 0.6
        base_plot_config["dpi"] = plot_dpi

        # Get extra kwargs for specific plot types
        if plot_type == "Raw Intensity Changes":
            hysteresis_kwargs = create_hysteresis_kwargs(widget_instance=self)
            hysteresis_kwargs.pop("merged_results", None)
        elif plot_type == "Sleep Quality":
            sleep_metric = "sleep_minutes"
            if hasattr(self, "sleep_quality_metric_combo"):
                metric_text = self.sleep_quality_metric_combo.currentText()
                if "Transitions" in metric_text:
                    sleep_metric = "transitions"
                elif "Bout" in metric_text:
                    sleep_metric = "bout_length"
            hysteresis_kwargs = {"sleep_metric": sleep_metric}
        else:
            hysteresis_kwargs = {}

        try:
            for i, frame_idx in enumerate(range(start_frame, end_frame)):
                if progress.wasCanceled():
                    self._log_message("⚠️ Export cancelled by user")
                    break

                progress.setValue(i)

                # Log progress every 10 frames
                if i % 10 == 0:
                    self._log_message(f"   Processing frame {i + 1}/{n_frames}...")
                    QApplication.processEvents()

                # Get frame
                frame = self._get_viewer_frame(frame_idx)

                # Add time text to frame
                time_seconds = frame_idx * self.viewer_frame_interval
                time_minutes = time_seconds / 60.0
                time_text = f"t = {time_seconds:.1f}s ({time_minutes:.2f}min)"

                frame_bgr = frame.copy()
                if len(frame_bgr.shape) == 2:
                    frame_bgr = cv2.cvtColor(frame_bgr, cv2.COLOR_GRAY2BGR)
                elif len(frame_bgr.shape) == 3:
                    if frame_bgr.shape[2] == 4:
                        frame_bgr = cv2.cvtColor(frame_bgr, cv2.COLOR_RGBA2BGR)
                    elif frame_bgr.shape[2] == 1:
                        frame_bgr = cv2.cvtColor(frame_bgr[:, :, 0], cv2.COLOR_GRAY2BGR)
                    # else: assume already BGR with 3 channels

                cv2.putText(
                    frame_bgr, time_text, (10, frame_height - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2, cv2.LINE_AA
                )

                # Draw ROI overlay (circles + numbers) if enabled
                if hasattr(self, "viewer_show_roi_overlay") and self.viewer_show_roi_overlay.isChecked():
                    self._draw_roi_overlay(frame_bgr)

                # Create plot using PlotGenerator (same as Analysis tab)
                plot_fig = Figure(figsize=(fig_width, fig_height), dpi=plot_dpi)
                plot_generator = PlotGenerator(plot_fig)

                # Generate plot with same settings as Analysis tab
                plot_generator.generate_plot(
                    plot_type, data_dict, roi_colors, base_plot_config, **hysteresis_kwargs
                )

                # Add red time marker to all subplots
                for ax in plot_fig.get_axes():
                    ax.axvline(
                        x=time_minutes, color="red", linewidth=2,
                        linestyle="--", alpha=0.9, zorder=100
                    )

                # Render plot to image (suppress tight_layout warning)
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    plot_fig.tight_layout()
                plot_canvas = FigureCanvasAgg(plot_fig)
                plot_canvas.draw()

                # Get image from canvas
                buf = plot_canvas.buffer_rgba()
                plot_img = np.asarray(buf)
                plot_img_bgr = cv2.cvtColor(plot_img, cv2.COLOR_RGBA2BGR)

                # Resize plot to match frame width if needed
                if plot_img_bgr.shape[1] != frame_width or plot_img_bgr.shape[0] != plot_height:
                    plot_img_bgr = cv2.resize(plot_img_bgr, (frame_width, plot_height))

                # Close figure to free memory
                import matplotlib.pyplot as plt
                plt.close(plot_fig)

                # Combine frame and plot vertically
                combined = np.vstack([frame_bgr, plot_img_bgr])

                # Write frame
                out.write(combined)

            progress.setValue(n_frames)
            out.release()

            self._log_message(f"✅ Video with plots exported: {file_path}")
            self._log_message(f"   Dimensions: {output_width}x{output_height}, {n_frames} frames at {export_fps} FPS")

        except Exception as e:
            out.release()
            self._log_message(f"❌ Export failed: {e}")
            import traceback
            traceback.print_exc()

    def _get_viewer_frame(self, frame_idx):
        """Get a single frame from the viewer data source."""
        import numpy as np

        if hasattr(self, "viewer_frame_cache") and self.viewer_frame_cache:
            return self.viewer_frame_cache[frame_idx].copy()

        if hasattr(self, "viewer_file_handle") and self.viewer_file_handle:
            if hasattr(self, "viewer_frame_names"):
                frame_name = self.viewer_frame_names[frame_idx]
                frame = self.viewer_file_handle[f"{self.viewer_dataset_name}/{frame_name}"][()]
            else:
                frame = self.viewer_file_handle[self.viewer_dataset_name][frame_idx]
        else:
            frame = self.viewer_frames[frame_idx]

        # Normalize to uint8 if needed
        frame = np.array(frame)
        if frame.dtype != np.uint8:
            if frame.max() > 0:
                frame = ((frame - frame.min()) / (frame.max() - frame.min()) * 255).astype(np.uint8)
            else:
                frame = np.zeros_like(frame, dtype=np.uint8)

        return frame

    def _viewer_goto_frame(self, frame_idx):
        """Go to specific frame."""
        if frame_idx < 0:
            frame_idx = self.viewer_n_frames - 1
        frame_idx = max(0, min(self.viewer_n_frames - 1, frame_idx))
        self.viewer_frame_slider.setValue(frame_idx)

    def _viewer_step_frame(self, step):
        """Step forward or backward by n frames."""
        new_idx = self.viewer_current_frame + step
        self._viewer_goto_frame(new_idx)

    def _viewer_toggle_play(self):
        """Toggle playback on/off."""
        if self.btn_viewer_play.isChecked():
            # Start playing
            self.viewer_is_playing = True
            self.btn_viewer_play.setText("⏸ Pause")
            interval = int(1000 / self.viewer_fps_spin.value())
            self.viewer_timer.start(interval)
            self._log_message(f"▶ Playing at {self.viewer_fps_spin.value()} FPS")
        else:
            # Stop playing
            self.viewer_is_playing = False
            self.btn_viewer_play.setText("▶ Play")
            self.viewer_timer.stop()
            self._log_message("⏸ Paused")

    def _viewer_play_next_frame(self):
        """Advance to next frame during playback."""
        if self.viewer_is_playing:
            next_idx = self.viewer_current_frame + 1
            if next_idx >= self.viewer_n_frames:
                # Loop back to start
                next_idx = 0
            self._viewer_goto_frame(next_idx)

    def _viewer_update_timer_interval(self):
        """Update playback timer interval when FPS changes."""
        if self.viewer_is_playing:
            interval = int(1000 / self.viewer_fps_spin.value())
            self.viewer_timer.setInterval(interval)

    def _export_video(self):
        """Export selected frame range as MP4 video."""
        try:
            import cv2
            import numpy as np
            from qtpy.QtWidgets import QFileDialog

            # Get frame range
            start_frame = self.export_start_frame.value()
            end_frame = self.export_end_frame.value()

            if start_frame >= end_frame:
                self._log_message("❌ Start frame must be less than end frame")
                return

            # Get export FPS
            export_fps = self.export_fps_spin.value()

            # Ask user for save location
            default_name = f"export_frames_{start_frame}-{end_frame}.mp4"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Video",
                default_name,
                "MP4 Video (*.mp4);;All Files (*)",
            )

            if not file_path:
                return  # User cancelled

            self._log_message(
                f"🎬 Exporting frames {start_frame}-{end_frame} as video ({export_fps} FPS)..."
            )

            n_frames = end_frame - start_frame + 1
            from qtpy.QtWidgets import QProgressDialog
            from qtpy.QtCore import Qt
            progress_dlg = QProgressDialog(
                f"Exporting video ({n_frames} frames)...", "Cancel", 0, n_frames, self
            )
            progress_dlg.setWindowModality(Qt.WindowModal)
            progress_dlg.setWindowTitle("Export Progress")
            progress_dlg.setMinimumDuration(0)
            progress_dlg.setValue(0)
            progress_dlg.show()

            # Get first frame to determine dimensions
            if hasattr(self, "viewer_file_handle") and self.viewer_file_handle:
                if hasattr(self, "viewer_frame_names"):
                    frame_name = self.viewer_frame_names[start_frame]
                    first_frame = self.viewer_file_handle[
                        f"{self.viewer_dataset_name}/{frame_name}"
                    ][()]
                else:
                    first_frame = self.viewer_file_handle[self.viewer_dataset_name][
                        start_frame
                    ]
            else:
                first_frame = self.viewer_frames[start_frame]

            # Prepare frame with text overlay — use float32 to halve memory usage
            first_frame = np.array(first_frame, copy=True)
            if first_frame.dtype != np.uint8:
                frame_min = first_frame.min()
                frame_max = first_frame.max()
                if frame_max > frame_min:
                    first_frame = (
                        (first_frame.astype(np.float32) - frame_min)
                        / (frame_max - frame_min)
                        * 255
                    ).astype(np.uint8)
                else:
                    first_frame = np.zeros_like(first_frame, dtype=np.uint8)

            if first_frame.ndim == 3 and first_frame.shape[2] == 1:
                first_frame = first_frame[:, :, 0]

            if len(first_frame.shape) == 2:
                first_frame = cv2.cvtColor(first_frame, cv2.COLOR_GRAY2BGR)

            height, width = first_frame.shape[:2]

            # Initialize video writer
            fourcc = cv2.VideoWriter_fourcc(*"mp4v")
            out = cv2.VideoWriter(file_path, fourcc, export_fps, (width, height))

            if not out.isOpened():
                self._log_message(f"❌ Failed to create video file: {file_path}")
                return

            # Process each frame
            cancelled = False
            for idx, frame_idx in enumerate(range(start_frame, end_frame + 1)):
                if progress_dlg.wasCanceled():
                    cancelled = True
                    break

                # Get frame data
                if hasattr(self, "viewer_file_handle") and self.viewer_file_handle:
                    if hasattr(self, "viewer_frame_names"):
                        frame_name = self.viewer_frame_names[frame_idx]
                        frame_data = self.viewer_file_handle[
                            f"{self.viewer_dataset_name}/{frame_name}"
                        ][()]
                    else:
                        frame_data = self.viewer_file_handle[self.viewer_dataset_name][
                            frame_idx
                        ]
                else:
                    frame_data = self.viewer_frames[frame_idx]

                # Prepare frame — use float32 to halve memory usage
                frame = np.array(frame_data, copy=True)
                if frame.dtype != np.uint8:
                    frame_min = frame.min()
                    frame_max = frame.max()
                    if frame_max > frame_min:
                        frame = (
                            (frame.astype(np.float32) - frame_min)
                            / (frame_max - frame_min)
                            * 255
                        ).astype(np.uint8)
                    else:
                        frame = np.zeros_like(frame, dtype=np.uint8)

                if frame.ndim == 3 and frame.shape[2] == 1:
                    frame = frame[:, :, 0]

                if len(frame.shape) == 2:
                    frame = cv2.cvtColor(frame, cv2.COLOR_GRAY2BGR)

                # Add time text
                frame_time_seconds = frame_idx * self.viewer_frame_interval
                frame_time_minutes = frame_time_seconds / 60.0
                time_text = (
                    f"t = {frame_time_seconds:.1f}s ({frame_time_minutes:.2f}min)"
                )

                text_position = (10, height - 10)
                font = cv2.FONT_HERSHEY_SIMPLEX
                font_scale = 1.2
                color = (255, 255, 255)  # White
                thickness = 2

                cv2.putText(
                    frame,
                    time_text,
                    text_position,
                    font,
                    font_scale,
                    color,
                    thickness,
                    cv2.LINE_AA,
                )

                # Write frame
                out.write(frame)
                progress_dlg.setValue(idx + 1)

            # Finalize
            progress_dlg.setValue(n_frames)
            out.release()

            if cancelled:
                import os as _os
                try:
                    _os.remove(file_path)
                except OSError:
                    pass
                self._log_message("⚠️ Video export cancelled.")
                return

            self._log_message(f"✅ Video exported successfully: {file_path}")
            self._log_message(
                f"   {n_frames} frames at {export_fps} FPS ({n_frames/export_fps:.1f}s duration)"
            )

        except Exception as e:
            self._log_message(f"❌ Video export failed: {e}")
            import traceback

            traceback.print_exc()

    def _export_gif(self):
        """Export selected frame range as animated GIF."""
        try:
            import cv2
            import numpy as np
            from qtpy.QtWidgets import QFileDialog
            from PIL import Image

            # Get frame range
            start_frame = self.export_start_frame.value()
            end_frame = self.export_end_frame.value()

            if start_frame >= end_frame:
                self._log_message("❌ Start frame must be less than end frame")
                return

            # Pre-flight: GIF stores all frames in RAM simultaneously.
            # 1024×1224 RGB uint8 ≈ 3.6 MB per frame → 200 frames ≈ 720 MB.
            MAX_GIF_FRAMES = 200
            n_frames_requested = end_frame - start_frame + 1
            if n_frames_requested > MAX_GIF_FRAMES:
                self._log_message(
                    f"⚠️ GIF export blocked: {n_frames_requested} frames requested, "
                    f"maximum is {MAX_GIF_FRAMES} (GIF loads all frames into RAM at once). "
                    f"Use 'Export as Video (MP4)' for long sequences."
                )
                return

            # Get export FPS
            export_fps = self.export_fps_spin.value()
            frame_duration = int(1000 / export_fps)  # Duration in milliseconds

            # Ask user for save location
            default_name = f"export_frames_{start_frame}-{end_frame}.gif"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export GIF",
                default_name,
                "GIF Animation (*.gif);;All Files (*)",
            )

            if not file_path:
                return  # User cancelled

            n_frames = end_frame - start_frame + 1
            self._log_message(
                f"🎞️ Exporting frames {start_frame}-{end_frame} as GIF ({export_fps} FPS)..."
            )

            from qtpy.QtWidgets import QProgressDialog
            from qtpy.QtCore import Qt
            progress_dlg = QProgressDialog(
                f"Exporting GIF ({n_frames} frames)...", "Cancel", 0, n_frames, self
            )
            progress_dlg.setWindowModality(Qt.WindowModal)
            progress_dlg.setWindowTitle("Export Progress")
            progress_dlg.setMinimumDuration(0)
            progress_dlg.setValue(0)
            progress_dlg.show()

            frames_for_gif = []
            cancelled = False

            # Process each frame
            for idx, frame_idx in enumerate(range(start_frame, end_frame + 1)):
                if progress_dlg.wasCanceled():
                    cancelled = True
                    break

                # Get frame data
                if hasattr(self, "viewer_file_handle") and self.viewer_file_handle:
                    if hasattr(self, "viewer_frame_names"):
                        frame_name = self.viewer_frame_names[frame_idx]
                        frame_data = self.viewer_file_handle[
                            f"{self.viewer_dataset_name}/{frame_name}"
                        ][()]
                    else:
                        frame_data = self.viewer_file_handle[self.viewer_dataset_name][
                            frame_idx
                        ]
                else:
                    frame_data = self.viewer_frames[frame_idx]

                # Prepare frame
                try:
                    frame = np.array(frame_data, copy=True)
                    if frame.dtype != np.uint8:
                        frame_min = frame.min()
                        frame_max = frame.max()
                        if frame_max > frame_min:
                            frame = (
                                (frame.astype(np.float32) - frame_min)
                                / (frame_max - frame_min)
                                * 255
                            ).astype(np.uint8)
                        else:
                            frame = np.zeros_like(frame, dtype=np.uint8)
                except MemoryError:
                    self._log_message(
                        f"⚠️ GIF export ran out of memory at frame {idx + 1}/{n_frames}. "
                        f"Use 'Export as Video (MP4)' instead."
                    )
                    return

                if frame.ndim == 3 and frame.shape[2] == 1:
                    frame = frame[:, :, 0]

                if len(frame.shape) == 2:
                    frame = cv2.cvtColor(frame, cv2.COLOR_GRAY2BGR)

                # Add time text
                frame_time_seconds = frame_idx * self.viewer_frame_interval
                frame_time_minutes = frame_time_seconds / 60.0
                time_text = (
                    f"t = {frame_time_seconds:.1f}s ({frame_time_minutes:.2f}min)"
                )

                height, width = frame.shape[:2]
                text_position = (10, height - 10)
                font = cv2.FONT_HERSHEY_SIMPLEX
                font_scale = 1.2
                color = (255, 255, 255)  # White
                thickness = 2

                cv2.putText(
                    frame,
                    time_text,
                    text_position,
                    font,
                    font_scale,
                    color,
                    thickness,
                    cv2.LINE_AA,
                )

                # Convert BGR to RGB for PIL
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                try:
                    pil_image = Image.fromarray(frame_rgb)
                    frames_for_gif.append(pil_image)
                except MemoryError:
                    self._log_message(
                        f"⚠️ GIF export ran out of memory at frame {idx + 1}/{n_frames}. "
                        f"Use 'Export as Video (MP4)' instead."
                    )
                    return
                progress_dlg.setValue(idx + 1)

            progress_dlg.setValue(n_frames)

            if cancelled:
                self._log_message("⚠️ GIF export cancelled.")
                return

            # Save as GIF
            self._log_message("   Saving GIF file...")
            frames_for_gif[0].save(
                file_path,
                save_all=True,
                append_images=frames_for_gif[1:],
                duration=frame_duration,
                loop=0,
                optimize=False,
            )

            self._log_message(f"✅ GIF exported successfully: {file_path}")
            self._log_message(
                f"   {n_frames} frames at {export_fps} FPS ({n_frames/export_fps:.1f}s duration)"
            )

        except ImportError:
            self._log_message(
                "❌ PIL (Pillow) is required for GIF export. Install with: pip install Pillow"
            )
        except Exception as e:
            self._log_message(f"❌ GIF export failed: {e}")
            import traceback

            traceback.print_exc()

    def _export_gif_with_plots(self):
        """Export animated GIF with synchronized analysis plots (same style as Analysis tab)."""
        import cv2
        import numpy as np
        from qtpy.QtWidgets import QFileDialog, QProgressDialog
        from qtpy.QtCore import Qt
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        from ._plot import PlotGenerator, create_plot_config, create_hysteresis_kwargs
        from PIL import Image

        # Check prerequisites
        if not hasattr(self, "viewer_n_frames") or self.viewer_n_frames == 0:
            self._log_message("❌ No frames loaded. Load data into Frame Viewer first.")
            return

        if not hasattr(self, "merged_results") or not self.merged_results:
            self._log_message("❌ No analysis data. Run Main Analysis first.")
            return

        # Get frame range from export controls
        start_frame = self.export_start_frame.value()
        end_frame = self.export_end_frame.value()

        if start_frame >= end_frame:
            self._log_message("❌ Start frame must be less than end frame")
            return

        # Ask for save location
        default_name = f"gif_with_plots_{start_frame}-{end_frame}.gif"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export GIF with Plots",
            default_name,
            "GIF Animation (*.gif);;All Files (*)",
        )

        if not file_path:
            return

        export_fps = self.export_fps_spin.value()
        frame_duration = int(1000 / export_fps)  # Duration in milliseconds
        n_frames = end_frame - start_frame

        self._log_message(f"🎞️ Exporting {n_frames} frames with plots as GIF at {export_fps} FPS...")

        # Setup progress dialog
        progress = QProgressDialog("Exporting GIF with plots...", "Cancel", 0, n_frames, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setWindowTitle("Export Progress")
        progress.show()

        # Process events to show dialog
        from qtpy.QtWidgets import QApplication
        QApplication.processEvents()

        plot_type, data_dict = self._get_sync_plot_data()

        if not data_dict:
            self._log_message(f"❌ No {self.sync_plot_type.currentText()} data available")
            return

        roi_colors = getattr(self, "roi_colors", {})
        n_rois = len(data_dict)

        # Get first frame to determine dimensions
        first_frame = self._get_viewer_frame(start_frame)
        frame_height, frame_width = first_frame.shape[:2]

        # Create plot figure dimensions matching Analysis tab style
        plot_dpi = 100
        plot_width = frame_width
        height_per_roi = 60  # pixels per ROI
        plot_height = max(200, height_per_roi * n_rois)
        fig_width = plot_width / plot_dpi
        fig_height = plot_height / plot_dpi

        # Total output dimensions: frame on top, plots below
        output_width = frame_width
        output_height = frame_height + plot_height

        # Create plot configuration from widget settings (same as Analysis tab)
        base_plot_config = create_plot_config(self)
        base_plot_config["fig_width"] = fig_width
        base_plot_config["height_per_roi"] = fig_height / n_rois if n_rois > 0 else 0.6
        base_plot_config["dpi"] = plot_dpi

        # Get extra kwargs for specific plot types
        if plot_type == "Raw Intensity Changes":
            hysteresis_kwargs = create_hysteresis_kwargs(widget_instance=self)
            hysteresis_kwargs.pop("merged_results", None)
        elif plot_type == "Sleep Quality":
            sleep_metric = "sleep_minutes"
            if hasattr(self, "sleep_quality_metric_combo"):
                metric_text = self.sleep_quality_metric_combo.currentText()
                if "Transitions" in metric_text:
                    sleep_metric = "transitions"
                elif "Bout" in metric_text:
                    sleep_metric = "bout_length"
            hysteresis_kwargs = {"sleep_metric": sleep_metric}
        else:
            hysteresis_kwargs = {}

        frames_for_gif = []

        try:
            for i, frame_idx in enumerate(range(start_frame, end_frame)):
                if progress.wasCanceled():
                    self._log_message("⚠️ Export cancelled by user")
                    return

                progress.setValue(i)

                # Log progress every 10 frames
                if i % 10 == 0:
                    self._log_message(f"   Processing frame {i + 1}/{n_frames}...")
                    QApplication.processEvents()

                # Get frame
                frame = self._get_viewer_frame(frame_idx)

                # Add time text to frame
                time_seconds = frame_idx * self.viewer_frame_interval
                time_minutes = time_seconds / 60.0
                time_text = f"t = {time_seconds:.1f}s ({time_minutes:.2f}min)"

                frame_bgr = frame.copy()
                if len(frame_bgr.shape) == 2:
                    frame_bgr = cv2.cvtColor(frame_bgr, cv2.COLOR_GRAY2BGR)
                elif len(frame_bgr.shape) == 3:
                    if frame_bgr.shape[2] == 4:
                        frame_bgr = cv2.cvtColor(frame_bgr, cv2.COLOR_RGBA2BGR)
                    elif frame_bgr.shape[2] == 1:
                        frame_bgr = cv2.cvtColor(frame_bgr[:, :, 0], cv2.COLOR_GRAY2BGR)
                    # else: assume already BGR with 3 channels

                cv2.putText(
                    frame_bgr, time_text, (10, frame_height - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2, cv2.LINE_AA
                )

                # Draw ROI overlay (circles + numbers) if enabled
                if hasattr(self, "viewer_show_roi_overlay") and self.viewer_show_roi_overlay.isChecked():
                    self._draw_roi_overlay(frame_bgr)

                # Create plot using PlotGenerator (same as Analysis tab)
                plot_fig = Figure(figsize=(fig_width, fig_height), dpi=plot_dpi)
                plot_generator = PlotGenerator(plot_fig)

                # Generate plot with same settings as Analysis tab
                plot_generator.generate_plot(
                    plot_type, data_dict, roi_colors, base_plot_config, **hysteresis_kwargs
                )

                # Add red time marker to all subplots
                for ax in plot_fig.get_axes():
                    ax.axvline(
                        x=time_minutes, color="red", linewidth=2,
                        linestyle="--", alpha=0.9, zorder=100
                    )

                # Render plot to image (suppress tight_layout warning)
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    plot_fig.tight_layout()
                plot_canvas = FigureCanvasAgg(plot_fig)
                plot_canvas.draw()

                # Get image from canvas
                buf = plot_canvas.buffer_rgba()
                plot_img = np.asarray(buf)
                plot_img_bgr = cv2.cvtColor(plot_img, cv2.COLOR_RGBA2BGR)

                # Resize plot to match frame width if needed
                if plot_img_bgr.shape[1] != frame_width or plot_img_bgr.shape[0] != plot_height:
                    plot_img_bgr = cv2.resize(plot_img_bgr, (frame_width, plot_height))

                # Close figure to free memory
                import matplotlib.pyplot as plt
                plt.close(plot_fig)

                # Combine frame and plot vertically
                combined = np.vstack([frame_bgr, plot_img_bgr])

                # Convert BGR to RGB for PIL
                combined_rgb = cv2.cvtColor(combined, cv2.COLOR_BGR2RGB)
                pil_image = Image.fromarray(combined_rgb)
                frames_for_gif.append(pil_image)

            progress.setValue(n_frames)

            # Save as GIF
            self._log_message("   Saving GIF file...")
            frames_for_gif[0].save(
                file_path,
                save_all=True,
                append_images=frames_for_gif[1:],
                duration=frame_duration,
                loop=0,
                optimize=False,
            )

            self._log_message(f"✅ GIF with plots exported: {file_path}")
            self._log_message(f"   Dimensions: {output_width}x{output_height}, {n_frames} frames at {export_fps} FPS")

        except ImportError as e:
            if "PIL" in str(e) or "Pillow" in str(e):
                self._log_message("❌ PIL (Pillow) is required for GIF export. Install with: pip install Pillow")
            else:
                self._log_message(f"❌ Import error: {e}")
        except Exception as e:
            self._log_message(f"❌ GIF export failed: {e}")
            import traceback
            traceback.print_exc()

    # ===================================================================
    # UTILITY METHODS
    # ===================================================================

    def _get_recording_start_str(self) -> str:
        """Return formatted recording start datetime for plot titles, or empty string."""
        dt = getattr(self, "recording_start_datetime", None)
        if dt is None:
            return ""
        return f"  [{dt.strftime('%d.%m.%Y  %H:%M')}]"

    def _log_message(self, message: str):
        """Add message to analysis log with proper Qt handling."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}"

        # Use QTimer.singleShot to ensure this runs in the main thread
        from qtpy.QtCore import QTimer

        def append_to_log():
            try:
                self.log_text.append(formatted_message)
                # Auto-scroll to bottom - use moveCursor instead of setTextCursor
                cursor = self.log_text.textCursor()
                cursor.movePosition(cursor.End)
                # Don't connect the cursor, just move to end
                self.log_text.moveCursor(cursor.End)
                self.log_text.ensureCursorVisible()
            except Exception as e:
                print(f"Logging error: {e}")

        # Execute in main thread
        QTimer.singleShot(0, append_to_log)

    def cleanup_resources(self):
        """Clean up resources when widget is destroyed."""
        if self.current_worker:
            self._cancel_requested = True

        if hasattr(self, "performance_timer"):
            self.performance_timer.stop()

    def __del__(self):
        """Destructor to ensure proper cleanup."""
        self.cleanup_resources()


# ===================================================================
# HELPER FUNCTIONS (MOVED FROM OUTSIDE CLASS)
# ===================================================================


def prepare_analysis_parameters(widget, method):
    """
    Prepare parameters for analysis based on widget state and method.
    """
    base_params = {
        "enable_matlab_norm": True,
        "enable_detrending": getattr(widget, "enable_detrending", None),
        "frame_interval": getattr(widget, "frame_interval", None),
    }

    # Extract values safely
    if hasattr(base_params["enable_detrending"], "isChecked"):
        base_params["enable_detrending"] = base_params["enable_detrending"].isChecked()
    else:
        base_params["enable_detrending"] = True

    if hasattr(base_params["frame_interval"], "value"):
        base_params["frame_interval"] = base_params["frame_interval"].value()
    else:
        base_params["frame_interval"] = 5.0

    # Method-specific parameters
    if method == "baseline":
        try:
            baseline_duration_minutes = widget.baseline_duration_minutes.value()
            multiplier = widget.threshold_multiplier.value()
            frame_interval = base_params["frame_interval"]

            base_params.update(
                {
                    "threshold_block_count": int(
                        (baseline_duration_minutes * 60) / frame_interval
                    ),
                    "multiplier": multiplier,
                    "enable_jump_correction": getattr(
                        widget, "enable_jump_correction", None
                    ),
                }
            )

            if hasattr(base_params["enable_jump_correction"], "isChecked"):
                base_params["enable_jump_correction"] = base_params[
                    "enable_jump_correction"
                ].isChecked()
            else:
                base_params["enable_jump_correction"] = True

        except Exception as e:
            print(f"Warning: Could not extract baseline parameters: {e}")

    elif method == "adaptive":
        try:
            duration_minutes = widget.adaptive_duration_minutes.value()
            frame_interval = base_params["frame_interval"]

            base_params.update(
                {
                    "analysis_duration_frames": int(
                        (duration_minutes * 60) / frame_interval
                    ),
                    "base_multiplier": widget.adaptive_base_multiplier.value(),
                }
            )
        except Exception as e:
            print(f"Warning: Could not extract adaptive parameters: {e}")

    elif method == "calibration":
        try:
            calibration_file = widget.calibration_file_path.property("full_path")

            base_params.update(
                {
                    "calibration_file_path": calibration_file,
                    "masks": getattr(widget, "masks", []),
                    "calibration_multiplier": widget.calibration_multiplier.value(),
                }
            )
        except Exception as e:
            print(f"Warning: Could not extract calibration parameters: {e}")

    return base_params


# Provide the dock widget to Napari
def napari_provide_dock_widget(viewer):
    return HDF5AnalysisWidget(viewer)
